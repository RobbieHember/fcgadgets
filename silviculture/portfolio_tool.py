'''
Portfolio Tool

'''

#%% Import modules

import numpy as np
import pandas as pd
import shutil
import openpyxl
from fcgadgets.macgyver import utilities_general as gu
from fcgadgets.macgyver import utilities_gis as gis
from fcgadgets.macgyver import utilities_inventory as invu
from fcgadgets.cbrunner import cbrun_utilities as cbu

meta={}
meta['Paths']={}
meta['Paths']['Project']=r'C:\Users\rhember\Documents\Data\FCI_Projects\FCI_Portfolio'
meta['Paths']['Model Code']=r'C:\Users\rhember\Documents\Code_Python\fcgadgets\cbrunner'

#%% Import portfolio inputs from spreadsheet

def ImportPortfolio(meta):

    meta['N Scenario']=2
    
    meta['Portfolio']={}
    
    #--------------------------------------------------------------------------
    # Import activity types
    #--------------------------------------------------------------------------
    
    d=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\Portfolio Inputs.xlsx',sheet_name='Activity Types',skiprows=0).to_dict('split')
    meta['Portfolio']['Activities']={}
    
    # Index to columns to keep
    a=np.array(d['data'][1])
    ind=np.where( (a!='Activity description') & (a!='nan') )[0]
    for i in range(len(d['data'])):
        tmp=np.array([])
        for j in range(len(ind)):
            tmp=np.append(tmp,d['data'][i][ind[j]])
        meta['Portfolio']['Activities'][d['data'][i][0]]=tmp
    
    meta['Portfolio']['Activities']['Activity ID']=meta['Portfolio']['Activities']['Activity ID'].astype(int)
    
    #--------------------------------------------------------------------------
    # Import implementation levels
    #--------------------------------------------------------------------------
    
    d=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\Portfolio Inputs.xlsx',sheet_name='Annual Implementation Level',skiprows=1).to_dict('split')
    meta['Portfolio']['AIL']={}
    
    meta['Portfolio']['AIL']['Year']=np.arange(d['data'][1][0],d['data'][-1][0]+1,1,dtype=int)
    
    meta['Portfolio']['AIL']['BAU']=np.zeros((meta['Portfolio']['AIL']['Year'].size,12))
    meta['Portfolio']['AIL']['CAP']=np.zeros((meta['Portfolio']['AIL']['Year'].size,12))
    for i in range(0,meta['Portfolio']['AIL']['Year'].size):
        cnt=0
        for j in range(1,12):
            meta['Portfolio']['AIL']['BAU'][i,cnt]=d['data'][i+1][j]
            cnt=cnt+1
        cnt=0
        for j in range(12,22):    
            meta['Portfolio']['AIL']['CAP'][i,cnt]=d['data'][i+1][j]
            cnt=cnt+1
    
    meta['Portfolio']['AIL']['BAU']=np.nan_to_num(meta['Portfolio']['AIL']['BAU'])
    meta['Portfolio']['AIL']['CAP']=np.nan_to_num(meta['Portfolio']['AIL']['CAP'])
    meta['Portfolio']['AIL']['BAU']=meta['Portfolio']['AIL']['BAU'].astype(int)
    meta['Portfolio']['AIL']['CAP']=meta['Portfolio']['AIL']['CAP'].astype(int)
    
    indBAU=np.where(np.sum(meta['Portfolio']['AIL']['BAU'],axis=0)>0)[0]
    indCAP=np.where(np.sum(meta['Portfolio']['AIL']['CAP'],axis=0)>0)[0]
    aid=np.array(d['data'][0][1:25])
    meta['Portfolio']['AIL']['Activity ID']=np.append(aid[indBAU],aid[indCAP+11])
    meta['Portfolio']['AIL']['Activity ID']=meta['Portfolio']['AIL']['Activity ID'].astype(int)
    
    #--------------------------------------------------------------------------
    # Index to activity ID and Year for each unique stand
    #--------------------------------------------------------------------------
    
    # Number of unique activities
    meta['Portfolio']['N AT BAU']=np.sum(np.sum(meta['Portfolio']['AIL']['BAU'],axis=0)>0)
    meta['Portfolio']['N AT CAP']=np.sum(np.sum(meta['Portfolio']['AIL']['CAP'],axis=0)>0)
    meta['Portfolio']['N AT']=meta['Portfolio']['N AT BAU']+meta['Portfolio']['N AT CAP']
    
    # Number of years with activities
    meta['Portfolio']['N Year']=meta['Portfolio']['AIL']['Year'].size
    
    # Total number of stands
    meta['N Stand']=meta['Portfolio']['N AT']*meta['Portfolio']['N Year']
    
    meta['Portfolio']['Indices']={}
    meta['Portfolio']['Indices']['Activity']=np.zeros(meta['N Stand'],dtype=int)
    meta['Portfolio']['Indices']['Activity Unique']=np.zeros(meta['N Stand'],dtype=int)
    meta['Portfolio']['Indices']['Year']=np.zeros(meta['N Stand'],dtype=int)
    cnt=0
    for iA in range(meta['Portfolio']['N AT']):
        for iY in range(meta['Portfolio']['N Year']):
            meta['Portfolio']['Indices']['Activity'][cnt]=meta['Portfolio']['AIL']['Activity ID'][iA]-1
            meta['Portfolio']['Indices']['Activity Unique'][cnt]=iA
            meta['Portfolio']['Indices']['Year'][cnt]=iY
            cnt=cnt+1
    
    meta['Portfolio']['Indices']['BAU Activity']=indBAU
    meta['Portfolio']['Indices']['CAP Activity']=np.arange(indBAU[-1]+1,meta['Portfolio']['N AT'])
    
    return meta

#%% Prepare inputs for BatchTIPSY.exe

def PrepareInputsForBatchTIPSY(meta):
    
    # Create a function that will return the column corresponding to a variable name
    fin=meta['Paths']['Model Code'] + '\\Parameters\\GrowthCurvesTIPSY_Parameters_Template.xlsx'
    df_frmt=pd.read_excel(fin,sheet_name='Sheet1')
    gy_labels=df_frmt.loc[5,:].values

    def GetColumn(lab):
        ind=np.where(gy_labels==lab)[0]
        return int(ind+1)

    def isnan(x):        
        if (x.dtype!='float') | (x.dtype!='float'):
            y=False
        else:
            y=np.isnan(x)
        return y

    # Generate new spreadsheet from template
    fout=meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx'
    shutil.copy(fin,fout)
    
    # Load file, delete everything - start clean!
    xfile=openpyxl.load_workbook(fout)
    sheet=xfile['Sheet1']
    N_headers=7

    # Initiate counter
    cnt=1

    # meta['N Stand Full']=5

    # Define growth curve 1 (pre-project curves)
    for iStand in range(meta['N Stand Full']): 
        
        # Index to activity
        iActivity=meta['Portfolio']['Indices']['Activity'][iStand]
        
        for iScn in range(meta['N Scenario']):
        
            sheet.cell(row=cnt+N_headers,column=1).value=cnt  
            sheet.cell(row=cnt+N_headers,column=2).value=iStand+1 # Stand ID
            sheet.cell(row=cnt+N_headers,column=3).value=iScn+1 # Scenario ID
            sheet.cell(row=cnt+N_headers,column=4).value=1 # Growth curve
            
            vnam='regeneration_method'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value='N' # Regeneration type (N, C, P)                             
            
            vnam='s1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['Previous_Spc1_CD'][iActivity] # Species 1 code            
            
            vnam='p1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['Previous_Spc1_PCT'][iActivity] # Species 1 percent           
            
            vnam='i1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['Site index (m)'][iActivity] # Species 1 site index            
            
            if isnan(meta['Portfolio']['Activities']['Previous_Spc2_CD'][iActivity])==False:
                vnam='s2'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['Previous_Spc2_CD'][iActivity] # Species 2 code             
            
                vnam='p2'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['Previous_Spc2_PCT'][iActivity] # Species 2 percent            
            
            if isnan(meta['Portfolio']['Activities']['Previous_Spc3_CD'][iActivity])==False:
                vnam='s3'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['Previous_Spc3_CD'][iActivity] # Species 3 code             
                
                vnam='p3'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['Previous_Spc3_PCT'][iActivity] # Species 3 percent
            
            if isnan(meta['Portfolio']['Activities']['Previous_Spc4_CD'][iActivity])==False:
                vnam='s4'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['Previous_Spc4_CD'][iActivity] # Species 3 code             
                
                vnam='p4'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['Previous_Spc4_PCT'][iActivity] # Species 3 percent
            
            vnam='init_density'; vc=GetColumn(vnam)            
            sheet.cell(row=cnt+N_headers,column=vc).value=int(meta['Portfolio']['Activities']['Previous initial density (SPH)'][iActivity]) # Planting density    
            
            vnam='regen_delay'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=0 # Regeneration delay            
            
            vnam='oaf1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['OAF1'][iActivity] # OAF1                        
            
            vnam='oaf2'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['OAF2'][iActivity] # OAF2
            
            vnam='bec_zone'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['BGC Zone'][iActivity] # BEC zone
            
            vnam='FIZ'; vc=GetColumn(vnam)
            if (meta['Portfolio']['Activities']['BGC Zone'][iActivity]=='CWH') | (meta['Portfolio']['Activities']['BGC Zone'][iActivity]=='CDF'):
                fiz='C'
            else:
                fiz='I'
            sheet.cell(row=cnt+N_headers,column=vc).value=fiz # FIZ
        
            # Update counter
            cnt=cnt+1 

    # Define growth curve 2 (baseline or project scenario)
    for iStand in range(meta['N Stand Full']):
        
        # Index to activity
        iActivity=meta['Portfolio']['Indices']['Activity'][iStand]
        
        for iScn in range(meta['N Scenario']):
    
            sheet.cell(row=cnt+N_headers,column=1).value=cnt
            sheet.cell(row=cnt+N_headers,column=2).value=iStand+1 # Stand ID
            sheet.cell(row=cnt+N_headers,column=3).value=iScn+1 # Scenario ID
            sheet.cell(row=cnt+N_headers,column=4).value=2 # Growth curve
        
            if meta['Portfolio']['Activities']['New regeneration method'][iActivity]=='NAT':
                rm='N'
            else:
                rm='P'        
            vnam='regeneration_method'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=rm # Regeneration type (N, C, P)                             
            
            vnam='s1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['New_Spc1_CD'][iActivity] # Species 1 code            
            
            vnam='p1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['New_Spc1_PCT'][iActivity] # Species 1 percent           
            
            vnam='i1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['Site index (m)'][iActivity] # Species 1 site index            
            
            if isnan(meta['Portfolio']['Activities']['New_Spc2_CD'][iActivity])==False:
                vnam='s2'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['New_Spc2_CD'][iActivity] # Species 2 code             
            
                vnam='p2'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['New_Spc2_PCT'][iActivity] # Species 2 percent            
            
            if isnan(meta['Portfolio']['Activities']['New_Spc3_CD'][iActivity])==False:
                vnam='s3'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['New_Spc3_CD'][iActivity] # Species 3 code             
                
                vnam='p3'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['New_Spc3_PCT'][iActivity] # Species 3 percent
            
            if isnan(meta['Portfolio']['Activities']['New_Spc4_CD'][iActivity])==False:
                vnam='s4'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['New_Spc4_CD'][iActivity] # Species 3 code             
                
                vnam='p4'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['New_Spc4_PCT'][iActivity] # Species 3 percent
            
            vnam='init_density'; vc=GetColumn(vnam)            
            sheet.cell(row=cnt+N_headers,column=vc).value=int(meta['Portfolio']['Activities']['New initial density (SPH)'][iActivity]) # Planting density    
            
            vnam='regen_delay'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['New regeneration delay (years)'][iActivity] # Regeneration delay            
            
            vnam='oaf1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['OAF1'][iActivity] # OAF1                        
            
            vnam='oaf2'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['OAF2'][iActivity] # OAF2
            
            vnam='bec_zone'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Portfolio']['Activities']['BGC Zone'][iActivity] # BEC zone
            
            vnam='FIZ'; vc=GetColumn(vnam)
            if (meta['Portfolio']['Activities']['BGC Zone'][iActivity]=='CWH') | (meta['Portfolio']['Activities']['BGC Zone'][iActivity]=='CDF'):
                fiz='C'
            else:
                fiz='I'
            sheet.cell(row=cnt+N_headers,column=vc).value=fiz # FIZ
    
        # Update counter
        cnt=cnt+1 

    # Save to spreadsheet
    xfile.save(fout)

    # Populate BatchTIPSY.exe input variable (.dat) file 
    cbu.Write_BatchTIPSY_Input_File(meta)
    
    return

#%% Prepare inventory

def PrepareInventory(meta):
    
    for iScn in range(meta['N Scenario']):
    
        # Loop through batches, saving inventory to file
        for iBat in range(meta['N Batch']):
      
            inv={}
    
            # Index to batch
            indBat=cbu.IndexToBatch(meta,iBat)    
            N_StandsInBatch=len(indBat)
    
            # Initialize inventory variables
            inv['Lat']=np.zeros((1,N_StandsInBatch))
            inv['Lon']=np.zeros((1,N_StandsInBatch))
            inv['X']=inv['Lat']
            inv['Y']=inv['Lon']
        
            # BEC zone
            inv['ID_BECZ']=np.zeros((1,N_StandsInBatch),dtype=np.int)
            for i in range(N_StandsInBatch):
                cd=meta['Portfolio']['Activities']['BGC Zone'][meta['Portfolio']['Indices']['Activity'][indBat[i]]]
                inv['ID_BECZ'][0,indBat[i]]=meta['LUT']['VRI']['BEC_ZONE_CODE'][cd]
    
            # Timber harvesting landbase (1=yes, 0=no)
            inv['THLB']=1*np.ones((meta['Year'].size,N_StandsInBatch))
        
            # Temperature will be updated automatically
            inv['MAT']=4*np.ones((1,N_StandsInBatch))
            
            if meta['Biomass Module']=='Sawtooth':
                inv['Srs1_ID']=meta['LUT']['Spc'][meta['Scenario'][iScn]['SRS1_CD']]*np.ones((1,N_StandsInBatch),dtype=np.int)
            else:
                inv['Srs1_ID']=9999*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Srs1_Pct']=100*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Srs2_ID']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Srs2_Pct']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Srs3_ID']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Srs3_Pct']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc1_ID']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc1_Pct']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc2_ID']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc2_Pct']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc3_ID']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc3_Pct']=0*np.ones((1,N_StandsInBatch),dtype=np.int)

            # Save
            gu.opickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + cbu.FixFileNum(iBat) + '.pkl',inv)
    
    return

#%% Prepare event chronology

def PrepareEventChronology(meta):
    
    #--------------------------------------------------------------------------
    # Import custom harvest inputs
    #--------------------------------------------------------------------------
    
    meta['Harvest Custom']={}
    for i in range(meta['Portfolio']['N AT']):
        
        iA=meta['Portfolio']['Indices']['Activity'][i]
        
        d={}
        d['BiomassMerch_Affected']=meta['Portfolio']['Activities']['% of overstory that is felled'][iA]
        d['BiomassMerch_Removed']=meta['Portfolio']['Activities']['% removed (merch biomass)'][iA]
        d['BiomassMerch_Burned']=meta['Portfolio']['Activities']['% burned on site (merch biomass)'][iA]
        d['BiomassMerch_LeftOnSite']=meta['Portfolio']['Activities']['% left to decay on site (merch biomass)'][iA]
        d['BiomassNonMerch_Affected']=meta['Portfolio']['Activities']['% of overstory that is felled'][iA]
        d['BiomassNonMerch_Removed']=meta['Portfolio']['Activities']['% removed (non merch biomass)'][iA]
        d['BiomassNonMerch_Burned']=meta['Portfolio']['Activities']['% burned on site (non merch biomass)'][iA]
        d['BiomassNonMerch_LeftOnSite']=meta['Portfolio']['Activities']['% left to decay on site (non merch biomass)'][iA]
        d['Snags_Affected']=meta['Portfolio']['Activities']['% of overstory that is felled'][iA]
        d['Snags_Removed']=meta['Portfolio']['Activities']['% removed (snags)'][iA]
        d['Snags_Burned']=meta['Portfolio']['Activities']['% burned on site (snags)'][iA]
        d['Snags_LeftOnSite']=meta['Portfolio']['Activities']['% left to decay on site (snags)'][iA]
        
        d['RemovedMerchToPulp']=5
        d['RemovedMerchToFuel']=5
        d['RemovedMerchToLumber']=5
        d['RemovedMerchToPlywood']=5
        d['RemovedMerchToOSB']=5
        d['RemovedMerchToMDF']=5
        d['RemovedMerchToCants']=5
        d['RemovedMerchToFirewood']=5
        d['RemovedNonMerchToFuel']=5
        d['RemovedNonMerchToLumber']=5
        d['RemovedNonMerchToPlywood']=5
        d['RemovedNonMerchToOSB']=5
        d['RemovedNonMerchToMDF']=5
        d['RemovedNonMerchToPulp']=5
        d['RemovedNonMerchToCants']=5
        d['RemovedNonMerchToFirewood']=5
        d['RemovedSnagStemToFuel']=5
        d['RemovedSnagStemToLumber']=5
        d['RemovedSnagStemToPlywood']=5
        d['RemovedSnagStemToOSB']=5
        d['RemovedSnagStemToMDF']=5
        d['RemovedSnagStemToPulp']=5
        d['RemovedSnagStemToCants']=5
        d['RemovedSnagStemToFirewood']=5
    
        # Populate custom harvest event
        meta['Harvest Custom'][int(i+1)]=d
    
    #--------------------------------------------------------------------------
    # Generate event chronology
    #--------------------------------------------------------------------------    
    
    for iScn in range(meta['N Scenario']):
        
        for iEns in range(meta['N Ensemble']):        
            
            for iBat in range(meta['N Batch']):
    
                # Index to batch
                indBat=cbu.IndexToBatch(meta,iBat)
                N_StandsInBatch=len(indBat)
    
                tv=np.arange(meta['Year Start'],meta['Year End']+1,1)
    
                # Initialize dictionary
                ec={}
                ec['ID_Type']=np.zeros((meta['Year'].size,indBat.size,meta['Max Events Per Year']),dtype='int16')
                ec['MortalityFactor']=np.zeros((meta['Year'].size,indBat.size,meta['Max Events Per Year']),dtype='int16')
                ec['GrowthFactor']=np.zeros((meta['Year'].size,indBat.size,meta['Max Events Per Year']),dtype='int16')
                ec['ID_GrowthCurve']=np.zeros((meta['Year'].size,indBat.size,meta['Max Events Per Year']),dtype='int16')
            
                for iS in range(N_StandsInBatch): 
                    
                    # Index to portfolio tables
                    iY=meta['Portfolio']['Indices']['Year'][indBat[iS]]
                    iA=meta['Portfolio']['Indices']['Activity'][indBat[iS]]
                    
                    #----------------------------------------------------------
                    # Add spinup events
                    #----------------------------------------------------------
                    
                    ivl_spin=meta['Spinup Disturbance Return Inverval']                
                    YearRef=meta['Portfolio']['AIL']['Year'][iY]
                    AgeRef=meta['Portfolio']['Activities']['Mean stand age at time of inciting disturbance (years)'][iA]
                    if AgeRef>=0:
                        Year=np.arange(YearRef-AgeRef-100*ivl_spin,YearRef-AgeRef+ivl_spin,ivl_spin)        
                    else:
                        Year1=meta['Year Start']+ivl_spin
                        Year2=meta['Spinup Year End']
                        Year=np.arange(Year1,Year2+1,meta['Spinup Disturbance Return Inverval'])
                    
                    for iYr in range(Year.size):
                        iT=np.where(tv==Year[iYr])[0]
                        ec['ID_Type'][iT,iS,0]=meta['LUT']['Dist'][meta['Spinup Disturbance Type']]
                        ec['MortalityFactor'][iT,iS,0]=100
                        ec['GrowthFactor'][iT,iS,0]=0
                        ec['ID_GrowthCurve'][iT,iS,0]=meta['Spinup Growth Curve ID']
                    
                    #----------------------------------------------------------
                    # Add prescribed modern events
                    #----------------------------------------------------------
                    
                    if iScn>0:
                        YearRef=meta['Portfolio']['AIL']['Year'][iY]
                        iT=np.where(tv==YearRef)[0]
                        ec['ID_Type'][iT,iS,0]=meta['LUT']['Dist']['Harvest Custom ' + str(iA+1)]
                        ec['MortalityFactor'][iT,iS,0]=100
                        ec['GrowthFactor'][iT,iS,0]=0
                        ec['ID_GrowthCurve'][iT,iS,0]=2
            
                #--------------------------------------------------------------
                # Save to file            
                #--------------------------------------------------------------
                
                gu.opickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + cbu.FixFileNum(iEns) + '_Bat' + cbu.FixFileNum(iBat) + '.pkl',ec)
    
    return

#%% Import results
    
def ImportResults(meta):
    
    # Import cbrunner outputs
    v1=cbu.LoadScenarioResults(meta,[0,1])
    v2,meta=cbu.CalculateGHGBalance(v1,meta)
    
    # Extract annual implementation level (ha/year)
    iBAU=np.where(np.sum(meta['Portfolio']['AIL']['BAU'],axis=0)>0)[0]
    iCAP=np.where(np.sum(meta['Portfolio']['AIL']['CAP'],axis=0)>0)[0]
    ail=np.column_stack([meta['Portfolio']['AIL']['BAU'][:,iBAU],meta['Portfolio']['AIL']['CAP'][:,iCAP]])


    # Initialize GHG balance for each activity type
    vAT=[None]*meta['N Scenario']
    for iScn in range(meta['N Scenario']):
        vAT[iScn]={}
        vAT[iScn]['Sum']={}
        vAT[iScn]['Per Ha']={}
        for k in v2[0].keys():
            if k=='Year':
                continue
            vAT[iScn]['Sum'][k]=np.zeros((v2[iScn]['Year'].size,meta['Portfolio']['N AT']))
            vAT[iScn]['Per Ha'][k]=np.zeros((v2[iScn]['Year'].size,meta['Portfolio']['N AT']))
            
            for iAT in range(meta['Portfolio']['N AT']):
                ind=np.where(meta['Portfolio']['Indices']['Activity Unique']==iAT)[0]
                y=v2[iScn][k][:,ind]
                y_x_area=v2[iScn][k][:,ind]
                for iYr in range(y.shape[1]):
                    y_x_area[:,iYr]=ail[iYr,iAT]*y[:,iYr]
                vAT[iScn]['Sum'][k][:,iAT]=np.sum(y_x_area,axis=1)
                vAT[iScn]['Per Ha'][k][:,iAT]=np.mean(y,axis=1)

    # Results by portfolio
    iBAU=meta['Portfolio']['Indices']['BAU Activity']
    iCAP=meta['Portfolio']['Indices']['CAP Activity']
    
    vBAU=[None]*meta['N Scenario']
    vCAP=[None]*meta['N Scenario']
    for iScn in range(meta['N Scenario']):
        vBAU[iScn]={}
        vBAU[iScn]['Sum']={}
        vBAU[iScn]['W Ave']={}
        vCAP[iScn]={}
        vCAP[iScn]['Sum']={}
        vCAP[iScn]['W Ave']={}
        for k in v2[0].keys():
            if k=='Year':
                continue
            vBAU[iScn]['Sum'][k]=np.sum(vAT[iScn]['Sum'][k][:,iBAU],axis=1)
            vBAU[iScn]['W Ave'][k]=0
            vCAP[iScn]['Sum'][k]=np.sum(vAT[iScn]['Sum'][k][:,iCAP],axis=1)
            vCAP[iScn]['W Ave'][k]=0

    return v1,v2,vAT,vBAU,vCAP,meta