'''
Portfolio Utilities

'''

#%% Import modules

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import shutil
import openpyxl
import gc as garc
import copy
import matplotlib as mpl
from matplotlib.patches import Rectangle
from fcgadgets.macgyver import utilities_general as gu
from fcgadgets.cbrunner import cbrun_utilities as cbu
from fcgadgets.hardhat import economics as econo
from fcgadgets.taz import aspatial_stat_models as asm

#%% Plotting parameters

gp=gu.SetGraphics('Manuscript')

#%% Import portfolio inputs from spreadsheet

def ImportPortfolio(meta):

    meta['Project']={}
    meta['Project']['N Scenario']=2
    meta['Project']['N Portfolio']=4

    #--------------------------------------------------------------------------
    # Import portfolio-level parameters
    #--------------------------------------------------------------------------

    d=gu.ReadExcel(meta['Paths']['Project'] + '\\Inputs\\ProjectConfig.xlsx','Portfolios')

    meta['Project']['Portfolio']={}
    meta['Project']['Portfolio']['Raw']=d

    #--------------------------------------------------------------------------
    # Import activity types
    #--------------------------------------------------------------------------

    d=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\ProjectConfig.xlsx',sheet_name='Activity Types',skiprows=0).to_dict('split')

    meta['Project']['Activities']={}

    # Index to columns to keep
    a=np.array(d['data'][1])
    ind=np.where( (a!='Activity description') & (a!='nan') )[0]
    for i in range(len(d['data'])):
        if d['data'][i][0][-1]!=':':
            tmp=np.array([])
            for j in range(len(ind)):
                tmp=np.append(tmp,d['data'][i][ind[j]])
            meta['Project']['Activities'][d['data'][i][0]]=tmp

    meta['Project']['Activities']['Activity ID']=meta['Project']['Activities']['Activity ID'].astype(int)

    #--------------------------------------------------------------------------
    # Import implementation levels
    #--------------------------------------------------------------------------

    d=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\ProjectConfig.xlsx',sheet_name='AIL',skiprows=1).to_dict('split')

    meta['Project']['AIL']={}

    meta['Project']['AIL']['Year']=np.arange(d['data'][1][0],d['data'][-1][0]+1,1,dtype=int)

    meta['Project']['AIL']['ID Portfolio']=np.zeros(meta['Project']['N Portfolio']*12,dtype=int)
    meta['Project']['AIL']['ID Portfolio'][12:]=1
    meta['Project']['AIL']['ID Portfolio'][24:]=2
    meta['Project']['AIL']['ID Portfolio'][36:]=3

    meta['Project']['AIL']['Area']=np.zeros((meta['Project']['AIL']['Year'].size,meta['Project']['N Portfolio']*12))
    for i in range(0,meta['Project']['AIL']['Year'].size):
        cnt=0
        for j in range(1,meta['Project']['N Portfolio']*12+1):
            meta['Project']['AIL']['Area'][i,cnt]=d['data'][i+1][j]
            cnt=cnt+1

    meta['Project']['AIL']['Area']=np.nan_to_num(meta['Project']['AIL']['Area'])

    meta['Project']['AIL']['Area']=meta['Project']['AIL']['Area'].astype(float)

    # Activity ID
    meta['Project']['AIL']['ID AT']=np.array(d['data'][0][1:meta['Project']['N Portfolio']*12+1],dtype=int)

    # Activity ID Unique
    meta['Project']['AIL']['ID AT Unique']=np.arange(1,meta['Project']['N Portfolio']*12+1,1).astype(int)

    #--------------------------------------------------------------------------
    # Number of activities and years
    #--------------------------------------------------------------------------

    # Index to rows with implementation
    ind=np.where(np.sum(meta['Project']['AIL']['Area'],axis=0)>0)[0]

    meta['Project']['AIL']['N AT']=ind.size

    # Number of years with activities
    meta['Project']['AIL']['N Years']=meta['Project']['AIL']['Year'].size

    return meta

#%% Prepare inputs for BatchTIPSY.exe

def PrepareInputsForBatchTIPSY(meta):

    # Create a function that will return the column corresponding to a variable name
    fin=meta['Paths']['Model Code'] + '\\Parameters\\GrowthCurvesTIPSY_Parameters_Template.xlsx'
    df_frmt=pd.read_excel(fin,sheet_name='Sheet1')
    gy_labels=df_frmt.loc[5,:].values

    def GetColumn(lab):
        ind=np.where(gy_labels==lab)[0]
        return int(ind+1)

    def isnan(x):
        if (x.dtype!='float') | (x.dtype!='float'):
            y=False
        else:
            y=np.isnan(x)
        return y

    # Generate new spreadsheet from template
    fout=meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx'
    shutil.copy(fin,fout)

    # Load file, delete everything - start clean!
    xfile=openpyxl.load_workbook(fout)
    sheet=xfile['Sheet1']
    N_headers=7

    # Overwrite existing data entries with empty cells
    # *** This is really important - failing to wipe it clean first will lead to
    # weird parameters ***
    for i in range(10000):
        for j in range(len(gy_labels)):
            sheet.cell(row=i+1+N_headers,column=j+1).value=''

    # Initiate counter
    cnt=1

    # Define growth curve 1 (pre-project curves)
    for iStand in range(meta['Project']['N Stand']):

        # Index to activity
        iAT=np.where(meta['Project']['Activities']['Activity ID']==meta['Project']['Portfolio']['ID AT'][iStand])[0][0]

        for iScn in range(meta['Project']['N Scenario']):

            sheet.cell(row=cnt+N_headers,column=1).value=cnt
            sheet.cell(row=cnt+N_headers,column=2).value=iStand+1 # Stand ID
            sheet.cell(row=cnt+N_headers,column=3).value=iScn+1 # Scenario ID
            sheet.cell(row=cnt+N_headers,column=4).value=1 # Growth curve

            vnam='regeneration_method'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value='N' # Regeneration type (N, C, P)

            vnam='s1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['Historical_Spc1_CD'][iAT] # Species 1 code

            vnam='p1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['Historical_Spc1_PCT'][iAT] # Species 1 percent

            vnam='i1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['Site index (m)'][iAT] # Species 1 site index

            if isnan(meta['Project']['Activities']['Historical_Spc2_CD'][iAT])==False:
                vnam='s2'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['Historical_Spc2_CD'][iAT] # Species 2 code

                vnam='p2'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['Historical_Spc2_PCT'][iAT] # Species 2 percent

            if isnan(meta['Project']['Activities']['Historical_Spc3_CD'][iAT])==False:
                vnam='s3'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['Historical_Spc3_CD'][iAT] # Species 3 code

                vnam='p3'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['Historical_Spc3_PCT'][iAT] # Species 3 percent

            if isnan(meta['Project']['Activities']['Historical_Spc4_CD'][iAT])==False:
                vnam='s4'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['Historical_Spc4_CD'][iAT] # Species 3 code

                vnam='p4'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['Historical_Spc4_PCT'][iAT] # Species 3 percent

            vnam='init_density'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=int(meta['Project']['Activities']['Historical initial density (SPH)'][iAT]) # Planting density

            vnam='regen_delay'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=0 # Regeneration delay

            vnam='oaf1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['OAF1'][iAT] # OAF1

            vnam='oaf2'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['OAF2'][iAT] # OAF2

            vnam='bec_zone'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['BGC Zone'][iAT] # BEC zone

            vnam='FIZ'; vc=GetColumn(vnam)
            if (meta['Project']['Activities']['BGC Zone'][iAT]=='CWH') | (meta['Project']['Activities']['BGC Zone'][iAT]=='CDF'):
                fiz='C'
            else:
                fiz='I'
            sheet.cell(row=cnt+N_headers,column=vc).value=fiz # FIZ

            # Update counter
            cnt=cnt+1

    # Define growth curve 2 (baseline or project scenario)
    for iStand in range(meta['Project']['N Stand']):

        # Index to activity
        iAT=np.where(meta['Project']['Activities']['Activity ID']==meta['Project']['Portfolio']['ID AT'][iStand])[0][0]

        for iScn in range(meta['Project']['N Scenario']):

            if iScn==0:
                scn_nam='Baseline'
            else:
                scn_nam='Project'

            sheet.cell(row=cnt+N_headers,column=1).value=cnt
            sheet.cell(row=cnt+N_headers,column=2).value=iStand+1 # Stand ID
            sheet.cell(row=cnt+N_headers,column=3).value=iScn+1 # Scenario ID
            sheet.cell(row=cnt+N_headers,column=4).value=2 # Growth curve

            vnam='regeneration_method'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities'][scn_nam + ' regeneration method'][iAT] # Regeneration type (N, C, P)

            vnam='s1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities'][scn_nam + '_Spc1_CD'][iAT] # Species 1 code

            vnam='p1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities'][scn_nam + '_Spc1_PCT'][iAT] # Species 1 percent

            vnam='i1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['Site index (m)'][iAT] # Species 1 site index

            if isnan(meta['Project']['Activities'][scn_nam + '_Spc2_CD'][iAT])==False:
                vnam='s2'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities'][scn_nam + '_Spc2_CD'][iAT] # Species 2 code

                vnam='p2'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities'][scn_nam + '_Spc2_PCT'][iAT] # Species 2 percent

            if isnan(meta['Project']['Activities'][scn_nam + '_Spc3_CD'][iAT])==False:
                vnam='s3'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities'][scn_nam + '_Spc3_CD'][iAT] # Species 3 code

                vnam='p3'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities'][scn_nam + '_Spc3_PCT'][iAT] # Species 3 percent

            if isnan(meta['Project']['Activities'][scn_nam + '_Spc4_CD'][iAT])==False:
                vnam='s4'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities'][scn_nam + '_Spc4_CD'][iAT] # Species 3 code

                vnam='p4'; vc=GetColumn(vnam)
                sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities'][scn_nam + '_Spc4_PCT'][iAT] # Species 3 percent

            vnam='init_density'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=int(meta['Project']['Activities'][scn_nam + ' initial density (SPH)'][iAT]) # Planting density

            vnam='regen_delay'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities'][scn_nam + ' regeneration delay (years)'][iAT] # Regeneration delay

            vnam='oaf1'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['OAF1'][iAT] # OAF1

            vnam='oaf2'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['OAF2'][iAT] # OAF2

            vnam='bec_zone'; vc=GetColumn(vnam)
            sheet.cell(row=cnt+N_headers,column=vc).value=meta['Project']['Activities']['BGC Zone'][iAT] # BEC zone

            vnam='FIZ'; vc=GetColumn(vnam)
            if (meta['Project']['Activities']['BGC Zone'][iAT]=='CWH') | (meta['Project']['Activities']['BGC Zone'][iAT]=='CDF'):
                fiz='C'
            else:
                fiz='I'
            sheet.cell(row=cnt+N_headers,column=vc).value=fiz # FIZ

            # Update counter
            cnt=cnt+1

    # Save to spreadsheet
    xfile.save(fout)

    # Populate BatchTIPSY.exe input variable (.dat) file
    cbu.Write_BatchTIPSY_Input_File(meta)

    return

#%% Prepare inventory

def PrepareInventory(meta):

    for iScn in range(meta['Project']['N Scenario']):

        # Loop through batches, saving inventory to file
        for iBat in range(meta['Project']['N Batch']):

            inv={}

            # Index to batch
            indBat=cbu.IndexToBatch(meta,iBat)
            N_StandsInBatch=len(indBat)

            # Initialize inventory variables
            inv['Lat']=np.zeros((1,N_StandsInBatch))
            inv['Lon']=np.zeros((1,N_StandsInBatch))
            inv['X']=inv['Lat']
            inv['Y']=inv['Lon']

            # BEC zone
            inv['ID_BECZ']=np.zeros((1,N_StandsInBatch),dtype=np.int)
            for iS in range(N_StandsInBatch):

                # Index to activity
                iAT=np.where(meta['Project']['Activities']['Activity ID']==meta['Project']['Portfolio']['ID AT'][indBat[iS]])[0][0]

                cd=meta['Project']['Activities']['BGC Zone'][iAT]
                inv['ID_BECZ'][0,iS]=meta['LUT']['VRI']['BEC_ZONE_CODE'][cd]

            # Timber harvesting landbase (1=yes, 0=no)
            inv['THLB']=1*np.ones((meta['Year'].size,N_StandsInBatch))

            # Temperature will be updated automatically
            inv['MAT']=4*np.ones((1,N_StandsInBatch))

            if meta['Project']['Biomass Module']=='Sawtooth':
                inv['Srs1_ID']=meta['LUT']['Spc'][meta['Scenario'][iScn]['SRS1_CD']]*np.ones((1,N_StandsInBatch),dtype=np.int)
            else:
                inv['Srs1_ID']=9999*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Srs1_Pct']=100*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Srs2_ID']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Srs2_Pct']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Srs3_ID']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Srs3_Pct']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc1_ID']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc1_Pct']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc2_ID']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc2_Pct']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc3_ID']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['Spc3_Pct']=0*np.ones((1,N_StandsInBatch),dtype=np.int)

            # Save
            gu.opickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + cbu.FixFileNum(iBat) + '.pkl',inv)

    return

#%% Prepare event chronology

def PrepareEventChronology(meta):

    #--------------------------------------------------------------------------
    # Simulate wildfire
    #--------------------------------------------------------------------------

    iScn=0
    iBat=0
    inv=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + cbu.FixFileNum(iBat) + '.pkl')

    if (meta['Scenario'][iScn]['Wildfire Status Pre-modern']=='On') | (meta['Scenario'][iScn]['Wildfire Status Modern']=='On') | (meta['Scenario'][iScn]['Wildfire Status Future']=='On'):
        asm.SimulateWildfireFromAAO(meta,inv)

    #--------------------------------------------------------------------------
    # Generate event chronology
    #--------------------------------------------------------------------------

    for iScn in range(meta['Project']['N Scenario']):

        for iEns in range(meta['Project']['N Ensemble']):

            # Import wildfire simulations from Taz
            if (meta['Scenario'][iScn]['Wildfire Status Pre-modern']=='On') | (meta['Scenario'][iScn]['Wildfire Status Modern']=='On') | (meta['Scenario'][iScn]['Wildfire Status Future']=='On'):

                wf_sim=gu.ipickle(meta['Paths']['Project'] + '\\Inputs\\Ensembles\\wf_sim_Scn' + cbu.FixFileNum(iScn) + '_Ens' + cbu.FixFileNum(iEns) + '.pkl')
                if 'idx' in wf_sim:
                    idx=wf_sim['idx']
                    tmp=wf_sim.copy()
                    for v in ['Occurrence','Mortality']:
                        wf_sim[v]=np.zeros((meta['Project']['N Time'],meta['Project']['N Stand']),dtype='int16')
                        wf_sim[v][idx[0],idx[1]]=tmp[v]
                    del tmp

            for iBat in range(meta['Project']['N Batch']):

                # Index to batch
                indBat=cbu.IndexToBatch(meta,iBat)

                N_StandsInBatch=len(indBat)

                tv=np.arange(meta['Project']['Year Start'],meta['Project']['Year End']+1,1)

                # Initialize dictionary
                ec={}
                ec['ID_Type']=np.zeros((meta['Year'].size,indBat.size,meta['Core']['Max Events Per Year']),dtype='int16')
                ec['MortalityFactor']=np.zeros((meta['Year'].size,indBat.size,meta['Core']['Max Events Per Year']),dtype='int16')
                ec['GrowthFactor']=np.zeros((meta['Year'].size,indBat.size,meta['Core']['Max Events Per Year']),dtype='int16')
                ec['ID_GrowthCurve']=np.zeros((meta['Year'].size,indBat.size,meta['Core']['Max Events Per Year']),dtype='int16')

                for iS in range(N_StandsInBatch):

                    # Index to portfolio year
                    iY=np.where(meta['Project']['AIL']['Year']==meta['Project']['Portfolio']['Year'][indBat[iS]])[0][0]

                    # Index to portfolio activity
                    iAT=np.where(meta['Project']['Activities']['Activity ID']==meta['Project']['Portfolio']['ID AT'][indBat[iS]])[0][0]

                    #----------------------------------------------------------
                    # Add spinup events
                    #----------------------------------------------------------

                    ivl_spin=meta['Project']['Activities']['Spinup Disturbance Return Inverval'][iAT]
                    type_spin=meta['Project']['Activities']['Spinup Disturbance Type'][iAT]
                    gcid_spin=meta['Project']['Activities']['Spinup Growth Curve ID'][iAT]

                    if meta['Project']['Activities']['Lag2 event type'][iAT]!='None':
                        YearRef=meta['Project']['AIL']['Year'][iY]-meta['Project']['Activities']['Years between Lag2 and Current event'][iAT]
                        AgeRef=meta['Project']['Activities']['Stand age at time of Lag2 event'][iAT]
                    else:
                        YearRef=meta['Project']['AIL']['Year'][iY]-meta['Project']['Activities']['Years between Lag1 and Current event'][iAT]
                        AgeRef=meta['Project']['Activities']['Stand age at time of Lag1 event'][iAT]

                    Year=np.arange(YearRef-AgeRef-100*ivl_spin,YearRef-AgeRef+ivl_spin,ivl_spin)

                    for iYr in range(Year.size):
                        iT=np.where(tv==Year[iYr])[0]
                        ec['ID_Type'][iT,iS,0]=meta['LUT']['Dist'][type_spin]
                        ec['MortalityFactor'][iT,iS,0]=100
                        ec['GrowthFactor'][iT,iS,0]=0
                        ec['ID_GrowthCurve'][iT,iS,0]=gcid_spin

                    #----------------------------------------------------------
                    # Add Lag2 event
                    #----------------------------------------------------------

                    if meta['Project']['Activities']['Lag2 event type'][iAT]!='None':
                        iT=np.where(tv==meta['Project']['AIL']['Year'][iY]-meta['Project']['Activities']['Years between Lag2 and Current event'][iAT])[0]
                        ec['ID_Type'][iT,iS,0]=meta['LUT']['Dist'][ meta['Project']['Activities']['Lag2 event type'][iAT] ]
                        ec['MortalityFactor'][iT,iS,0]=meta['Project']['Activities']['% of biomass affected by Lag2 event'][iAT]
                        ec['GrowthFactor'][iT,iS,0]=0
                        ec['ID_GrowthCurve'][iT,iS,0]=meta['Project']['Activities']['Growth curve ID following Lag2 event'][iAT]

                    #----------------------------------------------------------
                    # Add Lag1 event
                    #----------------------------------------------------------

                    yr=meta['Project']['AIL']['Year'][iY]-meta['Project']['Activities']['Years between Lag1 and Current event'][iAT]
                    iT=np.where(tv==yr)[0]
                    iOpenSpot=np.where(ec['ID_Type'][iT,iS,:]==0)[0]
                    if iOpenSpot.size>0:
                        iOpenSpot=iOpenSpot[0]
                        ec['ID_Type'][iT,iS,iOpenSpot]=meta['LUT']['Dist'][ meta['Project']['Activities']['Lag1 event type'][iAT] ]
                        ec['MortalityFactor'][iT,iS,iOpenSpot]=meta['Project']['Activities']['% of biomass affected by Lag1 event'][iAT]
                        ec['GrowthFactor'][iT,iS,iOpenSpot]=0
                        if (iScn==0):
                            ec['ID_GrowthCurve'][iT,iS,iOpenSpot]=meta['Project']['Activities']['Growth curve ID following Lag1 event (baseline)'][iAT]
                        else:
                            ec['ID_GrowthCurve'][iT,iS,iOpenSpot]=meta['Project']['Activities']['Growth curve ID following Lag1 event (project)'][iAT]

                    #----------------------------------------------------------
                    # Add current event
                    #----------------------------------------------------------

                    type=meta['Project']['Activities']['Current event type'][iAT]
                    if (iScn>0) & (type!='None'):
                        yr=meta['Project']['AIL']['Year'][iY]
                        iT=np.where(tv==yr)[0]
                        iOpenSpot=np.where(ec['ID_Type'][iT,iS,:]==0)[0]
                        if iOpenSpot.size>0:
                            iOpenSpot=iOpenSpot[0]
                            ec['ID_Type'][iT,iS,iOpenSpot]=meta['LUT']['Dist'][type]
                            ec['MortalityFactor'][iT,iS,iOpenSpot]=meta['Project']['Activities']['% of biomass affected by Current event'][iAT]
                            ec['GrowthFactor'][iT,iS,iOpenSpot]=meta['Project']['Activities']['Growth modifier for current event (%)'][iAT]
                            ec['ID_GrowthCurve'][iT,iS,iOpenSpot]=meta['Project']['Activities']['Growth curve ID following Current event (project)'][iAT]

                    #----------------------------------------------------------
                    # Post 1 event
                    #----------------------------------------------------------

                    type=meta['Project']['Activities']['Post1 event type'][iAT]
                    if (type!='None'):

                        if meta['Project']['Activities']['Post1 scenarios affected (All or Project Only)'][iAT]=='Project Only':

                            # Only add for project scenario
                            if (iScn>0):
                                yr=meta['Project']['AIL']['Year'][iY]+meta['Project']['Activities']['Years between Current and Post1 event'][iAT]
                                iT=np.where(tv==yr)[0]
                                if iT.size>0:
                                    iOpenSpot=np.where(ec['ID_Type'][iT,iS,:]==0)[0]
                                    if iOpenSpot.size>0:
                                        iOpenSpot=iOpenSpot[0]
                                        ec['ID_Type'][iT,iS,iOpenSpot]=meta['LUT']['Dist'][type]
                                        ec['MortalityFactor'][iT,iS,iOpenSpot]=meta['Project']['Activities']['% of biomass affected by Post1 event'][iAT]
                                        ec['GrowthFactor'][iT,iS,iOpenSpot]=0
                                        ec['ID_GrowthCurve'][iT,iS,iOpenSpot]=meta['Project']['Activities']['Growth curve ID following Post1 event (project)'][iAT]
                        else:

                            # Add for baseline and project scenario
                            yr=meta['Project']['AIL']['Year'][iY]+meta['Project']['Activities']['Years between Current and Post1 event'][iAT]
                            iT=np.where(tv==yr)[0]
                            if iT.size>0:
                                iOpenSpot=np.where(ec['ID_Type'][iT,iS,:]==0)[0]
                                if iOpenSpot.size>0:
                                    iOpenSpot=iOpenSpot[0]
                                    ec['ID_Type'][iT,iS,iOpenSpot]=meta['LUT']['Dist'][type]
                                    ec['MortalityFactor'][iT,iS,iOpenSpot]=meta['Project']['Activities']['% of biomass affected by Post1 event'][iAT]
                                    ec['GrowthFactor'][iT,iS,iOpenSpot]=0
                                    ec['ID_GrowthCurve'][iT,iS,iOpenSpot]=meta['Project']['Activities']['Growth curve ID following Post1 event (project)'][iAT]

                    #----------------------------------------------------------
                    # Post 2 event
                    #----------------------------------------------------------

                    type=meta['Project']['Activities']['Post2 event type'][iAT]
                    if (type!='None'):

                        if meta['Project']['Activities']['Post2 scenarios affected (All or Project Only)'][iAT]=='Project Only':

                            # Only add for project scenario
                            if (iScn>0):
                                yr=meta['Project']['AIL']['Year'][iY]+meta['Project']['Activities']['Years between Current and Post2 event'][iAT]
                                iT=np.where(tv==yr)[0]
                                if iT.size>0:
                                    iOpenSpot=np.where(ec['ID_Type'][iT,iS,:]==0)[0]
                                    if iOpenSpot.size>0:
                                        iOpenSpot=iOpenSpot[0]
                                        ec['ID_Type'][iT,iS,iOpenSpot]=meta['LUT']['Dist'][type]
                                        ec['MortalityFactor'][iT,iS,iOpenSpot]=meta['Project']['Activities']['% of biomass affected by Post2 event'][iAT]
                                        ec['GrowthFactor'][iT,iS,iOpenSpot]=0
                                        ec['ID_GrowthCurve'][iT,iS,iOpenSpot]=meta['Project']['Activities']['Growth curve ID following Post2 event (project)'][iAT]

                        else:

                            # Add for baseline and project scenario
                            yr=meta['Project']['AIL']['Year'][iY]+meta['Project']['Activities']['Years between Current and Post2 event'][iAT]
                            iT=np.where(tv==yr)[0]
                            if iT.size>0:
                                iOpenSpot=np.where(ec['ID_Type'][iT,iS,:]==0)[0]
                                if iOpenSpot.size>0:
                                    iOpenSpot=iOpenSpot[0]
                                    ec['ID_Type'][iT,iS,iOpenSpot]=meta['LUT']['Dist'][type]
                                    ec['MortalityFactor'][iT,iS,iOpenSpot]=meta['Project']['Activities']['% of biomass affected by Post2 event'][iAT]
                                    ec['GrowthFactor'][iT,iS,iOpenSpot]=0
                                    ec['ID_GrowthCurve'][iT,iS,iOpenSpot]=meta['Project']['Activities']['Growth curve ID following Post2 event (project)'][iAT]

                    #----------------------------------------------------------
                    # Post 3 event
                    #----------------------------------------------------------

                    type=meta['Project']['Activities']['Post3 event type'][iAT]
                    if (type!='None'):

                        if meta['Project']['Activities']['Post3 scenarios affected (All or Project Only)'][iAT]=='Project Only':

                            # Only add for project scenario
                            if (iScn>0):
                                yr=meta['Project']['AIL']['Year'][iY]+meta['Project']['Activities']['Years between Current and Post3 event'][iAT]
                                iT=np.where(tv==yr)[0]
                                if iT.size>0:
                                    iOpenSpot=np.where(ec['ID_Type'][iT,iS,:]==0)[0]
                                    if iOpenSpot.size>0:
                                        iOpenSpot=iOpenSpot[0]
                                        ec['ID_Type'][iT,iS,iOpenSpot]=meta['LUT']['Dist'][type]
                                        ec['MortalityFactor'][iT,iS,iOpenSpot]=meta['Project']['Activities']['% of biomass affected by Post3 event'][iAT]
                                        ec['GrowthFactor'][iT,iS,iOpenSpot]=0
                                        ec['ID_GrowthCurve'][iT,iS,iOpenSpot]=meta['Project']['Activities']['Growth curve ID following Post3 event (project)'][iAT]

                        else:

                            # Add for baseline and project scenario
                            yr=meta['Project']['AIL']['Year'][iY]+meta['Project']['Activities']['Years between Current and Post3 event'][iAT]
                            iT=np.where(tv==yr)[0]
                            if iT.size>0:
                                iOpenSpot=np.where(ec['ID_Type'][iT,iS,:]==0)[0]
                                if iOpenSpot.size>0:
                                    iOpenSpot=iOpenSpot[0]
                                    ec['ID_Type'][iT,iS,iOpenSpot]=meta['LUT']['Dist'][type]
                                    ec['MortalityFactor'][iT,iS,iOpenSpot]=meta['Project']['Activities']['% of biomass affected by Post3 event'][iAT]
                                    ec['GrowthFactor'][iT,iS,iOpenSpot]=0
                                    ec['ID_GrowthCurve'][iT,iS,iOpenSpot]=meta['Project']['Activities']['Growth curve ID following Post3 event (project)'][iAT]

                    #----------------------------------------------------------
                    # Add simulated wildfire from Taz
                    #----------------------------------------------------------

                    ind=np.array([],dtype=int)
                    if meta['Scenario'][iScn]['Wildfire Status Pre-modern']=='On':
                        ind0=np.where( (wf_sim['Occurrence'][:,iS]==1) & (meta['Year']<1920) )[0]
                        ind=np.append(ind,ind0)
                    if meta['Scenario'][iScn]['Wildfire Status Modern']=='On':
                        ind0=np.where( (wf_sim['Occurrence'][:,iS]==1) & (meta['Year']>=1920) & (meta['Year']<meta['Project']['Year Project']) )[0]
                        ind=np.append(ind,ind0)
                    if meta['Scenario'][iScn]['Wildfire Status Future']=='On':
                        ind0=np.where( (wf_sim['Occurrence'][:,iS]==1) & (meta['Year']>=meta['Project']['Year Project']) )[0]
                        ind=np.append(ind,ind0)

                    if ind.size>0:

                        ID_Type=meta['LUT']['Dist']['Wildfire']*np.ones(ind.size)
                        Year=tv[ind]
                        MortF=wf_sim['Mortality'][ind,iS]
                        GrowthF=0*np.ones(ind.size)
                        ID_GrowthCurve=1*np.ones(ind.size)

                        for iYr in range(Year.size):
                            iT=np.where(tv==Year[iYr])[0]

                            if ec['ID_Type'][iT,iS,0]!=0:
                                continue

                            ec['ID_Type'][iT,iS,0]=ID_Type[iYr]
                            ec['MortalityFactor'][iT,iS,0]=MortF[iYr]
                            ec['GrowthFactor'][iT,iS,0]=GrowthF[iYr]
                            ec['ID_GrowthCurve'][iT,iS,0]=ID_GrowthCurve[iYr]

                #--------------------------------------------------------------
                # Save to file
                #--------------------------------------------------------------

                gu.opickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + cbu.FixFileNum(iEns) + '_Bat' + cbu.FixFileNum(iBat) + '.pkl',ec)

    return

#%% Export tabular results

# *** Out of date ***

#def ExportResultsTabular(meta,vAT,tv,t_start,t_end):
#
#    #vList=['Year Calendar','Sec_NGHGB','Revenue Net','Revenue Gross','Cost Total']
#    sList=['BAU','CAP']
#
#    iT=np.where( (tv>=t_start) & (tv<=t_end) )[0]
#
#    for iScn in range(meta['Project']['N Scenario']):
#
#        for iAT in range(meta['Project']['N AT']):
#            d={}
#            d['Year Calendar']=tv[iT]
#            for k in vAT[iScn]['Sum'].keys():
#                d[k]=vAT[iScn]['Sum'][k][iT,iAT]
#            df=pd.DataFrame.from_dict(d)
#            df.to_excel(meta['Paths']['Project'] + '\\Outputs\\Scenario_' + sList[iScn] + '_Activity' + str(iAT+1) + '_Sum.xlsx')
#
#            d={}
#            d['Year Calendar']=tv[iT]
#            for k in vAT[iScn]['Mean'].keys():
#                d[k]=vAT[iScn]['Mean'][k][iT,iAT]
#            df=pd.DataFrame.from_dict(d)
#            df.to_excel(meta['Paths']['Project'] + '\\Outputs\\Scenario_' + sList[iScn] + '_Activity' + str(iAT+1) + '_PerHa.xlsx')
#
#    return


#%% Model Output Statistics (by portfolio)

def ModelOutputStatsByPortfolio(meta,**kwargs):

    #--------------------------------------------------------------------------
    # Kewword argumenst
    #--------------------------------------------------------------------------

    flag_save=1

    #--------------------------------------------------------------------------
    # Time
    #--------------------------------------------------------------------------

    tv=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)

    #--------------------------------------------------------------------------
    # Initialize structure
    #--------------------------------------------------------------------------

    mos=[]

    for iPort in range(meta['Project']['N Portfolio']):

        ListS=[]

        for iScn in range(meta['Project']['N Scenario']):

            d0={}
            d0['Sum']={}
            d0['Mean']={}

            if (iPort==0) & (iScn==0):

                # Import example output
                d1=cbu.LoadSingleOutputFile(meta,0,0,0)

                # Import event chronology
                if (meta['Scenario'][iScn]['Harvest Status Future']=='On') | (meta['Scenario'][iScn]['Breakup Status']=='On'):
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Modified_Events_Ens' + cbu.FixFileNum(0) + '_Bat' + cbu.FixFileNum(0) + '.pkl')
                else:
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + cbu.FixFileNum(0) + '_Bat' + cbu.FixFileNum(0) + '.pkl')

                # Uncompress event chronology if it has been compressed
                ec=cbu.EventChronologyDecompress(meta,ec,0,0,0)

                # Inventory
                inv=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + cbu.FixFileNum(0) + '.pkl')

                # Cashflow
                econ=econo.CalculateNetRevenue(meta,0,0,0,inv,ec,d1)
                d1.update(econ)

            # Initialize basic variables
            for k in d1.keys():

                if (k=='Year') | (k=='C_M_ByAgent'):
                    continue

                d0['Sum'][k]={}
                d0['Sum'][k]['Ensembles']=np.zeros((tv.size,meta['Project']['N Ensemble']))
                d0['Sum'][k]['Ensemble Mean']=np.zeros(tv.size)
                d0['Sum'][k]['Ensemble SD']=np.zeros(tv.size)
                d0['Sum'][k]['Ensemble P1']=np.zeros(tv.size)
                d0['Sum'][k]['Ensemble P10']=np.zeros(tv.size)
                d0['Sum'][k]['Ensemble CIL']=np.zeros(tv.size)
                d0['Sum'][k]['Ensemble CIU']=np.zeros(tv.size)
                d0['Sum'][k]['Ensemble P90']=np.zeros(tv.size)
                d0['Sum'][k]['Ensemble P99']=np.zeros(tv.size)

                d0['Mean'][k]={}
                d0['Mean'][k]['Ensembles']=np.zeros((tv.size,meta['Project']['N Ensemble']))
                d0['Mean'][k]['Ensemble Mean']=np.zeros(tv.size)
                d0['Mean'][k]['Ensemble SD']=np.zeros(tv.size)
                d0['Mean'][k]['Ensemble P1']=np.zeros(tv.size)
                d0['Mean'][k]['Ensemble P10']=np.zeros(tv.size)
                d0['Mean'][k]['Ensemble CIL']=np.zeros(tv.size)
                d0['Mean'][k]['Ensemble CIU']=np.zeros(tv.size)
                d0['Mean'][k]['Ensemble P90']=np.zeros(tv.size)
                d0['Mean'][k]['Ensemble P99']=np.zeros(tv.size)

            # Append to list of scenarios
            ListS.append(copy.deepcopy(d0))

        # Append to list of portoflios
        mos.append(copy.deepcopy(ListS))

    #----------------------------------------------------------------------
    # Loop through ensembles
    #----------------------------------------------------------------------

    for iScn in range(meta['Project']['N Scenario']):

        for iEns in range(meta['Project']['N Ensemble']):

            # Keep track of means to create area-weighted averages
            dmu0={}
            dmu1={}
            dmu2={}
            dmu3={}
            for k in d0['Sum'].keys():
                dmu0[k]=np.array([])
                dmu1[k]=np.array([])
                dmu2[k]=np.array([])
                dmu3[k]=np.array([])
            Area0=np.array([])
            Area1=np.array([])
            Area2=np.array([])
            Area3=np.array([])

            for iBat in range(meta['Project']['N Batch']):

                # Index to batch
                indBat=cbu.IndexToBatch(meta,iBat)

                # Area by year and activity type
                Area=np.tile(meta['Project']['Portfolio']['Area'][indBat],(tv.size,1))

                # Load basic output
                d1=cbu.LoadSingleOutputFile(meta,iScn,iEns,iBat)

                # Some projects may elect to keep each biomass pool in output ("Save Biomass Pools"=On")
                # This script doesn't handle it so they need to be summed first.
                if meta['Project']['Save Biomass Pools']=='On':
                    List=['C_Eco_Pools','C_Pro_Pools','C_G_Gross','C_G_Net','C_M_Reg','C_LF','C_RH']
                    for iList in range(len(List)):
                        nam=List[iList]
                        d1[nam]=np.sum(d1[nam],axis=2)

                # Import event chronology
                if (meta['Scenario'][iScn]['Harvest Status Future']=='On') | (meta['Scenario'][iScn]['Breakup Status']=='On'):
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Modified_Events_Ens' + cbu.FixFileNum(iEns) + '_Bat' + cbu.FixFileNum(iBat) + '.pkl')
                else:
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + cbu.FixFileNum(iEns) + '_Bat' + cbu.FixFileNum(iBat) + '.pkl')

                # Uncompress event chronology if it has been compressed
                ec=cbu.EventChronologyDecompress(meta,ec,iScn,iEns,iBat)

                # Inventory
                inv=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + cbu.FixFileNum(iBat) + '.pkl')

                # Cashflow
                econ=econo.CalculateNetRevenue(meta,iScn,iEns,iBat,inv,ec,d1)

                # Add to d1 structure
                d1.update(econ)

                # Index to portfolio
                indBAU=np.where(meta['Project']['Portfolio']['ID Portfolio'][indBat]==0)[0]
                indCAPa=np.where(meta['Project']['Portfolio']['ID Portfolio'][indBat]==1)[0]
                indCAPb=np.where(meta['Project']['Portfolio']['ID Portfolio'][indBat]==2)[0]
                indCAPc=np.where(meta['Project']['Portfolio']['ID Portfolio'][indBat]==3)[0]

                # Populate mos structure
                for k in d1.keys():

                    if (k=='Year') | (k=='C_M_ByAgent'):
                        continue

                    y=copy.copy(d1[k])

                    if indBAU.size>0:

                        mos[0][iScn]['Sum'][k]['Ensembles'][:,iEns]=mos[0][iScn]['Sum'][k]['Ensembles'][:,iEns]+np.sum(Area[:,indBAU]*y[:,indBAU],axis=1)/meta['Project']['N Stand per Activity Type']

                        try:
                            dmu0[k]=np.append(dmu0[k],y[:,indBAU],axis=1)
                        except:
                            dmu0[k]=y[:,indBAU]

                    if indCAPa.size>0:

                        mos[1][iScn]['Sum'][k]['Ensembles'][:,iEns]=mos[1][iScn]['Sum'][k]['Ensembles'][:,iEns]+np.sum(Area[:,indCAPa]*y[:,indCAPa],axis=1)/meta['Project']['N Stand per Activity Type']

                        try:
                            dmu1[k]=np.append(dmu1[k],y[:,indCAPa],axis=1)
                        except:
                            dmu1[k]=y[:,indCAPa]

                    if indCAPb.size>0:

                        mos[2][iScn]['Sum'][k]['Ensembles'][:,iEns]=mos[2][iScn]['Sum'][k]['Ensembles'][:,iEns]+np.sum(Area[:,indCAPb]*y[:,indCAPb],axis=1)/meta['Project']['N Stand per Activity Type']

                        try:
                            dmu2[k]=np.append(dmu2[k],y[:,indCAPb],axis=1)
                        except:
                            dmu2[k]=y[:,indCAPb]

                    if indCAPc.size>0:

                        mos[3][iScn]['Sum'][k]['Ensembles'][:,iEns]=mos[3][iScn]['Sum'][k]['Ensembles'][:,iEns]+np.sum(Area[:,indCAPc]*y[:,indCAPc],axis=1)/meta['Project']['N Stand per Activity Type']

                        try:
                            dmu3[k]=np.append(dmu3[k],y[:,indCAPc],axis=1)
                        except:
                            dmu3[k]=y[:,indCAPc]

                # Delete variables from workspace
                del d1,ec,inv,econ
                garc.collect()

                # Add areas
                if indBAU.size>0:
                    Area0=np.append(Area0,meta['Project']['Portfolio']['Area'][indBat[indBAU]])

                if indCAPa.size>0:
                    Area1=np.append(Area1,meta['Project']['Portfolio']['Area'][indBat[indCAPa]])

                if indCAPb.size>0:
                    Area2=np.append(Area2,meta['Project']['Portfolio']['Area'][indBat[indCAPb]])

                if indCAPc.size>0:
                    Area3=np.append(Area3,meta['Project']['Portfolio']['Area'][indBat[indCAPc]])

            # Area weighted average
            for k in dmu0.keys():

                if (k=='Year') | (k=='C_M_ByAgent'):
                    continue

                y=copy.copy(dmu0[k])
                mos[0][iScn]['Mean'][k]['Ensembles'][:,iEns]=np.sum(y*np.tile(Area0,(tv.size,1)),axis=1)/np.sum(Area0) #/meta['Project']['N Stand per Activity Type']

                y=copy.copy(dmu1[k])
                if y.size!=0:
                    mos[1][iScn]['Mean'][k]['Ensembles'][:,iEns]=np.sum(y*np.tile(Area1,(tv.size,1)),axis=1)/np.sum(Area1) #/meta['Project']['N Stand per Activity Type']

                y=copy.copy(dmu2[k])
                if y.size!=0:
                    mos[2][iScn]['Mean'][k]['Ensembles'][:,iEns]=np.sum(y*np.tile(Area2,(tv.size,1)),axis=1)/np.sum(Area2) #/meta['Project']['N Stand per Activity Type']
            #del dmu0,dmu1
            #print(dmu0['A'].shape)
            #print(dmu1['A'].shape)

    #----------------------------------------------------------------------
    # Calculate statistics
    #----------------------------------------------------------------------

    for iPort in range(meta['Project']['N Portfolio']):

        for iScn in range(meta['Project']['N Scenario']):

            for k in mos[iPort][iScn]['Sum'].keys():

                if (k=='Year') | (k=='C_M_ByAgent') | (k=='C_Eco_Pools') | (k=='C_Pro_Pools'):
                    continue

                mu=copy.copy(np.mean(mos[iPort][iScn]['Sum'][k]['Ensembles'],axis=1))
                sd=copy.copy(np.std(mos[iPort][iScn]['Sum'][k]['Ensembles'],axis=1))
                mos[iPort][iScn]['Sum'][k]['Ensemble Mean']=mu
                mos[iPort][iScn]['Sum'][k]['Ensemble SD']=sd
                mos[iPort][iScn]['Sum'][k]['Ensemble CIL']=mu-2*sd/np.sqrt(meta['Project']['N Ensemble'])
                mos[iPort][iScn]['Sum'][k]['Ensemble CIU']=mu+2*sd/np.sqrt(meta['Project']['N Ensemble'])
                mos[iPort][iScn]['Sum'][k]['Ensemble P1']=copy.copy(np.percentile(mos[iPort][iScn]['Sum'][k]['Ensembles'],1,axis=1))
                mos[iPort][iScn]['Sum'][k]['Ensemble P10']=copy.copy(np.percentile(mos[iPort][iScn]['Sum'][k]['Ensembles'],10,axis=1))
                mos[iPort][iScn]['Sum'][k]['Ensemble P90']=copy.copy(np.percentile(mos[iPort][iScn]['Sum'][k]['Ensembles'],90,axis=1))
                mos[iPort][iScn]['Sum'][k]['Ensemble P99']=copy.copy(np.percentile(mos[iPort][iScn]['Sum'][k]['Ensembles'],99,axis=1))

                mu=copy.copy(np.mean(mos[iPort][iScn]['Mean'][k]['Ensembles'],axis=1))
                sd=copy.copy(np.std(mos[iPort][iScn]['Mean'][k]['Ensembles'],axis=1))
                mos[iPort][iScn]['Mean'][k]['Ensemble Mean']=mu
                mos[iPort][iScn]['Mean'][k]['Ensemble SD']=sd
                mos[iPort][iScn]['Mean'][k]['Ensemble CIL']=mu-2*sd/np.sqrt(meta['Project']['N Ensemble'])
                mos[iPort][iScn]['Mean'][k]['Ensemble CIU']=mu+2*sd/np.sqrt(meta['Project']['N Ensemble'])
                mos[iPort][iScn]['Mean'][k]['Ensemble P1']=copy.copy(np.percentile(mos[iPort][iScn]['Mean'][k]['Ensembles'],1,axis=1))
                mos[iPort][iScn]['Mean'][k]['Ensemble P10']=copy.copy(np.percentile(mos[iPort][iScn]['Mean'][k]['Ensembles'],10,axis=1))
                mos[iPort][iScn]['Mean'][k]['Ensemble P90']=copy.copy(np.percentile(mos[iPort][iScn]['Mean'][k]['Ensembles'],90,axis=1))
                mos[iPort][iScn]['Mean'][k]['Ensemble P99']=copy.copy(np.percentile(mos[iPort][iScn]['Mean'][k]['Ensembles'],99,axis=1))

    #----------------------------------------------------------------------
    # Delete individual ensembles
    #----------------------------------------------------------------------

    #    for iPort in range(meta['Project']['N Portfolio']):
    #        for iScn in range(meta['Project']['N Scenario']):
    #            for k in mos[iPort][iScn].keys():
    #                try:
    #                    del mos[iPort][iScn][k]['Ensembles']
    #                except:
    #                    print(k)

    #--------------------------------------------------------------------------
    # Save
    #--------------------------------------------------------------------------

    if flag_save==1:
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\MOS.pkl',mos)

    return mos

#%% Model Output Statistics (by portfolio)

def ModelOutputStatsByAT(meta,**kwargs):

    #--------------------------------------------------------------------------
    # Time
    #--------------------------------------------------------------------------

    tv=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)

    #--------------------------------------------------------------------------
    # Initialize structure
    #--------------------------------------------------------------------------

    mos=[]

    for iPort in range(meta['Project']['N Portfolio']):

        ListAT=[]

        # Unique ATs for this portfolio
        indP=np.where( (meta['Project']['Portfolio']['ID Portfolio']==iPort) )[0]
        uAT=np.unique(meta['Project']['Portfolio']['ID AT'][indP])

        for iAT in range(uAT.size):

            ListS=[]

            for iScn in range(meta['Project']['N Scenario']):

                d0={}
                d0['Sum']={}
                d0['Mean']={}

                if (iAT==0) & (iScn==0):

                    # Import example output
                    d1=cbu.LoadSingleOutputFile(meta,0,0,0)

                    # Import event chronology
                    if (meta['Scenario'][iScn]['Harvest Status Future']=='On') | (meta['Scenario'][iScn]['Breakup Status']=='On'):
                        ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Modified_Events_Ens' + cbu.FixFileNum(0) + '_Bat' + cbu.FixFileNum(0) + '.pkl')
                    else:
                        ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + cbu.FixFileNum(0) + '_Bat' + cbu.FixFileNum(0) + '.pkl')

                    # Uncompress event chronology if it has been compressed
                    ec=cbu.EventChronologyDecompress(meta,ec,0,0,0)

                    # Inventory
                    inv=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + cbu.FixFileNum(0) + '.pkl')

                    # Cashflow
                    econ=econo.CalculateNetRevenue(meta,0,0,0,inv,ec,d1)
                    d1.update(econ)

                # Initialize basic variables
                for k in d1.keys():

                    if (k=='Year') | (k=='C_M_ByAgent'):
                        continue

                    d0['Sum'][k]={}
                    d0['Sum'][k]['Ensembles']=np.zeros((tv.size,meta['Project']['N Ensemble']))
                    d0['Sum'][k]['Ensemble Mean']=np.zeros(tv.size)

                    d0['Mean'][k]={}
                    d0['Mean'][k]['Ensembles']=np.zeros((tv.size,meta['Project']['N Ensemble']))
                    d0['Mean'][k]['Ensemble Mean']=np.zeros(tv.size)

                # Add cumulative variables
                v_cumu=['E_CO2e_AGHGB_WSub','E_CO2e_AGHGB_WOSub']
                #'Cost Total','Revenue Gross','Revenue Net','Cost Total Disc','Revenue Gross Disc','Revenue Net Disc'
                for k in v_cumu:
                    d0['Sum'][k + '_cumu']={}
                    d0['Sum'][k + '_cumu']['Ensembles']=np.zeros((tv.size,meta['Project']['N Ensemble']))
                    d0['Sum'][k + '_cumu']['Ensemble Mean']=np.zeros((tv.size,meta['Project']['N Portfolio']))

                    d0['Mean'][k + '_cumu']={}
                    d0['Mean'][k + '_cumu']['Ensembles']=np.zeros((tv.size,meta['Project']['N Ensemble']))
                    d0['Mean'][k + '_cumu']['Ensemble Mean']=np.zeros((tv.size,meta['Project']['N Portfolio']))

                # Append to list of scenarios
                ListS.append(copy.deepcopy(d0))

            # Append to list of
            ListAT.append(copy.deepcopy(ListS))

        # Append to list of portoflios
        mos.append(copy.deepcopy(ListAT))

    #----------------------------------------------------------------------
    # Loop through ensembles
    #----------------------------------------------------------------------

    for iScn in range(meta['Project']['N Scenario']):

        for iEns in range(meta['Project']['N Ensemble']):

            # Save area for calculating mean
            #d_ForMean=[]
            #Area_ForMean=[]
            for iPort in range(meta['Project']['N Portfolio']):

                # Index to portfolio
                indP=np.where(meta['Project']['Portfolio']['ID Portfolio']==iPort)[0]

                # Unique ATs for this portfolio
                uAT=np.unique(meta['Project']['Portfolio']['ID AT'][indP])

                #d_ForMean.append(copy.deepcopy(d_ForMean0))
                #Area_ForMean.append(copy.deepcopy(Area_ForMean0))

            for iBat in range(meta['Project']['N Batch']):

                # Index to batch
                indBat=cbu.IndexToBatch(meta,iBat)

                # Area by year and activity type
                Area=np.tile(meta['Project']['Portfolio']['Area'][indBat],(tv.size,1))

                # Load basic output
                d1=cbu.LoadSingleOutputFile(meta,iScn,iEns,iBat)

                # Some projects may elect to keep each biomass pool in output ("Save Biomass Pools"=On")
                # This script doesn't handle it so they need to be summed first.
                if meta['Project']['Save Biomass Pools']=='On':
                    List=['C_Eco_Pools','C_Pro_Pools','C_G_Gross','C_G_Net','C_M_Reg','C_LF','C_RH']
                    for iList in range(len(List)):
                        nam=List[iList]
                        d1[nam]=np.sum(d1[nam],axis=2)

                # Import event chronology
                if (meta['Scenario'][iScn]['Harvest Status Future']=='On') | (meta['Scenario'][iScn]['Breakup Status']=='On'):
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Modified_Events_Ens' + cbu.FixFileNum(iEns) + '_Bat' + cbu.FixFileNum(iBat) + '.pkl')
                else:
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + cbu.FixFileNum(iEns) + '_Bat' + cbu.FixFileNum(iBat) + '.pkl')

                # Uncompress event chronology if it has been compressed
                ec=cbu.EventChronologyDecompress(meta,ec,iScn,iEns,iBat)

                # Inventory
                inv=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + cbu.FixFileNum(iBat) + '.pkl')

                # Cashflow
                econ=econo.CalculateNetRevenue(meta,iScn,iEns,iBat,inv,ec,d1)

                # Add to d1 structure
                d1.update(econ)

                # Add cumulative
                for k in v_cumu:
                    d1[k + '_cumu']=np.cumsum(d1[k],axis=0)

                # Loop through combinations of portfolio and activity
                for iPort in range(meta['Project']['N Portfolio']):

                    # Index to portfolio
                    indP=np.where(meta['Project']['Portfolio']['ID Portfolio'][indBat]==iPort)[0]

                    # Unique ATs for this portfolio
                    uAT=np.unique(meta['Project']['Portfolio']['ID AT'][indP])

                    for iAT in range(uAT.size):

                        indAT=np.where( (meta['Project']['Portfolio']['ID Portfolio'][indBat]==iPort) & (meta['Project']['Portfolio']['ID AT'][indBat]==uAT[iAT]) )[0]

                        # Populate mos structure
                        for k in d1.keys():

                            if (k=='Year') | (k=='C_M_ByAgent'):
                                continue

                            if indP.size>0:

                                mos[iPort][iAT][iScn]['Sum'][k]['Ensembles'][:,iEns]=mos[iPort][iAT][iScn]['Sum'][k]['Ensembles'][:,iEns]+np.sum(Area[:,indAT]*d1[k][:,indAT],axis=1)/meta['Project']['N Stand per Activity Type']

#                                try:
#                                    d_ForMean[iPort][iAT][k]=np.append(d_ForMean[iPort][iAT][k],d1[k][:,indAT],axis=1)
#                                except:
#                                    d_ForMean[iPort][iAT][k]=d1[k][:,indAT]

                        #Area_ForMean[iPort][iAT]=np.append(Area_ForMean[iPort][iAT],meta['Project']['Portfolio']['Area'][indBat[indAT]])

                # Delete variables from workspace
                del d1,ec,inv,econ
                garc.collect()


    #----------------------------------------------------------------------
    # Calculate statistics
    #----------------------------------------------------------------------

    for iPort in range(meta['Project']['N Portfolio']):

        # Index to portfolio
        indP=np.where(meta['Project']['Portfolio']['ID Portfolio'][indBat]==iPort)[0]

        # Unique ATs for this portfolio
        uAT=np.unique(meta['Project']['Portfolio']['ID AT'][indP])

        for iAT in range(uAT.size):

            iAT_FromAIL=np.where( (meta['Project']['AIL']['ID Portfolio']==iPort) & (meta['Project']['AIL']['ID AT']==uAT[iAT]) )[0]
            TotalArea_AT=np.sum(meta['Project']['AIL']['Area'][:,iAT_FromAIL])

            for iScn in range(meta['Project']['N Scenario']):

                for k in mos[iPort][iAT][iScn]['Sum'].keys():

                    if (k=='Year') | (k=='C_M_ByAgent') | (k=='C_Eco_Pools') | (k=='C_Pro_Pools'):
                        continue

                    mos[iPort][iAT][iScn]['Sum'][k]['Ensemble Mean']=np.mean(mos[iPort][iAT][iScn]['Sum'][k]['Ensembles'],axis=1)

                    #mos[iPort][iAT][iScn]['Mean'][k]['Ensemble Mean']=np.mean(mos[iPort][iAT][iScn]['Mean'][k]['Ensembles'],axis=1)
                    mos[iPort][iAT][iScn]['Mean'][k]['Ensemble Mean']=mos[iPort][iAT][iScn]['Sum'][k]['Ensemble Mean']/TotalArea_AT

    #----------------------------------------------------------------------
    # Delete individual ensembles
    #----------------------------------------------------------------------

#    for iPort in range(meta['Project']['N Portfolio']):
#        for iScn in range(meta['Project']['N Scenario']):
#            for k in mos[iPort][iAT][iScn].keys():
#                try:
#                    del mos[iPort][iAT][iScn][k]['Ensembles']
#                except:
#                    print(k)

    #--------------------------------------------------------------------------
    # Save
    #--------------------------------------------------------------------------

    #if flag_save==1:
    gu.opickle(meta['Paths']['Project'] + '\\Outputs\\MOS_ByAT.pkl',mos)

    return mos

#%% Plot results

def Plot_TimeSeries(meta,mos,t_start,t_end,iBAU,iCAPa,iCAPb):

    # Scenarios
    iB=0; iP=1

    # Import ylabels
    ylabs=gu.ReadExcel(meta['Paths']['Model Code'] + '\\Parameters\\LabelYAxis.xlsx')

    tv=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)

    iT=np.where( (tv>=t_start) & (tv<=t_end) )[0]

    cl_b=[0,0,0.5];
    cl_p=[1,0.2,0.2];

    cl_b2=[0,0,0.5];
    cl_p2=[1,0.2,0.2];

    cl_b3=[0,0,0.5];
    cl_p3=[1,0.2,0.2];

    cl_d=[0.24,0.49,0.77]
    cl_d2=[0.45,0.85,0]
    cl_d3=[0,0.55,0]

    Alpha=0.08

    #ListOp=['Sum','Mean']
    ListOp=['Sum']

    for op in ListOp:

        if op=='Sum':
            ddb=meta['Project']['Display divide by']
        else:
            ddb=1.0

        # Full list of variables
        #ListV=['A','V_MerchLive','V_MerchDead','V_MerchTotal','V_ToMillMerchLive','V_ToMillMerchDead','V_ToMillMerchTotal','V_ToMillNonMerch','LogSizeEnhancement','C_Biomass_Tot','C_Piled_Tot','C_Litter_Tot','C_DeadWood_Tot','C_Soil_Tot','C_InUse_Tot','DM_Buildings_Tot','C_DumpLandfill_Tot','C_M_Dist','C_G_Gross_Tot','C_G_Net_Tot','C_M_Reg_Tot','C_LF_Tot','C_RH_Tot','C_ToMillMerch','C_ToMillNonMerch','C_ToMillSnagStem','C_ToSlashpileBurn','C_ToLumber','C_ToPlywood','C_ToOSB','C_ToMDF','C_ToPaper','C_ToPowerFacilityDom','C_ToPowerFacilityFor','C_ToPowerGrid','C_ToPellets','C_ToFirewoodDom','C_ToFirewoodFor','C_ToLogExport','E_CO2e_LULUCF_NEE','E_CO2e_LULUCF_Wildfire','E_CO2e_LULUCF_OpenBurning','E_CO2e_LULUCF_EcoOther','E_CO2e_LULUCF_HWP','E_CO2e_ESC_Comb','E_CO2e_ESC_SubE','E_CO2e_ESC_SubBM','E_CO2e_ET_Comb','E_CO2e_IPPU_Comb','C_NPP_Tot','C_ToMill','C_Forest','C_HWP','DM_Production_Sawnwood','DM_Production_Panels','E_CO2e_LULUCF_Fire','E_CO2e_AGHGB_WSub','E_CO2e_AGHGB_WOSub','E_CO2e_AGHGB_WSub_cumu','E_CO2e_AGHGB_WOSub_cumu','E_CO2e_AGHGB_WSub_cumu_from_tref','E_CO2e_AGHGB_WOSub_cumu_from_tref','Yield Lumber','Yield Plywood','Yield OSB','Yield MDF','Yield Paper','Yield Pellets','Yield PowerGrid','Yield PowerFacilityDom','Yield FirewoodDom','Yield LogExport','Price Lumber','Price Plywood','Price OSB','Price MDF','Price Newsprint','Price PowerFacilityDom','Price PowerGrid','Price Pellets','Price LogExport','Price FirewoodDom','Exchange Rate US','Exchange Rate Euro','Cost Roads','Cost Harvest Overhead','Cost Harvest Felling and Piling','Cost Harvest Hauling','Cost Harvest Residuals','Cost Milling','Cost Nutrient Management','Cost Planting','Cost Survey','Cost Knockdown','Cost Ripping','Cost Slashpile Burn','Harvest Vol Merch','Harvest Vol Resid','Cost Total','Revenue Lumber','Revenue Plywood','Revenue OSB','Revenue MDF','Revenue Paper','Revenue PowerFacilityDom','Revenue PowerGrid','Revenue Pellets','Revenue FirewoodDom','Revenue LogExport','Revenue Gross','Revenue Net','Revenue Net Disc','Revenue Gross Disc','Cost Total Disc','Revenue Gross Disc_cumu','Revenue Net Disc_cumu','Cost Total Disc_cumu','Cost Total_cumu','Revenue Gross_cumu','Revenue Net_cumu',
        #'Production Concrete','Production Steel','Production Aluminum','Production Plastic','E_CO2e_Concrete','E_CO2e_Steel','E_CO2e_Aluminum','E_CO2e_Plastic','E_CO2e_NRBM']

        # Condensed list of variables
        # ListV=['A','V_MerchTotal','V_ToMillMerchTotal','V_ToMillNonMerch','C_Biomass_Tot',
        #        'C_ToMillMerch','C_ToMillNonMerch','C_ToMillSnagStem',
        #        'C_Buildings_Tot','C_NonBuildings_Tot','C_ToSlashpileBurnTot','C_ToPowerFacilityDom',
        #        'E_CO2e_ESC_Bioenergy','E_CO2e_LULUCF_NEE','E_CO2e_LULUCF_Wildfire',
        #        'E_CO2e_LULUCF_OpenBurning','C_NPP_Tot','E_CO2e_SUB_E','E_CO2e_OPER','E_CO2e_SUB_M',
        #        'E_CO2e_Coal','E_CO2e_Oil','E_CO2e_Gas','E_CO2e_SUB_Calcination',
        #        'E_CO2e_SUB_Concrete','E_CO2e_SUB_Steel','E_CO2e_SUB_Aluminum',
        #        'E_CO2e_SUB_Plastic','E_CO2e_AGHGB_WSub','E_CO2e_AGHGB_WOSub',
        #        'E_CO2e_AGHGB_WSub_cumu_from_tref','E_CO2e_AGHGB_WOSub_cumu_from_tref',
        #        'Yield Sawnwood','Yield Panels','Yield Concrete','Yield Steel','Yield Aluminum',
        #        'Yield Plastic','Yield Lumber','Yield Plywood','Yield OSB','Yield MDF',
        #        'Yield Paper','Yield Pellets','Yield PowerGrid','Yield PowerFacilityDom',
        #        'Yield Firewood','Yield LogExport',
        #        'Revenue Paper',
        #        'Revenue Gross','Revenue Net','Revenue Net Disc','Revenue Gross Disc',
        #        'Revenue Gross Disc_cumu','Revenue Net Disc_cumu',
        #        'Revenue Gross_cumu','Revenue Net_cumu',
        #        'Cost Total','Cost Total Disc','Cost Total Disc_cumu','Cost Total_cumu',
        #        'Cost Silviculture Total','Cost Silviculture Total Disc','Cost Silviculture Total Disc_cumu']
        ListV=meta['Core']['Output Variable List']

        for iV in range(len(ListV)):

            #if iV>=1:
            #    continue

            k=ListV[iV]

            # BAU
            be_b_bau=mos[iBAU][iB][op][k]['Ensemble Mean'][iT]/ddb
            lo_b_bau=mos[iBAU][iB][op][k]['Ensemble CIL'][iT]/ddb
            hi_b_bau=mos[iBAU][iB][op][k]['Ensemble CIU'][iT]/ddb

            be_p_bau=mos[iBAU][iP][op][k]['Ensemble Mean'][iT]/ddb
            lo_p_bau=mos[iBAU][iP][op][k]['Ensemble CIL'][iT]/ddb
            hi_p_bau=mos[iBAU][iP][op][k]['Ensemble CIU'][iT]/ddb

            be_d_bau=be_p_bau-be_b_bau
            lo_d_bau,hi_d_bau=gu.GetCIsFromDifference(lo_b_bau,hi_b_bau,lo_p_bau,hi_p_bau)

            # CAP-A
            be_b_cap_a=mos[iCAPa][iB][op][k]['Ensemble Mean'][iT]/ddb
            lo_b_cap_a=mos[iCAPa][iB][op][k]['Ensemble CIL'][iT]/ddb
            hi_b_cap_a=mos[iCAPa][iB][op][k]['Ensemble CIU'][iT]/ddb

            be_p_cap_a=mos[iCAPa][iP][op][k]['Ensemble Mean'][iT]/ddb
            lo_p_cap_a=mos[iCAPa][iP][op][k]['Ensemble CIL'][iT]/ddb
            hi_p_cap_a=mos[iCAPa][iP][op][k]['Ensemble CIU'][iT]/ddb

            be_d_cap_a=be_p_cap_a-be_b_cap_a
            lo_d_cap_a,hi_d_cap_a=gu.GetCIsFromDifference(lo_b_cap_a,hi_b_cap_a,lo_p_cap_a,hi_p_cap_a)

            be_p_d_capa=be_p_cap_a-be_p_bau
            lo_p_d_capa,hi_p_d_capa=gu.GetCIsFromDifference(lo_p_bau,hi_p_bau,lo_p_cap_a,hi_p_cap_a)

            # CAP-B
            be_b_cap_b=mos[iCAPb][iB][op][k]['Ensemble Mean'][iT]/ddb
            lo_b_cap_b=mos[iCAPb][iB][op][k]['Ensemble CIL'][iT]/ddb
            hi_b_cap_b=mos[iCAPb][iB][op][k]['Ensemble CIU'][iT]/ddb

            be_p_cap_b=mos[iCAPb][iP][op][k]['Ensemble Mean'][iT]/ddb
            lo_p_cap_b=mos[iCAPb][iP][op][k]['Ensemble CIL'][iT]/ddb
            hi_p_cap_b=mos[iCAPb][iP][op][k]['Ensemble CIU'][iT]/ddb

            be_d_cap_b=be_p_cap_b-be_b_cap_b
            lo_d_cap_b,hi_d_cap_b=gu.GetCIsFromDifference(lo_b_cap_b,hi_b_cap_b,lo_p_cap_b,hi_p_cap_b)

            be_p_d_capb=be_p_cap_b-be_p_bau
            lo_p_d_capb,hi_p_d_capb=gu.GetCIsFromDifference(lo_p_bau,hi_p_bau,lo_p_cap_b,hi_p_cap_b)

            # y-axis label
            ind=np.where(ylabs['Name']==k)[0]
            if ind.size==1:
                if (op=='Sum') & (ddb==1e6):
                    lab=ylabs['Y Label Sum Mt'][ind[0]]
                elif (op=='Sum') & (ddb==1e9):
                    lab=ylabs['Y Label Sum Gt'][ind[0]]
                else:
                    lab=ylabs['Y Label Mean'][ind[0]]
            else:
                lab=k

            # Plot
            plt.close('all')
            fig,ax=plt.subplots(1,3,figsize=gu.cm2inch(17,7.5))

            ax[0].plot(tv[iT],np.zeros(iT.size),'-',lw=2,color=[0.8,0.8,0.8])
            #ax[0].fill_between(tv[iT],lo_b_bau,hi_b_bau,color=cl_b,alpha=Alpha,lw=0)
            #ax[0].fill_between(tv[iT],lo_p_bau,hi_p_bau,color=cl_p,alpha=Alpha,lw=0)
            #ax[0].fill_between(tv[iT],lo_b_cap_a,hi_b_cap_a,color=cl_b,alpha=Alpha,lw=0)
            #ax[0].fill_between(tv[iT],lo_p_cap_a,hi_p_cap_a,color=cl_p,alpha=Alpha,lw=0)
            ax[0].plot(tv[iT],be_b_bau,'-',color=cl_b,label='BAU baseline')
            ax[0].plot(tv[iT],be_p_bau,'-',color=cl_p,label='BAU project')
            ax[0].plot(tv[iT],be_b_cap_a,'--',color=cl_b,label='CAP-A baseline')
            ax[0].plot(tv[iT],be_p_cap_a,'--',color=cl_p,label='CAP-A project')
            ax[0].plot(tv[iT],be_b_cap_b,'-.',color=cl_b,label='CAP-B baseline')
            ax[0].plot(tv[iT],be_p_cap_b,'-.',color=cl_p,label='CAP-B project')
            ax[0].set(position=[0.06,0.08,0.26,0.68],xlim=[tv[iT[0]],tv[iT[-1]]],ylabel=lab)
            ax[0].legend(loc='upper center',ncol=2,bbox_to_anchor=(0.5,1.21),frameon=False,facecolor=None,fontsize=gp['fs3'])
            ax[0].yaxis.set_ticks_position('both'); ax[0].xaxis.set_ticks_position('both')

            ax[1].plot(tv[iT],np.zeros(iT.size),'-',lw=2,color=[0.8,0.8,0.8])
            #ax[1].fill_between(tv[iT],lo_d_bau,hi_d_bau,color=cl_d,alpha=Alpha,lw=0)
            #ax[1].fill_between(tv[iT],lo_d_cap_a,hi_d_cap_a,color=cl_d2,alpha=Alpha,lw=0)
            ax[1].plot(tv[iT],be_d_bau,'-',color=cl_d,label='BAU project minus BAU baseline ($\Delta$BAU)')
            ax[1].plot(tv[iT],be_d_cap_a,'--',color=cl_d2,label='CAP-A project minus CAP-A baseline ($\Delta$CAP-A)')
            ax[1].plot(tv[iT],be_d_cap_b,'-.',color=cl_d3,label='CAP-B project minus CAP-B baseline ($\Delta$CAP-B)')
            ax[1].legend(loc='upper center',bbox_to_anchor=(0.5,1.27),frameon=False,facecolor=None,fontsize=gp['fs3'])
            ax[1].set(position=[0.41,0.08,0.26,0.68],xlim=[tv[iT[0]],tv[iT[-1]]],ylabel='$\Delta$ ' + lab)
            ax[1].yaxis.set_ticks_position('both'); ax[1].xaxis.set_ticks_position('both')

            ax[2].plot(tv[iT],np.zeros(iT.size),'-',lw=2,color=[0.8,0.8,0.8])
            #ax[2].fill_between(tv[iT],lo_p_d_capa,hi_p_d_capa,color=cl_d,alpha=Alpha,lw=0)
            #ax[2].plot(tv[iT],be_p_d_capa,'-',color=cl_d,label='CAP Project minus BAU Project')
            ax[2].plot(tv[iT],be_d_cap_a-be_d_bau,'--',color=cl_d2,label='$\Delta$CAP-A minus $\Delta$BAU')
            ax[2].plot(tv[iT],be_d_cap_b-be_d_bau,'-.',color=cl_d3,label='$\Delta$CAP-B minus $\Delta$BAU')
            ax[2].legend(loc='upper center',bbox_to_anchor=(0.5,1.15),frameon=False,facecolor=None,fontsize=gp['fs3'])
            ax[2].set(position=[0.76,0.08,0.26,0.68],xlim=[tv[iT[0]],tv[iT[-1]]],ylabel='$\Delta$ ' + lab)
            ax[2].yaxis.set_ticks_position('both'); ax[2].xaxis.set_ticks_position('both')

            gu.PrintFig(meta['Paths']['Figures'] + '\\ByPortfolio_' + op + '_' + k,'png',200)

    plt.close('all')

    return

#%% Plot implementation level of each activity type

def Plot_AIL_WithAccounting(meta):

    # Hatch width
    mpl.rcParams['hatch.linewidth']=1.3

    tv=meta['Project']['AIL']['Year']

    # Mean AIL of the reference period
    iRP=np.where( (tv>=meta['Project']['Year Reference Period Start']) & (tv<=meta['Project']['Year Reference Period End']) )[0]

    for iActivity in range(meta['Project']['Activities']['Activity ID'].size):

        #if iActivity>0:
        #    continue

        idAT=meta['Project']['Activities']['Activity ID'][iActivity]
        nameAT=meta['Project']['Activities']['Activity description'][iActivity]

        # Index to activities for each portfolio
        iAT1=np.where( (meta['Project']['AIL']['ID Portfolio']==1) & (meta['Project']['AIL']['ID AT']==idAT) )[0][0]
        iAT2=np.where( (meta['Project']['AIL']['ID Portfolio']==2) & (meta['Project']['AIL']['ID AT']==idAT) )[0][0]

        # AIL
        ail1=meta['Project']['AIL']['Area'][:,iAT1]/1000
        ail2=meta['Project']['AIL']['Area'][:,iAT2]/1000

        # Mean implementation of the reference period
        mirp=np.mean(ail1[iRP])*np.ones(tv.size)

        iSpan=np.where( (ail1>0) | (ail2>0) )[0]
        if iSpan.size==0:
            continue
        iComp=np.where( (tv>=tv[0]) & (tv<=2020) )[0]
        iPlanned=np.where( (tv>2020) & (tv<=tv[iSpan[-1]]) )[0]

        iInc=np.where( (tv>=meta['Project']['Time Start Accounting']) & (ail1>mirp) )[0]
        ail_ni1=ail1.copy();
        ail_ni1[iInc]=mirp[iInc]
        ail_i1=np.zeros(tv.size);
        ail_i1[iInc]=ail1[iInc]-mirp[iInc]

        iInc=np.where( (tv>=meta['Project']['Time Start Accounting']) & (ail2>mirp) )[0]
        ail_ni2=ail2.copy();
        ail_ni2[iInc]=mirp[iInc]
        ail_i2=np.zeros(tv.size);
        ail_i2[iInc]=ail2[iInc]-mirp[iInc]

        plt.close('all');
        ymx=1.2*np.max(ail1)

        fig,ax=plt.subplots(1,figsize=gu.cm2inch(15.5,6));

        # Reference period
        ax.add_patch(Rectangle([tv[iRP[0]]-0.5,0],iRP.size,ymx,fc=[1,1,0.9],ec="none"))
        ax.text(np.mean(tv[iRP]),0.9*ymx,'Reference Period',color=[0.6,0.6,0.2],ha='center',fontsize=7,fontweight='bold')
        ax.plot(tv,mirp[0]*np.ones(tv.size),'k--',color=[0.6,0.6,0.2],linewidth=1.5,label='Mean implementation of reference period')

        # Completed
        ax.bar(tv[iComp],ail_ni1[iComp],0.8,facecolor=[0.6,0.6,0.6],label='Completed total implementation')
        ax.bar(tv[iComp],ail_i1[iComp],0.8,bottom=ail_ni1[iComp],facecolor=[0,0.9,0.9],label='Completed incremental implementation')

        # Planned spend
        cl1=[0.45,0.85,0]
        cl2=[0,0.55,0]
        ax.plot(tv[iPlanned],ail1[iPlanned],'-ko',ms=4,color=cl1,mfc=cl1,mec=cl1,lw=1,label='CAP-A')
        ax.plot(tv[iPlanned],ail2[iPlanned],'-ks',ms=4,color=cl2,mfc=cl2,mec=cl2,lw=1,label='CAP-B')

        ax.set(position=[0.055,0.12,0.93,0.86],xlim=[tv[0]-0.5,tv[iSpan[-1]]+0.5],
               xticks=np.arange(tv[0],tv[iSpan[-1]]+0.5,5),xlabel='Time, years',
               ylim=[0,ymx],ylabel='Annual implementation level (hectares x 1000)')

        # Activity type label
        #ax.text(np.mean(tv[0]+2),16,'Reference Period',ha='center',fontsize=7,fontweight='bold')

        ax.legend(loc='upper left',facecolor=[1,1,1],frameon=False)
        ax.yaxis.set_ticks_position('both'); ax.xaxis.set_ticks_position('both')

        gu.PrintFig(meta['Paths']['Figures'] + '\\AIL_' + nameAT,'png',500)

    return

#%% Plot results by AT

def PlotResultsByAT(meta,mosAT,t_start,t_end,iBAU,iCAP):

    iB=0
    iP=1

    # Import ylabels
    ylabs=gu.ReadExcel(meta['Paths']['Model Code'] + '\\Parameters\\LabelYAxis.xlsx')

    tv=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)

    iT=np.where( (tv>=t_start) & (tv<=t_end) )[0]

    cl_b=[0,0,0.5];
    cl_p=[1,0.2,0.2];

    cl_b2=[0,0,0.5];
    cl_p2=[1,0.2,0.2];

    cl_d=[0.4,0.8,0]
    cl_d2=[0.2,0.4,0]

    Alpha=0.08

    ListOp=['Sum','Mean']
    #ListOp=['Sum']

    for op in ListOp:

        if op=='Sum':
            ddb=meta['Project']['Display divide by']
        else:
            ddb=1.0

        # Full list of variables
        #ListV=['A', 'V_StemMerch', 'V_StemMerchToMill', 'LogSizeEnhancement', 'C_Biomass_Tot', 'C_Piled_Tot', 'C_Litter_Tot', 'C_DeadWood_Tot', 'C_Soil_Tot', 'C_InUse_Tot', 'C_DumpLandfill_Tot', 'C_M_Dist', 'C_G_Gross_Tot', 'C_G_Net_Tot', 'C_M_Reg_Tot', 'C_LF_Tot', 'C_RH_Tot', 'C_ToMillMerch', 'C_ToMillNonMerch', 'C_ToMillSnagStem', 'C_ToSlashpileBurn', 'C_ToLumber', 'C_ToPlywood', 'C_ToOSB', 'C_ToMDF', 'C_ToPaper', 'C_ToPowerFacilityDom', 'C_ToPowerFacilityFor', 'C_ToPowerGrid', 'C_ToPellets', 'C_ToFirewoodDom', 'C_ToFirewoodFor', 'C_ToLogExport', 'E_CO2e_LULUCF_NEE', 'E_CO2e_LULUCF_Wildfire', 'E_CO2e_LULUCF_OpenBurning', 'E_CO2e_LULUCF_EcoOther', 'E_CO2e_LULUCF_HWP', 'E_CO2e_ESC_Comb', 'E_CO2e_ESC_SubE', 'E_CO2e_ESC_SubBM', 'E_CO2e_ET_Comb', 'E_CO2e_IPPU_Comb', 'C_NPP_Tot', 'C_ToMill', 'E_CO2e_LULUCF_Fire', 'E_CO2e_AGHGB_WSub', 'E_CO2e_AGHGB_WOSub', 'E_CO2e_AGHGB_WSub_cumu', 'E_CO2e_AGHGB_WOSub_cumu', 'E_CO2e_AGHGB_WSub_cumu_from_tproj', 'E_CO2e_AGHGB_WOSub_cumu_from_tproj', 'Yield Lumber', 'Yield Plywood', 'Yield OSB', 'Yield MDF', 'Yield Paper', 'Yield Pellets', 'Yield PowerGrid', 'Yield PowerFacilityDom', 'Yield FirewoodDom', 'Yield LogExport', 'Price Lumber', 'Price Plywood', 'Price OSB', 'Price MDF', 'Price Newsprint', 'Price PowerFacilityDom', 'Price PowerGrid', 'Price Pellets', 'Price LogExport', 'Price FirewoodDom', 'Exchange Rate US', 'Exchange Rate Euro', 'Cost Roads', 'Cost Harvest Overhead', 'Cost Harvest Felling and Piling', 'Cost Harvest Hauling', 'Cost Harvest Residuals', 'Cost Milling', 'Cost Nutrient Management', 'Cost Planting', 'Cost Survey', 'Cost Knockdown', 'Cost Ripping', 'Cost Slashpile Burn', 'Harvest Vol Merch', 'Harvest Vol Resid', 'Cost Total', 'Revenue Lumber', 'Revenue Plywood', 'Revenue OSB', 'Revenue MDF', 'Revenue Paper', 'Revenue PowerFacilityDom', 'Revenue PowerGrid', 'Revenue Pellets', 'Revenue FirewoodDom', 'Revenue LogExport', 'Revenue Gross', 'Revenue Net', 'Revenue Net Disc', 'Revenue Gross Disc', 'Cost Total Disc', 'Revenue Gross Disc_cumu', 'Revenue Net Disc_cumu', 'Cost Total Disc_cumu', 'Cost Total_cumu', 'Revenue Gross_cumu', 'Revenue Net_cumu']

        # Condensed list of variables
        ListV=['A', 'V_StemMerch', 'V_StemMerchToMill','C_Biomass_Tot','C_ToSlashpileBurn','E_CO2e_LULUCF_NEE', 'E_CO2e_LULUCF_Wildfire', 'E_CO2e_LULUCF_OpenBurning','C_NPP_Tot','E_CO2e_AGHGB_WSub','E_CO2e_AGHGB_WOSub', 'E_CO2e_AGHGB_WSub_cumu', 'E_CO2e_AGHGB_WOSub_cumu', 'E_CO2e_AGHGB_WSub_cumu_from_tproj', 'E_CO2e_AGHGB_WOSub_cumu_from_tproj', 'Harvest Vol Merch', 'Harvest Vol Resid', 'Cost Total','Revenue Gross', 'Revenue Net', 'Revenue Net Disc', 'Revenue Gross Disc', 'Cost Total Disc', 'Revenue Gross Disc_cumu', 'Revenue Net Disc_cumu', 'Cost Total Disc_cumu', 'Cost Total_cumu', 'Revenue Gross_cumu', 'Revenue Net_cumu','Cost Silviculture Total','Cost Silviculture Total Disc','Cost Silviculture Total Disc_cumu']

        for iV in range(len(ListV)):

            k=ListV[iV]

            for iAT in range(meta['Project']['Activities']['Activity ID'].size):

                nameAT=meta['Project']['Activities']['Activity description'][iAT]

                # An activity may not be in BAU portfolio
                try:
                    be_b_bau=mosAT[iBAU][iAT][iB][op][k]['Ensemble Mean'][iT]/ddb
                    #lo_b_bau=mosAT[iBAU][iAT][iB][op][k]['Ensemble CIL'][iT]/ddb
                    #hi_b_bau=mosAT[iBAU][iAT][iB][op][k]['Ensemble CIU'][iT]/ddb

                    be_p_bau=mosAT[iBAU][iAT][iP][op][k]['Ensemble Mean'][iT]/ddb
                    #lo_p_bau=mosAT[iBAU][iAT][iP][op][k]['Ensemble CIL'][iT]/ddb
                    #hi_p_bau=mosAT[iBAU][iAT][iP][op][k]['Ensemble CIU'][iT]/ddb

                    be_d_bau=be_p_bau-be_b_bau
                    #lo_d_bau,hi_d_bau=gu.GetCIsFromDifference(lo_b_bau,hi_b_bau,lo_p_bau,hi_p_bau)
                except:
                    be_b_bau=np.zeros(iT.size)
                    be_p_bau=np.zeros(iT.size)
                    be_d_bau=np.zeros(iT.size)

                # An activity may not be in CAP portfolio
                try:
                    be_b_cap=mosAT[iCAP][iAT][iB][op][k]['Ensemble Mean'][iT]/ddb
                    #lo_b_cap=mosAT[iCAP][iAT][iB][op][k]['Ensemble CIL'][iT]/ddb
                    #hi_b_cap=mosAT[iCAP][iAT][iB][op][k]['Ensemble CIU'][iT]/ddb

                    be_p_cap=mosAT[iCAP][iAT][iP][op][k]['Ensemble Mean'][iT]/ddb
                    #lo_p_cap=mosAT[iCAP][iAT][iP][op][k]['Ensemble CIL'][iT]/ddb
                    #hi_p_cap=mosAT[iCAP][iAT][iP][op][k]['Ensemble CIU'][iT]/ddb

                    be_d_cap=be_p_cap-be_b_cap
                    #lo_d_cap,hi_d_cap=gu.GetCIsFromDifference(lo_b_cap,hi_b_cap,lo_p_cap,hi_p_cap)

                    be_p_d=be_p_cap-be_p_bau
                    #lo_p_d,hi_p_d=gu.GetCIsFromDifference(lo_p_bau,hi_p_bau,lo_p_cap,hi_p_cap)
                except:
                    be_b_bau=np.zeros(iT.size)
                    be_p_bau=np.zeros(iT.size)
                    be_d_bau=np.zeros(iT.size)

                # y-axis label
                ind=np.where(ylabs['Name']==k)[0]
                if ind.size==1:
                    if (op=='Sum') & (ddb==1e6):
                        lab=ylabs['Y Label Sum Mt'][ind[0]]
                    elif (op=='Sum') & (ddb==1e9):
                        lab=ylabs['Y Label Sum Gt'][ind[0]]
                    else:
                        lab=ylabs['Y Label Mean'][ind[0]]
                else:
                    lab=k

                # Plot
                plt.close('all')
                fig,ax=plt.subplots(1,3,figsize=gu.cm2inch(17,7.5))

                ax[0].plot(tv[iT],np.zeros(iT.size),'-',lw=2,color=[0.8,0.8,0.8])
                #ax[0].fill_between(tv[iT],lo_b_bau,hi_b_bau,color=cl_b,alpha=Alpha,lw=0)
                #ax[0].fill_between(tv[iT],lo_p_bau,hi_p_bau,color=cl_p,alpha=Alpha,lw=0)
                #ax[0].fill_between(tv[iT],lo_b_cap,hi_b_cap,color=cl_b,alpha=Alpha,lw=0)
                #ax[0].fill_between(tv[iT],lo_p_cap,hi_p_cap,color=cl_p,alpha=Alpha,lw=0)
                ax[0].plot(tv[iT],be_b_bau,'-',color=cl_b,label='BAU Baseline')
                ax[0].plot(tv[iT],be_p_bau,'-.',color=cl_p,label='BAU Project')
                ax[0].plot(tv[iT],be_b_cap,'--',color=cl_b2,label='CAP Baseline')
                ax[0].plot(tv[iT],be_p_cap,':',color=cl_p2,label='CAP Project')
                ax[0].set(position=[0.06,0.08,0.27,0.7],xlim=[tv[iT[0]],tv[iT[-1]]],ylabel=lab)
                ax[0].legend(loc='upper center',ncol=2,bbox_to_anchor=(0.5,1.14),frameon=False,facecolor=None)
                ax[0].yaxis.set_ticks_position('both');
                ax[0].xaxis.set_ticks_position('both')

                ax[1].plot(tv[iT],np.zeros(iT.size),'-',lw=2,color=[0.8,0.8,0.8])
                #ax[1].fill_between(tv[iT],lo_d_bau,hi_d_bau,color=cl_d,alpha=Alpha,lw=0)
                #ax[1].fill_between(tv[iT],lo_d_cap,hi_d_cap,color=cl_d2,alpha=Alpha,lw=0)
                ax[1].plot(tv[iT],be_d_bau,'-',color=cl_d,label='BAU Project minus BAU Baseline ')
                ax[1].plot(tv[iT],be_d_cap,'--',color=cl_d2,label='CAP Project minus CAP Baseline')
                ax[1].legend(loc='upper center',bbox_to_anchor=(0.5,1.15),frameon=False,facecolor=None)
                ax[1].set(position=[0.4,0.08,0.27,0.7],xlim=[tv[iT[0]],tv[iT[-1]]],ylabel='$\Delta$ ' + lab)
                ax[1].yaxis.set_ticks_position('both');
                ax[1].xaxis.set_ticks_position('both')

                ax[2].plot(tv[iT],np.zeros(iT.size),'-',lw=2,color=[0.8,0.8,0.8])
                #ax[2].fill_between(tv[iT],lo_p_d,hi_p_d,color=cl_d,alpha=Alpha,lw=0)
                #ax[2].plot(tv[iT],be_p_d,'-',color=cl_d,label='CAP Project minus BAU Project')
                ax[2].plot(tv[iT],be_d_cap-be_d_bau,'-',color=cl_d,label='$\Delta$ CAP minus $\Delta$ BAU')
                ax[2].legend(loc='upper center',bbox_to_anchor=(0.5,1.15),frameon=False,facecolor=None)
                ax[2].set(position=[0.75,0.08,0.27,0.7],xlim=[tv[iT[0]],tv[iT[-1]]],ylabel='$\Delta$ ' + lab)
                ax[2].yaxis.set_ticks_position('both');
                ax[2].xaxis.set_ticks_position('both')

                gu.PrintFig(meta['Paths']['Figures'] + '\\ByAT_' + nameAT + '_' + op + '_' + k,'png',200)
                #gu.PrintFig(meta['Paths']['Figures'] + '\\ByAT_' + op + '_' + k,'png',200)
    plt.close('all')

    return