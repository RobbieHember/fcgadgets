#%% Import modules
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import geopandas as gpd
import os
import glob
import openpyxl
import copy
import gc as garc
import time
from shapely.geometry import Point
import scipy.stats as stats
from fcgadgets.macgyver import util_general as gu
from fcgadgets.macgyver import util_gis as gis
import fcgadgets.cbrunner.cbrun_postprocess as post
from fcgadgets.macgyver import util_fcs_graphs as ufc
from fcgadgets.hardhat import economics as econo
from fcgadgets.taz import default_stat_event_sim as asm

#%% Configure project
def ImportProjectConfig(meta,pNam,**kwargs):
	#--------------------------------------------------------------------------
	# Initialize nested dictionaries
	#--------------------------------------------------------------------------
	if 'Core' not in meta:
		meta['Core']={}
	if 'Param' not in meta:
		meta['Param']={}
	if 'Modules' not in meta:
		meta['Modules']={}

	if pNam not in meta:
		meta[pNam]={}

	if 'Project' not in meta[pNam]:
		meta[pNam]['Project']={}

	if 'Scenario' not in meta[pNam]:
		meta[pNam]['Scenario']={}

	#--------------------------------------------------------------------------
	# Import project parameters from spreadsheet
	#--------------------------------------------------------------------------
	df=pd.read_excel(meta['Paths'][pNam]['Data'] + '\\Inputs\\ProjectConfig.xlsx',sheet_name='Project')
	for i in range(len(df)):
		Name=df['Name'].iloc[i]
		Value=df['Value'].iloc[i]
		if Name[-1]==':':
			# Exclude headers
			continue
		meta[pNam]['Project'][Name]=Value

	#--------------------------------------------------------------------------
	# Import look-up tables
	#--------------------------------------------------------------------------
	meta=Load_LUTs_Modelling(meta)

	#--------------------------------------------------------------------------
	# Define pool names
	#--------------------------------------------------------------------------
	# Pool names (ecosystem)
	# *** If you change this, you need to change the same list in "Update Parameters" ***
	meta['Core']['Name Pools Eco']=['StemMerch','StemNonMerch','Foliage','Branch','Bark','RootCoarse','RootFine', \
		'PiledStemMerch','PiledStemNonMerch','PiledBranch','PiledBark','PiledFoliage','PiledDeadStem','PiledDeadBranch', \
		'LitterVF','LitterF','LitterM','LitterS', \
		'DeadStemMerch','DeadStemNonMerch','DeadBark','DeadBranch', \
		'SoilVF','SoilF','SoilS']

	meta['Core']['Name Pools Dead']=['DeadStemMerch','DeadStemNonMerch','DeadBark','DeadBranch',
		'LitterVF','LitterF','LitterM','LitterS','SoilVF','SoilF','SoilS',
		'PiledStemMerch','PiledStemNonMerch','PiledBranch','PiledBark',
		'PiledFoliage','PiledDeadStem','PiledDeadBranch']

	# Number of ecosystem pools
	meta['Core']['N Pools Eco']=len(meta['Core']['Name Pools Eco'])

	# Pool names (products)
	meta['Core']['Name Pools Pro']=['SFH','MFH','Comm','Furn','Ship','Repairs', \
		'Other','Paper','EffluentPulp','DumpWood','DumpPaper', \
		'LandfillWoodDegradable','LandfillWoodNonDegradable', \
		'LandfillPaperDegradable','LandfillPaperNonDegradable']

	# Number of product pools
	meta['Core']['N Pools Pro']=len(meta['Core']['Name Pools Pro'])

	#--------------------------------------------------------------------------
	# Define indices to each pool
	#--------------------------------------------------------------------------
	# Indices to ecosystem pools pools
	meta['Core']['iEP']={}; cnt=0
	for nam in meta['Core']['Name Pools Eco']:
		meta['Core']['iEP'][nam]=cnt
		cnt=cnt+1
	iEP=meta['Core']['iEP']
	meta['Core']['iEP']['StemTotal']=np.array([iEP['StemMerch'],iEP['StemNonMerch']])
	meta['Core']['iEP']['BiomassTotal']=np.array([iEP['StemMerch'],iEP['StemNonMerch'],iEP['Foliage'],iEP['Branch'],iEP['Bark'],iEP['RootCoarse'],iEP['RootFine']])
	meta['Core']['iEP']['BiomassAboveground']=np.array([iEP['StemMerch'],iEP['StemNonMerch'],iEP['Foliage'],iEP['Branch'],iEP['Bark']])
	meta['Core']['iEP']['BiomassBelowground']=np.array([iEP['RootCoarse'],iEP['RootFine']])
	meta['Core']['iEP']['DeadWood']=np.array([iEP['DeadStemMerch'],iEP['DeadStemNonMerch'],iEP['DeadBark'],iEP['DeadBranch'],iEP['LitterM'],iEP['SoilF']])
	meta['Core']['iEP']['Litter']=np.array([iEP['LitterVF'],iEP['LitterF'],iEP['LitterS']])
	meta['Core']['iEP']['Piled']=np.array([iEP['PiledStemMerch'],iEP['PiledStemNonMerch'],iEP['PiledBranch'],iEP['PiledBark'],iEP['PiledFoliage'],iEP['PiledDeadStem'],iEP['PiledDeadBranch'],])
	#meta['Core']['iEP']['Soil']=np.array([iEP['SoilVF'],iEP['SoilF'],iEP['SoilS']])
	meta['Core']['iEP']['Soil']=np.array([iEP['SoilVF'],iEP['SoilS']])
	meta['Core']['iEP']['Dead']=np.array([iEP['DeadStemMerch'],iEP['DeadStemNonMerch'],iEP['DeadBark'],iEP['DeadBranch'],
									   iEP['PiledStemMerch'],iEP['PiledStemNonMerch'],iEP['PiledBranch'],iEP['PiledBark'],
									   iEP['LitterVF'],iEP['LitterF'],iEP['LitterM'],iEP['LitterS'],
									   iEP['SoilVF'],iEP['SoilF'],iEP['SoilS']])

	# Indices to produce pools pools
	meta['Core']['iPP']={}; cnt=0
	for nam in meta['Core']['Name Pools Pro']:
		meta['Core']['iPP'][nam]=cnt
		cnt=cnt+1
	iPP=meta['Core']['iPP']
	meta['Core']['iPP']['InUse']=np.array([ iPP['SFH'],iPP['MFH'],iPP['Comm'],iPP['Furn'],iPP['Ship'],iPP['Repairs'],iPP['Other'],iPP['Paper'] ])
	meta['Core']['iPP']['Buildings']=np.array([ iPP['SFH'],iPP['MFH'],iPP['Comm'] ])
	meta['Core']['iPP']['WasteSystems']=np.array([ iPP['DumpWood'],iPP['DumpPaper'], \
		iPP['LandfillWoodDegradable'],iPP['LandfillWoodNonDegradable'], \
		iPP['LandfillPaperDegradable'],iPP['LandfillPaperNonDegradable'],
		iPP['EffluentPulp'] ])

	#--------------------------------------------------------------------------
	# Maximum number of events per year
	# 8 appears to be sufficient but this may need to be changed for some
	# special projects
	#--------------------------------------------------------------------------
	meta['Core']['Max Events Per Year']=8

	#--------------------------------------------------------------------------
	# If Early Record Recycling is on, include a buffer before t_StartSaving
	#--------------------------------------------------------------------------
	meta['Core']['Recycle Early Record Buffer']=200

	#--------------------------------------------------------------------------
	# Define time
	#--------------------------------------------------------------------------
	# Calendar year
	meta[pNam]['Year']=np.arange(meta[pNam]['Project']['Year Start'],meta[pNam]['Project']['Year End']+1,1)

	meta[pNam]['Project']['N Time']=meta[pNam]['Year'].size

	#--------------------------------------------------------------------------
	# Define spatial domain for geospatial projects
	#--------------------------------------------------------------------------
	if meta[pNam]['Project']['Scenario Source']=='Script':
		
		# Import land cover class
		zLCC1=gis.OpenGeoTiff(meta['Paths']['bc1ha'] + '\\LandCoverUse\\LandCover_Comp1_2019.tif')

		#gdf_tsa=gpd.read_file(meta['Paths']['GDB']['GDB'] + '\\LandUse\\tsa.geojson')
		gdf_bcb=gpd.read_file(meta['Paths']['GDB']['LandUse'],layer='NRC_POLITICAL_BOUNDARIES_1M_SP')

		# Import mask (0=excluded, 1=included)
		if meta[pNam]['Project']['ROI Source']=='Province':
			# Land mask for BC
			zMask=gis.OpenGeoTiff(meta['Paths']['bc1ha Ref Grid'])

		elif meta[pNam]['Project']['ROI Source']=='TSA':
			# Mask from timber supply area
			zTSA=gis.OpenGeoTiff(meta['Paths']['bc1ha'] + '\\FADM_TSA\\TSA_NUMBER_DESCRIPTION.tif')
			zMask=copy.deepcopy(zLCC1)
			zMask['Data']=np.zeros(zLCC1['Data'].shape,'int8')
			ind=np.where( (zTSA['Data']==meta['LUT']['FADM_TSA']['TSA_NUMBER_DESCRIPTION'][ meta[pNam]['Project']['ROI Elements'] ]) )
			zMask['Data'][ind]=1

		elif meta[pNam]['Project']['ROI Source']=='Regional District':
			# Mask from regional district
			zMask=gis.OpenGeoTiff(meta['Paths']['bc1ha'] + '\\TA_REGIONAL_DISTRICTS_SVW\\REGIONAL_DISTRICT_NAME.tif')
			ind=np.where( (zMask['Data']!=meta['LUT']['TA_REGIONAL_DISTRICTS_SVW']['REGIONAL_DISTRICT_NAME'][meta[pNam]['Project']['ROI Elements']]) )
			zMask['Data'][ind]=0
			ind=np.where( (zMask['Data']==meta['LUT']['TA_REGIONAL_DISTRICTS_SVW']['REGIONAL_DISTRICT_NAME'][meta[pNam]['Project']['ROI Elements']]) )
			zMask['Data'][ind]=1

		elif meta[pNam]['Project']['ROI Source']=='BGC Zone':
			# Mask from BGC Zone
			zMask=gis.OpenGeoTiff(meta['Paths']['bc1ha'] + '\\BEC_BIOGEOCLIMATIC_POLY\\ZONE.tif')
			ind=np.where( (zMask['Data']!=meta['LUT']['BEC_BIOGEOCLIMATIC_POLY']['ZONE'][meta[pNam]['Project']['ROI Elements']]) )
			zMask['Data'][ind]=0
			ind=np.where( (zMask['Data']==meta['LUT']['BEC_BIOGEOCLIMATIC_POLY']['ZONE'][meta[pNam]['Project']['ROI Elements']]) )
			zMask['Data'][ind]=1

		elif meta[pNam]['Project']['ROI Source']=='BCFCS_LUC':
			# Land use change
			zMask=gis.OpenGeoTiff(meta['Paths']['bc1ha'] + '\\Disturbances\\LandCover_Comp1_DeforestationMaskAll.tif')
		
		elif meta[pNam]['Project']['ROI Source']=='BCFCS_NMC':
			# Nutrient management
			zMask=gis.OpenGeoTiff(meta['Paths']['bc1ha'] + '\\RSLT_ACTIVITY_TREATMENT_SVW\\FE-CA_MaskAll.tif')

		elif meta[pNam]['Project']['ROI Source']=='BCFCS_NOSE':
			# Non-obligation stand establishment
			zMask=gis.OpenGeoTiff(meta['Paths']['bc1ha'] + '\\RSLT_ACTIVITY_TREATMENT_SVW\\PL_NonOb_MaskAll.tif')

		elif meta[pNam]['Project']['ROI Source']=='BCFCS_NMC':
			# Nutrient management
			zMask=gis.OpenGeoTiff(meta['Paths']['bc1ha'] + '\\RSLT_ACTIVITY_TREATMENT_SVW\\FE-CA_MaskAll.tif')

		elif meta[pNam]['Project']['ROI Source']=='Landscape_NicolaRiverWatershed':
			zMask=gis.OpenGeoTiff(meta['Paths']['bc1ha'] + '\\Masks\\Nicola River.tif')

		elif meta[pNam]['Project']['ROI Source']=='EvalAtPlots':
			gplts=gu.ipickle(r'C:\Users\rhember\Documents\Data\GroundPlots\PSP-NADB2\Processed\L2\L2_BC.pkl')
			soils=gu.ipickle(r'C:\Users\rhember\Documents\Data\Soils\Shaw et al 2018 Database\SITES.pkl')
			x=np.append(gplts['sobs']['X'],soils['x'])
			y=np.append(gplts['sobs']['Y'],soils['y'])
			xy=np.unique(np.column_stack((x,y)),axis=0)
			zMask=zLCC1.copy()
			zMask['Data']=np.zeros(zLCC1['Data'].shape,dtype='int16')
			ind=gis.GetGridIndexToPoints(zLCC1,xy[:,0],xy[:,1])
			zMask['Data'][ind]=1
		
		elif meta[pNam]['Project']['ROI Source']=='EvalAtCN':
			gplts=gu.ipickle(r'C:\Users\rhember\Documents\Data\GroundPlots\PSP-NADB2\Processed\L2\L2_BC.pkl')
			ind=np.where(gplts['sobs']['PTF CN']==1)[0]
			x=gplts['sobs']['X'][ind]
			y=gplts['sobs']['Y'][ind]
			xy=np.unique(np.column_stack((x,y)),axis=0)
			zMask=zLCC1.copy()
			zMask['Data']=np.zeros(zLCC1['Data'].shape,dtype='int16')
			ind=gis.GetGridIndexToPoints(zLCC1,xy[:,0],xy[:,1])
			zMask['Data'][ind]=1
		
		elif meta[pNam]['Project']['ROI Source']=='EvalCoast':
			
			gplts=gu.ipickle(r'C:\Users\rhember\Documents\Data\GroundPlots\PSP-NADB2\Processed\L2\L2_BC.pkl')
			ind=np.where( (gplts['sobs']['PTF CN']==1) & (gplts['sobs']['Ecozone BC L1']==meta['LUT']['GP']['Ecozone BC L1']['CWH']) | \
						 (gplts['sobs']['PTF CN']==1) & (gplts['sobs']['Ecozone BC L1']==meta['LUT']['GP']['Ecozone BC L1']['MH']) )[0]
			x=gplts['sobs']['X'][ind]
			y=gplts['sobs']['Y'][ind]
			xy=np.unique(np.column_stack((x,y)),axis=0)
			zMask=zLCC1.copy()
			zMask['Data']=np.zeros(zLCC1['Data'].shape,dtype='int16')
			ind=gis.GetGridIndexToPoints(zLCC1,xy[:,0],xy[:,1])
			zMask['Data'][ind]=1
		
		elif meta[pNam]['Project']['ROI Source']=='EvalInterior':
			gplts=gu.ipickle(r'C:\Users\rhember\Documents\Data\GroundPlots\PSP-NADB2\Processed\L2\L2_BC.pkl')
			ind=np.where( (gplts['sobs']['PTF CN']==1) & (gplts['sobs']['Ecozone BC L1']!=meta['LUT']['GP']['Ecozone BC L1']['CWH']) & (gplts['sobs']['Ecozone BC L1']!=meta['LUT']['GP']['Ecozone BC L1']['MH']) )[0]
			x=gplts['sobs']['X'][ind]
			y=gplts['sobs']['Y'][ind]
			xy=np.unique(np.column_stack((x,y)),axis=0)
			zMask=zLCC1.copy()
			zMask['Data']=np.zeros(zLCC1['Data'].shape,dtype='int16')
			ind=gis.GetGridIndexToPoints(zLCC1,xy[:,0],xy[:,1])
			zMask['Data'][ind]=1
		
		elif meta[pNam]['Project']['ROI Source']=='BCFCS_Eval':
			# Add grid
			zMask0=gis.OpenGeoTiff(meta['Paths']['bc1ha Ref Grid'])
			zMask0['Data'][0::200,0::200]=zMask0['Data'][0::200,0::200]+1
			zMask=gis.OpenGeoTiff(meta['Paths']['bc1ha Ref Grid'])
			zMask['Data']=0*zMask['Data']
			ind=np.where(zMask0['Data']==2)
			zMask['Data'][ind]=1
			
			# Add plots
			gpt=gu.ipickle(r'C:\Users\rhember\Documents\Data\GroundPlots\PSP-NADB2\Processed\L2\L2_BC.pkl')['sobs']
			ikp=np.where(gpt['PTF CN']==1)[0]
			x=gpt['X'][ikp]
			y=gpt['Y'][ikp]
			#soc=gu.ipickle(r'C:\Users\rhember\Documents\Data\Soils\Shaw et al 2018 Database\SITES.pkl')
			#x=np.append(gpt['X'],soc['x'])
			#y=np.append(gpt['Y'],soc['y'])
			xy=np.unique(np.column_stack((x,y)),axis=0)
			ind=gis.GetGridIndexToPoints(zMask,xy[:,0],xy[:,1])
			zMask['Data'][ind]=1
			MaskGP=np.zeros(zMask['Data'].shape,dtype='int8')
			MaskGP[ind]=1
			
			zLCC1['Data'][ind]=meta['LUT']['Derived']['lc_comp1']['Forest']
			
		else:
			print('ROI Source not recognized.')
			pass

		# Initialize geosptial info
		if 'Geos' not in meta.keys():
			meta['Geos']={}

		meta['Geos']['RGSF']=meta[pNam]['Project']['Regular grid sampling frequency (ha)']

		# Extract subgrid
		meta['Geos']['Grid']=zMask.copy()
		meta['Geos']['Grid']['Data']=np.zeros((zMask['Data'].shape),dtype='int8')
		meta['Geos']['Grid']=gis.UpdateGridCellsize(meta['Geos']['Grid'],meta['Geos']['RGSF'])

		# Resample required grids
		zMask_r=gis.UpdateGridCellsize(zMask,meta['Geos']['RGSF'])
		zLCC1_r=gis.UpdateGridCellsize(zLCC1,meta['Geos']['RGSF'])

		# Define additional sampling criteria
		if meta[pNam]['Project']['Land Cover Scope']=='Forest':

			iMask_Full=np.where( (zMask['Data']==1) & (zLCC1['Data']==meta['LUT']['Derived']['lc_comp1']['Forest']) )
			meta['Geos']['iMask']=np.where( (zMask_r['Data']==1) & (zLCC1_r['Data']==meta['LUT']['Derived']['lc_comp1']['Forest']) )

		elif meta[pNam]['Project']['Land Cover Scope']=='All Land':

			iMask_Full=np.where( (zMask['Data']==1) & (zLCC1['Data']>0) & (zLCC1['Data']!=meta['LUT']['Derived']['lc_comp1']['Water']) )
			meta['Geos']['iMask']=np.where( (zMask_r['Data']==1) & (zLCC1_r['Data']>0) & (zLCC1_r['Data']!=meta['LUT']['Derived']['lc_comp1']['Water']) )

		else:
			print('Land Cover Scope not recognized.')

		# Area expansion factor
		meta['Geos']['AEF']=iMask_Full[0].size/meta['Geos']['iMask'][0].size
		#meta['Geos']['AEF']=bc_grid_size/(meta['Geos']['Grid']['m']*meta['Geos']['Grid']['n'])
		meta[pNam]['Project']['AEF']=meta['Geos']['AEF']

		# Revise mask
		meta['Geos']['Grid']['Data'][meta['Geos']['iMask']]=1

		# Generate sparse grid
		meta['Geos']['Sparse']={}
		meta['Geos']['Sparse']['X']=meta['Geos']['Grid']['X'][meta['Geos']['iMask']]
		meta['Geos']['Sparse']['Y']=meta['Geos']['Grid']['Y'][meta['Geos']['iMask']]
		meta['Geos']['Sparse']['ID_Admin']=zMask_r['Data'][meta['Geos']['iMask']]
		if meta[pNam]['Project']['ROI Source']=='BCFCS_Eval':
			meta['Geos']['Sparse']['GP']=MaskGP[meta['Geos']['iMask']]

		# Save to pickle file
		# Flatten coordinate matrices first to save space
		meta['Geos']['Grid']['X']=meta['Geos']['Grid']['X'][0,:]
		meta['Geos']['Grid']['Y']=meta['Geos']['Grid']['Y'][:,0]
		gu.opickle(meta['Paths'][pNam]['Data'] + '\\geos.pkl',meta['Geos'])

		print('Number of stands = ' + str(meta['Geos']['Sparse']['X'].size))

		# Save sparse points to geojson
		flg=1
		if flg==1:
			points=[]
			for k in range(meta['Geos']['Sparse']['X'].size):
				points.append(Point(meta['Geos']['Sparse']['X'][k],meta['Geos']['Sparse']['Y'][k]))
			gdf_sxy=gpd.GeoDataFrame({'geometry':points,'ID_Admin':meta['Geos']['Sparse']['ID_Admin']})
			gdf_sxy.crs=meta['Geos']['crs']
			gdf_sxy.to_file(meta['Paths'][pNam]['Data'] + '\\geos.geojson',driver='GeoJSON')

		# Plot map
		ufc.PlotMapOfSparseXYSample(meta,pNam)

	else:
		# From spreadsheet - a few graphics require AEF, set to 1.0
		meta[pNam]['Project']['AEF']=1

	#--------------------------------------------------------------------------
	# Dimensions of simulation
	#--------------------------------------------------------------------------
	# Number of stands
	if meta[pNam]['Project']['Scenario Source']=='Spreadsheet':
		meta[pNam]['Project']['N Stand']=meta[pNam]['Project']['Batch Interval']
	elif meta[pNam]['Project']['Scenario Source']=='Portfolio':
		meta[pNam]['Project']['N Stand']=meta[pNam]['Project']['N Stand per Activity Type']*meta[pNam]['Project']['AIL']['N AT']*meta[pNam]['Project']['AIL']['N Years']
	elif meta[pNam]['Project']['Scenario Source']=='Script':
		meta[pNam]['Project']['N Stand']=meta['Geos']['Sparse']['X'].size

	# Number of batches
	meta[pNam]['Project']['N Batch']=np.ceil(meta[pNam]['Project']['N Stand']/meta[pNam]['Project']['Batch Interval']).astype(int)

	# Initialize list that can keep track of batch sizes
	meta[pNam]['Project']['Batch Size']=[None]*meta[pNam]['Project']['N Batch']
	for iBat in range(meta[pNam]['Project']['N Batch']):
		meta[pNam]['Project']['Batch Size'][iBat]=IndexToBatch(meta[pNam],iBat).size

	#--------------------------------------------------------------------------
	# Import model parameters
	#--------------------------------------------------------------------------
	meta=ImportParameters(meta)

	#--------------------------------------------------------------------------
	# Define scenario parameters
	#--------------------------------------------------------------------------
	if meta[pNam]['Project']['Scenario Source']!='Portfolio':
		df=pd.read_excel(meta['Paths'][pNam]['Data'] + '\\Inputs\\ProjectConfig.xlsx',sheet_name='Scenarios') # ,usecols='A:OM'
		df=df.iloc[:,df.iloc[0,:].isnull().values==False]
		meta[pNam]['Scenario']=list()
		for i in range(1,df.shape[1]):
			pScn0={}
			for j in range(df.shape[0]):
				if df.iloc[j,0][-1]==':':
					# Exclude headers
					continue
				pScn0.update({df.iloc[j,0]:df.iat[j,i]})
			meta[pNam]['Scenario'].append(pScn0)

		# Number of scenarios
		meta[pNam]['Project']['N Scenario']=np.sum([i['Scenario Status']=='On' for i in meta[pNam]['Scenario']])

	# #--------------------------------------------------------------------------
	# # Number of Land Surface Scenarios
	# #--------------------------------------------------------------------------
	# meta[pNam]['Project']['LSC']={}
	# if meta[pNam]['Project']['Scenario Source']=='Script':
	#	 meta[pNam]['Project']['LSC']['Scenario Names Unique']=np.unique(np.array([meta[pNam]['Scenario'][i]['Land Surface Scenario'] for i in range(len(meta[pNam]['Scenario']))],dtype=object))
	#	 meta[pNam]['Project']['LSC']['N Scenario']=meta[pNam]['Project']['LSC']['Scenario Names Unique'].size
	# else:
	#	 meta[pNam]['Project']['LSC']['Scenario Names Unique']=0
	#	 meta[pNam]['Project']['LSC']['N Scenario']=0

	#--------------------------------------------------------------------------
	# Define strata for analyzing results (optional)
	#--------------------------------------------------------------------------
	meta[pNam]['Project']['Strata']={}
	meta[pNam]['Project']['Strata']['Project Type']={}
	meta[pNam]['Project']['Strata']['Project Type']['Unique CD']=np.array(['All'],dtype=object)
	meta[pNam]['Project']['Strata']['Project Type']['Unique ID']=np.array([0],dtype='int32')
	meta[pNam]['Project']['Strata']['Project Type']['ID']=np.zeros(meta[pNam]['Project']['N Stand'],dtype='int32')
	meta[pNam]['Project']['Strata']['Spatial']={}
	meta[pNam]['Project']['Strata']['Spatial']['Unique CD']=np.array(['All'],dtype=object)
	meta[pNam]['Project']['Strata']['Spatial']['Unique ID']=np.array([0],dtype='int32')
	meta[pNam]['Project']['Strata']['Spatial']['ID']=np.zeros(meta[pNam]['Project']['N Stand'],dtype='int32')
	meta[pNam]['Project']['Strata']['Year']={}
	meta[pNam]['Project']['Strata']['Year']['Unique CD']=np.array(['All'],dtype=object)
	meta[pNam]['Project']['Strata']['Year']['Unique ID']=np.array([0],dtype='int32')
	meta[pNam]['Project']['Strata']['Year']['ID']=np.zeros(meta[pNam]['Project']['N Stand'],dtype='int32')
	meta[pNam]['Project']['Strata']['Other']={}
	meta[pNam]['Project']['Strata']['Other']['Unique CD']=np.array(['All'],dtype=object)
	meta[pNam]['Project']['Strata']['Other']['Unique ID']=np.array([0],dtype='int32')
	meta[pNam]['Project']['Strata']['Other']['ID']=np.zeros(meta[pNam]['Project']['N Stand'],dtype='int32')

	#--------------------------------------------------------------------------
	# Initialize project folders if they do not exist
	#--------------------------------------------------------------------------
	meta['Paths'][pNam]['Input Scenario']=[]
	meta['Paths'][pNam]['Output Scenario']=[]
	for iScn in range(0,meta[pNam]['Project']['N Scenario']):
		meta['Paths'][pNam]['Input Scenario'].append(meta['Paths'][pNam]['Data'] + '\\Inputs\\Scenario' + FixFileNum(iScn))
		if os.path.exists(meta['Paths'][pNam]['Input Scenario'][iScn])==False:
			os.mkdir(meta['Paths'][pNam]['Input Scenario'][iScn])
		meta['Paths'][pNam]['Output Scenario'].append(meta['Paths'][pNam]['Data'] + '\\Outputs\\Scenario' + FixFileNum(iScn))
		if os.path.exists(meta['Paths'][pNam]['Output Scenario'][iScn])==False:
			os.mkdir(meta['Paths'][pNam]['Output Scenario'][iScn])

	#--------------------------------------------------------------------------
	# Scale factors
	#--------------------------------------------------------------------------
	# *** Scale factor for saving results (this needs to be 100, 10 does not
	# capture carbon fluxes and it will affect GHG benefit estimates) ***
	# One variable ('CO2e_E_Products') requires the big one
	meta['Core']['Scale Factor Export Small']=0.001
	meta['Core']['Scale Factor Export Big']=0.001
	meta['Core']['Scale Factor C_M_DistByAgent']=0.1

	#--------------------------------------------------------------------------
	# Define strings that frequently need to be populated with zeros
	#--------------------------------------------------------------------------
	meta['Core']['StringsToFill']=['Month','Day','SILV_FUND_SOURCE_CODE','FIA_PROJECT_ID','OPENING_ID','ACTUAL_TREATMENT_AREA','ACTUAL_PLANTED_NUMBER', \
			'PL_SPECIES_CD1','PL_SPECIES_PCT1','PL_SPECIES_GW1','PL_SPECIES_CD2','PL_SPECIES_PCT2','PL_SPECIES_GW2', \
			'PL_SPECIES_CD3','PL_SPECIES_PCT3','PL_SPECIES_GW3','PL_SPECIES_CD4','PL_SPECIES_PCT4','PL_SPECIES_GW4', \
			'PL_SPECIES_CD5','PL_SPECIES_PCT5','PL_SPECIES_GW5']

	#--------------------------------------------------------------------------
	# Harvested wood product information
	# Year to start calling annual HWP methods - running it before 1800 is a
	# waste of time.
	#--------------------------------------------------------------------------
	meta['Core']['HWP Year Start']=1850

	#--------------------------------------------------------------------------
	# For tracking heterotrophic consumption
	#--------------------------------------------------------------------------
	meta[pNam]['Project']['HC']={}

	#--------------------------------------------------------------------------
	# Growth curve information
	#--------------------------------------------------------------------------
	meta['Modules']['GYM']={}
	meta['Modules']['GYM']['N Growth Curves']=5
	meta['Modules']['GYM']['ID GC Unique']=np.array([1,2,3,4,5])
	meta['Modules']['GYM']['BatchTIPSY Maximum Age']=200
	meta['Modules']['GYM']['BatchTIPSY Column Names']=['Age','VolTot0','VolMerch125',
		'VolMerch175','ODT_Bark','ODT_Branch','ODT_Foliage','ODT_Roots',
		'ODT_Stem','MortalityVolumeTotal']

	# GC Input file variables
	meta['Modules']['GYM']['GC Input Indices']={'StemMerch':0,'StemNonMerch':1,'Bark':2,'Branch':3,'Foliage':4,'StemMerchV':5}

	# Scale factor for growth curves
	# Note: Do not change this to 0.1 - aerial Aerial BTK Spray response will not work properly at 0.1
	meta['Modules']['GYM']['Scale Factor']=0.001

	meta['Modules']['GYM']['GC_Variable_List']=['ID_Stand','ID_Scn','ID_GC','regeneration_method','s1','p1','i1','s2', \
		'p2','s3','p3','s4','p4','s5','p5','gain1','selage1','gain2','selage2', \
		'gain3','selage3','gain4','selage4','gain5','selage5', \
		'init_density','regen_delay','oaf1','oaf2','bec_zone','FIZ', \
		'fert_age1','fert_age2','fert_age3','fert_age4','fert_age5']

	# Operational adjustment factors (see TIPSY documentation)
	meta['Modules']['GYM']['OAF1 Default']=0.85
	meta['Modules']['GYM']['OAF2 Default']=0.95

	#--------------------------------------------------------------------------
	# TAZ module
	#--------------------------------------------------------------------------
	meta['Modules']['Taz']={}
	
	# Wildfire from Pareto distribution
	meta['Modules']['Taz']['wfss']=gu.ipickle(meta['Paths']['DB']['Taz'] + '\\Wildfire\\Wildfire_Stats_Scenarios_By_BGCZ.pkl')

	# Mountain pine beetle from Pareto distribution
	meta['Modules']['Taz']['IBM Pareto']={}
	beta=np.array([meta['Param']['BE']['OnTheFly']['IBM Pareto Shape'],
				meta['Param']['BE']['OnTheFly']['IBM Pareto Loc'],
				meta['Param']['BE']['OnTheFly']['IBM Pareto Scale']])
	meta['Modules']['Taz']['IBM Pareto']['Mortality Fraction']=meta['Param']['BE']['OnTheFly']['IBM Pareto Mort']
	meta['Modules']['Taz']['IBM Pareto']['Po']=np.zeros((meta[pNam]['Project']['N Time'],meta[pNam]['Project']['N Ensemble']))
	for i in range(meta[pNam]['Project']['N Ensemble']):
		meta['Modules']['Taz']['IBM Pareto']['Po'][:,i]=stats.pareto.rvs(beta[0],loc=beta[1],scale=beta[2],size=meta[pNam]['Project']['N Time'])
		
	# Frost from Pareto distribution (DISCONTINUED)
	# 	meta['Modules']['Taz']['Frost Pareto']={}
	# 	beta=np.array([meta['Param']['BE']['OnTheFly']['Frost Pareto Shape'],
	# 				meta['Param']['BE']['OnTheFly']['Frost Pareto Loc'],
	# 				meta['Param']['BE']['OnTheFly']['Frost Pareto Scale']])
	# 	meta['Modules']['Taz']['Frost Pareto']['Mortality Fraction']=meta['Param']['BE']['OnTheFly']['Frost Pareto Mort']
	# 	meta['Modules']['Taz']['Frost Pareto']['Po']=np.zeros((meta[pNam]['Project']['N Time'],meta[pNam]['Project']['N Ensemble']))
	# 	for i in range(meta[pNam]['Project']['N Ensemble']):
	# 		meta['Modules']['Taz']['Frost Pareto']['Po'][:,i]=stats.pareto.rvs(beta[0],loc=beta[1],scale=beta[2],size=meta[pNam]['Project']['N Time'])

	#--------------------------------------------------------------------------
	# Growth factor information
	# *** Not currently used ***
	#--------------------------------------------------------------------------
	#	# Default status of growth factors
	#	meta['Scenario Switch']['Net Growth Factor Status']=[None]*meta[pNam]['Project']['N Scenario']
	#	meta['Scenario Switch']['Mortality Factor Status']=[None]*meta[pNam]['Project']['N Scenario']
	#	for iScn in range(0,meta[pNam]['Project']['N Scenario']):
	#		meta['Scenario Switch']['Net Growth Factor Status'][iScn]='Off'
	#		meta['Scenario Switch']['Mortality Factor Status'][iScn]='Off'
	#		#meta['Scenario Switch'][iScn]['Status Net Growth Factor']='Off'
	#		#meta[pNam]['Scenario'][iScn]['Status Mortality Factor']='Off'

	#--------------------------------------------------------------------------
	# Nutrient management information (for compatibility with "silviculture" module)
	#--------------------------------------------------------------------------
	# Initialize dictionary
	meta['Modules']['NutrientApplication']={}

	# Initialize index to stands affected by nutrient application
	# This needs to be populated with an empty array for when Sawtooth is used.
	meta['Modules']['NutrientApplication']['iApplication']=np.array([])

	# BGC zone exclusions (for on-the-fly application scheduler)
	meta['Modules']['NutrientApplication']['BGC Zone Exclusion CD']=['PP','MH','BAFA','BG','CMA','IMA']
	meta['Modules']['NutrientApplication']['BGC Zone Exclusion ID']=np.zeros(len(meta['Modules']['NutrientApplication']['BGC Zone Exclusion CD']))
	for iZ in range(len(meta['Modules']['NutrientApplication']['BGC Zone Exclusion CD'])):
		meta['Modules']['NutrientApplication']['BGC Zone Exclusion ID'][iZ]=meta['LUT']['BEC_BIOGEOCLIMATIC_POLY']['ZONE'][ meta['Modules']['NutrientApplication']['BGC Zone Exclusion CD'][iZ] ]

	# Coastal zones used to make prob of occurrence region-specific
	meta['Modules']['NutrientApplication']['Coastal Zones CD']=['CWH','CDF']
	meta['Modules']['NutrientApplication']['Coastal Zones ID']=np.zeros(len(meta['Modules']['NutrientApplication']['Coastal Zones CD']))
	for iZ in range(len(meta['Modules']['NutrientApplication']['Coastal Zones CD'])):
		meta['Modules']['NutrientApplication']['Coastal Zones ID'][iZ]=meta['LUT']['BEC_BIOGEOCLIMATIC_POLY']['ZONE'][ meta['Modules']['NutrientApplication']['Coastal Zones CD'][iZ] ]

	#--------------------------------------------------------------------------
	# FAIR (github.com/OMS-NetZero/FAIR-pro)
	#--------------------------------------------------------------------------
	meta['Modules']['FAIR']={}
	meta['Modules']['FAIR']['Forcings']=['CO2','CH4','N2O','CO','Land use','Volcanic','Solar','Aerosol-cloud interactions','Aerosol-radiation interactions','Ozone','Stratospheric water vapour']

	#--------------------------------------------------------------------------
	# Grassland module
	#--------------------------------------------------------------------------
	for iScn in range(meta[pNam]['Project']['N Scenario']):
		if 'Grassland Module Status' not in meta[pNam]['Scenario'][iScn]:
			# Default is off
			meta[pNam]['Scenario'][iScn]['Grass Module Status']='Off'
			meta[pNam]['Scenario'][iScn]['Grass Module Year Start']=meta[pNam]['Project']['Year Project']+1

	#--------------------------------------------------------------------------
	# Parameter uncertainty
	#--------------------------------------------------------------------------

	if meta[pNam]['Project']['Scenario Source']=='Spreadsheet':

		meta['Param']['By Ensemble']={}

		#----------------------------------------------------------------------
		# Nutrient management
		#----------------------------------------------------------------------
		if meta[pNam]['Project']['Uncertainty Status Nutrient Application']!='On':
			# If uncertainty turned off, use best estimates
			meta['Param']['By Ensemble']['NutrientApplication']={}
			for k in meta['Param']['BE']['NutrientApplication'].keys():
				mu=meta['Param']['BE']['NutrientApplication'][k]
				meta['Param']['By Ensemble']['NutrientApplication'][k]=mu*np.ones(meta[pNam]['Project']['N Stand'])
		else:
			meta['Param']['By Ensemble']['NutrientApplication']={}
			for k in meta['Param']['BE']['NutrientApplication'].keys():
				mu=meta['Param']['BE']['NutrientApplication'][k]
				sig=meta['Param']['Sigma']['NutrientApplication'][k]
				r=np.random.normal(loc=mu,scale=mu*sig,size=meta[pNam]['Project']['N Stand'])
				r=np.maximum(0,r)
				meta['Param']['By Ensemble']['NutrientApplication'][k]=r

		#if meta[pNam]['Project']['Uncertainty Status Nutrient Application']=='On':
# 			#----------------------------------------------------------------------
# 			# Biomass turnover
# 			#----------------------------------------------------------------------
# 			if meta[pNam]['Project']['Uncertainty Status Biomass Turnover']=='On':
# 				meta['Param']['By Ensemble'][iStand]['BiomassTurnover']={}
# 				for k in meta['Param']['BE']['BiomassTurnover'].keys():
# 					mu=meta['Param']['BE']['BiomassTurnover'][k]
# 					sig=meta['Param']['Sigma']['BiomassTurnover'][k]
# 					bl=meta['Param']['BL']['BiomassTurnover'][k]
# 					bu=meta['Param']['BU']['BiomassTurnover'][k]
# 					r=np.random.normal(loc=mu,scale=mu*sig)
# 					if (bl!=-9999) & (bu!=-9999):
# 						r=gu.Clamp(r,bl,bu)
# 					elif (bl!=-9999) & (bu==-9999):
# 						r=np.maximum(bl,r)
# 					elif (bl==-9999) & (bu!=-9999):
# 						r=np.minimum(bu,r)
# 					else:
# 						r=r
# 					meta['Param']['By Ensemble'][iStand]['BiomassTurnover'][k]=r
# 	
# 			#----------------------------------------------------------------------
# 			# Decomposition
# 			#----------------------------------------------------------------------
# 			if meta[pNam]['Project']['Uncertainty Status Decomposition']=='On':
# 				meta['Param']['By Ensemble'][iStand]['Decomposition']={}
# 				for k in meta['Param']['BE']['Decomposition'].keys():
# 					mu=meta['Param']['BE']['Decomposition'][k]
# 					sig=meta['Param']['Sigma']['Decomposition'][k]
# 					bl=meta['Param']['BL']['Decomposition'][k]
# 					bu=meta['Param']['BU']['Decomposition'][k]
# 					r=np.random.normal(loc=mu,scale=mu*sig)
# 					if (bl!=-9999) & (bu!=-9999):
# 						r=gu.Clamp(r,bl,bu)
# 					elif (bl!=-9999) & (bu==-9999):
# 						r=np.maximum(bl,r)
# 					elif (bl==-9999) & (bu!=-9999):
# 						r=np.minimum(bu,r)
# 					else:
# 						r=r
# 	
# 					meta['Param']['By Ensemble'][iStand]['Decomposition'][k]=r
# 	
# 			#----------------------------------------------------------------------
# 			# Harvesting
# 			#----------------------------------------------------------------------
# 			if meta[pNam]['Project']['Uncertainty Status Harvest Utilization']=='On':
# 				meta['Param']['By Ensemble'][iStand]['Event']={}
# 				EventList=['Harvest','Harvest Salvage']
# 				for Event in EventList:
# 					ID_Type=meta['LUT']['Event'][Event]
# 					meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]={}
# 	
# 					#--------------------------------------------------------------
# 					# Biomass merch
# 					#--------------------------------------------------------------
# 					# Removed fraction
# 					mu=meta['Param']['BE']['Event'][ID_Type]['StemwoodMerch_Removed']
# 					sig=np.array([0.1])
# 					bl=np.array([0.0])
# 					bu=np.array([1.0])
# 					r_Removed=np.random.normal(loc=mu,scale=mu*sig)
# 					r_Removed=gu.Clamp(r_Removed,bl,bu)
# 					meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['StemwoodMerch_Removed']=r_Removed
# 	
# 					# Total fraction that is piled and dispersed
# 					r_PiledAndDispersed=1.0-r_Removed
# 	
# 					# Specific piled fraction
# 					mu=np.array([0.60]) # Specific fraction that is piled
# 					sig=np.array([0.1])
# 					bl=np.array([0.0])
# 					bu=np.array([1.0])
# 					rSpecific_Piled=np.random.normal(loc=mu,scale=mu*sig)
# 					rSpecific_Piled=gu.Clamp(rSpecific_Piled,bl,bu)
# 	
# 					# Piled fraction
# 					meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['StemwoodMerch_Piled']=r_PiledAndDispersed*rSpecific_Piled
# 	
# 					# Specific dispersed fraction
# 					rSpecific_Dispersed=1.0-rSpecific_Piled
# 	
# 					# Dispersed fraction
# 					meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['StemwoodMerch_LeftOnSite']=r_PiledAndDispersed*rSpecific_Dispersed
# 	
# 					#print(meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['BiomassMerch_Removed']+meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['BiomassMerch_Piled']+meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['BiomassMerch_LeftOnSite'])
# 	
# 					#--------------------------------------------------------------
# 					# Biomass non-merch
# 					#--------------------------------------------------------------
# 	
# 					# Removed fraction
# 					mu=meta['Param']['BE']['Event'][ID_Type]['StemwoodNonMerch_Removed']
# 					sig=np.array([0.1])
# 					bl=np.array([0.0])
# 					bu=np.array([1.0])
# 					r_Removed=np.random.normal(loc=mu,scale=mu*sig)
# 					r_Removed=gu.Clamp(r_Removed,bl,bu)
# 					meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['StemwoodNonMerch_Removed']=r_Removed
# 	
# 					# Total fraction that is piled and dispersed
# 					r_PiledAndDispersed=1.0-r_Removed
# 	
# 					# Specific piled fraction
# 					mu=np.array([0.60]) # Specific fraction that is piled
# 					sig=np.array([0.1])
# 					bl=np.array([0.0])
# 					bu=np.array([1.0])
# 					rSpecific_Piled=np.random.normal(loc=mu,scale=mu*sig)
# 					rSpecific_Piled=gu.Clamp(rSpecific_Piled,bl,bu)
# 	
# 					# Piled fraction
# 					meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['StemwoodNonMerch_Piled']=r_PiledAndDispersed*rSpecific_Piled
# 	
# 					# Specific dispersed fraction
# 					rSpecific_Dispersed=1.0-rSpecific_Piled
# 	
# 					# Dispersed fraction
# 					meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['StemwoodNonMerch_LeftOnSite']=r_PiledAndDispersed*rSpecific_Dispersed
# 	
# 					#print(meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['BiomassNonMerch_Removed']+meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['BiomassNonMerch_Piled']+meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['BiomassNonMerch_LeftOnSite'])
# 	
# 					#--------------------------------------------------------------
# 					# DeadStems
# 					#--------------------------------------------------------------
# 					# Removed fraction
# 					mu=meta['Param']['BE']['Event'][ID_Type]['DeadStemRemoved']
# 					sig=np.array([0.1])
# 					bl=np.array([0.0])
# 					bu=np.array([1.0])
# 					r_Removed=np.random.normal(loc=mu,scale=mu*sig)
# 					r_Removed=gu.Clamp(r_Removed,bl,bu)
# 					meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['DeadStemRemoved']=r_Removed
# 	
# 					# Total fraction that is piled and dispersed
# 					r_PiledAndDispersed=1.0-r_Removed
# 	
# 					# Specific piled fraction
# 					mu=np.array([0.60]) # Specific fraction that is piled
# 					sig=np.array([0.1])
# 					bl=np.array([0.0])
# 					bu=np.array([1.0])
# 					rSpecific_Piled=np.random.normal(loc=mu,scale=mu*sig)
# 					rSpecific_Piled=gu.Clamp(rSpecific_Piled,bl,bu)
# 	
# 					# Piled fraction
# 					meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['DeadStemPiled']=r_PiledAndDispersed*rSpecific_Piled
# 	
# 					# Specific dispersed fraction
# 					rSpecific_Dispersed=1.0-rSpecific_Piled
# 	
# 					# Dispersed fraction
# 					meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['DeadStemLeftOnSite']=r_PiledAndDispersed*rSpecific_Dispersed
# 	
# 					#print(meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['DeadStems_Removed']+meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['DeadStems_Piled']+meta['Param']['By Ensemble'][iStand]['Event'][ID_Type]['DeadStems_LeftOnSite'])
# 	
# 			#----------------------------------------------------------------------
# 			# Substitution effects
# 			#----------------------------------------------------------------------
# 			if meta[pNam]['Project']['Uncertainty Status Substitution']=='On':
# 	
# 				meta['Param']['By Ensemble'][iStand]['Substitution']={}
# 	
# 				#vL=['LumberDisplacementFactor','PanelDisplacementFactor']
# 				vL=['SawnwoodFracDisplacingConcrete','SawnwoodFracDisplacingSteel','SawnwoodFracDisplacingAluminum',
# 					'SawnwoodFracDisplacingPlastic','SawnwoodFracDisplacingTextile','PanelFracDisplacingConcrete','PanelFracDisplacingSteel',
# 					'PanelFracDisplacingAluminum','PanelFracDisplacingPlastic','PanelFracDisplacingTextile','ResidualsFracDisplacingConcrete',
# 					'ResidualsFracDisplacingSteel','ResidualsFracDisplacingAluminum','ResidualsFracDisplacingPlastic','ResidualsFracDisplacingTextile',
# 					'DisplacementRatio_ConcreteForSawnwood','DisplacementRatio_ConcreteForPanel','DisplacementRatio_ConcreteForResiduals',
# 					'DisplacementRatio_SteelForSawnwood','DisplacementRatio_SteelForPanel','DisplacementRatio_SteelForResiduals',
# 					'DisplacementRatio_AluminumForSawnwood','DisplacementRatio_AluminumForPanel','DisplacementRatio_AluminumForResiduals',
# 					'DisplacementRatio_PlasticForSawnwood','DisplacementRatio_PlasticForPanel','DisplacementRatio_PlasticForResiduals',
# 					'DisplacementRatio_TextileForSawnwood','DisplacementRatio_TextileForPanel','DisplacementRatio_TextileForResiduals']
# 				for k in vL:
# 					mu=meta['Param']['BE']['Substitution'][k]
# 					sig=meta['Param']['Sigma']['Substitution'][k]
# 					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
# 					r=np.maximum(0,r)
# 					meta['Param']['By Ensemble'][iStand]['Substitution'][k]=r
# 	
# 				lv0=np.array(['PowerFacilityDomFracDisplacingRenewables','PowerFacilityDomFracDisplacingCoal','PowerFacilityDomFracDisplacingDiesel','PowerFacilityDomFracDisplacingNaturalGas','PowerFacilityDomFracDisplacingOil'])
# 				lv=lv0[np.argsort(np.random.random(len(lv0)))]
# 				r_remain=1.0
# 				for k in lv:
# 					mu=meta['Param']['BE']['Substitution'][k]
# 					sig=meta['Param']['Sigma']['Substitution'][k]
# 					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
# 					r=gu.Clamp(r,0.0,r_remain)
# 					meta['Param']['By Ensemble'][iStand]['Substitution'][k]=r
# 					r_remain=r_remain-r
# 	
# 				lv0=np.array(['PowerFacilityExportFracDisplacingRenewables','PowerFacilityExportFracDisplacingCoal','PowerFacilityExportFracDisplacingDiesel','PowerFacilityExportFracDisplacingNaturalGas','PowerFacilityExportFracDisplacingOil'])
# 				lv=lv0[np.argsort(np.random.random(len(lv0)))]
# 				r_remain=1.0
# 				for k in lv:
# 					mu=meta['Param']['BE']['Substitution'][k]
# 					sig=meta['Param']['Sigma']['Substitution'][k]
# 					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
# 					r=gu.Clamp(r,0.0,r_remain)
# 					meta['Param']['By Ensemble'][iStand]['Substitution'][k]=r
# 					r_remain=r_remain-r
# 	
# 				lv0=np.array(['PelletExportFracDisplacingRenewables','PelletExportFracDisplacingCoal','PelletExportFracDisplacingDiesel','PelletExportFracDisplacingNaturalGas','PelletExportFracDisplacingOil'])
# 				lv=lv0[np.argsort(np.random.random(len(lv0)))]
# 				r_remain=1.0
# 				for k in lv:
# 					mu=meta['Param']['BE']['Substitution'][k]
# 					sig=meta['Param']['Sigma']['Substitution'][k]
# 					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
# 					r=gu.Clamp(r,0.0,r_remain)
# 					meta['Param']['By Ensemble'][iStand]['Substitution'][k]=r
# 					r_remain=r_remain-r
# 	
# 				lv0=np.array(['PelletDomGridFracDisplacingRenewables','PelletDomGridFracDisplacingCoal','PelletDomGridFracDisplacingDiesel','PelletDomGridFracDisplacingNaturalGas','PelletDomGridFracDisplacingOil'])
# 				lv=lv0[np.argsort(np.random.random(len(lv0)))]
# 				r_remain=1.0
# 				for k in lv:
# 					mu=meta['Param']['BE']['Substitution'][k]
# 					sig=meta['Param']['Sigma']['Substitution'][k]
# 					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
# 					r=gu.Clamp(r,0.0,r_remain)
# 					meta['Param']['By Ensemble'][iStand]['Substitution'][k]=r
# 					r_remain=r_remain-r
# 	
# 				lv0=np.array(['PelletDomRNGFracDisplacingRenewables','PelletDomRNGFracDisplacingCoal','PelletDomRNGFracDisplacingDiesel','PelletDomRNGFracDisplacingNaturalGas','PelletDomRNGFracDisplacingOil'])
# 				lv=lv0[np.argsort(np.random.random(len(lv0)))]
# 				r_remain=1.0
# 				for k in lv:
# 					mu=meta['Param']['BE']['Substitution'][k]
# 					sig=meta['Param']['Sigma']['Substitution'][k]
# 					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
# 					r=gu.Clamp(r,0.0,r_remain)
# 					meta['Param']['By Ensemble'][iStand]['Substitution'][k]=r
# 					r_remain=r_remain-r
# 	
# 				lv0=np.array(['FirewoodDomFracDisplacingRenewables','FirewoodDomFracDisplacingCoal','FirewoodDomFracDisplacingDiesel','FirewoodDomFracDisplacingNaturalGas','FirewoodDomFracDisplacingOil'])
# 				lv=lv0[np.argsort(np.random.random(len(lv0)))]
# 				r_remain=1.0
# 				for k in lv:
# 					mu=meta['Param']['BE']['Substitution'][k]
# 					sig=meta['Param']['Sigma']['Substitution'][k]
# 					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
# 					r=gu.Clamp(r,0.0,r_remain)
# 					meta['Param']['By Ensemble'][iStand]['Substitution'][k]=r
# 					r_remain=r_remain-r
# 	
# 				lv0=np.array(['FirewoodExportFracDisplacingRenewables','FirewoodExportFracDisplacingCoal','FirewoodExportFracDisplacingDiesel','FirewoodExportFracDisplacingNaturalGas','FirewoodExportFracDisplacingOil'])
# 				lv=lv0[np.argsort(np.random.random(len(lv0)))]
# 				r_remain=1.0
# 				for k in lv:
# 					mu=meta['Param']['BE']['Substitution'][k]
# 					sig=meta['Param']['Sigma']['Substitution'][k]
# 					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
# 					r=gu.Clamp(r,0.0,r_remain)
# 					meta['Param']['By Ensemble'][iStand]['Substitution'][k]=r
# 					r_remain=r_remain-r
# 	
# 				lv0=np.array(['PowerGridFracDisplacingRenewables','PowerGridFracDisplacingCoal','PowerGridFracDisplacingDiesel','PowerGridFracDisplacingNaturalGas','PowerGridFracDisplacingOil'])
# 				lv=lv0[np.argsort(np.random.random(len(lv0)))]
# 				r_remain=1.0
# 				for k in lv:
# 					mu=meta['Param']['BE']['Substitution'][k]
# 					sig=meta['Param']['Sigma']['Substitution'][k]
# 					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
# 					r=gu.Clamp(r,0.0,r_remain)
# 					meta['Param']['By Ensemble'][iStand]['Substitution'][k]=r
# 					r_remain=r_remain-r

	else:

		meta['Param']['By Ensemble']=[None]*meta[pNam]['Project']['N Ensemble']
		for iEns in range(meta[pNam]['Project']['N Ensemble']):
		
			# Initialize dictionary
			meta['Param']['By Ensemble'][iEns]={}
		
			#----------------------------------------------------------------------
			# Biomass turnover
			#----------------------------------------------------------------------
			if meta[pNam]['Project']['Uncertainty Status Biomass Turnover']=='On':
				meta['Param']['By Ensemble'][iEns]['BiomassTurnover']={}
				for k in meta['Param']['BE']['BiomassTurnover'].keys():
					mu=meta['Param']['BE']['BiomassTurnover'][k]
					sig=meta['Param']['Sigma']['BiomassTurnover'][k]
					bl=meta['Param']['BL']['BiomassTurnover'][k]
					bu=meta['Param']['BU']['BiomassTurnover'][k]
					r=np.random.normal(loc=mu,scale=mu*sig)
					if (bl!=-9999) & (bu!=-9999):
						r=gu.Clamp(r,bl,bu)
					elif (bl!=-9999) & (bu==-9999):
						r=np.maximum(bl,r)
					elif (bl==-9999) & (bu!=-9999):
						r=np.minimum(bu,r)
					else:
						r=r
					meta['Param']['By Ensemble'][iEns]['BiomassTurnover'][k]=r
		
			#----------------------------------------------------------------------
			# Decomposition
			#----------------------------------------------------------------------
			if meta[pNam]['Project']['Uncertainty Status Decomposition']=='On':
				meta['Param']['By Ensemble'][iEns]['Decomposition']={}
				for k in meta['Param']['BE']['Decomposition'].keys():
					mu=meta['Param']['BE']['Decomposition'][k]
					sig=meta['Param']['Sigma']['Decomposition'][k]
					bl=meta['Param']['BL']['Decomposition'][k]
					bu=meta['Param']['BU']['Decomposition'][k]
					r=np.random.normal(loc=mu,scale=mu*sig)
					if (bl!=-9999) & (bu!=-9999):
						r=gu.Clamp(r,bl,bu)
					elif (bl!=-9999) & (bu==-9999):
						r=np.maximum(bl,r)
					elif (bl==-9999) & (bu!=-9999):
						r=np.minimum(bu,r)
					else:
						r=r
		
					meta['Param']['By Ensemble'][iEns]['Decomposition'][k]=r
		
			#----------------------------------------------------------------------
			# Harvesting
			#----------------------------------------------------------------------
			if meta[pNam]['Project']['Uncertainty Status Harvest Utilization']=='On':
				meta['Param']['By Ensemble'][iEns]['Event']={}
				EventList=['Harvest','Harvest Salvage']
				for Event in EventList:
					ID_Type=meta['LUT']['Event'][Event]
					meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]={}
		
					#--------------------------------------------------------------
					# Biomass merch
					#--------------------------------------------------------------
					# Removed fraction
					mu=meta['Param']['BE']['Event'][ID_Type]['StemwoodMerch_Removed']
					sig=np.array([0.1])
					bl=np.array([0.0])
					bu=np.array([1.0])
					r_Removed=np.random.normal(loc=mu,scale=mu*sig)
					r_Removed=gu.Clamp(r_Removed,bl,bu)
					meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['StemwoodMerch_Removed']=r_Removed
		
					# Total fraction that is piled and dispersed
					r_PiledAndDispersed=1.0-r_Removed
		
					# Specific piled fraction
					mu=np.array([0.60]) # Specific fraction that is piled
					sig=np.array([0.1])
					bl=np.array([0.0])
					bu=np.array([1.0])
					rSpecific_Piled=np.random.normal(loc=mu,scale=mu*sig)
					rSpecific_Piled=gu.Clamp(rSpecific_Piled,bl,bu)
		
					# Piled fraction
					meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['StemwoodMerch_Piled']=r_PiledAndDispersed*rSpecific_Piled
		
					# Specific dispersed fraction
					rSpecific_Dispersed=1.0-rSpecific_Piled
		
					# Dispersed fraction
					meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['StemwoodMerch_LeftOnSite']=r_PiledAndDispersed*rSpecific_Dispersed
		
					#print(meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['BiomassMerch_Removed']+meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['BiomassMerch_Piled']+meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['BiomassMerch_LeftOnSite'])
		
					#--------------------------------------------------------------
					# Biomass non-merch
					#--------------------------------------------------------------
		
					# Removed fraction
					mu=meta['Param']['BE']['Event'][ID_Type]['StemwoodNonMerch_Removed']
					sig=np.array([0.1])
					bl=np.array([0.0])
					bu=np.array([1.0])
					r_Removed=np.random.normal(loc=mu,scale=mu*sig)
					r_Removed=gu.Clamp(r_Removed,bl,bu)
					meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['StemwoodNonMerch_Removed']=r_Removed
		
					# Total fraction that is piled and dispersed
					r_PiledAndDispersed=1.0-r_Removed
		
					# Specific piled fraction
					mu=np.array([0.60]) # Specific fraction that is piled
					sig=np.array([0.1])
					bl=np.array([0.0])
					bu=np.array([1.0])
					rSpecific_Piled=np.random.normal(loc=mu,scale=mu*sig)
					rSpecific_Piled=gu.Clamp(rSpecific_Piled,bl,bu)
		
					# Piled fraction
					meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['StemwoodNonMerch_Piled']=r_PiledAndDispersed*rSpecific_Piled
		
					# Specific dispersed fraction
					rSpecific_Dispersed=1.0-rSpecific_Piled
		
					# Dispersed fraction
					meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['StemwoodNonMerch_LeftOnSite']=r_PiledAndDispersed*rSpecific_Dispersed
		
					#print(meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['BiomassNonMerch_Removed']+meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['BiomassNonMerch_Piled']+meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['BiomassNonMerch_LeftOnSite'])
		
					#--------------------------------------------------------------
					# DeadStems
					#--------------------------------------------------------------
					# Removed fraction
					mu=meta['Param']['BE']['Event'][ID_Type]['DeadStemRemoved']
					sig=np.array([0.1])
					bl=np.array([0.0])
					bu=np.array([1.0])
					r_Removed=np.random.normal(loc=mu,scale=mu*sig)
					r_Removed=gu.Clamp(r_Removed,bl,bu)
					meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['DeadStemRemoved']=r_Removed
		
					# Total fraction that is piled and dispersed
					r_PiledAndDispersed=1.0-r_Removed
		
					# Specific piled fraction
					mu=np.array([0.60]) # Specific fraction that is piled
					sig=np.array([0.1])
					bl=np.array([0.0])
					bu=np.array([1.0])
					rSpecific_Piled=np.random.normal(loc=mu,scale=mu*sig)
					rSpecific_Piled=gu.Clamp(rSpecific_Piled,bl,bu)
		
					# Piled fraction
					meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['DeadStemPiled']=r_PiledAndDispersed*rSpecific_Piled
		
					# Specific dispersed fraction
					rSpecific_Dispersed=1.0-rSpecific_Piled
		
					# Dispersed fraction
					meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['DeadStemLeftOnSite']=r_PiledAndDispersed*rSpecific_Dispersed
		
					#print(meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['DeadStems_Removed']+meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['DeadStems_Piled']+meta['Param']['By Ensemble'][iEns]['Event'][ID_Type]['DeadStems_LeftOnSite'])
		
			#----------------------------------------------------------------------
			# Substitution effects
			#----------------------------------------------------------------------
			if meta[pNam]['Project']['Uncertainty Status Substitution']=='On':
		
				meta['Param']['By Ensemble'][iEns]['Substitution']={}
		
				#vL=['LumberDisplacementFactor','PanelDisplacementFactor']
				vL=['SawnwoodFracDisplacingConcrete','SawnwoodFracDisplacingSteel','SawnwoodFracDisplacingAluminum',
					'SawnwoodFracDisplacingPlastic','SawnwoodFracDisplacingTextile','PanelFracDisplacingConcrete','PanelFracDisplacingSteel',
					'PanelFracDisplacingAluminum','PanelFracDisplacingPlastic','PanelFracDisplacingTextile','ResidualsFracDisplacingConcrete',
					'ResidualsFracDisplacingSteel','ResidualsFracDisplacingAluminum','ResidualsFracDisplacingPlastic','ResidualsFracDisplacingTextile',
					'DisplacementRatio_ConcreteForSawnwood','DisplacementRatio_ConcreteForPanel','DisplacementRatio_ConcreteForResiduals',
					'DisplacementRatio_SteelForSawnwood','DisplacementRatio_SteelForPanel','DisplacementRatio_SteelForResiduals',
					'DisplacementRatio_AluminumForSawnwood','DisplacementRatio_AluminumForPanel','DisplacementRatio_AluminumForResiduals',
					'DisplacementRatio_PlasticForSawnwood','DisplacementRatio_PlasticForPanel','DisplacementRatio_PlasticForResiduals',
					'DisplacementRatio_TextileForSawnwood','DisplacementRatio_TextileForPanel','DisplacementRatio_TextileForResiduals']
				for k in vL:
					mu=meta['Param']['BE']['Substitution'][k]
					sig=meta['Param']['Sigma']['Substitution'][k]
					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
					r=np.maximum(0,r)
					meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
		
				lv0=np.array(['PowerFacilityDomFracDisplacingRenewables','PowerFacilityDomFracDisplacingCoal','PowerFacilityDomFracDisplacingDiesel','PowerFacilityDomFracDisplacingNaturalGas','PowerFacilityDomFracDisplacingOil'])
				lv=lv0[np.argsort(np.random.random(len(lv0)))]
				r_remain=1.0
				for k in lv:
					mu=meta['Param']['BE']['Substitution'][k]
					sig=meta['Param']['Sigma']['Substitution'][k]
					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
					r=gu.Clamp(r,0.0,r_remain)
					meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
					r_remain=r_remain-r
		
				lv0=np.array(['PowerFacilityExportFracDisplacingRenewables','PowerFacilityExportFracDisplacingCoal','PowerFacilityExportFracDisplacingDiesel','PowerFacilityExportFracDisplacingNaturalGas','PowerFacilityExportFracDisplacingOil'])
				lv=lv0[np.argsort(np.random.random(len(lv0)))]
				r_remain=1.0
				for k in lv:
					mu=meta['Param']['BE']['Substitution'][k]
					sig=meta['Param']['Sigma']['Substitution'][k]
					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
					r=gu.Clamp(r,0.0,r_remain)
					meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
					r_remain=r_remain-r
		
				lv0=np.array(['PelletExportFracDisplacingRenewables','PelletExportFracDisplacingCoal','PelletExportFracDisplacingDiesel','PelletExportFracDisplacingNaturalGas','PelletExportFracDisplacingOil'])
				lv=lv0[np.argsort(np.random.random(len(lv0)))]
				r_remain=1.0
				for k in lv:
					mu=meta['Param']['BE']['Substitution'][k]
					sig=meta['Param']['Sigma']['Substitution'][k]
					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
					r=gu.Clamp(r,0.0,r_remain)
					meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
					r_remain=r_remain-r
		
				lv0=np.array(['PelletDomGridFracDisplacingRenewables','PelletDomGridFracDisplacingCoal','PelletDomGridFracDisplacingDiesel','PelletDomGridFracDisplacingNaturalGas','PelletDomGridFracDisplacingOil'])
				lv=lv0[np.argsort(np.random.random(len(lv0)))]
				r_remain=1.0
				for k in lv:
					mu=meta['Param']['BE']['Substitution'][k]
					sig=meta['Param']['Sigma']['Substitution'][k]
					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
					r=gu.Clamp(r,0.0,r_remain)
					meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
					r_remain=r_remain-r
		
				lv0=np.array(['PelletDomRNGFracDisplacingRenewables','PelletDomRNGFracDisplacingCoal','PelletDomRNGFracDisplacingDiesel','PelletDomRNGFracDisplacingNaturalGas','PelletDomRNGFracDisplacingOil'])
				lv=lv0[np.argsort(np.random.random(len(lv0)))]
				r_remain=1.0
				for k in lv:
					mu=meta['Param']['BE']['Substitution'][k]
					sig=meta['Param']['Sigma']['Substitution'][k]
					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
					r=gu.Clamp(r,0.0,r_remain)
					meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
					r_remain=r_remain-r
		
				lv0=np.array(['FirewoodDomFracDisplacingRenewables','FirewoodDomFracDisplacingCoal','FirewoodDomFracDisplacingDiesel','FirewoodDomFracDisplacingNaturalGas','FirewoodDomFracDisplacingOil'])
				lv=lv0[np.argsort(np.random.random(len(lv0)))]
				r_remain=1.0
				for k in lv:
					mu=meta['Param']['BE']['Substitution'][k]
					sig=meta['Param']['Sigma']['Substitution'][k]
					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
					r=gu.Clamp(r,0.0,r_remain)
					meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
					r_remain=r_remain-r
		
				lv0=np.array(['FirewoodExportFracDisplacingRenewables','FirewoodExportFracDisplacingCoal','FirewoodExportFracDisplacingDiesel','FirewoodExportFracDisplacingNaturalGas','FirewoodExportFracDisplacingOil'])
				lv=lv0[np.argsort(np.random.random(len(lv0)))]
				r_remain=1.0
				for k in lv:
					mu=meta['Param']['BE']['Substitution'][k]
					sig=meta['Param']['Sigma']['Substitution'][k]
					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
					r=gu.Clamp(r,0.0,r_remain)
					meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
					r_remain=r_remain-r
		
				lv0=np.array(['PowerGridFracDisplacingRenewables','PowerGridFracDisplacingCoal','PowerGridFracDisplacingDiesel','PowerGridFracDisplacingNaturalGas','PowerGridFracDisplacingOil'])
				lv=lv0[np.argsort(np.random.random(len(lv0)))]
				r_remain=1.0
				for k in lv:
					mu=meta['Param']['BE']['Substitution'][k]
					sig=meta['Param']['Sigma']['Substitution'][k]
					r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
					r=gu.Clamp(r,0.0,r_remain)
					meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
					r_remain=r_remain-r
		
			#----------------------------------------------------------------------
			# Nutrient mmanagement
			#----------------------------------------------------------------------
			if meta[pNam]['Project']['Uncertainty Status Nutrient Application']=='On':
		
				meta['Param']['By Ensemble'][iEns]['NutrientApplication']={}
		
				for k in meta['Param']['BE']['NutrientApplication'].keys():
					mu=meta['Param']['BE']['NutrientApplication'][k]
					sig=meta['Param']['Sigma']['NutrientApplication'][k]
					r=np.random.normal(loc=mu,scale=mu*sig)
					r=np.maximum(0,r)
					meta['Param']['By Ensemble'][iEns]['NutrientApplication'][k]=r
		
			#----------------------------------------------------------------------
			# Mortality distribution
			#----------------------------------------------------------------------
			#meta[pNam]['Project']['Mortality Distribution']={}
			#meta[pNam]['Project']['Mortality Distribution']['Frequency (%)']=np.arange(0,101,1)
			#meta[pNam]['Project']['Mortality Distribution']['Severity (%)']=np.arange(0,101,1)
			#meta[pNam]['Project']['Mortality Distribution']['Variable List']=['Regular','Harvest','Wildfire','Beetles','Mechanical']
			#meta[pNam]['Project']['Mortality Distribution']['Data']={}
			#for v in meta[pNam]['Project']['Mortality Distribution']['Variable List']:
			#	meta[pNam]['Project']['Mortality Distribution']['Data'][v]=np.zeros((meta[pNam]['Project']['Mortality Distribution']['Frequency (%)'].size,meta[pNam]['Project']['Mortality Distribution']['Severity (%)'].size),dtype='int8')

	#--------------------------------------------------------------------------
	# Scenario info for portfolio projects
	# *** No longer supported ***
	#--------------------------------------------------------------------------

# 	if meta[pNam]['Project']['Scenario Source']=='Portfolio':
# 		# Index to rows with implementation
# 		indAT=np.where(np.sum(meta[pNam]['Project']['AIL']['Area'],axis=0)>0)[0]

# 		meta[pNam]['Project']['Portfolio']['ID Portfolio']=np.zeros(meta[pNam]['Project']['N Stand'],dtype=int)
# 		meta[pNam]['Project']['Portfolio']['ID AT']=np.zeros(meta[pNam]['Project']['N Stand'],dtype=int)
# 		meta[pNam]['Project']['Portfolio']['ID AT Unique']=np.zeros(meta[pNam]['Project']['N Stand'],dtype=int)
# 		meta[pNam]['Project']['Portfolio']['Area']=np.zeros(meta[pNam]['Project']['N Stand'])
# 		meta[pNam]['Project']['Portfolio']['Year']=np.zeros(meta[pNam]['Project']['N Stand'])
# 		meta[pNam]['Project']['Portfolio']['Region Code']=np.array(['' for _ in range(meta[pNam]['Project']['N Stand'])],dtype=object)
# 		meta[pNam]['Project']['Portfolio']['Felled Fate Scenario']=np.array(['' for _ in range(meta[pNam]['Project']['N Stand'])],dtype=object)
# 		meta[pNam]['Project']['Portfolio']['Removed Fate Scenario']=np.array(['' for _ in range(meta[pNam]['Project']['N Stand'])],dtype=object)
# 		meta[pNam]['Project']['Portfolio']['HWP End Use Scenario']=np.array(['' for _ in range(meta[pNam]['Project']['N Stand'])],dtype=object)
# 		cnt=0
# 		for iA in range(meta[pNam]['Project']['AIL']['N AT']):
# 			for iY in range(meta[pNam]['Project']['AIL']['N Years']):
# 				for iS in range(meta[pNam]['Project']['N Stand per Activity Type']):

# 					ID_Portfolio=meta[pNam]['Project']['AIL']['ID Portfolio'][indAT[iA]]

# 					meta[pNam]['Project']['Portfolio']['ID Portfolio'][cnt]=ID_Portfolio
# 					meta[pNam]['Project']['Portfolio']['ID AT'][cnt]=meta[pNam]['Project']['AIL']['ID AT'][indAT[iA]]
# 					meta[pNam]['Project']['Portfolio']['ID AT Unique'][cnt]=meta[pNam]['Project']['AIL']['ID AT Unique'][indAT[iA]]
# 					meta[pNam]['Project']['Portfolio']['Area'][cnt]=meta[pNam]['Project']['AIL']['Area'][iY,indAT[iA]]
# 					meta[pNam]['Project']['Portfolio']['Year'][cnt]=meta[pNam]['Project']['AIL']['Year'][iY]

# 					# Region from Activities table
# 					ind=np.where(meta[pNam]['Project']['Activities']['Activity ID']==meta[pNam]['Project']['Portfolio']['ID AT'][cnt])
# 					meta[pNam]['Project']['Portfolio']['Region Code'][cnt]=meta[pNam]['Project']['Activities']['Region Code'][ind][0]

# 					# Scenarios from Portfolio table
# 					iPortfolio=np.where(meta[pNam]['Project']['Portfolio']['Raw']['ID_Portfolio']==ID_Portfolio)[0]
# 					meta[pNam]['Project']['Portfolio']['Felled Fate Scenario'][cnt]=meta[pNam]['Project']['Portfolio']['Raw']['Felled Fate Scenario'][iPortfolio][0]
# 					meta[pNam]['Project']['Portfolio']['Removed Fate Scenario'][cnt]=meta[pNam]['Project']['Portfolio']['Raw']['Removed Fate Scenario'][iPortfolio][0]
# 					meta[pNam]['Project']['Portfolio']['HWP End Use Scenario'][cnt]=meta[pNam]['Project']['Portfolio']['Raw']['HWP End Use Scenario'][iPortfolio][0]
# 					cnt=cnt+1

# 		# Scenario information
# 		# Will this tool ever be used with on-the-fly disturbances? Current set
# 		# to "off".
# 		meta[pNam]['Scenario']=[None]*meta[pNam]['Project']['N Scenario']
# 		for iScn in range(meta[pNam]['Project']['N Scenario']):

# 			meta[pNam]['Scenario'][iScn]={}

# 			# *** This is super awkward - simulations can't change between activities or scenarios!!!
# 			# Do we need that type of functionality for the PT? ***
# 			meta[pNam]['Scenario'][iScn]['Wildfire Scenario ID']=meta[pNam]['Project']['Activities']['Wildfire Scenario ID'][0]
# 			meta[pNam]['Scenario'][iScn]['Wildfire Status Pre-modern']=meta[pNam]['Project']['Activities']['Wildfire Status Pre-modern'][0]
# 			meta[pNam]['Scenario'][iScn]['Wildfire Status Modern']=meta[pNam]['Project']['Activities']['Wildfire Status Modern'][0]
# 			meta[pNam]['Scenario'][iScn]['Wildfire Status Future']=meta[pNam]['Project']['Activities']['Wildfire Status Future'][0]

# 			meta[pNam]['Scenario'][iScn]['Harvest Status Historical']='Off'
# 			meta[pNam]['Scenario'][iScn]['Harvest Status Future']='Off'
# 			meta[pNam]['Scenario'][iScn]['Breakup Status Historical']='Off'
# 			meta[pNam]['Scenario'][iScn]['Breakup Status Future']='Off'
# 			meta[pNam]['Scenario'][iScn]['Nutrient Application Status']='Off'

	#--------------------------------------------------------------------------
	# Remove any modified event chronologies
	#--------------------------------------------------------------------------
	for iScn in range(meta[pNam]['Project']['N Scenario']):
		for iBat in range(meta[pNam]['Project']['N Batch']):
			for iEns in range(meta[pNam]['Project']['N Ensemble']):
				pth=meta['Paths'][pNam]['Input Scenario'][iScn] + '\\Modified_Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl'
				if os.path.exists(pth)==True:
					os.remove(pth)

	#--------------------------------------------------------------------------
	# Initialize run time tracking
	#--------------------------------------------------------------------------

	meta[pNam]['Project']['Run Time Summary']={}
	vL=['Stand initialization','Set location-specific parameters','Running biomass dynamics from sawtooth','Biomass from GY model','Biomass from GROMO','DOM','Events','HWP','Full annual loop','Run geological','Export results to file','Collect garbage']
	for v in vL:
		meta[pNam]['Project']['Run Time Summary'][v]=0.0

	return meta

#%% Look-up-table crosswalk (numbers to strings)
def lut_n2s(dc,numb):
	if numb!=9999:
		vals=np.fromiter(dc.values(),dtype=float)
		keys=np.fromiter(dc.keys(),dtype='<U70')
		ind=np.where(vals==numb)[0]
		s=keys[ind]
	else:
		s=np.array(['Unidentified'],ndmin=1)
	return s

#%% Index to batch
def IndexToBatch(m,iBat):
	iStart=m['Project']['Batch Interval']*iBat
	iStop=np.minimum(m['Project']['N Stand'],iStart+m['Project']['Batch Interval'])
	indBat=np.arange(iStart,iStop,1)
	return indBat

#%% Fix ensemble name and numbering
def FixFileNum(ind):
	indStrFixed=str(ind+1)
	if len(indStrFixed)==1:
		indStrFixed='000' + indStrFixed
	elif len(indStrFixed)==2:
		indStrFixed='00' + indStrFixed
	elif len(indStrFixed)==3:
		indStrFixed='0' + indStrFixed
	return indStrFixed

#%% Load look-up-tables
def Load_LUTs_Modelling(meta):

	# Initialize LUTs dictionary
	if 'LUT' not in meta:
		meta['LUT']={}

	# Import distubance type
	p=gu.ReadExcel(meta['Paths']['Model']['Code'] + '\\Parameters\\' + 'Parameters_Events.xlsx',sheet_name='Sheet1',skiprows=0)
	meta['LUT']['Event']={}
	for i in range(p['Name'].size):
		meta['LUT']['Event'][p['Name'][i]]=p['ID'][i]

	# Region
	meta['LUT']['Region']={'Coast':1,'Interior':2,'GFS22':3}

	# # Added this to accommodate jupyter notebook demos - will need updating periodically
	# if 'Results' not in meta['Paths']:
	#	 meta['Paths']['Results']=r'C:\Users\rhember\Documents\Data\ForestInventory\Results\20220422'
	# if 'VRI' not in meta['Paths']:
	#	 meta['Paths']['VRI']=r'C:\Users\rhember\Documents\Data\ForestInventory\VRI\20220404'
	# if 'Disturbances' not in meta['Paths']:
	#	 meta['Paths']['Disturbances']=r'C:\Users\rhember\Documents\Data\ForestInventory\Disturbances\20220422'
	# if 'LandUse' not in meta['Paths']:
	#	 meta['Paths']['LandUse']=r'C:\Users\rhember\Documents\Data\ForestInventory\LandUse\20220422'

	# meta['LUT']['ATU']=gu.ipickle(meta['Paths']['Results'] + '\\LUTs_RSLT_ACTIVITY_TREATMENT_SVW.pkl')
	# meta['LUT']['OP']=gu.ipickle(meta['Paths']['Results'] + '\\LUTs_RSLT_OPENING_SVW.pkl')
	# meta['LUT']['PL']=gu.ipickle(meta['Paths']['Results'] + '\\LUTs_RSLT_PLANTING_SVW.pkl')
	# meta['LUT']['FC_I']=gu.ipickle(meta['Paths']['Results'] + '\\LUTs_RSLT_FOREST_COVER_INV_SVW.pkl')
	# meta['LUT']['FC_S']=gu.ipickle(meta['Paths']['Results'] + '\\LUTs_RSLT_FOREST_COVER_SILV_SVW.pkl')
	# meta['LUT']['VEG_COMP_LYR_R1_POLY']=gu.ipickle(meta['Paths']['VRI'] + '\\LUTs_VEG_COMP_LYR_R1_POLY.pkl')
	# meta['LUT']['BS']=gu.ipickle(meta['Paths']['Disturbances'] + '\\LUTs_VEG_BURN_SEVERITY_SP.pkl')
	# meta['LUT']['Pest']=gu.ipickle(meta['Paths']['Disturbances'] + '\\LUTs_PEST_INFESTATION_POLY.pkl')
	# meta['LUT']['FC_R']=gu.ipickle(meta['Paths']['Results'] + '\\LUTs_RSLT_FOREST_COVER_RESERVE_SVW.pkl')
	# #meta['LUT']['LU NL']=gu.ipickle(meta['Paths']['LandUse'] + '\\LUTs_RMP_PLAN_NON_LEGAL_POLY_SVW.pkl')
	# meta['LUT']['LU L']=gu.ipickle(meta['Paths']['LandUse'] + '\\LUTs_RMP_PLAN_LEGAL_POLY_SVW.pkl')
	# meta['LUT']['PARK']=gu.ipickle(meta['Paths']['LandUse'] + '\\LUTs_TA_PARK_ECORES_PA_SVW.pkl')
	# meta['LUT']['OGMA']=gu.ipickle(meta['Paths']['LandUse'] + '\\LUTs_RMP_OGMA_LEGAL_ALL_SVW.pkl')
	# meta['LUT']['OGSR']=gu.ipickle(meta['Paths']['LandUse'] + '\\LUTs_OGSR_TAP_PRIORITY_DEF_AREA_SP.pkl')
	# meta['LUT']['UWR']=gu.ipickle(meta['Paths']['LandUse'] + '\\LUTs_WCP_UNGULATE_WINTER_RANGE_SP.pkl')

	meta['LUT']['TIPSY']={}
	meta['LUT']['TIPSY']['FIZ']={'C':np.array(1,dtype=int),'I':np.array(2,dtype=int)}
	meta['LUT']['TIPSY']['regeneration_method']={'C':np.array(1,dtype=int),'N':np.array(2,dtype=int),'P':np.array(3,dtype=int)}

	# # Land surface classification
	# meta['LUT']['LSC']={}
	# meta['LUT']['LSC']['Cover']={}
	# data=gu.ReadExcel(meta['Paths']['Model']['Code'] + '\\Parameters\\Parameters_LSC_Cover.xlsx')
	# for i in range(data['ID'].size):
	#	 meta['LUT']['LSC']['Cover'][data['Name'][i]]=data['ID'][i]
	# meta['LUT']['LSC']['Use']={}
	# data=gu.ReadExcel(meta['Paths']['Model']['Code'] + '\\Parameters\\Parameters_LSC_Use.xlsx')
	# for i in range(data['ID'].size):
	#	 meta['LUT']['LSC']['Use'][data['Name'][i]]=data['ID'][i]

	# Species (for Sawtooth)
	#meta['LUT']['SRS']={}
	#for i in range(len(par['SRS']['SRS_CD'])):
	#	meta['LUT']['SRS'][par['SRS']['SRS_CD'][i]]=par['SRS']['SRS_ID'][i]

	return meta



#%% Decompress event chronology
def EventChronologyDecompress(meta,pNam,ec,iScn,iEns,iBat):
	# Uncompress event chronology if it has been compressed
	if 'idx' in ec:
		idx=ec['idx']
		tmp=ec.copy()
		for v in ['ID Event Type','Mortality Factor','Growth Factor','ID Growth Curve','ASET']:
			ec[v]=np.zeros((meta[pNam]['Project']['N Time'],meta[pNam]['Project']['Batch Size'][iBat],meta['Core']['Max Events Per Year']),dtype='int16')
			try:
				ec[v][idx[0],idx[1],idx[2]]=tmp[v]
			except:
				pass
		del tmp
	return ec

#%% Load scenario results
# Return a list of dictionaries for each scenario. If multiple ensemble were run,
# the function will retun the average.
# e.g. d1=cbu.LoadSingleOutputFile(meta,pNam,0,0,0)
def LoadSingleOutputFile(meta,pNam,iScn,iEns,iBat):

	# Extract indices
	iEP=meta['Core']['iEP']

	# Extract parameters
	bB=meta['Param']['BE']['Biophysical']
	bS=meta['Param']['BEV']['Substitution']

	tv=np.arange(meta[pNam]['Project']['Year Start Saving'],meta[pNam]['Project']['Year End']+1,1)

	# Open batch results
	pth=meta['Paths'][pNam]['Data'] + '\\Outputs\\Scenario' + FixFileNum(iScn) + '\\Data_Scn' + FixFileNum(iScn) + '_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl'
	v0=gu.ipickle(pth)

	# Convert to float and apply scale factor
	for k in v0.keys():

		# Skip mortality summary by agent
		if (k=='C_M_DistByAgent') | (k=='C_M_DistByAgentPct'):
			continue

		# Skip integers (e.g. land cover)
		if v0[k].dtype=='int8':
			continue

		v0[k]=v0[k].astype(float)

		if (k=='E_RHE_ForestSector_Domestic') | (k=='E_EnergySC_Comb') | (k=='E_EnergyT_Comb') | (k=='E_IPPU_Comb'):
			v0[k]=v0[k]*meta['Core']['Scale Factor Export Big']
		else:
			v0[k]=v0[k]*meta['Core']['Scale Factor Export Small']

	# Add year
	it=np.where(meta[pNam]['Year']>=meta[pNam]['Project']['Year Start Saving'])[0]
	v0['Year']=meta[pNam]['Year'][it]

	v0['C_NPP']=v0['C_G_Gross']+v0['C_LF']

	v0['C_M']=v0['C_M_Reg']+v0['C_M_Dist']
	v0['C_M_Harv']=np.zeros(v0['C_NPP'].shape)
	id=meta['LUT']['Event']['Harvest']
	idx0=v0['C_M_DistByAgent'][id]['idx']
	if idx0[0].size>0:
		v0['C_M_Harv'][idx0[0],idx0[1]]=meta['Core']['Scale Factor C_M_DistByAgent']*v0['C_M_DistByAgent'][id]['M'].astype('float64')
	v0['C_M_Nat']=v0['C_M']-v0['C_M_Harv']
	v0['C_G_Net']=v0['C_G_Net_Reg']-v0['C_M_Dist']
	v0['C_ToMillTotal']=v0['C_ToMillMerchGreen']+v0['C_ToMillNonMerchGreen']+v0['C_ToMillMerchDead']+v0['C_ToMillNonMerchDead']
	v0['C_Forest']=v0['C_Biomass']+v0['C_DeadWood']+v0['C_Litter']+v0['C_Soil']+v0['C_Piles']
	v0['C_NonBuildings_Tot']=v0['C_InUse']-v0['C_Buildings']
	v0['C_HWP']=v0['C_InUse']+v0['C_WasteSystems']

	v0['ODT_Sawnwood']=v0['C_ToLumber']/bB['Carbon Content Wood']
	v0['ODT_Panel']=(v0['C_ToPlywood']+v0['C_ToOSB']+v0['C_ToMDF'])/bB['Carbon Content Wood']
	v0['ODT_Lumber']=v0['C_ToLumber']/bB['Carbon Content Wood']
	v0['ODT_LogExport']=v0['C_ToLogExport']/bB['Carbon Content Wood']
	v0['ODT_Plywood']=v0['C_ToPlywood']/bB['Carbon Content Wood']
	v0['ODT_OSB']=v0['C_ToOSB']/bB['Carbon Content Wood']
	v0['ODT_MDF']=v0['C_ToMDF']/bB['Carbon Content Wood']
	v0['ODT_Paper']=v0['C_ToPaper']/bB['Carbon Content Wood']
	v0['ODT_PelletExport']=v0['C_ToBBP_PelletExport']/bB['Carbon Content Wood']
	v0['ODT_PelletDomGrid']=v0['C_ToBBP_PelletDomGrid']/bB['Carbon Content Wood']
	v0['ODT_PelletDomRNG']=v0['C_ToBBP_PelletDomRNG']/bB['Carbon Content Wood']
	v0['ODT_ElectricityGrid']=v0['C_ToBBP_PowerGrid']/bB['Carbon Content Wood']
	v0['ODT_PowerFacilityDom']=v0['C_ToBBP_PowerFacilityDom']/bB['Carbon Content Wood']
	v0['ODT_FirewoodDom']=v0['C_ToBBP_FirewoodDom']/bB['Carbon Content Wood']
	v0['ODT_FirewoodTot']=(v0['C_ToBBP_FirewoodDom']+v0['C_ToBBP_FirewoodExport'])/bB['Carbon Content Wood']

	v0['ODT_Coal']=v0['E_Substitution_Domestic_CoalForBioenergy']/(bB['Emission Intensity Coal']/1000)/bB['Energy Content Coal']
	v0['ODT_Oil']=v0['E_Substitution_Domestic_OilForBioenergy']/(bB['Emission Intensity Oil']/1000)/bB['Energy Content Oil']
	v0['ODT_Gas']=v0['E_Substitution_Domestic_GasForBioenergy']/(bB['Emission Intensity Natural Gas']/1000)/bB['Energy Content Natural Gas']

	# Convert yield of bioenergy feedstock (ODT/ha) to energy (GJ/ha)
	v0['GJ_PelletExport']=v0['ODT_PelletExport']*bB['Energy Content Wood (0% moisture)']*bB['Electrical Conversion Efficiency of Pellet Electricity Plant (>25MW)']
	v0['GJ_PelletDomGrid']=v0['ODT_PelletDomGrid']*bB['Energy Content Wood (0% moisture)']*bB['Electrical Conversion Efficiency of Pellet Electricity Plant (>25MW)']
	v0['GJ_PelletDomRNG']=v0['ODT_PelletDomRNG']*bB['Energy Content Wood (0% moisture)']*bB['Electrical Conversion Efficiency of Pellet Electricity Plant (>25MW)']
	v0['GJ_PowerGrid']=v0['ODT_ElectricityGrid']*bB['Energy Content Wood (0% moisture)']*bB['Electrical Conversion Efficiency of Pellet Electricity Plant (>25MW)']
	v0['GJ_PowerFacilityDom']=v0['ODT_PowerFacilityDom']*bB['Energy Content Wood (0% moisture)']*bB['Electrical Conversion Efficiency of Pellet Electricity Plant (>25MW)']
	v0['GJ_FirewoodDom']=v0['ODT_FirewoodDom']*bB['Energy Content Wood (0% moisture)']*bB['Electrical Conversion Efficiency of Pellet Electricity Plant (>25MW)']
	v0['GJ_FirewoodTot']=v0['ODT_FirewoodTot']*bB['Energy Content Wood (0% moisture)']*bB['Electrical Conversion Efficiency of Pellet Electricity Plant (>25MW)']

	#--------------------------------------------------------------------------
	# Aggregate forest sector GHG emissions
	#--------------------------------------------------------------------------
	v0['E_NEE_ForestSector_Domestic']=-1*bB['Ratio_CO2_to_C']*(v0['C_NPP']-v0['C_RH'])
	v0['E_NPP_ForestSector_Domestic']=-1*bB['Ratio_CO2_to_C']*v0['C_NPP']
	v0['E_RHE_ForestSector_Domestic']=bB['Ratio_CO2_to_C']*v0['C_RH']

	v0['E_NEE_ForestSector_Internat']=np.zeros(v0['E_NEE_ForestSector_Domestic'].shape)
	v0['E_NPP_ForestSector_Internat']=np.zeros(v0['E_NEE_ForestSector_Domestic'].shape)
	v0['E_RHE_ForestSector_Internat']=np.zeros(v0['E_NEE_ForestSector_Domestic'].shape)

	v0['E_NEE_ForestSector_Total']=v0['E_NEE_ForestSector_Domestic']+v0['E_NEE_ForestSector_Internat']
	v0['E_NPP_ForestSector_Total']=v0['E_NPP_ForestSector_Domestic']+v0['E_NPP_ForestSector_Internat']
	v0['E_RHE_ForestSector_Total']=v0['E_RHE_ForestSector_Domestic']+v0['E_RHE_ForestSector_Internat']
	v0['E_Denit_ForestSector_Total']=v0['E_Denit_ForestSector_Domestic']+v0['E_Denit_ForestSector_Internat']
	v0['E_Volat_ForestSector_Total']=v0['E_Volat_ForestSector_Domestic']+v0['E_Volat_ForestSector_Internat']
	v0['E_Wildfire_ForestSector_Total']=v0['E_Wildfire_ForestSector_Domestic']+v0['E_Wildfire_ForestSector_Internat']
	v0['E_OpenBurning_ForestSector_Total']=v0['E_OpenBurning_ForestSector_Domestic']+v0['E_OpenBurning_ForestSector_Internat']
	v0['E_RHP_ForestSector_Total']=v0['E_RHP_ForestSector_Domestic']+v0['E_RHP_ForestSector_Internat']
	v0['E_BBP_ForestSector_Total']=v0['E_BBP_ForestSector_Domestic']+v0['E_BBP_ForestSector_Internat']
	v0['E_HWP_ForestSector_Total']=v0['E_BBP_ForestSector_Total']+v0['E_RHP_ForestSector_Total']

	#--------------------------------------------------------------------------
	# Aggregate forest operations emissions
	#--------------------------------------------------------------------------
	v0['E_ForestryOps_EnergySC_Domestic']=v0['E_ForestryOps_EnergySC_Domestic_Coal']+v0['E_ForestryOps_EnergySC_Domestic_Oil']+v0['E_ForestryOps_EnergySC_Domestic_Gas']
	v0['E_ForestryOps_EnergyT_Domestic']=v0['E_ForestryOps_EnergyT_Domestic_Coal']+v0['E_ForestryOps_EnergyT_Domestic_Oil']+v0['E_ForestryOps_EnergyT_Domestic_Gas']
	v0['E_ForestryOps_IPPU_Domestic']=v0['E_ForestryOps_IPPU_Domestic_Coal']+v0['E_ForestryOps_IPPU_Domestic_Oil']+v0['E_ForestryOps_IPPU_Domestic_Gas']

	v0['E_ForestryOps_Domestic_Coal']=v0['E_ForestryOps_EnergySC_Domestic_Coal']+v0['E_ForestryOps_EnergyT_Domestic_Coal']+v0['E_ForestryOps_IPPU_Domestic_Coal']
	v0['E_ForestryOps_Domestic_Oil']=v0['E_ForestryOps_EnergySC_Domestic_Oil']+v0['E_ForestryOps_EnergyT_Domestic_Oil']+v0['E_ForestryOps_IPPU_Domestic_Oil']
	v0['E_ForestryOps_Domestic_Gas']=v0['E_ForestryOps_EnergySC_Domestic_Gas']+v0['E_ForestryOps_EnergyT_Domestic_Gas']+v0['E_ForestryOps_IPPU_Domestic_Gas']
	v0['E_ForestryOps_Domestic_Total']=v0['E_ForestryOps_Domestic_Coal']+v0['E_ForestryOps_Domestic_Oil']+v0['E_ForestryOps_Domestic_Gas']

	v0['E_ForestryOps_Internat_Total']=np.zeros(v0['E_NEE_ForestSector_Domestic'].shape)

	v0['E_ForestryOps_EnergySC_Total']=v0['E_ForestryOps_EnergySC_Domestic']
	v0['E_ForestryOps_EnergyT_Total']=v0['E_ForestryOps_EnergyT_Domestic']
	v0['E_ForestryOps_IPPU_Total']=v0['E_ForestryOps_IPPU_Domestic']

	v0['E_ForestryOps_Total']=v0['E_ForestryOps_Domestic_Total']+v0['E_ForestryOps_Internat_Total']

	#--------------------------------------------------------------------------
	# Aggregate bioenergy combustion emissions
	#--------------------------------------------------------------------------

	v0['E_BBP_ForestSector_Total']=v0['E_BBP_ForestSector_Domestic']+v0['E_BBP_ForestSector_Internat']

	#--------------------------------------------------------------------------
	# Aggregate substitution emissions
	#--------------------------------------------------------------------------
	# Revise the sign of substitution effects
	vL=['E_Substitution_Domestic_CoalForBioenergy',
		'E_Substitution_Domestic_OilForBioenergy',
		'E_Substitution_Domestic_GasForBioenergy',
		'E_Substitution_Domestic_PowerFacility',
		'E_Substitution_Domestic_ElectricityGrid',
		'E_Substitution_Domestic_PelletElectricityGrid',
		'E_Substitution_Domestic_PelletRNG',
		'E_Substitution_Domestic_Firewood',
		'E_Substitution_Internat_CoalForBioenergy',
		'E_Substitution_Internat_OilForBioenergy',
		'E_Substitution_Internat_GasForBioenergy',
		'E_Substitution_Internat_PowerFacility',
		'E_Substitution_Internat_Pellet',
		'E_Substitution_Internat_Firewood',
		'E_Substitution_Internat_CoalForSolidWood',
		'E_Substitution_Internat_OilForSolidWood',
		'E_Substitution_Internat_GasForSolidWood',
		'E_Substitution_Internat_Sawnwood',
		'E_Substitution_Internat_Panel',
		'E_Substitution_Internat_Concrete',
		'E_Substitution_Internat_ConcreteFromCalcination',
		'E_Substitution_Internat_Steel',
		'E_Substitution_Internat_Aluminum',
		'E_Substitution_Internat_Plastic',
		'E_Substitution_Internat_Textile']
	for v in vL:
		v0[v]=-1*v0[v]

	# Reverse sign of building material production (saved as positive)
	vL=['ODT_Concrete','ODT_Steel','ODT_Aluminum','ODT_Plastic','ODT_Textile','ODT_Coal','ODT_Oil','ODT_Gas']
	for v in vL:
		v0[v]=-1*v0[v]

	v0['E_Substitution_Coal']=v0['E_Substitution_Domestic_CoalForBioenergy']+v0['E_Substitution_Internat_CoalForSolidWood']
	v0['E_Substitution_Oil']=v0['E_Substitution_Domestic_OilForBioenergy']+v0['E_Substitution_Internat_OilForSolidWood']
	v0['E_Substitution_Gas']=v0['E_Substitution_Domestic_GasForBioenergy']+v0['E_Substitution_Internat_GasForSolidWood']

	v0['E_Substitution_Energy']=v0['E_Substitution_Domestic_CoalForBioenergy']+ \
		v0['E_Substitution_Domestic_OilForBioenergy']+ \
		v0['E_Substitution_Domestic_GasForBioenergy']+ \
		v0['E_Substitution_Internat_CoalForBioenergy']+ \
		v0['E_Substitution_Internat_OilForBioenergy']+ \
		v0['E_Substitution_Internat_GasForBioenergy']

	v0['E_Substitution_Material']=v0['E_Substitution_Internat_CoalForSolidWood']+ \
		v0['E_Substitution_Internat_OilForSolidWood']+ \
		v0['E_Substitution_Internat_GasForSolidWood']+ \
		v0['E_Substitution_Internat_ConcreteFromCalcination']

	v0['E_Substitution_Total']=v0['E_Substitution_Material']+v0['E_Substitution_Energy']

	v0['E_Substitution_EnergySC_Total']=bS['Fraction_Energy_Subs_Counted_In_StationaryCombustion']*(v0['E_Substitution_Total']-v0['E_Substitution_Internat_ConcreteFromCalcination'])
	v0['E_Substitution_EnergyT_Total']=(1-bS['Fraction_Energy_Subs_Counted_In_StationaryCombustion'])*(v0['E_Substitution_Total']-v0['E_Substitution_Internat_ConcreteFromCalcination'])
	v0['E_Substitution_IPPU_Total']=v0['E_Substitution_Internat_ConcreteFromCalcination']

	#--------------------------------------------------------------------------
	# Summarize emissions from each type of fossil deposit
	#--------------------------------------------------------------------------
	v0['E_Coal']=v0['E_ForestryOps_Domestic_Coal']+v0['E_Substitution_Coal']
	v0['E_Oil']=v0['E_ForestryOps_Domestic_Oil']+v0['E_Substitution_Oil']
	v0['E_Gas']=v0['E_ForestryOps_Domestic_Gas']+v0['E_Substitution_Gas']

	#--------------------------------------------------------------------------
	# Lateral wood transfer (tCO2e/ha/yr)
	#--------------------------------------------------------------------------
	v0['E_LW_ForestSector_Domestic']=bB['Ratio_CO2_to_C']*v0['C_ToMillTotal']

	#--------------------------------------------------------------------------
	# Net ecosystem balance (tCO2e/ha/yr)
	#--------------------------------------------------------------------------
	v0['E_NEB']=v0['E_NEE_ForestSector_Domestic']+ \
		v0['E_Wildfire_ForestSector_Domestic']+ \
		v0['E_OpenBurning_ForestSector_Domestic']+ \
		v0['E_Denit_ForestSector_Domestic']+ \
		v0['E_Volat_ForestSector_Domestic']+ \
		v0['E_LW_ForestSector_Domestic']

	#--------------------------------------------------------------------------
	# Net sector balance (tCO2e/ha/yr)
	#--------------------------------------------------------------------------
	v0['E_NSB']=v0['E_NPP_ForestSector_Total']+ \
		v0['E_RHE_ForestSector_Total']+ \
		v0['E_Denit_ForestSector_Total']+ \
		v0['E_Volat_ForestSector_Total']+ \
		v0['E_Wildfire_ForestSector_Total']+ \
		v0['E_OpenBurning_ForestSector_Total']+ \
		v0['E_BBP_ForestSector_Total']+ \
		v0['E_RHP_ForestSector_Total']+ \
		v0['E_ForestryOps_Total']

	#--------------------------------------------------------------------------
	# Net atmospheric emissions (tCO2e/ha/yr)
	#--------------------------------------------------------------------------
	v0['E_NAB']=v0['E_NPP_ForestSector_Total']+ \
		v0['E_RHE_ForestSector_Total']+ \
		v0['E_Denit_ForestSector_Total']+ \
		v0['E_Volat_ForestSector_Total']+ \
		v0['E_Wildfire_ForestSector_Total']+ \
		v0['E_OpenBurning_ForestSector_Total']+ \
		v0['E_BBP_ForestSector_Total']+ \
		v0['E_RHP_ForestSector_Total']+ \
		v0['E_ForestryOps_Total']+ \
		v0['E_Substitution_Total']

	#--------------------------------------------------------------------------
	# Fluxes (following Byrne et al. 2023)
	#--------------------------------------------------------------------------
	flg=0
	if flg==1:
		v0['E_NetBiosphereExchange']=v0['E_NEE_ForestSector_Domestic']+\
			v0['E_Domestic_ForestSector_Fire']+\
			v0['E_Denit_ForestSector_Domestic']+ \
			v0['E_Volat_ForestSector_Domestic']+ \
			v0['E_RHE_ForestSector_Domestic']+ \
			v0['E_Domestic_Bioenergy']+\
			v0['E_NEE_ForestSector_Internat']+\
			v0['E_Internat_ForestSector_Fire']+\
			v0['E_Internat_ForestSector_HWP']+ \
			v0['E_Internat_Bioenergy']

		v0['E_FossilFuels']=v0['E_Substitution_Energy']+ \
			v0['E_Substitution_Material']+ \
			v0['E_ForestOperations']

	#--------------------------------------------------------------------------
	# Discounting
	#--------------------------------------------------------------------------
	r_disc=bB['Discount Rate Emissions']
	t_disc=np.maximum(0,tv-meta[pNam]['Project']['Year Project'])
	t_disc=np.tile(t_disc,(v0['E_NAB'].shape[1],1)).T

	v0['E_NSB_Disc']=v0['E_NSB'].copy()/((1+r_disc)**t_disc)
	v0['E_NAB_Disc']=v0['E_NAB'].copy()/((1+r_disc)**t_disc)

	#--------------------------------------------------------------------------
	# Add cumulative
	#--------------------------------------------------------------------------
	v0['E_NSB_Cumulative']=np.cumsum(v0['E_NSB'],axis=0)
	v0['E_NAB_Cumulative']=np.cumsum(v0['E_NAB'],axis=0)

	v0['E_NSB_Disc_Cumulative']=np.cumsum(v0['E_NSB_Disc'],axis=0)
	v0['E_NAB_Disc_Cumulative']=np.cumsum(v0['E_NAB_Disc'],axis=0)

	#--------------------------------------------------------------------------
	# Add cumulative (starting from a specified start year)
	#--------------------------------------------------------------------------
	indT=np.where(v0['Year']<meta[pNam]['Project']['Year Start Cumulative'])[0]

	v0['E_NSB_Cumulative_from_tref']=v0['E_NSB'].copy()
	v0['E_NSB_Cumulative_from_tref'][indT]=0
	v0['E_NSB_Cumulative_from_tref']=np.cumsum(v0['E_NSB_Cumulative_from_tref'],axis=0)

	v0['E_NSB_Disc_Cumulative_from_tref']=v0['E_NSB_Disc'].copy()
	v0['E_NSB_Disc_Cumulative_from_tref'][indT]=0
	v0['E_NSB_Disc_Cumulative_from_tref']=np.cumsum(v0['E_NSB_Disc_Cumulative_from_tref'],axis=0)

	v0['E_NAB_Cumulative_from_tref']=v0['E_NAB'].copy()
	v0['E_NAB_Cumulative_from_tref'][indT]=0
	v0['E_NAB_Cumulative_from_tref']=np.cumsum(v0['E_NAB_Cumulative_from_tref'],axis=0)

	v0['E_NAB_Disc_Cumulative_from_tref']=v0['E_NAB_Disc'].copy()
	v0['E_NAB_Disc_Cumulative_from_tref'][indT]=0
	v0['E_NAB_Disc_Cumulative_from_tref']=np.cumsum(v0['E_NAB_Disc_Cumulative_from_tref'],axis=0)

	#--------------------------------------------------------------------------
	# Back-calculate production of fossil fuel consumption from operational use
	# and substitution effects (tonnesC)
	#--------------------------------------------------------------------------

	E_Op=v0['E_ForestryOps_EnergySC_Domestic_Coal']+v0['E_ForestryOps_EnergyT_Domestic_Coal']+v0['E_ForestryOps_IPPU_Domestic_Coal']
	E_Substitution=meta['Param']['BEV']['Substitution']['Economic Contraction Fraction']*(v0['E_Substitution_Domestic_CoalForBioenergy']+v0['E_Substitution_Internat_CoalForSolidWood'])
	v0['C_FromCoal']=-1*(E_Op+E_Substitution)/(bB['Emission Intensity Coal']/1000)/bB['Energy Content Coal']*bB['Carbon Content Coal']

	E_Op=v0['E_ForestryOps_EnergySC_Domestic_Gas']+v0['E_ForestryOps_EnergyT_Domestic_Gas']+v0['E_ForestryOps_IPPU_Domestic_Gas']
	E_Substitution=meta['Param']['BEV']['Substitution']['Economic Contraction Fraction']*(v0['E_Substitution_Domestic_GasForBioenergy']+v0['E_Substitution_Internat_GasForSolidWood'])
	v0['C_FromGas']=-1*(E_Op+E_Substitution)/(bB['Emission Intensity Natural Gas']/1000)/bB['Energy Content Natural Gas']*bB['Carbon Content Natural Gas']

	E_Op=v0['E_ForestryOps_EnergySC_Domestic_Oil']+v0['E_ForestryOps_EnergyT_Domestic_Oil']+v0['E_ForestryOps_IPPU_Domestic_Oil']
	E_Substitution=meta['Param']['BEV']['Substitution']['Economic Contraction Fraction']*(v0['E_Substitution_Domestic_OilForBioenergy']+v0['E_Substitution_Internat_OilForSolidWood'])
	v0['C_FromOil']=-1*(E_Op+E_Substitution)/(bB['Emission Intensity Oil']/1000)/bB['Energy Content Oil']*bB['Carbon Content Oil']

	E_Substitution=meta['Param']['BEV']['Substitution']['Economic Contraction Fraction']*(v0['E_Substitution_Internat_Concrete'])
	v0['C_FromLimestone']=-1*(E_Substitution*bB['Ratio_C_to_CO2'])

	v0['C_Coal']=np.cumsum(v0['C_FromCoal'],axis=0)
	v0['C_Gas']=np.cumsum(v0['C_FromGas'],axis=0)
	v0['C_Oil']=np.cumsum(v0['C_FromOil'],axis=0)
	v0['C_Limestone']=np.cumsum(v0['C_FromLimestone'],axis=0)
	v0['C_Geological']=v0['C_Coal']+v0['C_Gas']+v0['C_Oil']+v0['C_Limestone']

	v0['RF_Biochem']=np.zeros(v0['A'].shape)
	#v0['RF_Biophys']=np.zeros(v0['A'].shape)

	# Track if radiative forcing status is on
	if meta[pNam]['Project']['Radiative Forcing Status']=='On':
		# Add NEE
		v0['E_CO2']=v0['E_CO2']+v0['E_NEE_ForestSector_Domestic']
		# Add denitrification
		v0['E_N2O']=v0['E_N2O']+v0['E_Denit_ForestSector_Domestic']/bB['GWP_N2O'] # Convert tCO2e back to tN2O
		# Add fossil fuels from operations
		v0['E_CO2']=v0['E_CO2']+v0['E_ForestryOps_Domestic_Coal']*bB['Emission Fraction Coal As CO2']
		v0['E_CO2']=v0['E_CO2']+v0['E_ForestryOps_Domestic_Oil']*bB['Emission Fraction Oil As CO2']
		v0['E_CO2']=v0['E_CO2']+v0['E_ForestryOps_Domestic_Gas']*bB['Emission Fraction Natural Gas As CO2']
		v0['E_CH4']=v0['E_CH4']+v0['E_ForestryOps_Domestic_Coal']*bB['Emission Fraction Coal As CH4']/bB['GWP_CH4']
		v0['E_CH4']=v0['E_CH4']+v0['E_ForestryOps_Domestic_Oil']*bB['Emission Fraction Oil As CH4']/bB['GWP_CH4']
		v0['E_CH4']=v0['E_CH4']+v0['E_ForestryOps_Domestic_Gas']*bB['Emission Fraction Natural Gas As CH4']/bB['GWP_CH4']
		v0['E_N2O']=v0['E_N2O']+v0['E_ForestryOps_Domestic_Coal']*bB['Emission Fraction Coal As N2O']/bB['GWP_N2O']
		v0['E_N2O']=v0['E_N2O']+v0['E_ForestryOps_Domestic_Oil']*bB['Emission Fraction Oil As N2O']/bB['GWP_N2O']
		v0['E_N2O']=v0['E_N2O']+v0['E_ForestryOps_Domestic_Gas']*bB['Emission Fraction Natural Gas As N2O']/bB['GWP_N2O']
		# Add fossil fuels from substitutions
		v0['E_CO2']=v0['E_CO2']+v0['E_Substitution_Coal']*bB['Emission Fraction Coal As CO2']
		v0['E_CO2']=v0['E_CO2']+v0['E_Substitution_Oil']*bB['Emission Fraction Oil As CO2']
		v0['E_CO2']=v0['E_CO2']+v0['E_Substitution_Gas']*bB['Emission Fraction Natural Gas As CO2']
		v0['E_CH4']=v0['E_CH4']+v0['E_Substitution_Coal']*bB['Emission Fraction Coal As CH4']/bB['GWP_CH4']
		v0['E_CH4']=v0['E_CH4']+v0['E_Substitution_Oil']*bB['Emission Fraction Oil As CH4']/bB['GWP_CH4']
		v0['E_CH4']=v0['E_CH4']+v0['E_Substitution_Gas']*bB['Emission Fraction Natural Gas As CH4']/bB['GWP_CH4']
		v0['E_N2O']=v0['E_N2O']+v0['E_Substitution_Coal']*bB['Emission Fraction Coal As N2O']/bB['GWP_N2O']
		v0['E_N2O']=v0['E_N2O']+v0['E_Substitution_Oil']*bB['Emission Fraction Oil As N2O']/bB['GWP_N2O']
		v0['E_N2O']=v0['E_N2O']+v0['E_Substitution_Gas']*bB['Emission Fraction Natural Gas As N2O']/bB['GWP_N2O']

	# Sawtooth variable adjustments
	if meta[pNam]['Project']['Biomass Module']=='Sawtooth':
		v0['N']=np.maximum(0,v0['N'])
		v0['N_R']=np.maximum(0,v0['N_R'])
		v0['N_M_Tot']=np.maximum(0,v0['N_M_Tot'])
		v0['TreeMean_D']=np.maximum(0,v0['TreeMean_D'])
		v0['TreeMean_Csw']=np.maximum(0,v0['TreeMean_Csw'])
		v0['TreeMean_Csw_G']=np.maximum(0,v0['TreeMean_Csw_G'])

	vToKeepBasic=['Year',
	 'A',
	 'C_Biomass',
	 'C_Buildings',
	 'C_DeadWood',
	 'C_WasteSystems',
	 'C_Felled',
	 'C_Forest',
	 'C_G_Gross',
	 'C_G_Net',
	 'C_G_Net_Reg',
	 'C_Geological',
	 'C_HWP',
	 'C_InUse',
	 'C_LF',
	 'C_Litter',
	 'C_M',
	 'C_M_Dist',
	 'C_M_DistByAgent',
	 'C_M_DistByAgentPct',
	 'C_M_Nat',
	 'C_M_Reg',
	 'C_NPP',
	 'C_Piles',
	 'C_RH',
	 'C_Soil',
	 'C_StemMerch',
	 'C_StemNonMerch',
	 'C_ToMillTotal',
	 'C_ToPileBurnTot',
	 'E_NPP_ForestSector_Total',
	 'E_RHE_ForestSector_Total',
	 'E_NEE_ForestSector_Total',
	 'E_Wildfire_ForestSector_Total',
	 'E_OpenBurning_ForestSector_Total',
	 'E_Denit_ForestSector_Total',
	 'E_Volat_ForestSector_Total',
	 'E_BBP_ForestSector_Total',
	 'E_RHP_ForestSector_Total',
	 'E_HWP_ForestSector_Total',
	 'E_ForestryOps_EnergySC_Total',
	 'E_ForestryOps_EnergyT_Total',
	 'E_ForestryOps_IPPU_Total',
	 'E_ForestryOps_Total',
	 'E_Substitution_Total',
	 'E_NEB',
	 'E_NSB',
	 'E_NSB_Cumulative_from_tref',
	 'E_NAB',
	 'E_NAB_Cumulative_from_tref',
	 'V_MerchDead',
	 'V_MerchLive',
	 'V_MerchTotal',
	 'V_ToMill_MerchDead',
	 'V_ToMill_MerchGreen',
	 'V_ToMill_MerchTotal',
	 'V_ToMill_NonMerchDead',
	 'V_ToMill_NonMerchGreen',
	 'V_ToMill_NonMerchTotal']

	vToKeepEcon=['ODT_Lumber','ODT_Plywood','ODT_OSB','ODT_MDF','ODT_Paper','ODT_FirewoodDom','ODT_LogExport',
			'GJ_PowerFacilityDom','GJ_PowerGrid','GJ_PelletExport','GJ_PelletDomGrid','GJ_PelletDomRNG','GJ_PowerGrid']

	if meta[pNam]['Project']['Save List Type']=='Basic':
		vToKeep=vToKeepBasic
		v1={}
		for k in v0.keys():
			if np.isin(k,vToKeep)==True:
				v1[k]=v0[k]
		v0=v1
	elif meta[pNam]['Project']['Save List Type']=='Basic+Econ':
		vToKeep=vToKeepBasic+vToKeepEcon
		v1={}
		for k in v0.keys():
			if np.isin(k,vToKeep)==True:
				v1[k]=v0[k]
		v0=v1

	# Put in alphabetical order
	vs={key: value for key, value in sorted(v0.items())}

	return vs

#%% LOAD SCENARIO RUSULTS
# Return a list of dictionaries for each scenario. If multiple ensemble were run,
# the function will retun the average.

def LoadScenarioResults(meta,pNam):

	# Initialize list that will contain scenarios
	v1=[]
	for iScn in range(meta[pNam]['Project']['N Scenario']):

		for iEns in range(meta[pNam]['Project']['N Ensemble']):

			for iBat in range(meta[pNam]['Project']['N Batch']):

				#--------------------------------------------------------------
				# Open batch results
				#--------------------------------------------------------------

				data_batch=LoadSingleOutputFile(meta,pNam,iScn,iEns,iBat)

				# Import event chronology
				if (meta[pNam]['Scenario'][iScn]['Harvest Status Future']=='On') | (meta[pNam]['Scenario'][iScn]['Breakup Status Historical']=='On') | (meta[pNam]['Scenario'][iScn]['Breakup Status Future']=='On'):
					ec=gu.ipickle(meta['Paths'][pNam]['Input Scenario'][iScn] + '\\Modified_Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
				else:
					ec=gu.ipickle(meta['Paths'][pNam]['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')

				# Uncompress event chronology if it has been compressed
				ec=EventChronologyDecompress(meta,ec,iScn,iEns,iBat)

				# Land surface attribues
				inv=gu.ipickle(meta['Paths'][pNam]['Input Scenario'][iScn] + '\\Inventory_Bat' + FixFileNum(iBat) + '.pkl')

				# Cashflow
				econ=econo.CalculateNetRevenue(meta,iScn,iEns,iBat,lsat,ec,data_batch)

				data_batch.update(econ)

				#--------------------------------------------------------------
				# Accumulate data in each batch
				#--------------------------------------------------------------

				if iBat==0:

					data_all=data_batch

				else:

					for key1 in data_batch.keys():

						if key1=='Year':
							# Only needed once
							continue

						elif (key1=='C_M_DistByAgent') | (key1=='C_M_DistByAgentPct'):
							# Nested dictionary
							for key2 in data_batch[key1].keys():
								data_all[key1][key2]=np.append(data_all[key1][key2],data_batch[key1][key2],axis=1)

						else:
							# No nested dictionary
							data_all[key1]=np.append(data_all[key1],data_batch[key1],axis=1)

			#------------------------------------------------------------------
			# Sum across ensembles
			#------------------------------------------------------------------

			if iEns==0:

				data_sum2ave=data_all

			else:

				for key1 in data_batch.keys():

					if (key1=='C_M_DistByAgent') | (key1=='C_M_DistByAgentPct'):

						# Nested dictionary
						for key2 in data_batch[key1].keys():
							data_sum2ave[key1][key2]=data_sum2ave[key1][key2]+data_all[key1][key2]

					else:

						# No nested dictionary
						data_sum2ave[key1]=data_sum2ave[key1]+data_all[key1]

		#----------------------------------------------------------------------
		# If the simulation includes ensembles, calculate average
		#----------------------------------------------------------------------

		for key1 in data_batch.keys():

			# Skip mortality summary by agent
			if (key1=='C_M_DistByAgent'):

				# Nested dictioanry
				for key2 in data_batch[key1].keys():
					data_sum2ave[key1][key2]=data_sum2ave[key1][key2]/meta[pNam]['Project']['N Ensemble']

			else:

				# No nested dictionary
				data_sum2ave[key1]=data_sum2ave[key1]/meta[pNam]['Project']['N Ensemble']

		#----------------------------------------------------------------------
		# Add year
		#----------------------------------------------------------------------

		it=np.where(meta[pNam]['Year']>=meta[pNam]['Project']['Year Start Saving'])[0]
		data_sum2ave['Year']=meta[pNam]['Year'][it]

		#----------------------------------------------------------------------
		# Append to list
		#----------------------------------------------------------------------

		v1.append(data_sum2ave)

	return v1



#%% GET TASS GROWTH CURVES
def GetTASSCurves(meta,iScn,iGC,fny,fnc,fnm):

	# Assume just one stand per scenario (i.e., demo)
	iS=0

	dfY=pd.read_csv(fny,header=30)

	# Initialize age response of net growth
	G=np.zeros((N_Age,1,6),dtype='int16')

	G[:,iS,0]=G_StemMerch[:,indTIPSY[0]]/meta['Modules']['GYM']['Scale Factor']
	G[:,iS,1]=G_StemNonMerch[:,indTIPSY[0]]/meta['Modules']['GYM']['Scale Factor']
	G[:,iS,2]=G_Bark[:,indTIPSY[0]]/meta['Modules']['GYM']['Scale Factor']
	G[:,iS,3]=G_Branch[:,indTIPSY[0]]/meta['Modules']['GYM']['Scale Factor']
	G[:,iS,4]=G_Foliage[:,indTIPSY[0]]/meta['Modules']['GYM']['Scale Factor']
	G[:,iS,5]=G_VStemMerch[:,indTIPSY[0]]/meta['Modules']['GYM']['Scale Factor']

	# Save data to file in input variables folder of project
	gu.opickle(meta['Paths'][pNam]['Data'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve' + str(iGC+1) + '_Bat' + FixFileNum(1) + '.pkl',G)

	return

#%%
def Import_BatchTIPSY_Output(meta,pNam,iScn,iGC):

	# Import unique growth curves
	ugc=gu.ipickle(meta['Paths'][pNam]['Data'] + '\\Inputs\\ugc.pkl')

	# TIPSY exports curves as MgDM/ha/yr, CBRunner expects inputs of MgC/ha/yr. Create
	# conversion factor.
	dm2c=0.5

	# Growth curve parameters and TIPSY outputs
	#dfPar=pd.read_excel(meta['Paths'][pNam]['Data'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx',sheet_name='Sheet1',skiprows=7)
	txtDat=np.loadtxt(meta['Paths'][pNam]['Data'] + '\\Inputs\\GrowthCurvesTIPSY_Output.out',skiprows=4)

	# TIPSY saves to text file -> convert to dataframe (column names must match TIPSY output file design)
	dfDat=pd.DataFrame(txtDat,columns=meta['Modules']['GYM']['BatchTIPSY Column Names'])

	del txtDat

	# Define age vector (must be consistent with how TIPSY was set up)
	Age=np.arange(0,meta['Modules']['GYM']['BatchTIPSY Maximum Age']+1,1)

	# Get dimensions of the TIPSY output file to reshape the data into Age x Stand
	N_Age=Age.size
	N_GC=int(dfDat.shape[0]/N_Age)

	# Define the fraction of merchantable stemwood
	fMerch=np.nan_to_num(np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')/np.reshape(dfDat['VolTot0'].values,(N_Age,N_GC),order='F'))
	fNonMerch=1-fMerch

	# Merchantable stemwood volume
	V_Merch=np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')
	G_VStemMerch=np.append(np.zeros((1,N_GC)),np.diff(V_Merch,axis=0),axis=0)

	# Extract age responses for each biomass pool
	C_Stem=dm2c*np.reshape(dfDat['ODT_Stem'].values,(N_Age,N_GC),order='F')
	C_StemMerch=fMerch*C_Stem
	C_StemNonMerch=fNonMerch*C_Stem
	C_Foliage=dm2c*np.reshape(dfDat['ODT_Foliage'].values,(N_Age,N_GC),order='F')
	C_Branch=dm2c*np.reshape(dfDat['ODT_Branch'].values,(N_Age,N_GC),order='F')
	C_Bark=dm2c*np.reshape(dfDat['ODT_Bark'].values,(N_Age,N_GC),order='F')

	# Calculate growth
	z=np.zeros((1,N_GC))
	G_StemMerch=np.append(z,np.diff(C_StemMerch,axis=0),axis=0)
	G_StemNonMerch=np.append(z,np.diff(C_StemNonMerch,axis=0),axis=0)
	G_Stem=G_StemMerch+G_StemNonMerch
	G_Foliage=np.append(z,np.diff(C_Foliage,axis=0),axis=0)
	G_Branch=np.append(z,np.diff(C_Branch,axis=0),axis=0)
	G_Bark=np.append(z,np.diff(C_Bark,axis=0),axis=0)

	# Fix growth of year zero
	G_Stem[0,:]=G_Stem[1,:]
	G_StemMerch[0,:]=G_StemMerch[1,:]
	G_StemNonMerch[0,:]=G_StemNonMerch[1,:]
	G_Foliage[0,:]=G_Foliage[1,:]
	G_Branch[0,:]=G_Branch[1,:]
	G_Bark[0,:]=G_Bark[1,:]

	#del C_Stem,C_StemMerch,C_StemNonMerch,C_Foliage,C_Branch,C_Bark,fMerch,fNonMerch

	# Index to the full set of growth curves for scenario iScn and growth curve iGC
	ind_ugc_ScnAndGc=np.where( (ugc['Full'][:,1]==iScn) & (ugc['Full'][:,2]==meta['Modules']['GYM']['ID GC Unique'][iGC]) )[0]

	# Extract the unique growth curve ID for scenario iScn and growth curve iGC
	ID_ugc_ScnAndGc=ugc['Full'][ind_ugc_ScnAndGc,0]

	# Extract the inverse index for scenario iScn and growth curve iGC
	Inverse_ugc_ScnAndGc=ugc['Inverse'][ind_ugc_ScnAndGc]

	# Intersect
	ind=np.arange(0,meta[pNam]['Project']['N Stand'],1,dtype=int)
	c,inda,indb=np.intersect1d(ID_ugc_ScnAndGc,ind,return_indices=True)

	Inverse_ugc_ScnAndGcAndBat=Inverse_ugc_ScnAndGc[inda]

	# Initialize array of growth data
	d={}
	d['Csw']=np.zeros((N_Age,ind.size))
	d['Gsw_Net']=np.zeros((N_Age,ind.size))
	for i in range(inda.size):
		iStand=indb[i]
		iGC_Unique=Inverse_ugc_ScnAndGcAndBat[i]
		d['Csw'][:,iStand]=C_Stem[:,iGC_Unique].T
		d['Gsw_Net'][:,iStand]=G_Stem[:,iGC_Unique].T

	return d

#%% POST-PROCESS TIPSY GROWTH CURVES
# Nested list, gc[Scenario][Stand][Growth Curve]
# *** This is problematic - I think it only works when total GCs = unique GCs ***
def Import_BatchTIPSY_Output_OLD(meta):

	# Growth curve parameters and TIPSY outputs
	dfPar=pd.read_excel(meta['Paths'][pNam]['Data'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx',sheet_name='Sheet1',skiprows=6)
	txtDat=np.loadtxt(meta['Paths'][pNam]['Data'] + '\\Inputs\\GrowthCurvesTIPSY_Output.out',skiprows=4)

	# TIPSY saves to text file -> convert to dataframe (column names must match TIPSY output file design)
	dfDat=pd.DataFrame(txtDat,columns=meta['Modules']['GYM']['BatchTIPSY Column Names'])

	# Define age vector (must be consistent with how TIPSY was set up)
	Age=np.arange(0,meta['Modules']['GYM']['BatchTIPSY Maximum Age']+1,1)

	# Get dimensions of the TIPSY output file to reshape the data into Age x Stand
	N_Age=Age.size
	N_GC=int(dfDat.shape[0]/N_Age)

	gc=[None]*N_GC
	for i in range(N_GC):
		gc[i]={}
	for i in range(len(meta['Modules']['GYM']['BatchTIPSY Column Names'])):
		data=np.reshape(dfDat[meta['Modules']['GYM']['BatchTIPSY Column Names'][i]].values,(N_Age,N_GC),order='F')
		for j in range(N_GC):
			gc[j][meta['Modules']['GYM']['BatchTIPSY Column Names'][i]]=data[:,j]

	gc2=[]
	uScn=np.unique(dfPar['ID_Scenario'])
	for iScn in range(uScn.size):
		ind=np.where(dfPar['ID_Scenario']==uScn[iScn])[0]
		uStand=np.unique(dfPar.loc[ind,'ID_Stand'])
		gc1=[]
		for iS in range(uStand.size):
			ind=np.where( (dfPar['ID_Scenario']==uScn[iScn]) & (dfPar['ID_Stand']==uStand[iS]) )[0]
			uGC=np.unique(dfPar.loc[ind,'ID_GC'])
			gc0=[]
			for iGC in range(uGC.size):
				ind=np.where( (dfPar['ID_Scenario']==uScn[iScn]) & (dfPar['ID_Stand']==uStand[iS]) & (dfPar['ID_GC']==uGC[iGC]) )[0]
				d={}
				for i in range(len(meta['Modules']['GYM']['BatchTIPSY Column Names'])):
					data=np.reshape(dfDat[meta['Modules']['GYM']['BatchTIPSY Column Names'][i]].values,(N_Age,N_GC),order='F')
					d[meta['Modules']['GYM']['BatchTIPSY Column Names'][i]]=data[:,ind]
				gc0.append(d)
			gc1.append(gc0)
		gc2.append(gc1)
	return gc2

#%% Import growth curves
# Nested list of: Scenario, growth curve ID, stand
def Import_CompiledGrowthCurves(meta,pNam,scn):
	gc=[]
	for iScn in range(len(scn)): #range(meta[pNam]['Project']['N Scenario']):
		gc0=[]

		gc1=[]
		for iBat in range(0,meta[pNam]['Project']['N Batch']):
			tmp=gu.ipickle(meta['Paths'][pNam]['Data'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve1_Bat' + FixFileNum(iBat) + '.pkl')
			tmp=tmp[:,:,0].astype(float)
			for iS in range(tmp.shape[1]):
				gc1.append(tmp[:,iS].copy()*meta['Modules']['GYM']['Scale Factor'])
		gc0.append(gc1.copy())

		gc1=[]
		for iBat in range(0,meta[pNam]['Project']['N Batch']):
			tmp=gu.ipickle(meta['Paths'][pNam]['Data'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve2_Bat' + FixFileNum(iBat) + '.pkl')
			tmp=tmp[:,:,0].astype(float)
			for iS in range(tmp.shape[1]):
				gc1.append(tmp[:,iS].copy()*meta['Modules']['GYM']['Scale Factor'])
		gc0.append(gc1.copy())

		gc1=[]
		for iBat in range(0,meta[pNam]['Project']['N Batch']):
			tmp=gu.ipickle(meta['Paths'][pNam]['Data'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve3_Bat' + FixFileNum(iBat) + '.pkl')
			tmp=tmp[:,:,0].astype(float)
			for iS in range(tmp.shape[1]):
				gc1.append(tmp[:,iS].copy()*meta['Modules']['GYM']['Scale Factor'])
		gc0.append(gc1.copy())
		gc.append(gc0.copy())
	return gc

#%% WRITE SPREADSHEET OF BatchTIPSY PARAMTERS
def Write_BatchTIPSY_Input_Spreadsheet(meta,pNam,ugc):

	# Create a function that will return the column corresponding to a variable name
	fin=meta['Paths']['Model']['Code'] + '\\Parameters\\Parameters_TemplateBatchTIPSY.xlsx'
	df_frmt=pd.read_excel(fin,sheet_name='Sheet1')
	gy_labels=df_frmt.loc[5,:].values
	def GetColumn(lab):
		ind=np.where(gy_labels==lab)[0]
		return int(ind+1)

	# Open spreadsheet
	#PathGrowthCurveParameters=meta['Paths'][pNam]['Data'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx'
	PathGrowthCurveParameters=fin
	xfile=openpyxl.load_workbook(PathGrowthCurveParameters)
	sheet=xfile.get_sheet_by_name('Sheet1')
	N_headers=7

	# # Overwrite existing data entries with empty cells
	# # *** This is really important - failing to wipe it clean first will lead to
	# # weird parameters ***
	# for i in range(int(1.5*ugc['Unique'].shape[0])):
	#	 for j in range(len(gy_labels)):
	#		 sheet.cell(row=i+1+N_headers,column=j+1).value=''

	# Initialize counter
	cnt=1

	# Loop through unique stand types
	for iUGC in range(ugc['Unique'].shape[0]):

		# It isn't necessary to populate these, as we will use the crosswalk in
		# Python session instead
		sheet.cell(row=cnt+N_headers,column=1).value=cnt
		sheet.cell(row=cnt+N_headers,column=2).value=0
		sheet.cell(row=cnt+N_headers,column=3).value=0
		sheet.cell(row=cnt+N_headers,column=4).value=0

		# Regeneration type (N, C, P)
		vnam='regeneration_method'
		id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		cd=lut_n2s(meta['LUT']['TIPSY'][vnam],id)[0]
		sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd

		# Species 1
		vnam='s1'
		id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		cd=lut_n2s(meta['LUT']['VEG_COMP_LYR_R1_POLY']['SPECIES_CD_1'],id)[0]
		sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd

		vnam='p1'
		dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)

		vnam='i1'
		dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)

		vnam='gain1'
		dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		if dat!=9999:
			sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
			sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(10)

		# Species 2
		vnam='s2'
		id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		if id!=9999:
			cd=lut_n2s(meta['LUT']['VEG_COMP_LYR_R1_POLY']['SPECIES_CD_1'],id)[0]
			sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd

			vnam='p2'
			dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
			sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)

			vnam='gain2'
			dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
			if dat!=9999:
				sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
				sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(10)

		# Species 3
		vnam='s3'
		id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		if id!=9999:
			cd=lut_n2s(meta['LUT']['VEG_COMP_LYR_R1_POLY']['SPECIES_CD_1'],id)[0]
			sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd

			vnam='p3'
			dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
			sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)

			vnam='gain3'
			dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
			if dat!=9999:
				sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
				sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(10)

		# Species 4
		vnam='s4'
		id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		if id!=9999:
			cd=lut_n2s(meta['LUT']['VEG_COMP_LYR_R1_POLY']['SPECIES_CD_1'],id)[0]
			sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd

			vnam='p4'
			dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
			sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)

			vnam='gain4'
			dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
			if dat!=9999:
				sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
				sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(10)

		# Species 5
		vnam='s5'
		id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		if id!=9999:
			cd=lut_n2s(meta['LUT']['VEG_COMP_LYR_R1_POLY']['SPECIES_CD_1'],id)[0]
			sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd

			vnam='p5'
			dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
			sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)

			vnam='gain5'
			dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
			if dat!=9999:
				sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
				sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(10)

		# Planting density
		vnam='init_density'
		dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)

		# Regeneration delay
		vnam='regen_delay'
		dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)

		# OAF1
		vnam='oaf1'
		dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=dat[0]

		# OAF2
		vnam='oaf2'
		dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=dat[0]

		# BEC zone
		vnam='bec_zone'
		id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		cd=lut_n2s(meta['LUT']['BEC_BIOGEOCLIMATIC_POLY']['ZONE'],id)[0]
		sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd

		# FIZ
		vnam='FIZ'
		id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		cd=lut_n2s(meta['LUT']['TIPSY']['FIZ'],id)[0]
		sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd

		# Age at Aerial BTK Spray
		# vnam='fert_age1'
		#dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
		#if dat!=9999:
		#	sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat[0])

		# Update counter
		cnt=cnt+1

	#------------------------------------------------------------------------------
	# Save to spreadsheet
	#------------------------------------------------------------------------------

	xfile.save(meta['Paths'][pNam]['Data'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx')



#%% Import disturbance history
# def GetDisturbanceHistory(meta):
# 	dh=[]
# 	for iScn in range(meta[pNam]['Project']['N Scenario']):
# 		dhB=[]
# 		for iBat in range(meta[pNam]['Project']['N Batch']):
# 			iEns=0
# 			dh0=gu.ipickle(meta['Paths'][pNam]['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
# 			if iBat==0:
# 				dhB=dh0
# 			else:
# 				for i in range(len(dh0)):
# 					dhB.append(dh0[i])
# 			dh.append(dhB)
# 	return dh

#%% Graphics

# Look at variables in rcParams:
#plt.rcParams.keys()
def Import_GraphicsParameters(x):

	params={}

	if x=='FCI_Demo':

		fs1=6
		fs2=7

		params={'font.sans-serif':'Arial',
				'font.size':fs1,
				'figure.titlesize':fs2,
				'figure.dpi':150,
				'figure.constrained_layout.use':True,
				'axes.edgecolor':'black',
				'axes.labelsize':fs1,
				'axes.labelcolor':'black',
				'axes.titlesize':fs2,
				'axes.titlepad':2,
				'axes.linewidth':0.5,
				'lines.linewidth':1,
				'text.color':'black',
				'xtick.color':'black',
				'xtick.labelsize':fs1,
				'xtick.major.width':0.5,
				'xtick.major.size':3,
				'xtick.direction':'in',
				'ytick.color':'black',
				'ytick.labelsize':fs1,
				'ytick.major.width':0.5,
				'ytick.major.size':3,
				'ytick.direction':'in',
				'legend.fontsize':fs1,
				'savefig.dpi':900,
				'savefig.transparent':True,
				'savefig.format':'png',
				'savefig.pad_inches':0.1,
				'savefig.bbox':'tight'}

	elif x=='bc1ha_1':

		params={'font.sans-serif':'Arial',
				'font.size':7,
				'axes.labelsize':7,
				'axes.titlesize':14,
				'axes.linewidth':0.5,
				'xtick.labelsize':7,
				'xtick.major.width':0.5,
				'xtick.major.size':5,
				'xtick.direction':'in',
				'ytick.labelsize':7,
				'ytick.major.width':0.5,
				'ytick.major.size':5,
				'ytick.direction':'in',
				'legend.fontsize':10,
				'savefig.dpi':150}

	elif (x=='Article') | (x=='article'):

		fs=6

		params={'font.sans-serif':'Arial','font.size':fs,'axes.edgecolor':'black','axes.labelsize':fs,'axes.labelcolor':'black',
				'axes.titlesize':fs,'axes.linewidth':0.5,'lines.linewidth':0.5,'text.color':'black','xtick.color':'black',
				'xtick.labelsize':fs,'xtick.major.width':0.5,'xtick.major.size':3,'xtick.direction':'in','ytick.color':'black',
				'ytick.labelsize':fs,'ytick.major.width':0.5,'ytick.major.size':3,'ytick.direction':'in','legend.fontsize':fs}

	else:

		params={}

	return params


#%%
# def PrepareInventoryFromSpreadsheet_OLD(meta,pNam):
# 	for iScn in range(0,meta[pNam]['Project']['N Scenario']):
# 		for iBat in range(0,meta[pNam]['Project']['N Batch']):
# 			lsat={}

# 			# Index to batch
# 			indBat=IndexToBatch(meta[pNam],iBat)
# 			N_StandsInBatch=len(indBat)

# 			# Initialize inventory variables
# 			lsat['Lat']=np.zeros((1,N_StandsInBatch))
# 			lsat['Lon']=np.zeros((1,N_StandsInBatch))
# 			lsat['X']=lsat['Lat']
# 			lsat['Y']=lsat['Lon']

# 			# BEC zone
# 			lsat['ID_BGCZ']=np.zeros((1,N_StandsInBatch),dtype='int16')
# 			lsat['ID_BGCZ'][0,:]=meta['LUT']['VEG_COMP_LYR_R1_POLY']['BEC_ZONE_CODE'][meta[pNam]['Scenario'][iScn]['BGC Zone Code']]

# 			# Timber harvesting landbase (1=yes, 0=no)
# 			lsat['THLB']=meta[pNam]['Scenario'][iScn]['THLB Status']*np.ones((meta[pNam]['Year'].size,N_StandsInBatch))

# 			# Temperature will be updated automatically
# 			lsat['MAT']=4.0*np.ones((1,N_StandsInBatch))
# 			#cd=meta[pNam]['Scenario'][iScn]['BGC Zone Code']
# 			#mat=meta['Param']['BE']['ByBGCZ'][cd]['MAT']
# 			#lsat['MAT']=mat*np.ones((1,N_StandsInBatch))

# 			lsat['Wood Density']=meta['Param']['BE']['Biophysical']['Density Wood Standard']*np.ones((1,N_StandsInBatch))

# 			if meta[pNam]['Project']['Biomass Module']=='Sawtooth':

# 				ind=np.where( meta['Param']['BE']['Sawtooth']['SRS Key']['SRS_CD']==meta[pNam]['Scenario'][iScn]['SRS1_CD'])[0]

# 				id=meta['Param']['BE']['Sawtooth']['SRS Key']['SRS_ID'][ind]

# 				lsat['SRS1_ID']=id*np.ones((1,N_StandsInBatch),dtype='int16')
# 				lsat['SRS1_PCT']=100*np.ones((1,N_StandsInBatch),dtype='int16')

# 				lsat['SRS2_ID']=np.ones((1,N_StandsInBatch),dtype='int16')
# 				lsat['SRS2_PCT']=0*np.ones((1,N_StandsInBatch),dtype='int16')

# 				lsat['SRS3_ID']=np.ones((1,N_StandsInBatch),dtype='int16')
# 				lsat['SRS3_PCT']=0*np.ones((1,N_StandsInBatch),dtype='int16')

# 			# Save
# 			gu.opickle(meta['Paths'][pNam]['Input Scenario'][iScn] + '\\Inventory_Bat' + FixFileNum(iBat) + '.pkl',lsat)

# 	return

#%% Compile events
def CompileEvents(ec,tv,iS,ID_Type,Year,MortalityFactor,GrowthFactor,ID_GrowthCurve,ASET):

	if Year.size>0:
		YearFloor=np.floor(Year)
		uYearFloor=np.unique(YearFloor)
		for iU in range(uYearFloor.size):
			iT=np.where(tv==uYearFloor[iU])[0]
			if iT.size==0:
				continue
			indYear=np.where(YearFloor==uYearFloor[iU])[0]
			for iY in range(indYear.size):
				indAvailable=np.where(ec['ID Event Type'][iT,iS,:].flatten()==0)[0]
				if indAvailable.size==0:
					print('Warning, more events per year than can be handled!')
				else:
					iE=indAvailable[0]
					ec['ID Event Type'][iT,iS,iE]=ID_Type[indYear[iY]]
					ec['Mortality Factor'][iT,iS,iE]=MortalityFactor[indYear[iY]]
					ec['Growth Factor'][iT,iS,iE]=GrowthFactor[indYear[iY]]
					ec['ID Growth Curve'][iT,iS,iE]=ID_GrowthCurve[indYear[iY]]
					ec['ASET'][iT,iS,iE]=ASET[indYear[iY]]
	return ec

#%% Mortality frequency distribution
def GetMortalityFrequencyDistribution(meta,pNam):

	iEns=0
	M=[None]*meta[pNam]['Project']['N Scenario']
	for iScn in range(meta[pNam]['Project']['N Scenario']):
		tv=np.arange(meta[pNam]['Project']['Year Start Saving'],meta[pNam]['Project']['Year End']+1,1)
		M[iScn]={}
		M[iScn]['Ma']={}
		M[iScn]['Mr']={}
		for k in meta['LUT']['Event'].keys():
			M[iScn]['Ma'][k]=np.zeros((tv.size,meta[pNam]['Project']['N Stand']))
			M[iScn]['Mr'][k]=np.zeros((tv.size,meta[pNam]['Project']['N Stand']))
		M[iScn]['Ma']['Reg']=np.zeros((tv.size,meta[pNam]['Project']['N Stand']))
		M[iScn]['Mr']['Reg']=np.zeros((tv.size,meta[pNam]['Project']['N Stand']))
		for iBat in range(meta[pNam]['Project']['N Batch']):
			indBat=IndexToBatch(meta[pNam],iBat)
			d1=LoadSingleOutputFile(meta,pNam,iScn,iEns,iBat)
			dh=gu.ipickle(meta['Paths'][pNam]['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
			for iStandInBat in range(len(indBat)):
				iStand=indBat[iStandInBat]
				for iYr in range(dh[iStandInBat]['Year'].size):
					it=np.where(tv==int(dh[iStandInBat]['Year'][iYr]))[0]
					if it.size==0:
						continue
					nam=lut_n2s(meta['LUT']['Event'],dh[iStandInBat]['ID Event Type'][iYr])[0]
					M[iScn]['Ma'][nam][it,iStand]=d1[0]['C_M_Dist'][it,iStandInBat]
					M[iScn]['Mr'][nam][it,iStand]=d1[0]['C_M_Dist'][it,iStandInBat]/np.maximum(0.0001,d1[0]['Eco_Biomass'][it-1,iStandInBat])
					M[iScn]['Ma']['Reg'][it,iStand]=d1[0]['C_M_Reg'][it,iStandInBat]
					M[iScn]['Mr']['Reg'][it,iStand]=d1[0]['C_M_Reg'][it,iStandInBat]/np.maximum(0.0001,d1[0]['Eco_Biomass'][it-1,iStandInBat])
			del d1,dh
			garc.collect()
	return M

#%% Summarize affected area due to natural disturbance and management
def SummarizeAreaAffected(meta,pNam,mos,tv,iScn,iPS,iSS,iYS,iOS,ivlT):
	A={}

	typ='Natural'
	dL=['Wildfire','Mountain Pine Beetle','Balsam Beetle','Douglas-fir Beetle','Spruce Beetle','Disease Root','Western Spruce Budworm','Wind','Frost Snow Ice Hail','Flooding Lightning Slides','Drought']
	A[typ]={}
	for d in dL:
		ind=np.where(meta['Param']['Raw']['Events']['Name']==d)[0]
		A[typ][d]={};
		A[typ][d]['Color']=[meta['Param']['Raw']['Events']['clr'][ind[0]],meta['Param']['Raw']['Events']['clg'][ind[0]],meta['Param']['Raw']['Events']['clb'][ind[0]]]
		A[typ][d]['Data']=post.GetMosScnVar(meta,pNam,mos,iScn,'Area_' + d,iPS,iSS,iYS,iOS)['Sum']['Ensemble Mean']
		#A[typ][d]['Data']=mos[pNam]['Scenarios'][iScn]['Sum']['Area_' + d]['Ensemble Mean'][:,iPS,iSS,iYS]
		if ivlT!=1:
			A[typ][d]['Data']=gu.BlockMean(A[typ][d]['Data'],ivlT)

	typ='Management'
	dL=['Harvest','Knockdown','Pile Burn','Prescribed Burn','Mechanical Site Prep','Thinning','Planting','Aerial BTK Spray','Nutrient App Aerial']
	A[typ]={}
	for d in dL:
		ind=np.where(meta['Param']['Raw']['Events']['Name']==d)[0]
		A[typ][d]={};
		A[typ][d]['Color']=[meta['Param']['Raw']['Events']['clr'][ind[0]],meta['Param']['Raw']['Events']['clg'][ind[0]],meta['Param']['Raw']['Events']['clb'][ind[0]]]
		#A[typ][d]['Data']=mos[pNam]['Scenarios'][iScn]['Sum']['Area_' + d]['Ensemble Mean'][:,iPS,iSS,iYS]
		A[typ][d]['Data']=post.GetMosScnVar(meta,pNam,mos,iScn,'Area_' + d,iPS,iSS,iYS,iOS)['Sum']['Ensemble Mean']
		if ivlT!=1:
			A[typ][d]['Data']=gu.BlockMean(A[typ][d]['Data'],ivlT)

	return A

#%% Area affected by individual multipolygon

# def AreaAffectedInSingleMultipolygon(meta,iScn,ivlT,tv,MosByMP,iMP):

#	 A={}
#	 # Ensemble mean (currently not equipped to give individual ensembles)
#	 A['Nat Dist']=[None]*10;
#	 c=-1
#	 c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Wildfire'; A['Nat Dist'][c]['Color']=[0.75,0,0]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Wildfire']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Mountain Pine beetle'; A['Nat Dist'][c]['Color']=[0,0.8,0]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Mountain Pine Beetle']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Douglas-fir beetle'; A['Nat Dist'][c]['Color']=[0.6,1,0]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['IBD']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Spruce beetle'; A['Nat Dist'][c]['Color']=[0.25,1,1]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['IBS']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='W. balsam beetle'; A['Nat Dist'][c]['Color']=[0,0.45,0]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['IBB']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Other pests'; A['Nat Dist'][c]['Color']=[0.8,1,0]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Beetles']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='W. spruce budworm'; A['Nat Dist'][c]['Color']=[0,0.75,1]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['IDW']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Rust'; A['Nat Dist'][c]['Color']=[0.75,0.5,1]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Rust Onset']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Dwarf Mistletoe'; A['Nat Dist'][c]['Color']=[1,0.5,0.25]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Dwarf Mistletoe Onset']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Mechanical'; A['Nat Dist'][c]['Color']=[0,0,0.6]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Mechanical']['Ensemble Mean'][:,iMP]
#	 A['Nat Dist']=A['Nat Dist'][0:c+1]

#	 A['Management']=[None]*10;
#	 c=-1
#	 c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Harvest'; A['Management'][c]['Color']=[0,0.75,1]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Harvest']['Ensemble Mean'][:,iMP]+MosByMP[iScn]['Area']['Harvest Salvage']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Slashpile burn'; A['Management'][c]['Color']=[0.75,0,0]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Slashpile Burn']['Ensemble Mean'][:,iMP]+MosByMP[iScn]['Area']['Harvest Salvage']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Thinning and knockdown'; A['Management'][c]['Color']=[0.2,0.4,0.7]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Knockdown']['Ensemble Mean'][:,iMP]+MosByMP[iScn]['Area']['Thinning']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Site preparation'; A['Management'][c]['Color']=[1,0.7,0.7]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Disc Trenching']['Ensemble Mean'][:,iMP]+MosByMP[iScn]['Area']['Ripping']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Prescribed burn'; A['Management'][c]['Color']=[0.5,0,0]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Prescribed Burn']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Dwarf Mistletoe control'; A['Management'][c]['Color']=[1,0.5,0]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Dwarf Mistletoe Control']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Planting'; A['Management'][c]['Color']=[0.3,0.8,0.2]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Planting']['Ensemble Mean'][:,iMP]+MosByMP[iScn]['Area']['Direct Seeding']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Foliage protection'; A['Management'][c]['Color']=[1,0.7,0]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['IDW Btk Spray']['Ensemble Mean'][:,iMP]
#	 c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Aerial nutrient application'; A['Management'][c]['Color']=[0.65,0,1]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Aerial BTK Spray Aerial']['Ensemble Mean'][:,iMP]
#	 A['Management']=A['Management'][0:c+1]

#	 # Convert to x-year intervals
#	 A['tv']=gu.BlockMean(tv,ivlT)
#	 for i in range(len(A['Nat Dist'])):
#		 A['Nat Dist'][i]['Data']=gu.BlockMean(A['Nat Dist'][i]['Data'],ivlT)
#	 for i in range(len(A['Management'])):
#		 A['Management'][i]['Data']=gu.BlockMean(A['Management'][i]['Data'],ivlT)

#	 return A

#%% Prepare growth curves (with early correction)
def PrepGrowthCurvesUniqueForCBR(meta,pNam,ugc):

	# *** This adjusts early net growth so that it is not zero. There is a
	# second version of this function that excludes the correction. ***

	# TIPSY Export curves as MgDM/ha/yr, CBRunner expects inputs of MgC/ha/yr. Create
	# conversion factor.
	dm2c=0.5

	# Growth curve parameters and TIPSY outputs
	#dfPar=pd.read_excel(meta['Paths'][pNam]['Data'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx',sheet_name='Sheet1',skiprows=7)
	txtDat=np.loadtxt(meta['Paths'][pNam]['Data'] + '\\Inputs\\GrowthCurvesTIPSY_Output.out',skiprows=4)

	# TIPSY saves to text file -> convert to dataframe (column names must match TIPSY output file design)
	dfDat=pd.DataFrame(txtDat,columns=meta['Modules']['GYM']['BatchTIPSY Column Names'])

	del txtDat

	# Define age vector (must be consistent with how TIPSY was set up)
	Age=np.arange(0,meta['Modules']['GYM']['BatchTIPSY Maximum Age']+1,1)

	# Get dimensions of the TIPSY output file to reshape the data into Age x Stand
	N_Age=Age.size
	N_GC=int(dfDat.shape[0]/N_Age)

	# Define the fraction of merchantable stemwood
	fMerch=np.nan_to_num(np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')/np.reshape(dfDat['VolTot0'].values,(N_Age,N_GC),order='F'))
	fNonMerch=1-fMerch

	## Merchantable stemwood volume
	#V_Merch=np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')
	#G_VStemMerch=np.append(np.zeros((1,N_GC)),np.diff(V_Merch,axis=0),axis=0)
	#
	## Extract age responses for each biomass pool
	#C_Stem=dm2c*np.reshape(dfDat['ODT_Stem'].values,(N_Age,N_GC),order='F')
	#C_StemMerch=fMerch*C_Stem
	#C_StemNonMerch=fNonMerch*C_Stem
	#C_Foliage=dm2c*np.reshape(dfDat['ODT_Foliage'].values,(N_Age,N_GC),order='F')
	#C_Branch=dm2c*np.reshape(dfDat['ODT_Branch'].values,(N_Age,N_GC),order='F')
	#C_Bark=dm2c*np.reshape(dfDat['ODT_Bark'].values,(N_Age,N_GC),order='F')
	#
	## Calculate growth
	#z=np.zeros((1,N_GC))
	#G_StemMerch=np.append(z,np.diff(C_StemMerch,axis=0),axis=0)
	#G_StemNonMerch=np.append(z,np.diff(C_StemNonMerch,axis=0),axis=0)
	#G_Foliage=np.append(z,np.diff(C_Foliage,axis=0),axis=0)
	#G_Branch=np.append(z,np.diff(C_Branch,axis=0),axis=0)
	#G_Bark=np.append(z,np.diff(C_Bark,axis=0),axis=0)
	#
	#plt.plot(np.mean(G_StemMerch+G_StemNonMerch,axis=1))
	##plt.plot(np.percentile(G_StemMerch+G_StemNonMerch,20,axis=1))
	##plt.plot(np.percentile(G_StemMerch+G_StemNonMerch,80,axis=1))

	# Function used to smooth curves
	def smooth(y, box_pts):
		box = np.ones(box_pts)/box_pts
		y_smooth = np.convolve(y, box, mode='same')
		return y_smooth

	# Stemwood

	# Merchantable stemwood volume
	V_Merch=np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')
	V_StemTot=np.reshape(dfDat['VolTot0'].values,(N_Age,N_GC),order='F')
	G_VStemMerch=np.append(np.zeros((1,N_GC)),np.diff(V_Merch,axis=0),axis=0)

	# Extract age responses for each biomass pool
	C_Stem=dm2c*np.reshape(dfDat['ODT_Stem'].values,(N_Age,N_GC),order='F')

	#import matplotlib.pyplot as plt
	#plt.close('all')
	#plt.plot(C_Stem[:,0])

	# Apply smoothing - it messes up the last ten years so don't smooth that part
	for j in range(C_Stem.shape[1]):
		a=smooth(C_Stem[:,j],10)
		C_Stem[:-10,j]=a[:-10]
		a=smooth(V_Merch[:,j],10)
		V_Merch[:-10,j]=a[:-10]
		a=smooth(V_StemTot[:,j],10)
		V_StemTot[:-10,j]=a[:-10]
		#plt.plot(C_Stem[:,0],'--')

	# Define the fraction of merchantable stemwood
	fMerch=np.nan_to_num(V_Merch/V_StemTot)
	fNonMerch=1-fMerch

	# Calculate growth
	z=np.zeros((1,N_GC))
	G_Stem=np.append(z,np.diff(C_Stem,axis=0),axis=0)

	# Adjust early net growth, but don't change the total stock change
	A_th=30
	ind=np.where(Age<=A_th)[0]
	bin=np.arange(0.005,0.15,0.005)
	x=np.arange(0,A_th+1,1)
	for j in range(N_GC):
		Gtot=C_Stem[ind[-1],j]
		y_th=G_Stem[ind[-1],j]
		Gtot_hat=1000*np.ones(bin.size)
		for k in range(bin.size):
			Gtot_hat[k]=np.sum(y_th*np.exp(bin[k]*(x-A_th)))
		ind1=np.where(np.abs(Gtot_hat-Gtot)==np.min(np.abs(Gtot_hat-Gtot)))[0]
		G_Stem[ind,j]=y_th*np.exp(bin[ind1[0]]*(x-A_th))

	# Update merch and nonmerch growth
	C_Stem=np.cumsum(G_Stem,axis=0)
	C_StemMerch=fMerch*C_Stem
	C_StemNonMerch=fNonMerch*C_Stem
	G_StemMerch=np.append(z,np.diff(C_StemMerch,axis=0),axis=0)
	G_StemNonMerch=np.append(z,np.diff(C_StemNonMerch,axis=0),axis=0)

	# Fix growth of year zero
	G_StemMerch[0,:]=G_StemMerch[1,:]
	G_StemNonMerch[0,:]=G_StemNonMerch[1,:]

	# Add negative nonmerch to merch
	ind=np.where(G_StemNonMerch<0)
	G_StemMerch[ind]=G_StemMerch[ind]+G_StemNonMerch[ind]
	G_StemNonMerch[ind]=0

	# Other pools - these are no longer being used. Net growth of non-stemwood
	# biomass is simulated in the annual loop based on allometric reatlionships
	# with stemwood net growth and stand age.

	# Foliage biomass is very low, revise
	#bF1=0.579
	#bF2=0.602
	#C_Foliage=np.maximum(0,bF1*C_Stem**bF2)

	C_Foliage=dm2c*np.reshape(dfDat['ODT_Foliage'].values,(N_Age,N_GC),order='F')
	C_Branch=dm2c*np.reshape(dfDat['ODT_Branch'].values,(N_Age,N_GC),order='F')
	C_Bark=dm2c*np.reshape(dfDat['ODT_Bark'].values,(N_Age,N_GC),order='F')

	G_Foliage=np.append(z,np.diff(C_Foliage,axis=0),axis=0)
	G_Branch=np.append(z,np.diff(C_Branch,axis=0),axis=0)
	G_Bark=np.append(z,np.diff(C_Bark,axis=0),axis=0)

	del C_Stem,C_StemMerch,C_StemNonMerch,C_Foliage,C_Branch,C_Bark,fMerch,fNonMerch

	for iScn in range(meta[pNam]['Project']['N Scenario']):

		for iGC in range(meta['Modules']['GYM']['N Growth Curves']):

			# Index to the full set of growth curves for scenario iScn and growth curve iGC
			ind_ugc_ScnAndGc=np.where( (ugc['Full'][:,1]==iScn) & (ugc['Full'][:,2]==meta['Modules']['GYM']['ID GC Unique'][iGC]) )[0]

			# Extract the unique growth curve ID for scenario iScn and growth curve iGC
			ID_ugc_ScnAndGc=ugc['Full'][ind_ugc_ScnAndGc,0]

			# Extract the inverse index for scenario iScn and growth curve iGC
			Inverse_ugc_ScnAndGc=ugc['Inverse'][ind_ugc_ScnAndGc]

			for iBat in range(0,meta[pNam]['Project']['N Batch']):

				# Index to batch
				indBat=IndexToBatch(meta[pNam],iBat)

				# Intersect
				c,inda,indb=np.intersect1d(ID_ugc_ScnAndGc,indBat,return_indices=True)

				Inverse_ugc_ScnAndGcAndBat=Inverse_ugc_ScnAndGc[inda]

				# Initialize array of growth data
				G=np.zeros((N_Age,indBat.size,6),dtype='int16')

				for i in range(inda.size):

					iStand=indb[i]
					iGC_Unique=Inverse_ugc_ScnAndGcAndBat[i]

					try:
						# This crashes often so leave it in try for troubleshooting
						# *** It's a sign that the BatchTIPSY inputs are pointing to the wrong directory ***
						G[:,iStand,0]=G_StemMerch[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']
					except:
						print(iStand)
						print(G.shape)
						print(iGC_Unique)
						print(G_StemMerch.shape)
					G[:,iStand,1]=G_StemNonMerch[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']
					G[:,iStand,2]=G_Bark[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']
					G[:,iStand,3]=G_Branch[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']
					G[:,iStand,4]=G_Foliage[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']
					G[:,iStand,5]=G_VStemMerch[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']

				# Save data to file in input variables folder of project
				gu.opickle(meta['Paths'][pNam]['Data'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve' + str(meta['Modules']['GYM']['ID GC Unique'][iGC]) + '_Bat' + FixFileNum(iBat) + '.pkl',G)

	return

#%% Prepare growth curves (without early correction)
# def PrepGrowthCurvesUniqueForCBR_WithoutEarlyCorrection(meta,pNam,ugc):

# 	# TIPSY exports curves as MgDM/ha/yr, CBRunner expects inputs of MgC/ha/yr. Create
# 	# conversion factor.
# 	dm2c=0.5

# 	# Growth curve parameters and TIPSY outputs
# 	#dfPar=pd.read_excel(meta['Paths'][pNam]['Data'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx',sheet_name='Sheet1',skiprows=7)
# 	txtDat=np.loadtxt(meta['Paths'][pNam]['Data'] + '\\Inputs\\GrowthCurvesTIPSY_Output.out',skiprows=4)

# 	# TIPSY saves to text file -> convert to dataframe (column names must match TIPSY output file design)
# 	dfDat=pd.DataFrame(txtDat,columns=meta['Modules']['GYM']['BatchTIPSY Column Names'])

# 	del txtDat

# 	# Define age vector (must be consistent with how TIPSY was set up)
# 	Age=np.arange(0,meta['Modules']['GYM']['BatchTIPSY Maximum Age']+1,1)

# 	# Get dimensions of the TIPSY output file to reshape the data into Age x Stand
# 	N_Age=Age.size
# 	N_GC=int(dfDat.shape[0]/N_Age)

# 	# Define the fraction of merchantable stemwood
# 	fMerch=np.nan_to_num(np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')/np.reshape(dfDat['VolTot0'].values,(N_Age,N_GC),order='F'))
# 	fNonMerch=1-fMerch

# 	# Merchantable stemwood volume
# 	V_Merch=np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')
# 	G_VStemMerch=np.append(np.zeros((1,N_GC)),np.diff(V_Merch,axis=0),axis=0)

# 	# Extract age responses for each biomass pool
# 	C_Stem=dm2c*np.reshape(dfDat['ODT_Stem'].values,(N_Age,N_GC),order='F')
# 	C_StemMerch=fMerch*C_Stem
# 	C_StemNonMerch=fNonMerch*C_Stem
# 	C_Foliage=dm2c*np.reshape(dfDat['ODT_Foliage'].values,(N_Age,N_GC),order='F')
# 	C_Branch=dm2c*np.reshape(dfDat['ODT_Branch'].values,(N_Age,N_GC),order='F')
# 	C_Bark=dm2c*np.reshape(dfDat['ODT_Bark'].values,(N_Age,N_GC),order='F')

# 	# Calculate growth
# 	z=np.zeros((1,N_GC))
# 	G_StemMerch=np.append(z,np.diff(C_StemMerch,axis=0),axis=0)
# 	G_StemNonMerch=np.append(z,np.diff(C_StemNonMerch,axis=0),axis=0)
# 	G_Foliage=np.append(z,np.diff(C_Foliage,axis=0),axis=0)
# 	G_Branch=np.append(z,np.diff(C_Branch,axis=0),axis=0)
# 	G_Bark=np.append(z,np.diff(C_Bark,axis=0),axis=0)

# 	# Fix growth of year zero
# 	G_StemMerch[0,:]=G_StemMerch[1,:]
# 	G_StemNonMerch[0,:]=G_StemNonMerch[1,:]
# 	G_Foliage[0,:]=G_Foliage[1,:]
# 	G_Branch[0,:]=G_Branch[1,:]
# 	G_Bark[0,:]=G_Bark[1,:]

# 	del C_Stem,C_StemMerch,C_StemNonMerch,C_Foliage,C_Branch,C_Bark,fMerch,fNonMerch

# 	for iScn in range(meta[pNam]['Project']['N Scenario']):
# 		for iGC in range(meta['Modules']['GYM']['N Growth Curves']):

# 			# Index to the full set of growth curves for scenario iScn and growth curve iGC
# 			ind_ugc_ScnAndGc=np.where( (ugc['Full'][:,1]==iScn) & (ugc['Full'][:,2]==meta['Modules']['GYM']['ID GC Unique'][iGC]) )[0]

# 			# Extract the unique growth curve ID for scenario iScn and growth curve iGC
# 			ID_ugc_ScnAndGc=ugc['Full'][ind_ugc_ScnAndGc,0]

# 			# Extract the inverse index for scenario iScn and growth curve iGC
# 			Inverse_ugc_ScnAndGc=ugc['Inverse'][ind_ugc_ScnAndGc]

# 			for iBat in range(0,meta[pNam]['Project']['N Batch']):

# 				# Index to batch
# 				indBat=IndexToBatch(meta[pNam],iBat)

# 				# Intersect
# 				c,inda,indb=np.intersect1d(ID_ugc_ScnAndGc,indBat,return_indices=True)

# 				Inverse_ugc_ScnAndGcAndBat=Inverse_ugc_ScnAndGc[inda]

# 				# Initialize array of growth data
# 				G=np.zeros((N_Age,indBat.size,6),dtype='int16')

# 				for i in range(inda.size):

# 					iStand=indb[i]
# 					iGC_Unique=Inverse_ugc_ScnAndGcAndBat[i]

# 					G[:,iStand,0]=G_StemMerch[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']
# 					G[:,iStand,1]=G_StemNonMerch[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']
# 					G[:,iStand,2]=G_Bark[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']
# 					G[:,iStand,3]=G_Branch[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']
# 					G[:,iStand,4]=G_Foliage[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']
# 					G[:,iStand,5]=G_VStemMerch[:,iGC_Unique].T/meta['Modules']['GYM']['Scale Factor']

# 				# Save data to file in input variables folder of project
# 				gu.opickle(meta['Paths'][pNam]['Data'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve' + str(meta['Modules']['GYM']['ID GC Unique'][iGC]) + '_Bat' + FixFileNum(iBat) + '.pkl',G)

# 	return

#%% Import parameters
def ImportParameters(meta):

	# Parameter files
	fileL=np.sort(os.listdir(meta['Paths']['Model']['Parameters']))

	# Initialize parameter structure
	if 'Param' not in meta:
		meta['Param']={}

	# Import raw tables
	meta['Param']['Raw']={}
	for file in fileL:
		if (file[0:10]=='Parameters') & (file[-4:]=='xlsx'):
			try:
				meta['Param']['Raw'][file[11:-5]]=gu.ReadExcel(meta['Paths']['Model']['Parameters'] + '\\' + file,sheet_name='Sheet1',skiprows=0)
			except:
				print(file)

	# Best estimates, variance, lower and upper confidence levels
	meta['Param']['BE']={}
	meta['Param']['Sigma']={}
	meta['Param']['BL']={}
	meta['Param']['BU']={}

	def ConvertParamRawToBE(meta,v):
		meta['Param']['BE'][v]={}
		meta['Param']['Sigma'][v]={}
		meta['Param']['BL'][v]={}
		meta['Param']['BU'][v]={}
		for i in range(meta['Param']['Raw'][v]['Name'].size):
			Name=meta['Param']['Raw'][v]['Name'][i]
			meta['Param']['BE'][v][Name]=meta['Param']['Raw'][v]['Best Estimate'][i]
			if 'Sigma' in meta['Param']['Raw'][v].keys():
				meta['Param']['Sigma'][v][Name]=meta['Param']['Raw'][v]['Sigma'][i]
			if 'BL' in meta['Param']['Raw'][v].keys():
				meta['Param']['BL'][v][Name]=meta['Param']['Raw'][v]['BL'][i]
			if 'BU' in meta['Param']['Raw'][v].keys():
				meta['Param']['BU'][v][Name]=meta['Param']['Raw'][v]['BU'][i]

			# Convert half life to turnover rate
			if v=='ProductDisposal':
				if Name[-2:]=='hl':
					Name=Name[0:-2] + 'tr'
					meta['Param']['BE'][v][Name]=1-np.exp(-np.log(2)/meta['Param']['Raw'][v]['Best Estimate'][i])
		return meta

	# Biophysical
	meta=ConvertParamRawToBE(meta,'Biophysical')

	# Biomass - allometry
	# Values are location specific so no specific processing is done here (see cbrun.py)
	meta['Param']['BE']['BiomassAllometrySL']=meta['Param']['Raw']['BiomassAllometrySL']

	# Biomass - roots
	meta=ConvertParamRawToBE(meta,'BiomassAllometryRoots')

	# Biomass - turnover
	meta=ConvertParamRawToBE(meta,'BiomassTurnover')
	# Uncertainty:
	# 	meta['Param']['BE']['BiomassTurnover']={}
	# 	meta['Param']['Sigma']['BiomassTurnover']={}
	# 	meta['Param']['BL']['BiomassTurnover']={}
	# 	meta['Param']['BU']['BiomassTurnover']={}
	# 	cnt=0
	# 	for Name in meta['Param']['Raw']['BiomassTurnover']['Name']:
	# 		meta['Param']['BE']['BiomassTurnover'][Name]=meta['Param']['Raw']['BiomassTurnover']['Best Estimate'][cnt]
	# 		meta['Param']['Sigma']['BiomassTurnover'][Name]=meta['Param']['Raw']['BiomassTurnover']['Sigma'][cnt]
	# 		meta['Param']['BL']['BiomassTurnover'][Name]=meta['Param']['Raw']['BiomassTurnover']['BL'][cnt]
	# 		meta['Param']['BU']['BiomassTurnover'][Name]=meta['Param']['Raw']['BiomassTurnover']['BU'][cnt]
	# 		cnt=cnt+1

	# By BGC zone
	meta['Param']['BE']['ByBGCZ']={}
	for i in range(meta['Param']['Raw']['ByBGCZ']['Name'].size):
		Name=meta['Param']['Raw']['ByBGCZ']['Name'][i]
		meta['Param']['BE']['ByBGCZ'][Name]={}
		for v in meta['Param']['Raw']['ByBGCZ'].keys():
			if v=='Name': continue
			meta['Param']['BE']['ByBGCZ'][Name][v]=meta['Param']['Raw']['ByBGCZ'][v][i]

	# Decomposition parameters (Kurz et al. 2009)
	meta=ConvertParamRawToBE(meta,'Decomposition')

	# Events
	str_to_exclude=['ID','Name','QA1','QA2','QA3','QA4','QA5','QA6','QA7','clr','clg','clb']
	meta['Param']['BE']['Events']={}
	for i in range(meta['Param']['Raw']['Events']['ID'].size):
		meta['Param']['BE']['Events'][meta['Param']['Raw']['Events']['ID'][i]]={}
		for k in meta['Param']['Raw']['Events'].keys():
			# Exclude some unneeded variables
			if np.isin(k,str_to_exclude)==True:
				continue
			id=meta['Param']['Raw']['Events']['ID'][i]
			meta['Param']['BE']['Events'][ id ][k]=np.nan_to_num(meta['Param']['Raw']['Events'][k][i])

	# Land use change
	meta['Param']['Raw']['LUC']={'LC Final':{},'LU Final':{}}
	for k in meta['LUT']['Derived']['lclu_chng_comp1'].keys():
		ind=np.where(meta['Param']['Raw']['LandUseChangeRules']['LUC']==k)[0]
		meta['Param']['Raw']['LUC']['LC Final'][k]=meta['Param']['Raw']['LandUseChangeRules']['LC Final'][ind][0]
		meta['Param']['Raw']['LUC']['LU Final'][k]=meta['Param']['Raw']['LandUseChangeRules']['LU Final'][ind][0]

	# Fate of felled material
	# The raw parameters are converted into time series by felled_fate_scenarios.py
	meta['Param']['BE']['FelledFate']=gu.ipickle(meta['Paths']['DB']['Taz'] + '\\Harvest\\Variables_FelledFate.pkl')

	# Fate of removals
	# The raw parameters are converted into time series by removed_fate_scenarios.py
	meta['Param']['BE']['RemovedFate']=gu.ipickle(meta['Paths']['DB']['Taz'] + '\\Harvest\\Variables_RemovedFate.pkl')

	# Product types
	# The raw parameters are converted into time series by product_end_use_scenarios.py
	meta['Param']['BE']['ProductTypes']=gu.ipickle(meta['Paths']['DB']['Taz'] + '\\Harvest\\Variables_ProductTypes.pkl')

	# Product disposal
	meta=ConvertParamRawToBE(meta,'ProductDisposal')

	# Disturbance - Wildfire aspatial (Taz-AAO) parameters
	meta['Param']['BE']['Taz']={'WF':{}}
	for i in range(meta['Param']['Raw']['WildfireStatsMod']['Name'].size):
		try:
			meta['Param']['BE']['Taz']['WF'][meta['Param']['Raw']['WildfireStatsMod']['Name'][i]]=meta['Param']['Raw']['WildfireStatsMod']['Value'][i].astype(float)
		except:
			meta['Param']['BE']['Taz']['WF'][meta['Param']['Raw']['WildfireStatsMod']['Name'][i]]=meta['Param']['Raw']['WildfireStatsMod']['Value'][i]

	# Disturbance - On the fly parameters
	meta=ConvertParamRawToBE(meta,'OnTheFly')

	# Disturbance - By severity class
	ind=np.where(meta['Param']['Raw']['BurnSev']['Zone']=='SBS')[0]
	meta['Param']['BE']['BurnSev']={}
	meta['Param']['BE']['BurnSev']['Po']=np.array([ meta['Param']['Raw']['BurnSev']['Po Unburned'][ind[0]],
												meta['Param']['Raw']['BurnSev']['Po Low'][ind[0]],
												meta['Param']['Raw']['BurnSev']['Po Medium'][ind[0]],
												meta['Param']['Raw']['BurnSev']['Po High'][ind[0]] ])
	meta['Param']['BE']['BurnSev']['PoC']=np.cumsum(meta['Param']['BE']['BurnSev']['Po'])
	meta['Param']['BE']['BurnSev']['M']=np.array([ meta['Param']['Raw']['BurnSev']['M Unburned'][ind[0]],
												meta['Param']['Raw']['BurnSev']['M Low'][ind[0]],
												meta['Param']['Raw']['BurnSev']['M Medium'][ind[0]],
												meta['Param']['Raw']['BurnSev']['M High'][ind[0]] ])

	# Economics
	meta=ConvertParamRawToBE(meta,'Economics')

	# Non-obligation funding source list
	ind=np.where(meta['Param']['Raw']['ByFSC']['Legal Obligation of Licensee To Fund Milestone']=='No')[0]
	meta['Param']['Raw']['FSC']={}
	meta['Param']['Raw']['FSC']['NO List Name']=meta['Param']['Raw']['ByFSC']['Name'][ind]
	meta['Param']['Raw']['FSC']['NO List ID']=meta['LUT']['RSLT_ACTIVITY_TREATMENT_SVW']['SILV_FUND_SOURCE_CODE']['FTM']*np.ones(meta['Param']['Raw']['FSC']['NO List Name'].size,dtype='int16')
	for i in range(meta['Param']['Raw']['FSC']['NO List Name'].size):
		try:
			meta['Param']['Raw']['FSC']['NO List ID'][i]=meta['LUT']['RSLT_ACTIVITY_TREATMENT_SVW']['SILV_FUND_SOURCE_CODE'][ meta['Param']['Raw']['FSC']['NO List Name'][i] ]
		except:
			pass

	# Growth trends
	#d=gu.ReadExcel(meta['Paths']['Model']['Code'] + '\\Parameters\\Parameters_GrowthModifier.xlsx')
	#meta['Param']['BE']['Growth Modifier']={}
	#meta['Param']['BE']['Growth Modifier']['Raw']=d
	#for i in range(d['Name'].size):
	#	meta['Param']['BE']['Growth Modifier'][d['Name'][i]]=d['Value'][i]

	# Inter-pool transfer parameters (Kurz et al. 2009)
	#meta=ConvertParamRawToBE(meta,'InterPoolFluxes')

	# Nutrient application paramaters
	meta=ConvertParamRawToBE(meta,'NutrientApplication')

	# Substitution effects (use of basic displacement factors)
	meta=ConvertParamRawToBE(meta,'Substitution')

	# Effects of disturbance on wildfire occurrence
	meta=ConvertParamRawToBE(meta,'WildfireDisturbanceEffects')

	# GROMO
	meta['Param']['Raw']['GROMO']={}
	meta['Param']['Raw']['GROMO']['gg1']={}

	# Growth (Chapman Richards)
	rsG=gu.ipickle(meta['Paths']['Model']['Code'] + '\\Parameters\\' + 'Parameters_gromo_gg1_ChapmanRichards.pkl')
	meta['Param']['Raw']['GROMO']['gg1']['ChapmanRichards']={}
	meta['Param']['Raw']['GROMO']['gg1']['ChapmanRichards']['Param']=rsG['Param']
	#meta['Param']['Raw']['GROMO']['gg1']['Zscore Stats']=rsG['Zscore Stats']

	# Growth (Chen)
	rsG=gu.ipickle(meta['Paths']['Model']['Code'] + '\\Parameters\\' + 'Parameters_gromo_gg1_Chen.pkl')
	meta['Param']['Raw']['GROMO']['gg1']['Chen']={}
	meta['Param']['Raw']['GROMO']['gg1']['Chen']['Param']=rsG['Param']
	#meta['Param']['Raw']['GROMO']['gg1']['Zscore Stats']=rsG['Zscore Stats']

	# Mortality
	rsM=gu.ipickle(meta['Paths']['Model']['Code'] + '\\Parameters\\' + 'Parameters_gromo_m1.pkl')
	meta['Param']['Raw']['GROMO']['m1']={}
	meta['Param']['Raw']['GROMO']['m1']['Param']=rsM['Param']
	meta['Param']['Raw']['GROMO']['m1']['Zscore Stats']=rsM['Zscore Stats']

	# 	#--------------------------------------------------------------------------
	# 	# Sawtooth parameters
	# 	#--------------------------------------------------------------------------
	# 	#if meta[pNam]['Project']['Biomass Module']=='Sawtooth':
	
	# 	ps={}
	# 	#pthin=r'C:\Users\rhember\Documents\Code_Python\fcgadgets\cbrunner\Parameters'
	
	# 	# Core sawtooth parameters
	# 	d=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','Core')
	# 	ps['Core']={}
	# 	for i in range(d['Name'].size):
	# 		ps['Core'][d['Name'][i]]=d['Value'][i]
	
	# 	# Sawtooth key between provincial species codes adn species-region sample
	# 	# codes
	# 	ps['SRS Key']=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','SRS Key')
	
	# 	# Tree-level allometry
	# 	ps['Allom']=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','Allometry')
	# 	ps['Eq R']={}
	# 	ps['Eq R']['Def1']=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','R_Def1')
	
	# 	ps['Eq M']={}
	# 	ps['Eq M']['Def1']=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','M_Def1')
	
	# 	ps['Eq G']={}
	# 	ps['Eq G']['Def1']=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','G_Def1')
	
	# 	meta['Param']['BE']['Sawtooth']=ps

	return meta

#%% Delete all output files
def DeleteAllOutputFiles(meta,pNam):

	for iScn in range(meta[pNam]['Project']['N Scenario']):
		files=glob.glob(meta['Paths'][pNam]['Output Scenario'][iScn] + '\\*')
		for f in files:
			os.remove(f)

	for iBat in range(meta[pNam]['Project']['N Batch']):
		try:
			pth=meta['Paths'][pNam]['Data'] + '\\Outputs\\WorkingOnBatch_' + FixFileNum(iBat) + '.pkl'
			os.remove(pth)
		except:
			pass

	return

#%% Net-down insect mortality to reflect spcecies composition
def PrepInsectMortalityPercentTreeSpeciesAffected(meta,pNam,lsat,iBat):
	lsat['Insect Mortality Percent Tree Species Affected']={}
	for iPest in range(meta['Param']['Raw']['DisturbanceSpeciesAffected']['Insect Name'].size):
		namPest=meta['Param']['Raw']['DisturbanceSpeciesAffected']['Insect Name'][iPest]
		idSpcL=[]
		for iSpc in range(1,9):
			cdSpc=meta['Param']['Raw']['DisturbanceSpeciesAffected']['Spc' + str(iSpc)][iPest]
			if cdSpc=='Empty':
				continue
			idSpc=meta['LUT']['VEG_COMP_LYR_R1_POLY']['SPECIES_CD_1'][cdSpc]
			idSpcL.append(idSpc)
		PercentAffected=np.zeros(meta[pNam]['Project']['Batch Size'][iBat])
		for iSpc in range(1,5):
			ind=np.where(np.isin(lsat['Spc' + str(iSpc) + '_ID'][0,:],idSpcL)==True)[0]
			PercentAffected[ind]=PercentAffected[ind]+lsat['Spc' + str(iSpc) + '_P'][0,ind]
		lsat['Insect Mortality Percent Tree Species Affected'][namPest]=np.minimum(100,PercentAffected)
	return lsat

#%%
# Example: L=cbu.ListVariables(meta,pNam,mos)
def ListVariables(meta,pNam,mos):
	keys=mos[pNam]['Scenarios'][0]['Mean'].keys()
	print(list(keys))
	return list(keys)
