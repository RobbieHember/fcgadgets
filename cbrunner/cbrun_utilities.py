
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
import glob
import openpyxl
import copy
import gc as garc
import time
from fcgadgets.macgyver import utilities_general as gu
from fcgadgets.macgyver import utilities_inventory as invu
from fcgadgets.hardhat import economics as econo
from fcgadgets.taz import aspatial_stat_models as asm

#%% CONVERT LUT NUMBER TO STRING NAME

def lut_n2s(dc,numb):
    if numb!=-999:
        vals=np.fromiter(dc.values(),dtype=float)
        keys=np.fromiter(dc.keys(),dtype='<U70')
        ind=np.where(vals==numb)[0]
        s=keys[ind]
    else:
        s=np.array(['Unidentified'],ndmin=1)
    return s

#%% Index to batch

def IndexToBatch(meta,iBat):
    iStart=meta['Project']['Batch Interval']*iBat
    iStop=np.minimum(meta['Project']['N Stand'],iStart+meta['Project']['Batch Interval'])
    indBat=np.arange(iStart,iStop,1)
    return indBat

#%% QUERY RESULTS CODES FOR MANAGEMENT ACTIVITY TYPES

def QueryResultsActivity(d):
    
    # Convert to arrays with at least 1d
    for key in d: 
        d[key]=np.array(d[key],ndmin=1)
    
    Name=[]
    for i in range(d['SILV_BASE_CODE'].size):
        
        if (d['SILV_BASE_CODE'][i]=='FE') & (d['SILV_TECHNIQUE_CODE'][i]=='CA'):
            Name.append('Fertilization Aerial')
        
        elif (d['SILV_BASE_CODE'][i]=='FE') & (d['SILV_TECHNIQUE_CODE'][i]=='CG') & (d['SILV_METHOD_CODE'][i]!='BAGS'):
            Name.append('Fertilization Hand')
        
        elif (d['SILV_BASE_CODE'][i]=='FE') & (d['SILV_TECHNIQUE_CODE'][i]=='CG') & (d['SILV_METHOD_CODE'][i]=='BAGS'):
            Name.append('Fertilization Teabag')
            
        elif (d['SILV_BASE_CODE'][i]=='FE') & (d['SILV_TECHNIQUE_CODE'][i]=='OG'):
            Name.append('Fertilization Organic')    
        
        elif (d['SILV_BASE_CODE'][i]=='PL') & (d['SILV_METHOD_CODE'][i]!='LAYOT'):
            # The planting in road rehab projects falls into this milestone type
            Name.append('Planting')            
        
        elif (d['SILV_BASE_CODE'][i]=='DS') & (d['SILV_TECHNIQUE_CODE'][i]!='GS'):
            # Everything except grass seeding
            Name.append('Direct Seeding')
            
        elif (d['SILV_BASE_CODE'][i]=='PC') & (d['SILV_TECHNIQUE_CODE'][i]=='MA') & (d['SILV_OBJECTIVE_CODE_1'][i]=='DM'):
            # This will exclude SILV_TECHNIQUE_CODE=BI. Virtually all of it is mechanical.
            Name.append('Dwarf Mistletoe Control')
        
        elif (d['SILV_BASE_CODE'][i]=='PC') & (d['SILV_TECHNIQUE_CODE'][i]=='CA') & (d['SILV_OBJECTIVE_CODE_1'][i]=='ID'):
            Name.append('IDW Control')
        
        elif (d['SILV_BASE_CODE'][i]=='RD') & (d['SILV_BASE_CODE'][i]=='UP'):
            Name.append('Road Rehab')
        
        elif (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='BU') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='ME') & (d['SILV_METHOD_CODE'][i]=='PBURN'):
            Name.append('Slashpile Burn')
            
        elif (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='MA') & (d['SILV_METHOD_CODE'][i]=='Unidentified') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='MA') & (d['SILV_METHOD_CODE'][i]=='GUARD') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='MA') & (d['SILV_METHOD_CODE'][i]=='HAND') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='MA') & (d['SILV_METHOD_CODE'][i]=='KNOCK') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='MA') & (d['SILV_METHOD_CODE'][i]=='POWER') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='MA') & (d['SILV_METHOD_CODE'][i]=='MANCT') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='MA') & (d['SILV_METHOD_CODE'][i]=='MDOWN') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='MA') & (d['SILV_METHOD_CODE'][i]=='PILE') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='MA') & (d['SILV_METHOD_CODE'][i]=='SNAG') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='ME') & (d['SILV_METHOD_CODE'][i]=='CABLE') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='ME') & (d['SILV_METHOD_CODE'][i]=='GUARD') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='ME') & (d['SILV_METHOD_CODE'][i]=='MDOWN') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='ME') & (d['SILV_METHOD_CODE'][i]=='PILE') | \
            (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='ME') & (d['SILV_METHOD_CODE'][i]=='PUSH'):
            Name.append('Knockdown')
        
        elif (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='ME') & (d['SILV_METHOD_CODE'][i]=='BRIP'):
            Name.append('Ripping')
        
        elif (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='ME') & (d['SILV_METHOD_CODE'][i]=='DISC'):
            Name.append('Disc Trenching')
        
        elif (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='ME') & (d['SILV_METHOD_CODE'][i]=='MULCH'):
            Name.append('Mulching')
            
        elif (d['SILV_BASE_CODE'][i]=='SP') & (d['SILV_TECHNIQUE_CODE'][i]=='ME') & (d['SILV_METHOD_CODE'][i]=='HARV'):
            Name.append('Harvest Salvage')
            
        elif (d['SILV_BASE_CODE'][i]=='LB') & (d['SILV_TECHNIQUE_CODE'][i]=='GR'):
            Name.append('LB-GR')
            
        elif (d['SILV_BASE_CODE'][i]=='SU'):
            Name.append('Surveys')
            
        else:
            Name.append('Undefined')
            
    return Name

#%% CONVERT DICTIONARY TO DATA STRUCTURE CLASS

class BunchDictionary(dict):
    def __init__(self, *args, **kwds):
        super(BunchDictionary, self).__init__(*args, **kwds)
        self.__dict__ = self


#%% Build event chronology from spreadsheet

def BuildEventChronologyFromSpreadsheet(meta):
    
    # Import inventory to get BGC zone
    iScn=0
    iBat=0
    inv=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + FixFileNum(iBat) + '.pkl')

    # Simulate wildfires
    if (meta['Scenario'][iScn]['Wildfire Status Pre-modern']=='On') | (meta['Scenario'][iScn]['Wildfire Status Modern']=='On') | (meta['Scenario'][iScn]['Wildfire Status Future']=='On'):
        asm.SimulateWildfireFromAAO(meta,inv)
    
    # Simulate MPB
    if (meta['Scenario'][iScn]['MPB Status Pre-modern']=='On') | (meta['Scenario'][iScn]['MPB Status Modern']=='On') | (meta['Scenario'][iScn]['MPB Status Future']=='On'):
        asm.SimulateIBMFromAAO(meta,inv)
    
    for iScn in range(meta['Project']['N Scenario']):
        
        for iEns in range(meta['Project']['N Ensemble']):
            
            # Import wildfire simulations from Taz   
            if (meta['Scenario'][iScn]['Wildfire Status Pre-modern']=='On') | (meta['Scenario'][iScn]['Wildfire Status Modern']=='On') | (meta['Scenario'][iScn]['Wildfire Status Future']=='On'):
                
                wf_sim=gu.ipickle(meta['Paths']['Project'] + '\\Inputs\\Ensembles\\wf_sim_Scn' + FixFileNum(iScn) + '_Ens' + FixFileNum(iEns) + '.pkl')
                if 'idx' in wf_sim:
                    idx=wf_sim['idx']
                    tmp=wf_sim.copy()
                    for v in ['Occurrence','Mortality']:
                        wf_sim[v]=np.zeros((meta['Project']['N Time'],meta['Project']['N Stand']),dtype='int16')
                        wf_sim[v][idx[0],idx[1]]=tmp[v]
                    del tmp
            
            # Import IBM  
            if (meta['Scenario'][iScn]['MPB Status Pre-modern']=='On') | (meta['Scenario'][iScn]['MPB Status Modern']=='On') | (meta['Scenario'][iScn]['MPB Status Future']=='On'):
                
                ibm_sim=gu.ipickle(meta['Paths']['Project'] + '\\Inputs\\Ensembles\\ibm_sim_Scn' + FixFileNum(iScn) + '_Ens' + FixFileNum(iEns) + '.pkl')
                if 'idx' in ibm_sim:
                    idx=ibm_sim['idx']
                    tmp=ibm_sim.copy()
                    for v in ['Occurrence','Mortality']:
                        ibm_sim[v]=np.zeros((meta['Project']['N Time'],meta['Project']['N Stand']),dtype='int16')
                        ibm_sim[v][idx[0],idx[1]]=tmp[v]
                    del tmp
            
            for iBat in range(meta['Project']['N Batch']):
    
                # Index to batch
                indBat=IndexToBatch(meta,iBat)
    
                # Always just one stand
                iS=0
    
                tv=np.arange(meta['Project']['Year Start'],meta['Project']['Year End']+1,1)
    
                # Initialize dictionary
                ec={}
                ec['ID_Type']=np.zeros((meta['Year'].size,indBat.size,meta['Core']['Max Events Per Year']),dtype='int16')
                ec['MortalityFactor']=np.zeros((meta['Year'].size,indBat.size,meta['Core']['Max Events Per Year']),dtype='int16')
                ec['GrowthFactor']=np.zeros((meta['Year'].size,indBat.size,meta['Core']['Max Events Per Year']),dtype='int16')
                ec['ID_GrowthCurve']=np.zeros((meta['Year'].size,indBat.size,meta['Core']['Max Events Per Year']),dtype='int16')
                
                #----------------------------------------------------------
                # Add spinup events
                #----------------------------------------------------------
                
                ivl_spin=meta['Project']['Spinup Disturbance Return Inverval']
                YearRef=meta['Scenario'][iScn]['Year1_DisFromInv']
                AgeRef=meta['Scenario'][iScn]['Age1_DisFromInv']
                if AgeRef>=0:
                    Year=np.arange(YearRef-AgeRef-100*ivl_spin,YearRef-AgeRef+ivl_spin,ivl_spin)
                else:
                    Year1=meta['Project']['Year Start']+ivl_spin
                    Year2=meta['Project']['Spinup Year End']
                    Year=np.arange(Year1,Year2+1,meta['Project']['Spinup Disturbance Return Inverval'])
                
                for iYr in range(Year.size):
                    iT=np.where(tv==Year[iYr])[0]
                    ec['ID_Type'][iT,:,0]=meta['LUT']['Dist'][meta['Project']['Spinup Disturbance Type']]
                    ec['MortalityFactor'][iT,:,0]=100
                    ec['GrowthFactor'][iT,:,0]=0
                    ec['ID_GrowthCurve'][iT,:,0]=meta['Project']['Spinup Growth Curve ID']
  
                #----------------------------------------------------------
                # Add events from inventory
                #----------------------------------------------------------
                
                for iYr in range(1,10):
                    
                    if ('Year' + str(iYr) + '_DisFromInv') not in meta['Scenario'][iScn]:
                        continue
                    
                    if np.isnan(meta['Scenario'][iScn]['Year' + str(iYr) + '_DisFromInv'])==True:
                        continue
        
                    # If IDW, convert IDW class to growth and mortality factor
                    sc=np.array(['IDW-T','IDW-L','IDW-M','IDW-S','IDW-V','IDW-MM','IDW-MS','IDW-MV','IDW-SS','IDW-SV','IDW-VV'])
                    flg_i=0
                    indSc=np.where(sc==meta['Scenario'][iScn]['Type' + str(iYr) + '_DisFromInv'])[0]
                    if indSc.size!=0:
                        if flg_i==0:
                            dfParDistBySC=pd.read_excel(meta['Paths']['Model Code'] + '\\Parameters\\Parameters_DisturbanceBySeverityClass.xlsx')
                            flg_i=1
                        indPar=np.where( (dfParDistBySC['Name']=='IDW') & (dfParDistBySC['SeverityCD']==sc[indSc[0]][4:]) )[0]
                        ID_TypeN=meta['LUT']['Dist']['IDW']
                        MF=dfParDistBySC.loc[indPar,'MortalityFactor']
                        GF=dfParDistBySC.loc[indPar,'GrowthFactor']
                    else:
                        ID_TypeS=meta['Scenario'][iScn]['Type' + str(iYr) + '_DisFromInv']
                        try:
                            ID_TypeN=meta['LUT']['Dist'][ID_TypeS]
                        except:
                            print(iScn)
                            print(iYr)
                            print(ID_TypeS)
                        MF=meta['Scenario'][iScn]['Severity' + str(iYr) + '_DisFromInv']
                        GF=0
        
                    Year=meta['Scenario'][iScn]['Year' + str(iYr) + '_DisFromInv']
                    iT=np.where(tv==Year)[0]
                    
                    if iT.size==0:
                        print('Warning: An event was scheduled outside the timeframe of the simulation.')
                    
                    iE=np.where(ec['ID_Type'][iT,:,:]==0)[1]
                    
                    ec['ID_Type'][iT,:,iE[0]]=ID_TypeN                    
                    ec['MortalityFactor'][iT,:,iE[0]]=MF
                    ec['GrowthFactor'][iT,:,iE[0]]=GF
                    ec['ID_GrowthCurve'][iT,:,iE[0]]=meta['Scenario'][iScn]['GrowthCurve' + str(iYr) + '_DisFromInv']  
        
                #----------------------------------------------------------
                # Add simulated wildfire from Taz
                #----------------------------------------------------------
                
                ind=np.array([],dtype=int)
                if meta['Scenario'][iScn]['Wildfire Status Pre-modern']=='On':            
                    ind0=np.where( (wf_sim['Occurrence'][:,iS]==1) & (meta['Year']<1920) )[0]
                    ind=np.append(ind,ind0)
                if meta['Scenario'][iScn]['Wildfire Status Modern']=='On':            
                    ind0=np.where( (wf_sim['Occurrence'][:,iS]==1) & (meta['Year']>=1920) & (meta['Year']<meta['Project']['Year Project']) )[0]
                    ind=np.append(ind,ind0)
                if meta['Scenario'][iScn]['Wildfire Status Future']=='On':            
                    ind0=np.where( (wf_sim['Occurrence'][:,iS]==1) & (meta['Year']>=meta['Project']['Year Project']) )[0]
                    ind=np.append(ind,ind0)
                    
                if ind.size>0:
                    
                    ID_Type=meta['LUT']['Dist']['Wildfire']*np.ones(ind.size)
                    Year=tv[ind]
                    MortF=wf_sim['Mortality'][ind,iS]
                    GrowthF=0*np.ones(ind.size)
                    ID_GrowthCurve=1*np.ones(ind.size)
                        
                    for iYr in range(Year.size):
                        iT=np.where(tv==Year[iYr])[0]
                        ec['ID_Type'][iT,:,0]=ID_Type[iYr]
                        ec['MortalityFactor'][iT,:,0]=MortF[iYr]
                        ec['GrowthFactor'][iT,:,0]=GrowthF[iYr]
                        ec['ID_GrowthCurve'][iT,:,0]=ID_GrowthCurve[iYr]
    
                    #----------------------------------------------------------
                    # Add simulated MPB from Taz
                    #----------------------------------------------------------

                    ind=np.array([],dtype=int)               
                    if meta['Scenario'][iScn]['MPB Status Pre-modern']=='On':            
                        ind0=np.where( (ibm_sim['Occurrence'][:,iS]==1) & (meta['Year']<1920) )[0]
                        ind=np.append(ind,ind0)
                    if meta['Scenario'][iScn]['MPB Status Modern']=='On':            
                        ind0=np.where( (ibm_sim['Occurrence'][:,iS]==1) & (meta['Year']>=1920) & (meta['Year']<meta['Project']['Year Project']) )[0]
                        ind=np.append(ind,ind0)
                    if meta['Scenario'][iScn]['MPB Status Future']=='On':            
                        ind0=np.where( (ibm_sim['Occurrence'][:,iS]==1) & (meta['Year']>=meta['Project']['Year Project']) )[0]
                        ind=np.append(ind,ind0)
                        
                    if ind.size>0:
                        
                        ID_Type=meta['LUT']['Dist']['IBM']*np.ones(ind.size)
                        Year=tv[ind]
                        MortF=ibm_sim['Mortality'][ind,iS]
                        GrowthF=0*np.ones(ind.size)
                        ID_GrowthCurve=1*np.ones(ind.size)
                            
                        for iYr in range(Year.size):
                            iT=np.where(tv==Year[iYr])[0]
                            ec['ID_Type'][iT,:,0]=ID_Type[iYr]
                            ec['MortalityFactor'][iT,:,0]=MortF[iYr]
                            ec['GrowthFactor'][iT,:,0]=GrowthF[iYr]
                            ec['ID_GrowthCurve'][iT,:,0]=ID_GrowthCurve[iYr]
            
                #--------------------------------------------------------------
                # Save to file            
                #--------------------------------------------------------------
                
                gu.opickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl',ec)

    return

#%% Decompress event chronology

def EventChronologyDecompress(meta,ec,iScn,iEns,iBat):

    # Uncompress event chronology if it has been compressed
    if 'idx' in ec:
        idx=ec['idx']
        tmp=ec.copy()
        for v in ['ID_Type','MortalityFactor','GrowthFactor','ID_GrowthCurve']:
            ec[v]=np.zeros((meta['Project']['N Time'],meta['Project']['Batch Size'][iBat],meta['Core']['Max Events Per Year']),dtype='int16')
            ec[v][idx[0],idx[1],idx[2]]=tmp[v]
        del tmp                
    
    return ec            

#%% Fix ensemble name and numbering

def FixFileNum(ind):
    indStrFixed=str(ind+1)
    if len(indStrFixed)==1:
        indStrFixed='000' + indStrFixed
    elif len(indStrFixed)==2:
        indStrFixed='00' + indStrFixed    
    elif len(indStrFixed)==3:
        indStrFixed='0' + indStrFixed
    return indStrFixed

#%% Configure project

def ImportProjectConfig(meta):
    
    #--------------------------------------------------------------------------
    # Initialize nested dictionaries
    #--------------------------------------------------------------------------
    
    if 'Project' not in meta:
        meta['Project']={}
    
    if 'Core' not in meta:
        meta['Core']={}
    
    #--------------------------------------------------------------------------
    # Import project parameters from spreadsheet
    #--------------------------------------------------------------------------
    
    df=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\ProjectConfig.xlsx',sheet_name='Project')
    for i in range(len(df)):
        Name=df['Name'].iloc[i]
        Value=df['Value'].iloc[i]
        if Name[-1]==':':
            # Exclude headers
            continue
        meta['Project'][Name]=Value
    
    #--------------------------------------------------------------------------
    # Import look-up tables
    #--------------------------------------------------------------------------
    
    meta=invu.Load_LUTs(meta)
    
    #--------------------------------------------------------------------------
    # Define pool names
    #--------------------------------------------------------------------------
    
    # Pool names (ecosystem)
    # *** If you change this, you need to change the same list in "Update Parameters" ***
    meta['Core']['Name Pools Eco']=['StemMerch','StemNonMerch','Foliage','Branch','Bark','RootCoarse','RootFine', \
        'PiledStemMerch','PiledStemNonMerch','PiledBranch','PiledBark','PiledSnagStem','PiledSnagBranch', \
        'LitterVF','LitterF','LitterM','LitterS','SnagStem','SnagBranch','SoilVF','SoilF','SoilS']
    
    # Used in fert work: ,'LitterDecomp'
    
    # Number of ecosystem pools
    meta['Core']['N Pools Eco']=len(meta['Core']['Name Pools Eco'])
    
    # Pool names (products)
    meta['Core']['Name Pools Pro']=['SFH','MFH','Comm','Furn','Ship','Repairs', \
        'Other','Paper','EffluentPulp','PowerFacilityDom','PowerFacilityFor','PowerGrid', \
        'Pellets','LogExport','FirewoodDom','FirewoodFor','DumpWood', \
        'DumpPaper','LandfillWoodDegradable','LandfillWoodNonDegradable', \
        'LandfillPaperDegradable','LandfillPaperNonDegradable','E_CO2','E_CH4']
    
    # Number of product pools
    meta['Core']['N Pools Pro']=len(meta['Core']['Name Pools Pro'])
    
    # List of output variables
    meta['Core']['Output Variable List']=['A',
        'LogSizeEnhancement',
        'V_MerchLive',
        'V_MerchDead',
        'V_MerchTotal',
        'V_ToMillMerchLive',
        'V_ToMillMerchDead',
        'V_ToMillMerchTotal',
        'C_Forest_Tot',
        'C_Biomass_Tot',
        'C_DeadWood_Tot',
        'C_DumpLandfill_Tot',
        'C_Piled_Tot',
        'C_G_Gross_Tot',
        'C_G_Net_Tot',
        'C_HWP_Tot',
        'C_InUse_Tot',
        'C_Buildings_Tot',
        'C_NonBuildings_Tot',
        'C_LF_Tot',
        'C_Litter_Tot',
        'C_M_Reg_Tot',
        'C_NPP_Tot',
        'C_RH_Tot',
        'C_Soil_Tot',
        'C_ToFirewoodDom',
        'C_ToFirewoodFor',
        'C_ToLogExport',
        'C_ToLumber',
        'C_ToMDF',
        'C_ToMill',
        'C_ToMillMerch',
        'C_ToMillNonMerch',
        'C_ToMillSnagStem',
        'C_ToOSB',
        'C_ToPaper',
        'C_ToPellets',
        'C_ToPlywood',
        'C_ToPowerFacilityDom',
        'C_ToPowerFacilityFor',
        'C_ToPowerGrid',
        'C_ToSlashpileBurn',
        'E_CO2e_LULUCF_NEE',
        'E_CO2e_LULUCF_Denit',
        'E_CO2e_LULUCF_Other',
        'E_CO2e_LULUCF_OpenBurning',
        'E_CO2e_LULUCF_Wildfire',
        'E_CO2e_LULUCF_Fire',
        'E_CO2e_LULUCF_HWP',
        'E_CO2e_ESC_Bioenergy',
        'E_CO2e_ESC_Operations',
        'E_CO2e_ET_Operations',
        'E_CO2e_IPPU_Operations',
        'E_CO2e_SUB_E',
        'E_CO2e_SUB_M',
        'E_CO2e_SUB_Coal',
        'E_CO2e_SUB_Oil',
        'E_CO2e_SUB_Gas',
        'E_CO2e_SUB_Calcination',
        'E_CO2e_AGHGB_WOSub',
        'E_CO2e_AGHGB_WOSub_cumu',
        'E_CO2e_AGHGB_WSub',
        'E_CO2e_AGHGB_WSub_cumu',
        'Yield Coal',
        'Yield Oil',
        'Yield Gas',
        'Yield Sawnwood',
        'Yield Panels',
        'Yield Concrete',
        'Yield Steel',
        'Yield Aluminum',
        'Yield Plastic',
        'Yield Textile',
        'Yield Firewood',
        'Yield FirewoodDom',
        'Yield LogExport',
        'Yield Lumber',
        'Yield MDF',
        'Yield OSB',
        'Yield Paper',
        'Yield Pellets',
        'Yield Plywood',
        'Yield PowerFacilityDom',
        'Yield PowerGrid',
        'Cost Roads',
        'Cost Knockdown',
        'Cost Ripping',
        'Cost Nutrient Management',
        'Cost PAS Deactivation',
        'Cost Harvest Felling and Piling',
        'Cost Harvest Hauling',
        'Cost Harvest Overhead',
        'Cost Harvest Residuals',
        'Cost Milling',
        'Cost Slashpile Burn',
        'Cost Planting',
        'Cost Survey',
        'Cost Silviculture Total',
        'Cost Total',            
        'Cost Total Disc',
        'Cost Total Disc_cumu',
        'Revenue FirewoodDom',
        'Revenue LogExport',
        'Revenue Lumber',
        'Revenue MDF',
        'Revenue OSB',
        'Revenue Paper',
        'Revenue Pellets',
        'Revenue Plywood',
        'Revenue PowerFacilityDom',
        'Revenue PowerGrid',
        'Revenue Gross',
        'Revenue Gross Disc',
        'Revenue Gross Disc_cumu',
        'Revenue Net',
        'Revenue Net Disc',
        'Revenue Net Disc_cumu']
    
    #--------------------------------------------------------------------------
    # Define indices to each pool
    #--------------------------------------------------------------------------
    
    # Indices to ecosystem pools pools
    meta['Core']['iEP']={}; cnt=0
    for nam in meta['Core']['Name Pools Eco']:
        meta['Core']['iEP'][nam]=cnt
        cnt=cnt+1
    iEP=meta['Core']['iEP']
    meta['Core']['iEP']['BiomassTotal']=np.array([iEP['StemMerch'],iEP['StemNonMerch'],iEP['Foliage'],iEP['Branch'],iEP['Bark'],iEP['RootCoarse'],iEP['RootFine']])
    meta['Core']['iEP']['BiomassAboveground']=np.array([iEP['StemMerch'],iEP['StemNonMerch'],iEP['Foliage'],iEP['Branch'],iEP['Bark']])
    meta['Core']['iEP']['BiomassBelowground']=np.array([iEP['RootCoarse'],iEP['RootFine']])
    meta['Core']['iEP']['DeadWood']=np.array([iEP['PiledStemMerch'],iEP['PiledStemNonMerch'],iEP['PiledBranch'],iEP['PiledBark'],iEP['SnagStem'],iEP['SnagBranch']])
    meta['Core']['iEP']['Litter']=np.array([iEP['LitterVF'],iEP['LitterF'],iEP['LitterM'],iEP['LitterS']])
    meta['Core']['iEP']['Piled']=np.array([iEP['PiledStemMerch'],iEP['PiledStemNonMerch'],iEP['PiledBranch'],iEP['PiledBark'],iEP['PiledSnagStem'],iEP['PiledSnagBranch']])
    meta['Core']['iEP']['Soil']=np.array([iEP['SoilVF'],iEP['SoilF'],iEP['SoilS']])
    
    # Indices to produce pools pools
    meta['Core']['iPP']={}; cnt=0
    for nam in meta['Core']['Name Pools Pro']:
        meta['Core']['iPP'][nam]=cnt
        cnt=cnt+1
    iPP=meta['Core']['iPP']
    meta['Core']['iPP']['InUse']=np.array([ iPP['SFH'],iPP['MFH'],iPP['Comm'],iPP['Furn'],iPP['Ship'],iPP['Repairs'],iPP['Other'],iPP['Paper'] ])
    meta['Core']['iPP']['Buildings']=np.array([ iPP['SFH'],iPP['MFH'],iPP['Comm'] ])
    meta['Core']['iPP']['DumpLandfill']=np.array([ iPP['DumpWood'],iPP['DumpPaper'],iPP['LandfillWoodDegradable'],iPP['LandfillWoodNonDegradable'],iPP['LandfillPaperDegradable'],iPP['LandfillPaperNonDegradable'] ])
    
    #--------------------------------------------------------------------------
    # Maximum number of events per year
    # 8 appears to be sufficient but this may need to be changed for some
    # special projects
    #--------------------------------------------------------------------------
    
    meta['Core']['Max Events Per Year']=8

    #--------------------------------------------------------------------------
    # Define time
    #--------------------------------------------------------------------------
    
    # Calendar year
    meta['Year']=np.arange(meta['Project']['Year Start'],meta['Project']['Year End']+1,1)
    
    meta['Project']['N Time']=meta['Year'].size
    
    #--------------------------------------------------------------------------
    # Dimensions of simulation
    #--------------------------------------------------------------------------
    
    # Number of stands 
    if meta['Project']['Scenario Source']=='Spreadsheet':
        
        if 'Number of stands' in meta['Project']:
            # Override!
            meta['Project']['N Stand']=meta['Project']['Number of stands']
        else:
            meta['Project']['N Stand']=1
    
    elif meta['Project']['Scenario Source']=='Portfolio':
        
        meta['Project']['N Stand']=meta['Project']['N Stand per Activity Type']*meta['Project']['AIL']['N AT']*meta['Project']['AIL']['N Years']
    
    # Number of batches
    meta['Project']['N Batch']=np.ceil(meta['Project']['N Stand']/meta['Project']['Batch Interval']).astype(int)

    # Initialize list that can keep track of batch sizes
    meta['Project']['Batch Size']=[None]*meta['Project']['N Batch']
    for iBat in range(meta['Project']['N Batch']):
        meta['Project']['Batch Size'][iBat]=IndexToBatch(meta,iBat).size
    
    #--------------------------------------------------------------------------
    # Import model parameters
    #--------------------------------------------------------------------------
    
    meta=ImportParameters(meta)
    
    #--------------------------------------------------------------------------
    # Define scenario parameters
    #--------------------------------------------------------------------------

    if meta['Project']['Scenario Source']!='Portfolio':
    
        df=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\ProjectConfig.xlsx',sheet_name='Scenarios',usecols='A:OM')
    
        df=df.iloc[:,df.iloc[0,:].isnull().values==False] 
        
        meta['Scenario']=list()
        for i in range(1,df.shape[1]):
            pScn0={}
            for j in range(df.shape[0]):
                if df.iloc[j,0][-1]==':':
                    # Exclude headers
                    continue
                pScn0.update({df.iloc[j,0]:df.iat[j,i]})
            meta['Scenario'].append(pScn0)
    
        # Number of scenarios
        meta['Project']['N Scenario']=np.sum([i['Scenario Status']=='On' for i in meta['Scenario']])    

    #--------------------------------------------------------------------------
    # Initialize project folders if they do not exist
    #--------------------------------------------------------------------------
    
    meta['Paths']['Input Scenario']=[]
    meta['Paths']['Output Scenario']=[]
    for iScn in range(0,meta['Project']['N Scenario']):    
        meta['Paths']['Input Scenario'].append(meta['Paths']['Project'] + '\\Inputs\\Scenario' + FixFileNum(iScn))
        if os.path.exists(meta['Paths']['Input Scenario'][iScn])==False:
            os.mkdir(meta['Paths']['Input Scenario'][iScn])
        meta['Paths']['Output Scenario'].append(meta['Paths']['Project'] + '\\Outputs\\Scenario' + FixFileNum(iScn))
        if os.path.exists(meta['Paths']['Output Scenario'][iScn])==False:
            os.mkdir(meta['Paths']['Output Scenario'][iScn])
    
    #--------------------------------------------------------------------------
    # Scale factors
    #--------------------------------------------------------------------------
    
    # *** Scale factor for saving results (this needs to be 100, 10 does not 
    # capture carbon fluxes and it will affect GHG benefit estimates) ***
    # One variable ('CO2e_E_Products') requires the big one
    meta['Core']['Scale Factor Export Small']=0.001
    meta['Core']['Scale Factor Export Big']=0.001
    
    #--------------------------------------------------------------------------
    # Define strings that frequently need to be populated with zeros
    #--------------------------------------------------------------------------
    
    meta['Core']['StringsToFill']=['Month','Day','SILV_FUND_SOURCE_CODE','FIA_PROJECT_ID','OPENING_ID','ACTUAL_TREATMENT_AREA','ACTUAL_PLANTED_NUMBER', \
               'PL_SPECIES_CD1','PL_SPECIES_PCT1','PL_SPECIES_GW1','PL_SPECIES_CD2','PL_SPECIES_PCT2','PL_SPECIES_GW2', \
               'PL_SPECIES_CD3','PL_SPECIES_PCT3','PL_SPECIES_GW3','PL_SPECIES_CD4','PL_SPECIES_PCT4','PL_SPECIES_GW4', \
               'PL_SPECIES_CD5','PL_SPECIES_PCT5','PL_SPECIES_GW5']
    
    #--------------------------------------------------------------------------
    # Growth curve information
    #--------------------------------------------------------------------------
    
    meta['GC']={}
    meta['GC']['N Growth Curves']=5
    meta['GC']['ID GC Unique']=np.array([1,2,3,4,5])
    meta['GC']['BatchTIPSY Maximum Age']=200
    meta['GC']['BatchTIPSY Column Names']=['Age','VolTot0','VolMerch125',
        'VolMerch175','ODT_Bark','ODT_Branch','ODT_Foliage','ODT_Roots',
        'ODT_Stem','MortalityVolumeTotal']
    
    # Scale factor for growth curves
    # Note: Do not change this to 0.1 - aerial fertilization response will not work properly at 0.1
    meta['GC']['Scale Factor']=0.001
    
    #--------------------------------------------------------------------------
    # Growth factor information
    # *** Not currently used ***
    #--------------------------------------------------------------------------
    
#    # Default status of growth factors
#    meta['Scenario Switch']['Net Growth Factor Status']=[None]*meta['Project']['N Scenario']
#    meta['Scenario Switch']['Mortality Factor Status']=[None]*meta['Project']['N Scenario']
#    for iScn in range(0,meta['Project']['N Scenario']): 
#        meta['Scenario Switch']['Net Growth Factor Status'][iScn]='Off'
#        meta['Scenario Switch']['Mortality Factor Status'][iScn]='Off'
#        #meta['Scenario Switch'][iScn]['Status Net Growth Factor']='Off'
#        #meta['Scenario'][iScn]['Status Mortality Factor']='Off'
    
    #--------------------------------------------------------------------------
    # Harvested wood product information
    # Year to start calling annual HWP methods - running it before 1800 is a 
    # waste of time.
    #--------------------------------------------------------------------------
    
    meta['Core']['HWP Year Start']=1850
    
    #--------------------------------------------------------------------------
    # Nutrient management information (for compatibility with "silviculture" module)
    #--------------------------------------------------------------------------
    
    # Initialize dictionary
    meta['Nutrient Management']={}
    
    # Initialize index to stands affected by nutrient application
    # This needs to be populated with an empty array for when Sawtooth is used.
    meta['Nutrient Management']['iApplication']=np.array([])
    
    # BGC zone exclusions (for on-the-fly application scheduler)
    meta['Nutrient Management']['BGC Zone Exclusion CD']=['PP','MH','BAFA','BG','CMA','IMA']    
    meta['Nutrient Management']['BGC Zone Exclusion ID']=np.zeros(len(meta['Nutrient Management']['BGC Zone Exclusion CD']))
    for iZ in range(len(meta['Nutrient Management']['BGC Zone Exclusion CD'])):
        meta['Nutrient Management']['BGC Zone Exclusion ID'][iZ]=meta['LUT']['VRI']['BEC_ZONE_CODE'][ meta['Nutrient Management']['BGC Zone Exclusion CD'][iZ] ]
    
    # Coastal zones used to make prob of occurrence region-specific
    meta['Nutrient Management']['Coastal Zones CD']=['CWH','CDF']    
    meta['Nutrient Management']['Coastal Zones ID']=np.zeros(len(meta['Nutrient Management']['Coastal Zones CD']))
    for iZ in range(len(meta['Nutrient Management']['Coastal Zones CD'])):
        meta['Nutrient Management']['Coastal Zones ID'][iZ]=meta['LUT']['VRI']['BEC_ZONE_CODE'][ meta['Nutrient Management']['Coastal Zones CD'][iZ] ]

    #--------------------------------------------------------------------------
    # Simulate random numbers that can be used for simulating harvest on the fly
    # The annual numbers will be the same among scenarios, but vary by ensemble
    #--------------------------------------------------------------------------
    
    # *** If you assign completely random numbers, random variation will occur among
    # scenarios, which can add considerable noise and demands many ensembles. 
    # Conversely if you assign these pre-set sequeences, the random component will
    # vary among ensembles, but not among scenarios.
    meta['Project']['On the Fly']={}    
    meta['Project']['On the Fly']['Random Numbers']={}
    meta['Project']['On the Fly']['Random Numbers']['Scale Factor']=0.0001
    
    # Only create these files if they will be used
    
    # Not needed for portfolio projects
    
    if meta['Project']['Scenario Source']!='Portfolio':
        
        flg_h=0    
        for iScn in range(meta['Project']['N Scenario']):
            if (meta['Scenario'][iScn]['Harvest Status Historical']=='On') | (meta['Scenario'][iScn]['Harvest Status Future']=='On'):
                flg_h=1
                break
        
        flg_b=0
        for iScn in range(meta['Project']['N Scenario']):
            if (meta['Scenario'][iScn]['Breakup Status']=='On'):
                flg_b=1
                break
        
        # Create random numbers and save them
        for iEns in range(meta['Project']['N Ensemble']):
            for iBat in range(meta['Project']['N Batch']):
            
                if flg_h==1:
                    rn=np.random.random( (meta['Project']['N Time'],meta['Project']['Batch Size'][iBat]) )
                    rn=rn/meta['Project']['On the Fly']['Random Numbers']['Scale Factor']
                    rn=rn.astype('int16')
                    gu.opickle(meta['Paths']['Project'] + '\\Inputs\\Ensembles\\RandomNumbers_Harvest_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl',rn)
                
                if flg_b==1:
                    rn=np.random.random( (meta['Project']['N Time'],meta['Project']['Batch Size'][iBat]) )
                    rn=rn/meta['Project']['On the Fly']['Random Numbers']['Scale Factor']
                    rn=rn.astype('int16')
                    gu.opickle(meta['Paths']['Project'] + '\\Inputs\\Ensembles\\RandomNumbers_Breakup_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl',rn)
    
    #meta['Project']['On the Fly']['Random Numbers']['Harvest']=np.random.random((meta['Project']['N Time'],meta['Project']['N Ensemble']))    
    #meta['Project']['On the Fly']['Random Numbers']['Breakup']=np.random.random((meta['Project']['N Time'],meta['Project']['N Ensemble']))
    
    #--------------------------------------------------------------------------
    # Parameter uncertainty by ensemble
    #--------------------------------------------------------------------------
    
    # Initialize list
    meta['Param']['By Ensemble']=[None]*meta['Project']['N Ensemble']
    
    for iEns in range(meta['Project']['N Ensemble']):
        
        # Initialize dictionary
        meta['Param']['By Ensemble'][iEns]={}
        
        #----------------------------------------------------------------------
        # Biomass turnover
        #----------------------------------------------------------------------
        
        if meta['Project']['Uncertainty Status Biomass Turnover']=='On':
            
            meta['Param']['By Ensemble'][iEns]['Biomass Turnover']={}            
            
            for k in meta['Param']['BE']['Biomass Turnover'].keys():            
                
                mu=meta['Param']['BE']['Biomass Turnover'][k]
                sig=meta['Param']['Sigma']['Biomass Turnover'][k]
                bl=meta['Param']['BL']['Biomass Turnover'][k]
                bu=meta['Param']['BU']['Biomass Turnover'][k]
                
                r=np.random.normal(loc=mu,scale=mu*sig)
                
                if (bl!=-9999) & (bu!=-9999):
                    r=gu.Clamp(r,bl,bu)
                elif (bl!=-9999) & (bu==-9999):
                    r=np.maximum(bl,r)
                elif (bl==-9999) & (bu!=-9999):
                    r=np.minimum(bu,r)
                else:
                    r=r
                    
                meta['Param']['By Ensemble'][iEns]['Biomass Turnover'][k]=r
        
        #----------------------------------------------------------------------
        # Decomposition
        #----------------------------------------------------------------------
        
        if meta['Project']['Uncertainty Status Decomposition']=='On':
            
            meta['Param']['By Ensemble'][iEns]['Decomp']={}            
            
            for k in meta['Param']['BE']['Decomp'].keys():            
                
                mu=meta['Param']['BE']['Decomp'][k]
                sig=meta['Param']['Sigma']['Decomp'][k]
                bl=meta['Param']['BL']['Decomp'][k]
                bu=meta['Param']['BU']['Decomp'][k]
                
                r=np.random.normal(loc=mu,scale=mu*sig)
                
                if (bl!=-9999) & (bu!=-9999):
                    r=gu.Clamp(r,bl,bu)
                elif (bl!=-9999) & (bu==-9999):
                    r=np.maximum(bl,r)
                elif (bl==-9999) & (bu!=-9999):
                    r=np.minimum(bu,r)
                else:
                    r=r
                    
                meta['Param']['By Ensemble'][iEns]['Decomp'][k]=r
        
        #----------------------------------------------------------------------
        # Harvesting
        #----------------------------------------------------------------------
        
        if meta['Project']['Uncertainty Status Harvest Utilization']=='On':
            
            meta['Param']['By Ensemble'][iEns]['Dist']={}            
            
            EventList=['Harvest','Harvest Salvage']
            
            for Event in EventList:
                
                ID_Type=meta['LUT']['Dist'][Event]
                meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]={}
                
                #--------------------------------------------------------------
                # Biomass merch
                #--------------------------------------------------------------
                
                # Removed fraction               
                mu=meta['Param']['BE']['Dist'][ID_Type]['BiomassMerch_Removed']
                sig=np.array([0.1])
                bl=np.array([0.0])
                bu=np.array([1.0])                
                r_Removed=np.random.normal(loc=mu,scale=mu*sig)
                r_Removed=gu.Clamp(r_Removed,bl,bu)                
                meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassMerch_Removed']=r_Removed
                
                # Total fraction that is piled and dispersed
                r_PiledAndDispersed=1.0-r_Removed
                
                # Specific piled fraction
                mu=np.array([0.60]) # Specific fraction that is piled
                sig=np.array([0.1])
                bl=np.array([0.0])
                bu=np.array([1.0])               
                rSpecific_Piled=np.random.normal(loc=mu,scale=mu*sig)
                rSpecific_Piled=gu.Clamp(rSpecific_Piled,bl,bu)
                
                # Piled fraction
                meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassMerch_Piled']=r_PiledAndDispersed*rSpecific_Piled
                
                # Specific dispersed fraction
                rSpecific_Dispersed=1.0-rSpecific_Piled
                
                # Dispersed fraction
                meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassMerch_LeftOnSite']=r_PiledAndDispersed*rSpecific_Dispersed
                
                #print(meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassMerch_Removed']+meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassMerch_Piled']+meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassMerch_LeftOnSite'])
                
                #--------------------------------------------------------------
                # Biomass non-merch
                #--------------------------------------------------------------
                
                # Removed fraction               
                mu=meta['Param']['BE']['Dist'][ID_Type]['BiomassNonMerch_Removed']
                sig=np.array([0.1])
                bl=np.array([0.0])
                bu=np.array([1.0])               
                r_Removed=np.random.normal(loc=mu,scale=mu*sig)
                r_Removed=gu.Clamp(r_Removed,bl,bu)
                meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassNonMerch_Removed']=r_Removed
                
                # Total fraction that is piled and dispersed
                r_PiledAndDispersed=1.0-r_Removed
                
                # Specific piled fraction
                mu=np.array([0.60]) # Specific fraction that is piled
                sig=np.array([0.1])
                bl=np.array([0.0])
                bu=np.array([1.0])               
                rSpecific_Piled=np.random.normal(loc=mu,scale=mu*sig)
                rSpecific_Piled=gu.Clamp(rSpecific_Piled,bl,bu)
                
                # Piled fraction
                meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassNonMerch_Piled']=r_PiledAndDispersed*rSpecific_Piled
                
                # Specific dispersed fraction
                rSpecific_Dispersed=1.0-rSpecific_Piled
                
                # Dispersed fraction
                meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassNonMerch_LeftOnSite']=r_PiledAndDispersed*rSpecific_Dispersed
                
                #print(meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassNonMerch_Removed']+meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassNonMerch_Piled']+meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['BiomassNonMerch_LeftOnSite'])
                
                #--------------------------------------------------------------
                # Snags
                #--------------------------------------------------------------
                
                # Removed fraction               
                mu=meta['Param']['BE']['Dist'][ID_Type]['Snags_Removed']
                sig=np.array([0.1])
                bl=np.array([0.0])
                bu=np.array([1.0])               
                r_Removed=np.random.normal(loc=mu,scale=mu*sig)
                r_Removed=gu.Clamp(r_Removed,bl,bu)
                meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['Snags_Removed']=r_Removed
                
                # Total fraction that is piled and dispersed
                r_PiledAndDispersed=1.0-r_Removed
                
                # Specific piled fraction
                mu=np.array([0.60]) # Specific fraction that is piled
                sig=np.array([0.1])
                bl=np.array([0.0])
                bu=np.array([1.0])               
                rSpecific_Piled=np.random.normal(loc=mu,scale=mu*sig)
                rSpecific_Piled=gu.Clamp(rSpecific_Piled,bl,bu)
                
                # Piled fraction
                meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['Snags_Piled']=r_PiledAndDispersed*rSpecific_Piled
                
                # Specific dispersed fraction
                rSpecific_Dispersed=1.0-rSpecific_Piled
                
                # Dispersed fraction
                meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['Snags_LeftOnSite']=r_PiledAndDispersed*rSpecific_Dispersed
        
                #print(meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['Snags_Removed']+meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['Snags_Piled']+meta['Param']['By Ensemble'][iEns]['Dist'][ID_Type]['Snags_LeftOnSite'])
        
        #----------------------------------------------------------------------
        # Substitution effects
        #----------------------------------------------------------------------
        
        if meta['Project']['Uncertainty Status Substitution']=='On':
            
            meta['Param']['By Ensemble'][iEns]['Substitution']={}            
            
            #vL=['LumberDisplacementFactor','PanelDisplacementFactor']            
            vL=['SawnwoodFracDisplacingConcrete','SawnwoodFracDisplacingSteel','SawnwoodFracDisplacingAluminum',
                'SawnwoodFracDisplacingPlastic','SawnwoodFracDisplacingTextile','PanelsFracDisplacingConcrete','PanelsFracDisplacingSteel',
                'PanelsFracDisplacingAluminum','PanelsFracDisplacingPlastic','PanelsFracDisplacingTextile','ResidualsFracDisplacingConcrete',
                'ResidualsFracDisplacingSteel','ResidualsFracDisplacingAluminum','ResidualsFracDisplacingPlastic','ResidualsFracDisplacingTextile',
                'DisplacementRatio_ConcreteForSawnwood','DisplacementRatio_ConcreteForPanels','DisplacementRatio_ConcreteForResiduals',
                'DisplacementRatio_SteelForSawnwood','DisplacementRatio_SteelForPanels','DisplacementRatio_SteelForResiduals',
                'DisplacementRatio_AluminumForSawnwood','DisplacementRatio_AluminumForPanels','DisplacementRatio_AluminumForResiduals',
                'DisplacementRatio_PlasticForSawnwood','DisplacementRatio_PlasticForPanels','DisplacementRatio_PlasticForResiduals',
                'DisplacementRatio_TextileForSawnwood','DisplacementRatio_TextileForPanels','DisplacementRatio_TextileForResiduals']
            for k in vL:
                mu=meta['Param']['BE']['Substitution'][k]
                sig=meta['Param']['Sigma']['Substitution'][k]
                r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
                r=np.maximum(0,r)            
                meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
            
            lv0=np.array(['PowerFacilityDomFracDisplacingRenewables','PowerFacilityDomFracDisplacingCoal','PowerFacilityDomFracDisplacingDiesel','PowerFacilityDomFracDisplacingNaturalGas','PowerFacilityDomFracDisplacingOil'])
            lv=lv0[np.argsort(np.random.random(len(lv0)))]
            r_remain=1.0
            for k in lv:
                mu=meta['Param']['BE']['Substitution'][k]
                sig=meta['Param']['Sigma']['Substitution'][k]
                r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
                r=gu.Clamp(r,0.0,r_remain)        
                meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
                r_remain=r_remain-r
            
            lv0=np.array(['PowerFacilityForFracDisplacingRenewables','PowerFacilityForFracDisplacingCoal','PowerFacilityForFracDisplacingDiesel','PowerFacilityForFracDisplacingNaturalGas','PowerFacilityForFracDisplacingOil'])
            lv=lv0[np.argsort(np.random.random(len(lv0)))]
            r_remain=1.0
            for k in lv:
                mu=meta['Param']['BE']['Substitution'][k]
                sig=meta['Param']['Sigma']['Substitution'][k]
                r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
                r=gu.Clamp(r,0.0,r_remain)        
                meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
                r_remain=r_remain-r
            
            lv0=np.array(['PelletFracDisplacingRenewables','PelletFracDisplacingCoal','PelletFracDisplacingDiesel','PelletFracDisplacingNaturalGas','PelletFracDisplacingOil'])
            lv=lv0[np.argsort(np.random.random(len(lv0)))]
            r_remain=1.0
            for k in lv:
                mu=meta['Param']['BE']['Substitution'][k]
                sig=meta['Param']['Sigma']['Substitution'][k]
                r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
                r=gu.Clamp(r,0.0,r_remain)        
                meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
                r_remain=r_remain-r
            
            lv0=np.array(['FirewoodDomFracDisplacingRenewables','FirewoodDomFracDisplacingCoal','FirewoodDomFracDisplacingDiesel','FirewoodDomFracDisplacingNaturalGas','FirewoodDomFracDisplacingOil'])
            lv=lv0[np.argsort(np.random.random(len(lv0)))]
            r_remain=1.0
            for k in lv:
                mu=meta['Param']['BE']['Substitution'][k]
                sig=meta['Param']['Sigma']['Substitution'][k]
                r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
                r=gu.Clamp(r,0.0,r_remain)        
                meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
                r_remain=r_remain-r
            
            lv0=np.array(['FirewoodForFracDisplacingRenewables','FirewoodForFracDisplacingCoal','FirewoodForFracDisplacingDiesel','FirewoodForFracDisplacingNaturalGas','FirewoodForFracDisplacingOil'])
            lv=lv0[np.argsort(np.random.random(len(lv0)))]
            r_remain=1.0
            for k in lv:
                mu=meta['Param']['BE']['Substitution'][k]
                sig=meta['Param']['Sigma']['Substitution'][k]
                r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
                r=gu.Clamp(r,0.0,r_remain)        
                meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
                r_remain=r_remain-r
            
            lv0=np.array(['PowerGridFracDisplacingRenewables','PowerGridFracDisplacingCoal','PowerGridFracDisplacingDiesel','PowerGridFracDisplacingNaturalGas','PowerGridFracDisplacingOil'])
            lv=lv0[np.argsort(np.random.random(len(lv0)))]
            r_remain=1.0
            for k in lv:
                mu=meta['Param']['BE']['Substitution'][k]
                sig=meta['Param']['Sigma']['Substitution'][k]
                r=np.array(np.random.normal(loc=mu,scale=mu*sig),dtype=float)
                r=gu.Clamp(r,0.0,r_remain)        
                meta['Param']['By Ensemble'][iEns]['Substitution'][k]=r
                r_remain=r_remain-r
            
        #----------------------------------------------------------------------
        # Nutrient mmanagement
        #----------------------------------------------------------------------
        
        if meta['Project']['Uncertainty Status Nutrient Application']=='On':
            
            meta['Param']['By Ensemble'][iEns]['Nutrient Management']={}            
            
            for k in meta['Param']['BE']['Nutrient Management'].keys():            
                mu=meta['Param']['BE']['Nutrient Management'][k]
                sig=meta['Param']['Sigma']['Nutrient Management'][k]
                r=np.random.normal(loc=mu,scale=mu*sig)
                r=np.maximum(0,r)            
                meta['Param']['By Ensemble'][iEns]['Nutrient Management'][k]=r
    
    #--------------------------------------------------------------------------
    # Scenario info for portfolio projects
    #--------------------------------------------------------------------------
    
    if meta['Project']['Scenario Source']=='Portfolio':
    
        # Index to rows with implementation
        indAT=np.where(np.sum(meta['Project']['AIL']['Area'],axis=0)>0)[0]
        
        meta['Project']['Portfolio']['ID Portfolio']=np.zeros(meta['Project']['N Stand'],dtype=int)
        meta['Project']['Portfolio']['ID AT']=np.zeros(meta['Project']['N Stand'],dtype=int)
        meta['Project']['Portfolio']['ID AT Unique']=np.zeros(meta['Project']['N Stand'],dtype=int)
        meta['Project']['Portfolio']['Area']=np.zeros(meta['Project']['N Stand'])
        meta['Project']['Portfolio']['Year']=np.zeros(meta['Project']['N Stand'])
        meta['Project']['Portfolio']['Region Code']=np.array(['' for _ in range(meta['Project']['N Stand'])],dtype=object)
        meta['Project']['Portfolio']['Felled Fate Scenario']=np.array(['' for _ in range(meta['Project']['N Stand'])],dtype=object)
        meta['Project']['Portfolio']['Removed Fate Scenario']=np.array(['' for _ in range(meta['Project']['N Stand'])],dtype=object)
        meta['Project']['Portfolio']['HWP End Use Scenario']=np.array(['' for _ in range(meta['Project']['N Stand'])],dtype=object)
        cnt=0
        for iA in range(meta['Project']['AIL']['N AT']):
            for iY in range(meta['Project']['AIL']['N Years']):                
                for iS in range(meta['Project']['N Stand per Activity Type']):
                    
                    ID_Portfolio=meta['Project']['AIL']['ID Portfolio'][indAT[iA]]
                    
                    meta['Project']['Portfolio']['ID Portfolio'][cnt]=ID_Portfolio
                    meta['Project']['Portfolio']['ID AT'][cnt]=meta['Project']['AIL']['ID AT'][indAT[iA]]
                    meta['Project']['Portfolio']['ID AT Unique'][cnt]=meta['Project']['AIL']['ID AT Unique'][indAT[iA]]
                    meta['Project']['Portfolio']['Area'][cnt]=meta['Project']['AIL']['Area'][iY,indAT[iA]]
                    meta['Project']['Portfolio']['Year'][cnt]=meta['Project']['AIL']['Year'][iY]
                    
                    # Region from Activities table
                    ind=np.where(meta['Project']['Activities']['Activity ID']==meta['Project']['Portfolio']['ID AT'][cnt])
                    meta['Project']['Portfolio']['Region Code'][cnt]=meta['Project']['Activities']['Region Code'][ind][0]
                    
                    # Scenarios from Portfolio table
                    iPortfolio=np.where(meta['Project']['Portfolio']['Raw']['ID_Portfolio']==ID_Portfolio)[0]
                    meta['Project']['Portfolio']['Felled Fate Scenario'][cnt]=meta['Project']['Portfolio']['Raw']['Felled Fate Scenario'][iPortfolio][0]
                    meta['Project']['Portfolio']['Removed Fate Scenario'][cnt]=meta['Project']['Portfolio']['Raw']['Removed Fate Scenario'][iPortfolio][0]
                    meta['Project']['Portfolio']['HWP End Use Scenario'][cnt]=meta['Project']['Portfolio']['Raw']['HWP End Use Scenario'][iPortfolio][0]
                    cnt=cnt+1

        # Scenario information
        # Will this tool ever be used with on-the-fly disturbances? Current set
        # to "off".
        meta['Scenario']=[None]*meta['Project']['N Scenario']
        for iScn in range(meta['Project']['N Scenario']):
            
            meta['Scenario'][iScn]={}
            
            # *** This is super awkward - simulations can't change between activities or scenarios!!!
            # Do we need that type of functionality for the PT? ***
            meta['Scenario'][iScn]['Wildfire Scenario ID']=meta['Project']['Activities']['Wildfire Scenario ID'][0]
            meta['Scenario'][iScn]['Wildfire Status Pre-modern']=meta['Project']['Activities']['Wildfire Status Pre-modern'][0]
            meta['Scenario'][iScn]['Wildfire Status Modern']=meta['Project']['Activities']['Wildfire Status Modern'][0]
            meta['Scenario'][iScn]['Wildfire Status Future']=meta['Project']['Activities']['Wildfire Status Future'][0]
            
            meta['Scenario'][iScn]['Harvest Status Historical']='Off'
            meta['Scenario'][iScn]['Harvest Status Future']='Off'
            meta['Scenario'][iScn]['Breakup Status']='Off'
            meta['Scenario'][iScn]['Nutrient Application Status']='Off'
    
    #--------------------------------------------------------------------------
    # Project type and region (not used in demos so initialize here)
    #--------------------------------------------------------------------------
    
    if meta['Project']['Scenario Source']=='Spreadsheet':
        meta['Project']['Project Stratum ID']=1.0*np.ones(meta['Project']['N Stand'])
        meta['Project']['Spatial Stratum ID']=1.0*np.ones(meta['Project']['N Stand'])
    
    #--------------------------------------------------------------------------
    # Grassland module
    #--------------------------------------------------------------------------
    
    for iScn in range(meta['Project']['N Scenario']):
        if 'Grassland Module Status' not in meta['Scenario'][iScn]:
            # Default is off
            meta['Scenario'][iScn]['Grass Module Status']='Off'
            meta['Scenario'][iScn]['Grass Module Year Start']=meta['Project']['Year Project']+1    
    
    return meta

#%% Load scenario results
# Return a list of dictionaries for each scenario. If multiple ensemble were run, 
# the function will retun the average.

def LoadSingleOutputFile(meta,iScn,iEns,iBat):
    
    # Extract indices
    iEP=meta['Core']['iEP']
    
    # Extract biophysical parameters
    bB=meta['Param']['BE']['Biophysical']
    
    # Open batch results
    pth=meta['Paths']['Project'] + '\\Outputs\\Scenario' + FixFileNum(iScn) + '\\Data_Scn' + FixFileNum(iScn) + '_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl'
    v0=gu.ipickle(pth)
    
    # Convert to float and apply scale factor    
    for k in v0.keys():
        
        # Skip mortality summary by agent
        if (k=='C_M_ByAgent'):
            continue                      
        
        v0[k]=v0[k].astype(float)
        
        if (k=='E_CO2e_LULUCF_HWP') | (k=='E_CO2e_ESC_Comb') | (k=='E_CO2e_ET_Comb') | (k=='E_CO2e_IPPU_Comb'):
            v0[k]=v0[k]*meta['Core']['Scale Factor Export Big']
        else:
            v0[k]=v0[k]*meta['Core']['Scale Factor Export Small']
    
    #--------------------------------------------------------------------------
    # Add derived variables
    #--------------------------------------------------------------------------
    
    # Add year        
    it=np.where(meta['Year']>=meta['Project']['Year Start Saving'])[0]
    v0['Year']=meta['Year'][it]
    
    if meta['Project']['Save Biomass Pools']=='On':
        
        # Aggregate variables not yet generated
        
        # Aggregate pools
        v0['C_Biomass_Tot']=np.sum(v0['C_Eco_Pools'][:,:,iEP['BiomassTotal']],axis=2)
        v0['C_Piled_Tot']=np.sum(v0['C_Eco_Pools'][:,:,iEP['Piled']],axis=2)
        v0['C_Litter_Tot']=np.sum(v0['C_Eco_Pools'][:,:,iEP['Litter']],axis=2)
        v0['C_DeadWood_Tot']=np.sum(v0['C_Eco_Pools'][:,:,iEP['DeadWood']],axis=2)
        v0['C_Soil_Tot']=np.sum(v0['C_Eco_Pools'][:,:,iEP['Soil']],axis=2)
        v0['C_InUse_Tot']=np.sum(v0['C_Pro_Pools'][:,:,meta['Core']['iPP']['InUse']],axis=2)
        v0['C_DumpLandfill_Tot']=np.sum(v0['C_Pro_Pools'][:,:,meta['Core']['iPP']['DumpLandfill']],axis=2)
    
        # Aggregate fluxes
        v0['C_G_Gross_Tot']=np.sum(v0['C_G_Gross'],axis=2)
        v0['C_G_Net_Tot']=np.sum(v0['C_G_Net'],axis=2)
        v0['C_M_Reg_Tot']=np.sum(v0['C_M_Reg'],axis=2)
        v0['C_LF_Tot']=np.sum(v0['C_LF'],axis=2)
        v0['C_RH_Tot']=np.sum(v0['C_RH'],axis=2)
    
    v0['C_NPP_Tot']=v0['C_G_Net_Tot']+v0['C_M_Reg_Tot']+v0['C_LF_Tot']
    v0['C_ToMill']=v0['C_ToMillMerch']+v0['C_ToMillNonMerch']+v0['C_ToMillSnagStem']    
    v0['C_Forest_Tot']=v0['C_Biomass_Tot']+v0['C_DeadWood_Tot']+v0['C_Litter_Tot']+v0['C_Soil_Tot']
    v0['C_NonBuildings_Tot']=v0['C_InUse_Tot']-v0['C_Buildings_Tot']
    v0['C_HWP_Tot']=v0['C_InUse_Tot']+v0['C_DumpLandfill_Tot']
    v0['E_CO2e_LULUCF_NEE']=-1*bB['Ratio_CO2_to_C']*(v0['C_NPP_Tot']-v0['C_RH_Tot'])
    v0['E_CO2e_LULUCF_Fire']=v0['E_CO2e_LULUCF_Wildfire']+v0['E_CO2e_LULUCF_OpenBurning']
    
    v0['Yield Firewood']=(v0['C_ToFirewoodDom']+v0['C_ToFirewoodFor'])/bB['Ratio_Wood_C_to_DM']
    
    # Aggregate operational emissions and substitution effects
    v0['E_CO2e_ESC_Operations']=v0['E_CO2e_ESC_OperationsBurnCoal']+v0['E_CO2e_ESC_OperationsBurnOil']+v0['E_CO2e_ESC_OperationsBurnGas']
    v0['E_CO2e_ET_Operations']=v0['E_CO2e_ET_OperationsBurnCoal']+v0['E_CO2e_ET_OperationsBurnOil']+v0['E_CO2e_ET_OperationsBurnGas']
    v0['E_CO2e_IPPU_Operations']=v0['E_CO2e_IPPU_BurningCoal']+v0['E_CO2e_IPPU_BurningOil']+v0['E_CO2e_IPPU_BurningGas']
    
    v0['E_CO2e_OPER_Coal']=v0['E_CO2e_ESC_OperationsBurnCoal']+v0['E_CO2e_ET_OperationsBurnCoal']+v0['E_CO2e_IPPU_BurningCoal']
    v0['E_CO2e_OPER_Oil']=v0['E_CO2e_ESC_OperationsBurnOil']+v0['E_CO2e_ET_OperationsBurnOil']+v0['E_CO2e_IPPU_BurningOil']
    v0['E_CO2e_OPER_Gas']=v0['E_CO2e_ESC_OperationsBurnGas']+v0['E_CO2e_ET_OperationsBurnGas']+v0['E_CO2e_IPPU_BurningGas']
    v0['E_CO2e_OPER']=v0['E_CO2e_OPER_Coal']+v0['E_CO2e_OPER_Oil']+v0['E_CO2e_OPER_Gas']
    
    # Aggregate substitution effects    
    # For unknown reasons, trying to reverse sign when calculated messes up the array
    # Do it here instead.
    vL=['E_CO2e_SUB_CoalForBioenergy','E_CO2e_SUB_OilForBioenergy','E_CO2e_SUB_GasForBioenergy','E_CO2e_SUB_CoalForWood',
        'E_CO2e_SUB_OilForWood','E_CO2e_SUB_GasForWood',
        'E_CO2e_SUB_Concrete','E_CO2e_SUB_Steel','E_CO2e_SUB_Aluminum','E_CO2e_SUB_Plastic','E_CO2e_SUB_Textile',
        'E_CO2e_SUB_Calcination']
    for v in vL:
        v0[v]=-1*v0[v]
    
    # Reverse sign of building material production (saved as positive)
    vL=['Yield Concrete','Yield Steel','Yield Aluminum','Yield Plastic','Yield Textile']
    for v in vL:
        v0[v]=-1*v0[v]
    
    v0['E_CO2e_SUB_Coal']=v0['E_CO2e_SUB_CoalForBioenergy']+v0['E_CO2e_SUB_CoalForWood']
    v0['E_CO2e_SUB_Oil']=v0['E_CO2e_SUB_OilForBioenergy']+v0['E_CO2e_SUB_OilForWood']
    v0['E_CO2e_SUB_Gas']=v0['E_CO2e_SUB_GasForBioenergy']+v0['E_CO2e_SUB_GasForWood']
    v0['E_CO2e_SUB_E']=v0['E_CO2e_SUB_CoalForBioenergy']+v0['E_CO2e_SUB_OilForBioenergy']+v0['E_CO2e_SUB_GasForBioenergy']
    v0['E_CO2e_SUB_M']=v0['E_CO2e_SUB_CoalForWood']+v0['E_CO2e_SUB_OilForWood']+v0['E_CO2e_SUB_GasForWood']+v0['E_CO2e_SUB_Calcination']
    v0['E_CO2e_SUB_Tot']=v0['E_CO2e_SUB_M']+v0['E_CO2e_SUB_E']
    
    v0['E_CO2e_Coal']=v0['E_CO2e_OPER_Coal']+v0['E_CO2e_SUB_Coal']
    v0['E_CO2e_Oil']=v0['E_CO2e_OPER_Oil']+v0['E_CO2e_SUB_Oil']
    v0['E_CO2e_Gas']=v0['E_CO2e_OPER_Gas']+v0['E_CO2e_SUB_Gas']
    
    # Atmospheric GHG balance (tCO2e/ha/yr)  
    v0['E_CO2e_AGHGB_WSub']=v0['E_CO2e_LULUCF_NEE']+v0['E_CO2e_LULUCF_Wildfire']+v0['E_CO2e_LULUCF_OpenBurning']+ \
        v0['E_CO2e_LULUCF_Denit']+v0['E_CO2e_LULUCF_Other']+v0['E_CO2e_LULUCF_HWP']+v0['E_CO2e_ESC_Bioenergy']+v0['E_CO2e_SUB_E']+v0['E_CO2e_SUB_M']+ \
        v0['E_CO2e_ESC_Operations']+v0['E_CO2e_ET_Operations']+v0['E_CO2e_IPPU_Operations']
    
    v0['E_CO2e_AGHGB_WOSub']=v0['E_CO2e_LULUCF_NEE']+v0['E_CO2e_LULUCF_Wildfire']+v0['E_CO2e_LULUCF_OpenBurning']+ \
        v0['E_CO2e_LULUCF_Denit']+v0['E_CO2e_LULUCF_Other']+v0['E_CO2e_LULUCF_HWP']+v0['E_CO2e_ESC_Bioenergy']+ \
        v0['E_CO2e_ESC_Operations']+v0['E_CO2e_ET_Operations']+v0['E_CO2e_IPPU_Operations']

    # Add cumulative
    v0['E_CO2e_AGHGB_WSub_cumu']=np.cumsum(v0['E_CO2e_AGHGB_WSub'],axis=0)
    v0['E_CO2e_AGHGB_WOSub_cumu']=np.cumsum(v0['E_CO2e_AGHGB_WOSub'],axis=0)
    
    # Add cumulative (starting from a specified start year)
    iT=np.where(v0['Year']>=meta['Project']['Year Start Cumulative'])[0]
    v0['E_CO2e_AGHGB_WSub_cumu_from_tref']=np.zeros(v0['A'].shape)
    v0['E_CO2e_AGHGB_WSub_cumu_from_tref'][iT,:]=np.cumsum(v0['E_CO2e_AGHGB_WSub'][iT,:],axis=0)
    v0['E_CO2e_AGHGB_WOSub_cumu_from_tref']=np.zeros(v0['A'].shape)    
    v0['E_CO2e_AGHGB_WOSub_cumu_from_tref'][iT,:]=np.cumsum(v0['E_CO2e_AGHGB_WOSub'][iT,:],axis=0)
        
    #--------------------------------------------------------------------------
    # Back-calculate production of fossil fuel consumption from operational use
    # and substitution effects (tonnesC)
    #--------------------------------------------------------------------------

    E_Operations=v0['E_CO2e_ESC_OperationsBurnCoal']+v0['E_CO2e_ET_OperationsBurnCoal']+v0['E_CO2e_IPPU_BurningCoal']
    E_Substitution=v0['E_CO2e_SUB_CoalForBioenergy']+v0['E_CO2e_SUB_CoalForWood']
    v0['Yield Coal']=(E_Operations+E_Substitution)/(bB['Emission Intensity Coal']/1000)/bB['Energy Content of Coal']*bB['Carbon Content of Coal']
    
    E_Operations=v0['E_CO2e_ESC_OperationsBurnOil']+v0['E_CO2e_ET_OperationsBurnOil']+v0['E_CO2e_IPPU_BurningOil']
    E_Substitution=v0['E_CO2e_SUB_OilForBioenergy']+v0['E_CO2e_SUB_OilForWood']
    v0['Yield Oil']=(E_Operations+E_Substitution)/(bB['Emission Intensity Oil']/1000)/bB['Energy Content of Oil']*bB['Carbon Content of Oil']
    
    E_Operations=v0['E_CO2e_ESC_OperationsBurnGas']+v0['E_CO2e_ET_OperationsBurnGas']+v0['E_CO2e_IPPU_BurningGas']
    E_Substitution=v0['E_CO2e_SUB_GasForBioenergy']+v0['E_CO2e_SUB_GasForWood']
    v0['Yield Gas']=(E_Operations+E_Substitution)/(bB['Emission Intensity Natural Gas']/1000)/bB['Energy Content of Natural Gas']*bB['Carbon Content of Natural Gas']
    
    #--------------------------------------------------------------------------
    # Sawtooth variable adjustments
    #--------------------------------------------------------------------------  
    
    if meta['Project']['Biomass Module']=='Sawtooth':
        v0['N']=np.maximum(0,v0['N'])
        v0['N_R']=np.maximum(0,v0['N_R'])
        v0['N_M_Tot']=np.maximum(0,v0['N_M_Tot'])
        v0['TreeMean_D']=np.maximum(0,v0['TreeMean_D'])
        v0['TreeMean_Csw']=np.maximum(0,v0['TreeMean_Csw'])
        v0['TreeMean_Csw_G']=np.maximum(0,v0['TreeMean_Csw_G'])
        
    return v0

#%% LOAD SCENARIO RUSULTS
# Return a list of dictionaries for each scenario. If multiple ensemble were run, 
# the function will retun the average.

def LoadScenarioResults(meta):
    
    # Initialize list that will contain scenarios
    v1=[]
    for iScn in range(meta['Project']['N Scenario']):
        
        for iEns in range(meta['Project']['N Ensemble']):            
            
            for iBat in range(meta['Project']['N Batch']):
                
                #--------------------------------------------------------------
                # Open batch results
                #--------------------------------------------------------------
                
                data_batch=LoadSingleOutputFile(meta,iScn,iEns,iBat)
                
                # Import event chronology                
                if (meta['Scenario'][iScn]['Harvest Status Future']=='On') | (meta['Scenario'][iScn]['Breakup Status']=='On'):
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Modified_Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
                else:
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
            
                # Uncompress event chronology if it has been compressed
                ec=EventChronologyDecompress(meta,ec,iScn,iEns,iBat) 
            
                # Inventory
                inv=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + FixFileNum(iBat) + '.pkl')
                    
                # Cashflow
                econ=econo.CalculateNetRevenue(meta,iScn,iEns,iBat,inv,ec,data_batch)
                
                data_batch.update(econ)
                
                #--------------------------------------------------------------
                # Accumulate data in each batch
                #--------------------------------------------------------------
                
                if iBat==0:                    
                    
                    data_all=data_batch
                
                else:
                    
                    for key1 in data_batch.keys():                        
                        
                        if key1=='Year':                            
                            # Only needed once
                            continue
                        
                        elif (key1=='C_M_ByAgent'):                            
                            # Nested dictionary
                            for key2 in data_batch[key1].keys():
                                data_all[key1][key2]=np.append(data_all[key1][key2],data_batch[key1][key2],axis=1)
                        
                        else:                            
                            # No nested dictionary
                            data_all[key1]=np.append(data_all[key1],data_batch[key1],axis=1)
            
            #------------------------------------------------------------------
            # Sum across ensembles
            #------------------------------------------------------------------
            
            if iEns==0:
                
                data_sum2ave=data_all
            
            else:                    
                
                for key1 in data_batch.keys():
                    
                    if (key1=='C_M_ByAgent'):
                        
                        # Nested dictionary
                        for key2 in data_batch[key1].keys():
                            data_sum2ave[key1][key2]=data_sum2ave[key1][key2]+data_all[key1][key2]
                    
                    else:
                        
                        # No nested dictionary
                        data_sum2ave[key1]=data_sum2ave[key1]+data_all[key1]
        
        #----------------------------------------------------------------------
        # If the simulation includes ensembles, calculate average
        #----------------------------------------------------------------------
        
        for key1 in data_batch.keys():
            
            # Skip mortality summary by agent
            if (key1=='C_M_ByAgent'):
                
                # Nested dictioanry
                for key2 in data_batch[key1].keys():
                    data_sum2ave[key1][key2]=data_sum2ave[key1][key2]/meta['Project']['N Ensemble']  
            
            else:
                
                # No nested dictionary
                data_sum2ave[key1]=data_sum2ave[key1]/meta['Project']['N Ensemble']        
        
        #----------------------------------------------------------------------
        # Add year
        #----------------------------------------------------------------------
        
        it=np.where(meta['Year']>=meta['Project']['Year Start Saving'])[0]
        data_sum2ave['Year']=meta['Year'][it]
        
        #----------------------------------------------------------------------
        # Append to list
        #----------------------------------------------------------------------
        
        v1.append(data_sum2ave)
        
    return v1

#%% Save output variables by multipolygon

# This is only designed for projects that apply inventory_from_polygons method.
# Not only is this outputting by project, but the project flux sums are accurately
# representing the given treatment area, rather than the area inferred from the 
# total number of sparse sample points within a project area, which can be wrong.
    
# *** Use this if you want to keep a time series for each MP (e.g. LCELF). If 
# you are content with summaries by project type and region, then use other 
# functions below. ***

def MosByMultipolygon(meta,switch_area,switch_cashflow):

    # Import multipolygons
    atu_multipolygons=gu.ipickle(meta['Paths']['Geospatial'] + '\\atu_multipolygons.pkl')

    # Import sxy
    geos=gu.ipickle(meta['Paths']['Geospatial'] + '\\geos.pkl')

    # Unique MPs
    uMP=np.unique(geos['Sparse']['ID_atu_multipolygons'])

    # Create listed index (faster than indexing on the fly)
    Crosswalk_sxy_to_mp=[None]*uMP.size
    for iMP in range(uMP.size):
        d={}
        d['Index']=np.where(geos['Sparse']['ID_atu_multipolygons']==uMP[iMP])[0]
        Crosswalk_sxy_to_mp[iMP]=d

    # Time series of saved results
    tv_full=np.arange(meta['Project']['Year Start'],meta['Project']['Year End']+1,1)
    tv_saving=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)
    it=np.where( (tv_full>=tv_saving[0]) & (tv_full<=tv_saving[-1]) )[0]
    
    # Variables to save
    d1=LoadSingleOutputFile(meta,0,0,0)
    nam1=list(d1.keys())
    nam1.remove('Year')
    nam1.remove('C_M_ByAgent')
    
    #nam1=['A','V_StemMerch','C_Biomass_Tot','C_DeadWood_Tot','C_Litter_Tot','C_Soil_Tot', \
    #      'C_InUse_Tot','C_DumpLandfill_Tot','C_ToMillMerch','C_ToMillNonMerch','C_ToMillSnagStem', \
    #      'C_ToLumber','C_ToPlywood','C_ToOSB','C_ToMDF','C_ToPaper','C_ToPowerFacilityDom','C_ToPowerGrid','C_ToPellets','C_ToFirewoodDom','C_ToLogExport', \
    #      'C_G_Net_Tot','C_LF_Tot','C_RH_Tot','E_CO2e_LULUCF_NEE','E_CO2e_LULUCF_Wildfire','E_CO2e_LULUCF_OpenBurning','CO2e_LULUCF_E_EcoOther', \
    #      'CO2e_LULUCF_E_HWP','E_CO2e_ESC_Comb','E_CO2e_ESC_SubE','E_CO2e_ESC_SubBM','E_CO2e_ET_Comb','E_CO2e_IPPU_Comb','E_CO2e_LULUCF_Fire','E_CO2e_AGHGB']
    
    nam_cashflow=['Cost Total','Revenue Gross']

    # Scale factor used to temporarily store data
    ScaleFactor=0.001

    #--------------------------------------------------------------------------
    # Initialize data by multipolygon structure    
    #--------------------------------------------------------------------------
    
    MosByMP=[None]*meta['Project']['N Scenario']
    for iScn in range(meta['Project']['N Scenario']):
        
        d={}
        
        d['v1']={}
        d['v1']['Mean']={}
        d['v1']['Sum']={}
        for iV in range(len(nam1)):
            d['v1']['Mean'][nam1[iV]]={}
            d['v1']['Mean'][nam1[iV]]['Ensemble Mean']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Mean'][nam1[iV]]['Ensemble SD']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Mean'][nam1[iV]]['Ensemble P10']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Mean'][nam1[iV]]['Ensemble P25']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Mean'][nam1[iV]]['Ensemble P75']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Mean'][nam1[iV]]['Ensemble P90']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Sum'][nam1[iV]]={}
            d['v1']['Sum'][nam1[iV]]['Ensemble Mean']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Sum'][nam1[iV]]['Ensemble SD']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Sum'][nam1[iV]]['Ensemble P10']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Sum'][nam1[iV]]['Ensemble P25']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Sum'][nam1[iV]]['Ensemble P75']=np.zeros((tv_saving.size,uMP.size))
            d['v1']['Sum'][nam1[iV]]['Ensemble P90']=np.zeros((tv_saving.size,uMP.size))
        
        if switch_area=='On':
            d['Area']={}
            for k in meta['LUT']['Dist'].keys():
                d['Area'][k]={}
                d['Area'][k]['Ensemble Mean']=np.zeros((tv_saving.size,uMP.size))
                #d['Area'][k]['Ensemble SD']=np.zeros((tv_saving.size,uMP.size))
        
        if switch_cashflow=='On':
            d['Cashflow']={}
            d['Cashflow']['Mean']={}
            d['Cashflow']['Sum']={}
            for iV in range(len(nam_cashflow)):
                d['Cashflow']['Mean'][nam_cashflow[iV]]={}
                d['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble Mean']=np.zeros((tv_saving.size,uMP.size))
                d['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble SD']=np.zeros((tv_saving.size,uMP.size))
                #d['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble P10']=np.zeros((tv_saving.size,uMP.size))
                #d['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble P25']=np.zeros((tv_saving.size,uMP.size))
                #d['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble P75']=np.zeros((tv_saving.size,uMP.size))
                #d['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble P90']=np.zeros((tv_saving.size,uMP.size))
                d['Cashflow']['Sum'][nam_cashflow[iV]]={}
                d['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble Mean']=np.zeros((tv_saving.size,uMP.size))
                d['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble SD']=np.zeros((tv_saving.size,uMP.size))
                d['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P10']=np.zeros((tv_saving.size,uMP.size))
                d['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P25']=np.zeros((tv_saving.size,uMP.size))
                d['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P75']=np.zeros((tv_saving.size,uMP.size))
                d['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P90']=np.zeros((tv_saving.size,uMP.size))
            
        MosByMP[iScn]=d
    
    #--------------------------------------------------------------------------
    # Loop through scenarios
    #--------------------------------------------------------------------------
    
    for iScn in range(meta['Project']['N Scenario']):
    
        for iEns in range(meta['Project']['N Ensemble']):
            
            #------------------------------------------------------------------
            # Initialize temporary data structure for full simulation
            #------------------------------------------------------------------
            
            Data={}
            
            Data['v1']={}
            for iV in range(len(nam1)):
                Data['v1'][nam1[iV]]=np.zeros((tv_saving.size,meta['Project']['N Stand']),dtype=int)
            
            if switch_area=='On':
                Data['Area']={}
                for k in MosByMP[iScn]['Area']:
                    Data['Area'][k]=np.zeros((tv_saving.size,meta['Project']['N Stand']),dtype=int)
                    
            if switch_cashflow=='On':
                Data['Cashflow']={}
                for iV in range(len(nam_cashflow)):
                    nam=nam_cashflow[iV]
                    Data['Cashflow'][nam]=np.zeros((tv_saving.size,meta['Project']['N Stand']),dtype=int)

            #------------------------------------------------------------------
            # Populate full simulation results
            #------------------------------------------------------------------
            
            for iBat in range(meta['Project']['N Batch']):
        
                indBat=IndexToBatch(meta,iBat)
            
                d1=LoadSingleOutputFile(meta,iScn,iEns,iBat)
                
                for iV in range(len(nam1)):
                    tmp=d1[nam1[iV]]/ScaleFactor
                    Data['v1'][nam1[iV]][:,indBat]=tmp.copy().astype(int)
            
                if (switch_area=='On' ) | (switch_cashflow=='On'):
                    
                    # Import event chronology
                    if (meta['Scenario'][iScn]['Harvest Status Future']=='On') | (meta['Scenario'][iScn]['Breakup Status']=='On'):
                        ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Modified_Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
                    else:
                        ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')

                    # Uncompress event chronology if it has been compressed
                    ec=EventChronologyDecompress(meta,ec,iScn,iEns,iBat) 
            
                if switch_area=='On':
                
                    for k in Data['Area'].keys():
                        Data0=np.zeros((tv_saving.size,indBat.size))
                        for iEY in range(meta['Core']['Max Events Per Year']):
                            ind=np.where(ec['ID_Type'][it,:,iEY]==meta['LUT']['Dist'][k])[0]
                            Data0[ind]=Data0[ind]+1
                        Data['Area'][k][:,indBat]=Data0
                
                if switch_cashflow=='On':
                    
                    inv=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + FixFileNum(iBat) + '.pkl')
                    
                    econ=econo.CalculateNetRevenue(meta,iScn,iEns,iBat,inv,ec,d1)
                    
                    for iV in range(len(nam_cashflow)):
                        tmp=econ[nam_cashflow[iV]]/ScaleFactor
                        Data['Cashflow'][nam_cashflow[iV]][:,indBat]=tmp.copy().astype(int)
                
                    del econ
                
                del d1
                garc.collect()
        
            if switch_area=='On':
                # Populating the final structure with area data is slow - get a flag
                # indicator of whether each event ID can be skipped because it has no
                # info
                flag_Area={}
                for k in Data['Area'].keys():            
                    if np.sum(Data['Area'][k])>0:
                        flag_Area[k]=1
                    else:
                        flag_Area[k]=0
            
            #------------------------------------------------------------------
            # Calculate stats and populate results for each treatment area    
            #------------------------------------------------------------------
            
            for iMP in range(uMP.size):
                
                ATA=atu_multipolygons[uMP[iMP]]['ACTUAL_TREATMENT_AREA']
                if ATA==None:
                    print('Encounterd no area, using zero')
                    ATA=0
                    
                ind=Crosswalk_sxy_to_mp[iMP]['Index']
                
                for iV in range(len(nam1)):
                    tmp=ScaleFactor*Data['v1'][nam1[iV]][:,ind].astype(float)
                    MosByMP[iScn]['v1']['Mean'][nam1[iV]]['Ensemble Mean'][:,iMP]=MosByMP[iScn]['v1']['Mean'][nam1[iV]]['Ensemble Mean'][:,iMP]+np.mean(tmp,axis=1)
                    MosByMP[iScn]['v1']['Mean'][nam1[iV]]['Ensemble SD'][:,iMP]=MosByMP[iScn]['v1']['Mean'][nam1[iV]]['Ensemble SD'][:,iMP]+np.std(tmp,axis=1)
                    MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble Mean'][:,iMP]=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble Mean'][:,iMP]+ATA*np.mean(tmp,axis=1)
                    MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble SD'][:,iMP]=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble SD'][:,iMP]+ATA*np.std(tmp,axis=1)
                    MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P10'][:,iMP]=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P10'][:,iMP]+ATA*np.percentile(tmp,10,axis=1)
                    MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P25'][:,iMP]=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P25'][:,iMP]+ATA*np.percentile(tmp,25,axis=1)
                    MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P75'][:,iMP]=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P75'][:,iMP]+ATA*np.percentile(tmp,75,axis=1)
                    MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P90'][:,iMP]=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P90'][:,iMP]+ATA*np.percentile(tmp,90,axis=1)
                
                if switch_cashflow=='On':
                    for iV in range(len(nam_cashflow)):
                        tmp=ScaleFactor*Data['Cashflow'][nam_cashflow[iV]][:,ind].astype(float)
                        MosByMP[iScn]['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble Mean'][:,iMP]=MosByMP[iScn]['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble Mean'][:,iMP]+np.mean(tmp,axis=1)
                        MosByMP[iScn]['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble SD'][:,iMP]=MosByMP[iScn]['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble SD'][:,iMP]+np.std(tmp,axis=1)
                        MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble Mean'][:,iMP]=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble Mean'][:,iMP]+ATA*np.mean(tmp,axis=1)
                        MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble SD'][:,iMP]=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble SD'][:,iMP]+ATA*np.std(tmp,axis=1)
                        MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P10'][:,iMP]=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P10'][:,iMP]+ATA*np.percentile(tmp,10,axis=1)
                        MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P25'][:,iMP]=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P25'][:,iMP]+ATA*np.percentile(tmp,25,axis=1)
                        MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P75'][:,iMP]=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P75'][:,iMP]+ATA*np.percentile(tmp,75,axis=1)
                        MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P90'][:,iMP]=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P90'][:,iMP]+ATA*np.percentile(tmp,90,axis=1)
                
                if switch_area=='On':
                    for k in MosByMP[iScn]['Area']:
                        if flag_Area[k]==1:
                            # Only continue if there are some events
                            MosByMP[iScn]['Area'][k]['Ensemble Mean'][:,iMP]=MosByMP[iScn]['Area'][k]['Ensemble Mean'][:,iMP]+np.sum(Data['Area'][k][:,ind],axis=1)

    #--------------------------------------------------------------------------
    # Divide by number of ensembles
    #--------------------------------------------------------------------------
    
    for iScn in range(meta['Project']['N Scenario']):
        
        for iV in range(len(nam1)):
            MosByMP[iScn]['v1']['Mean'][nam1[iV]]['Ensemble Mean']=MosByMP[iScn]['v1']['Mean'][nam1[iV]]['Ensemble Mean']/meta['Project']['N Ensemble']
            MosByMP[iScn]['v1']['Mean'][nam1[iV]]['Ensemble SD']=MosByMP[iScn]['v1']['Mean'][nam1[iV]]['Ensemble SD']/meta['Project']['N Ensemble']
            MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble Mean']=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble Mean']/meta['Project']['N Ensemble']
            MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble SD']=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble SD']/meta['Project']['N Ensemble']
            MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P10']=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P10']/meta['Project']['N Ensemble']
            MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P25']=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P25']/meta['Project']['N Ensemble']
            MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P75']=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P75']/meta['Project']['N Ensemble']
            MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P90']=MosByMP[iScn]['v1']['Sum'][nam1[iV]]['Ensemble P90']/meta['Project']['N Ensemble']

        if switch_cashflow=='On':
            for iV in range(len(nam_cashflow)):
                MosByMP[iScn]['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble Mean']=MosByMP[iScn]['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble Mean']/meta['Project']['N Ensemble']
                MosByMP[iScn]['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble SD']=MosByMP[iScn]['Cashflow']['Mean'][nam_cashflow[iV]]['Ensemble SD']/meta['Project']['N Ensemble']
                MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble Mean']=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble Mean']/meta['Project']['N Ensemble']
                MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble SD']=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble SD']/meta['Project']['N Ensemble']            
                MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P10']=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P10']/meta['Project']['N Ensemble']
                MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P25']=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P25']/meta['Project']['N Ensemble']
                MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P75']=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P75']/meta['Project']['N Ensemble']
                MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P90']=MosByMP[iScn]['Cashflow']['Sum'][nam_cashflow[iV]]['Ensemble P90']/meta['Project']['N Ensemble']    
        
        if switch_area=='On':
            for iV in MosByMP[iScn]['Area']:
                MosByMP[iScn]['Area'][iV]['Ensemble Mean'][:,iMP]=MosByMP[iScn]['Area'][iV]['Ensemble Mean'][:,iMP]/meta['Project']['N Ensemble']

    #--------------------------------------------------------------------------
    # Save
    #--------------------------------------------------------------------------
    
    gu.opickle(meta['Paths']['Project'] + '\\Outputs\\MosByMultipolygon.pkl',MosByMP)
    
    return

#%%

def MOS_FromPoints_ByProjAndReg_GHG(meta,**kwargs):

    t0=time.time()

    # Key word arguments
    if 'StandsToInclude' in kwargs.keys():
        flag_stands_to_include=kwargs['StandsToInclude']
    else:
        flag_stands_to_include=[]

    # Unique project types
    uPT=np.unique(meta['Project']['Project Stratum'])
    
    # Unique regions
    uReg=np.unique(meta['Project']['Spatial Stratum'])

    # Time series of saved results
    tv_saving=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)
    
    # All non-economic values
    d1=LoadSingleOutputFile(meta,0,0,0)
    del d1['C_M_ByAgent']
    del d1['Year']
    v2include=list(d1.keys())    
        
    # Add Sawtooth variables
    if meta['Project']['Biomass Module']=='Sawtooth':
        v2include=v2include+['N','N_R','N_M_Tot','N_M_Reg','TreeMean_A','TreeMean_H','TreeMean_D','TreeMean_Csw','TreeMean_Csw_G']
    
    # Loop through scenarios
    for iScn in range(meta['Project']['N Scenario']):
        
        #--------------------------------------------------------------------------
        # Initialize data structures        
        #--------------------------------------------------------------------------
        
        MoMeanByPT={}
        MoSumByPT={}
        for k in v2include:
            MoMeanByPT[k]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
            MoSumByPT[k]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
        
        # Loop through ensembles
        for iEns in range(meta['Project']['N Ensemble']):
            
            #--------------------------------------------------------------------------
            # Initialize temporary data structure for full simulation            
            #--------------------------------------------------------------------------
            
            DataSXY={}
            for k in v2include:
                DataSXY[k]=np.nan*np.empty( (tv_saving.size,meta['Project']['N Stand']) ,dtype=np.float)
            
            for iBat in range(meta['Project']['N Batch']):
        
                # Include specific subset of stands
                if len(flag_stands_to_include)==0:
                    iKeepStands=np.arange(0,meta['Project']['Batch Size'][iBat],1,dtype=int)
                else:
                    iKeepStands=np.where(flag_stands_to_include[iScn][iEns][iBat]==1)[0]
                
                # Index to batch
                indBat=IndexToBatch(meta,iBat)
            
                d1=LoadSingleOutputFile(meta,iScn,iEns,iBat)

                for k in v2include:
                    
                    DataSXY[k][:,indBat[iKeepStands]]=d1[k][:,iKeepStands].copy()
            
            del d1
            garc.collect()
            
            #--------------------------------------------------------------------------
            # Summarize by project type, region, and time
            #--------------------------------------------------------------------------
            
            for iPT in range(uPT.size):
                
                for iReg in range(uReg.size):
                
                    ind1=np.where( (meta['Project']['Project Stratum']==uPT[iPT]) & (meta['Project']['Spatial Stratum']==uReg[iReg]) )[0]
                
                    for k in v2include:
                    
                        Mu=np.nanmean(DataSXY[k][:,ind1],axis=1)
                        Sum=np.nansum(DataSXY[k][:,ind1],axis=1)
                        
                        MoMeanByPT[k][:,iEns,iPT,iReg]=Mu.copy()
                        MoSumByPT[k][:,iEns,iPT,iReg]=Sum.copy()

        #--------------------------------------------------------------------------
        # Save
        #--------------------------------------------------------------------------
        
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeAndReg_Mean_Scn' + str(iScn+1) + '_FromPoints.pkl',MoMeanByPT)
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeAndReg_Sum_Scn' + str(iScn+1) + '_FromPoints.pkl',MoSumByPT)
    
    t1=time.time()
    #print((t1-t0)/60)
    
    return

#%%
    
def MOS_FromPoints_ByProjAndReg_Econ(meta,**kwargs):

    t0=time.time()

    # Kewword argumenst
    
    if 'ScenariosToInclude' in kwargs.keys():
        ScenariosToInclude=kwargs['ScenariosToInclude']
    else:
        # Default is to not save individual ensembles
        ScenariosToInclude=np.arange(0,meta['Project']['N Scenario'])

    if 'StandsToInclude' in kwargs.keys():
        flag_stands_to_include=kwargs['StandsToInclude']
    else:
        flag_stands_to_include=[]

    # Import geos
    #geos=gu.ipickle(meta['Paths']['Geospatial'] + '\\geos.pkl')

    # Unique project types
    uPT=np.unique(meta['Project']['Project Stratum'])
    
    # Unique regions
    uReg=np.unique(meta['Project']['Spatial Stratum'])

    # Time series of saved results
    tv_saving=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)
    
    # All (excluding 'C_M_ByAgent' and 'Year','C_M_Dist')
    v2include=['Yield Lumber','Yield Plywood','Yield OSB','Yield MDF','Yield Paper','Yield Pellets','Yield PowerGrid',
               'Yield PowerFacilityDom','Yield FirewoodDom','Yield LogExport','Cost Roads','Cost Harvest Overhead',
               'Cost Harvest Felling and Piling','Cost Harvest Hauling','Cost Harvest Residuals',
               'Cost Milling','Cost Nutrient Management','Cost Planting','Cost Survey','Cost Knockdown','Cost Ripping','Cost PAS Deactivation',
               'Cost Slashpile Burn','Cost Total','Cost Silviculture Total','Revenue Lumber','Revenue Plywood',
               'Revenue OSB','Revenue MDF','Revenue Paper','Revenue PowerFacilityDom','Revenue PowerGrid','Revenue Pellets','Revenue FirewoodDom',
               'Revenue LogExport','Revenue Gross','Revenue Net','Revenue Net Disc','Revenue Gross Disc','Cost Total Disc','Cost Nutrient Management Disc',
               'Cost Silviculture Total Disc','Cost Total_cumu','Cost Silviculture Total_cumu','Cost Nutrient Management_cumu','Cost Total Disc_cumu',
               'Cost Silviculture Total Disc_cumu','Cost Nutrient Management Disc_cumu','Revenue Gross_cumu','Revenue Gross Disc_cumu','Revenue Net_cumu',
               'Revenue Net Disc_cumu']
    
    # Scale factor used to temporarily store data
    ScaleFactor=1.0
    
    # Loop through scenarios
    for iS in range(ScenariosToInclude.size):
        
        iScn=ScenariosToInclude[iS]
        
        #--------------------------------------------------------------------------
        # Initialize data structures        
        #--------------------------------------------------------------------------
        
        MoMeanByPT={}
        MoSumByPT={}
        for k in v2include:
            MoMeanByPT[k]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
            MoSumByPT[k]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
        
        # Loop through ensembles
        for iEns in range(meta['Project']['N Ensemble']):
            
            #--------------------------------------------------------------------------
            # Initialize temporary data structure for full simulation            
            #--------------------------------------------------------------------------
            
            DataSXY={}
            for k in v2include:
                DataSXY[k]=np.nan*np.empty( (tv_saving.size,meta['Project']['N Stand']) ,dtype=np.float)
            
            #--------------------------------------------------------------------------
            # Import batches
            #--------------------------------------------------------------------------
            
            for iBat in range(meta['Project']['N Batch']):
        
                # Include specific subset of stands
                if len(flag_stands_to_include)==0:
                    iKeepStands=np.arange(0,meta['Project']['Batch Size'][iBat],1,dtype=int)
                else:
                    iKeepStands=np.where(flag_stands_to_include[iScn][iEns][iBat]==1)[0]
                
                # Index to batch
                indBat=IndexToBatch(meta,iBat)
            
                d1=LoadSingleOutputFile(meta,iScn,iEns,iBat)
                
                if (meta['Scenario'][iScn]['Harvest Status Future']=='On') | (meta['Scenario'][iScn]['Breakup Status']=='On'):
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Modified_Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
                else:
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
            
                # Uncompress event chronology if it has been compressed
                ec=EventChronologyDecompress(meta,ec,iScn,iEns,iBat)
            
                # Inventory
                inv=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + FixFileNum(iBat) + '.pkl')
                    
                # Cashflow
                econ1=econo.CalculateNetRevenue(meta,iScn,iEns,iBat,inv,ec,d1) 
                
                for k in v2include:
                    
                    DataSXY[k][:,indBat[iKeepStands]]=econ1[k][:,iKeepStands].copy()
            
            del econ1,inv,ec
            #garc.collect()
            
            #--------------------------------------------------------------------------
            # Summarize by project type, region, and time
            #--------------------------------------------------------------------------
            
            for iPT in range(uPT.size):
                
                for iReg in range(uReg.size):
                
                    ind1=np.where( (meta['Project']['Project Stratum']==uPT[iPT]) & (meta['Project']['Spatial Stratum']==uReg[iReg]) )[0]
                
                    for k in v2include:
                    
                        MoMeanByPT[k][:,iEns,iPT,iReg]=ScaleFactor*np.nanmean(DataSXY[k][:,ind1],axis=1)
                        MoSumByPT[k][:,iEns,iPT,iReg]=ScaleFactor*np.nansum(DataSXY[k][:,ind1],axis=1)

        #--------------------------------------------------------------------------
        # Save
        #--------------------------------------------------------------------------
        
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeAndReg_Mean_Scn' + str(iScn+1) + '_FromPoints.pkl',MoMeanByPT)
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeAndReg_Sum_Scn' + str(iScn+1) + '_FromPoints.pkl',MoSumByPT)
    
    t1=time.time()
    #print((t1-t0)/60)
    
    return

#%% Summarize mortality from points
    
# *** This is not by project type and region because cbrunner is not currently 
# set up to save mortality by stand (too much data) ***

def MOS_FromPoints_Mortality(meta):

    # Time series of saved results
    tv_saving=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)
    
    # Initialize list
    mos=[None]*meta['Project']['N Scenario']
    
    # Loop through scenarios
    for iScn in range(meta['Project']['N Scenario']):

        #--------------------------------------------------------------------------
        # Initialize data structures        
        #--------------------------------------------------------------------------
        
        d0=LoadSingleOutputFile(meta,0,0,0)
        
        mos[iScn]={}
        for k in d0['C_M_ByAgent'].keys():
            mos[iScn][k]=np.zeros((tv_saving.size,meta['Project']['N Ensemble']))
        
        # Loop through ensembles
        for iEns in range(meta['Project']['N Ensemble']):
            
            # Initialize temporary data structure for full simulation                        
            Data={}
            for k in d0['C_M_ByAgent'].keys():
                Data[k]=np.zeros(tv_saving.size)
            
            # Loop through batches and add to Data structure           
            for iBat in range(meta['Project']['N Batch']):            
                d1=LoadSingleOutputFile(meta,iScn,iEns,iBat)
                for k in d0['C_M_ByAgent'].keys():
                    Data[k]=Data[k]+d1['C_M_ByAgent'][k].flatten()
                
                del d1
                garc.collect()
            
            # Divide by N stand to get mean
            for k in d0['C_M_ByAgent'].keys():
                Data[k]=Data[k]/meta['Project']['N Stand']
            
            # Populate ensembles
            for k in d0['C_M_ByAgent'].keys():
                mos[iScn][k][:,iEns]=Data[k].copy()

        # Average all ensembles
        for k in d0['C_M_ByAgent'].keys():
            mos[iScn][k]=np.nanmean(mos[iScn][k],axis=1)

    # Save
    gu.opickle(meta['Paths']['Project'] + '\\Outputs\\Mortality_Mean__FromPoints.pkl',mos)
    
    return

#%% Summarize area from points

def MOS_FromPoints_ByProjAndReg_Area(meta,**kwargs):

    t0=time.time()

    # Key word arguments
    if 'StandsToInclude' in kwargs.keys():
        flag_stands_to_include=kwargs['StandsToInclude']
    else:
        flag_stands_to_include=[]

    # Import geos
    #geos=gu.ipickle(meta['Paths']['Geospatial'] + '\\geos.pkl')

    # Unique project types
    uPT=np.unique(meta['Project']['Project Stratum'])
    
    # Unique regions
    uReg=np.unique(meta['Project']['Spatial Stratum'])

    tv=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)

    # Time series of saved results
    tv_saving=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)
    
    # Full time series (for event chronology)
    tv_full=np.arange(meta['Project']['Year Start'],meta['Project']['Year End']+1,1)
    
    # Loop through scenarios
    for iScn in range(meta['Project']['N Scenario']):
        
        #--------------------------------------------------------------------------
        # Initialize data structures        
        #--------------------------------------------------------------------------
        
        MoMeanByPT={}
        MoSumByPT={}
        for k in meta['LUT']['Dist'].keys():
            MoMeanByPT[k]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
            MoSumByPT[k]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
        
        # Loop through ensembles
        for iEns in range(meta['Project']['N Ensemble']):
            
            #--------------------------------------------------------------------------
            # Initialize temporary data structure for full simulation            
            #--------------------------------------------------------------------------
            
            DataSXY={}
            for k in meta['LUT']['Dist'].keys():
                DataSXY[k]=np.nan*np.empty( (tv_saving.size,meta['Project']['N Stand']) ,dtype=np.float)
                
            DataSXY_Full={}
            for k in meta['LUT']['Dist'].keys():
                DataSXY_Full[k]=np.zeros( (tv_saving.size,meta['Project']['N Stand']) ,dtype=np.float)
            
            #--------------------------------------------------------------------------
            # Import batches
            #--------------------------------------------------------------------------
            
            for iBat in range(meta['Project']['N Batch']):
        
                # Include specific subset of stands
                if len(flag_stands_to_include)==0:
                    iKeepStands=np.arange(0,meta['Project']['Batch Size'][iBat],1,dtype=int)
                else:
                    iKeepStands=np.where(flag_stands_to_include[iScn][iEns][iBat]==1)[0]
                
                # Index to batches
                indBat=IndexToBatch(meta,iBat)
            
                #d1=LoadSingleOutputFile(meta,iScn,iEns,iBat)
                
                # Import event chronology                
                if (meta['Scenario'][iScn]['Harvest Status Future']=='On') | (meta['Scenario'][iScn]['Breakup Status']=='On'):
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Modified_Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
                else:
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
            
                # Uncompress event chronology if it has been compressed
                ec=EventChronologyDecompress(meta,ec,iScn,iEns,iBat) 
                
                for iT in range(tv_full.size):
                    indT=np.where(tv==tv_full[iT])[0]
                    if indT.size==0:
                        continue
                    idDist0=np.squeeze(ec['ID_Type'][iT,:,:])
                    for iDistWithinYear in range(idDist0.shape[1]):
                        idDist1=idDist0[:,iDistWithinYear]
                        uidDist=np.unique(idDist1)
                        if np.sum(uidDist)==0:
                            continue
                        for iU in range(uidDist.size):
                            if uidDist[iU]==0:
                                continue
                            namDist=lut_n2s(meta['LUT']['Dist'],uidDist[iU])[0]
                            indDist=np.where(idDist1==uidDist[iU])[0]
                            
                            DataSXY_Full[namDist][indT,indBat[indDist]]=DataSXY_Full[namDist][indT,indBat[indDist]]+1
                
                # Pull results for subset of stands
                for namDist in meta['LUT']['Dist'].keys():
                    DataSXY[namDist][:,indBat[iKeepStands]]=DataSXY_Full[namDist][:,indBat[iKeepStands]]
                
            del ec
            garc.collect()
            
            #--------------------------------------------------------------------------
            # Summarize by project type, region, and time
            #--------------------------------------------------------------------------
            
            for iPT in range(uPT.size):
                
                for iReg in range(uReg.size):
                
                    ind1=np.where( (meta['Project']['Project Stratum']==uPT[iPT]) & (meta['Project']['Spatial Stratum']==uReg[iReg]) )[0]
                
                    for k in meta['LUT']['Dist'].keys():
                    
                        MoMeanByPT[k][:,iEns,iPT,iReg]=np.nanmean(DataSXY[k][:,ind1],axis=1)
                        MoSumByPT[k][:,iEns,iPT,iReg]=np.nansum(DataSXY[k][:,ind1],axis=1)

        #--------------------------------------------------------------------------
        # Save
        #--------------------------------------------------------------------------
        
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\Area_ByProjTypeAndReg_Mean_Scn' + str(iScn+1) + '_FromPoints.pkl',MoMeanByPT)
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\Area_ByProjTypeAndReg_Sum_Scn' + str(iScn+1) + '_FromPoints.pkl',MoSumByPT)
    
    t1=time.time()
    print((t1-t0)/60)
    
    return

#%% Summarize GHG for multipolygons

def MOS_FromMPs_ByProjTypeRegAndYear_GHG(meta):

    t0=time.time()

    # Import geos
    geos=gu.ipickle(meta['Paths']['Geospatial'] + '\\geos.pkl')

    # Unique MPs
    uMP=np.unique(geos['Sparse']['ID_atu_multipolygons'])

    # Unique project types
    uPT=np.unique(meta['Project']['ByMP']['Project Type'])
    
    # Unique regions
    uReg=np.unique(meta['Project']['ByMP']['Region ID'])

    # Create listed index (faster than indexing on the fly)
    Crosswalk_sxy_to_mp=[None]*uMP.size
    for iMP in range(uMP.size):
        d={}
        d['Index']=np.where(geos['Sparse']['ID_atu_multipolygons']==uMP[iMP])[0]
        Crosswalk_sxy_to_mp[iMP]=d

    # Pull out project type and year for unique MPs
    ProjectType=np.zeros(uMP.size)
    ProjectArea=np.zeros(uMP.size)
    ProjectYear=np.zeros(uMP.size)
    Region=np.zeros(uMP.size)
    for iMP in range(uMP.size):
        ProjectType[iMP]=meta['Project']['ByMP']['Project Type'][uMP[iMP]]
        ProjectArea[iMP]=meta['Project']['ByMP']['Project Area'][uMP[iMP]]
        ProjectYear[iMP]=meta['Project']['ByMP']['Project Year'][uMP[iMP]]
        Region[iMP]=meta['Project']['ByMP']['Region ID'][uMP[iMP]]

    # Time series of saved results
    tv_saving=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)
    
    # Years of implementation
    uT=np.unique(ProjectYear[ProjectYear!=0])
    tv_Implementation=np.arange(np.min(uT),np.max(uT)+1,1)
    
    # All (excluding 'C_M_ByAgent' and 'Year','C_M_Dist')
    v2include=['A','V_MerchLive','V_MerchDead','V_MerchTotal','V_ToMillMerchLive','V_ToMillMerchDead','V_ToMillMerchTotal','V_ToMillNonMerch',
               'LogSizeEnhancement','C_Forest_Tot','C_HWP_Tot','C_NPP_Tot','C_ToMill','C_Biomass_Tot','C_Piled_Tot','C_Litter_Tot','C_DeadWood_Tot',
               'C_Soil_Tot','C_InUse_Tot','C_DumpLandfill_Tot','C_G_Gross_Tot','C_G_Net_Tot','C_M_Reg_Tot','C_LF_Tot','C_RH_Tot',
               'C_ToMillMerch','C_ToMillNonMerch','C_ToMillSnagStem','C_ToSlashpileBurn','C_ToLumber','C_ToPlywood','C_ToOSB','C_ToMDF',
               'C_ToPaper','C_ToPowerFacilityDom','C_ToPowerFacilityFor','C_ToPowerGrid','C_ToPellets','C_ToFirewoodDom',
               'C_ToFirewoodFor','C_ToLogExport',
               'E_CO2e_LULUCF_NEE','E_CO2e_LULUCF_Denit','E_CO2e_LULUCF_Other','E_CO2e_LULUCF_Wildfire','E_CO2e_LULUCF_OpenBurning',
               'E_CO2e_LULUCF_Fire','E_CO2e_LULUCF_HWP','E_CO2e_ESC_Bioenergy',
               'E_CO2e_ESC_Operations','E_CO2e_ET_Operations','E_CO2e_IPPU_Operations',
               'E_CO2e_Coal','E_CO2e_Oil','E_CO2e_Gas',
               'E_CO2e_SUB_Coal','E_CO2e_SUB_Oil','E_CO2e_SUB_Gas','E_CO2e_SUB_Calcination','E_CO2e_SUB_E','E_CO2e_SUB_M','E_CO2e_SUB_Tot',
               'E_CO2e_AGHGB_WSub','E_CO2e_AGHGB_WOSub','E_CO2e_AGHGB_WSub_cumu','E_CO2e_AGHGB_WOSub_cumu',
               'E_CO2e_AGHGB_WSub_cumu_from_tref','E_CO2e_AGHGB_WOSub_cumu_from_tref',
               'Yield Coal','Yield Oil','Yield Gas','Yield Sawnwood','Yield Panels','Yield Concrete',
               'Yield Steel','Yield Aluminum','Yield Plastic','Yield Textile']
    
    # Scale factor used to temporarily store data
    #ScaleFactor=1.0
    
    # Loop through scenarios
    for iScn in range(meta['Project']['N Scenario']):
        
        #--------------------------------------------------------------------------
        # Initialize data structures        
        #--------------------------------------------------------------------------
        
        MoMeanByPT={}
        MoSumByPT={}
        MoMeanByPTAndYr={}
        MoSumByPTAndYr={}
        for k in v2include:
            MoMeanByPT[k]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
            MoSumByPT[k]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
        
            MoMeanByPTAndYr[k]={}
            MoSumByPTAndYr[k]={}
            for t in tv_Implementation:
                MoMeanByPTAndYr[k][t]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
                MoSumByPTAndYr[k][t]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
        
        # Loop through ensembles
        for iEns in range(meta['Project']['N Ensemble']):
            
            #--------------------------------------------------------------------------
            # Initialize temporary data structure for full simulation            
            #--------------------------------------------------------------------------
            
            DataSXY={}
            for k in v2include:
                DataSXY[k]=np.zeros( (tv_saving.size,meta['Project']['N Stand']) ,dtype=np.float)
            
            #--------------------------------------------------------------------------
            # Import batches
            #--------------------------------------------------------------------------
            
            for iBat in range(meta['Project']['N Batch']):
        
                indBat=IndexToBatch(meta,iBat)
            
                d1=LoadSingleOutputFile(meta,iScn,iEns,iBat)
                
                for k in v2include:
                    
                    DataSXY[k][:,indBat]=d1[k].copy()
            
            del d1
            garc.collect()
            
            #--------------------------------------------------------------------------
            # Convert from SXY results to MP averages
            #--------------------------------------------------------------------------
            
            MeanByMP={}
            SumByMP={}
            for k in v2include:
                MeanByMP[k]=np.zeros( (tv_saving.size,uMP.size) )
                SumByMP[k]=np.zeros( (tv_saving.size,uMP.size) )
            
            for iMP in range(uMP.size):                
                
                # Index to SXY for the ith multipolygon
                indSXY=Crosswalk_sxy_to_mp[iMP]['Index']
                
                for k in v2include:
                    MeanByMP[k][:,iMP]=MeanByMP[k][:,iMP]+np.mean(DataSXY[k][:,indSXY],axis=1)
                    SumByMP[k][:,iMP]=SumByMP[k][:,iMP]+np.sum(ProjectArea[iMP])*np.mean(DataSXY[k][:,indSXY],axis=1)
            
            #--------------------------------------------------------------------------
            # Summarize by project type, region, and time
            #--------------------------------------------------------------------------
            
            for iPT in range(uPT.size):
                
                for iReg in range(uReg.size):
                
                    ind1=np.where( (ProjectType==uPT[iPT]) & (Region==uReg[iReg]) )[0]
                
                    for k in v2include:
                    
                        MoMeanByPT[k][:,iEns,iPT,iReg]=np.mean(MeanByMP[k][:,ind1],axis=1)
                        MoSumByPT[k][:,iEns,iPT,iReg]=np.sum(SumByMP[k][:,ind1],axis=1)
                        
                        # Summmarize by PT and year
                        for t in tv_Implementation:
                        
                            ind2=np.where( (ProjectType==uPT[iPT]) & (Region==uReg[iReg]) & (ProjectYear==t) )[0]
                            
                            MoMeanByPTAndYr[k][t][:,iEns,iPT,iReg]=np.mean(MeanByMP[k][:,ind2],axis=1)
                            MoSumByPTAndYr[k][t][:,iEns,iPT,iReg]=np.sum(SumByMP[k][:,ind2],axis=1)

        #--------------------------------------------------------------------------
        # Save
        #--------------------------------------------------------------------------
        
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeReg_Mean_Scn' + str(iScn+1) + '_FromMPs.pkl',MoMeanByPT)
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeReg_Sum_Scn' + str(iScn+1) + '_FromMPs.pkl',MoSumByPT)
        
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeRegAndYear_Mean_Scn' + str(iScn+1) + '_FromMPs.pkl',MoMeanByPTAndYr)
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeRegAndYear_Sum_Scn' + str(iScn+1) + '_FromMPs.pkl',MoSumByPTAndYr)
    
    t1=time.time()
    print((t1-t0)/60)
    
    return

#%% Summarize economcis for multipolygons

def MOS_FromMPs_ByProjTypeRegAndYear_Econ(meta):

    t0=time.time()

    # Import geos
    geos=gu.ipickle(meta['Paths']['Geospatial'] + '\\geos.pkl')

    # Unique MPs
    uMP=np.unique(geos['Sparse']['ID_atu_multipolygons'])

    # Unique project types
    uPT=np.unique(meta['Project']['ByMP']['Project Type'])
    
    # Unique regions
    uReg=np.unique(meta['Project']['ByMP']['Region ID'])

    # Create listed index (faster than indexing on the fly)
    Crosswalk_sxy_to_mp=[None]*uMP.size
    for iMP in range(uMP.size):
        d={}
        d['Index']=np.where(geos['Sparse']['ID_atu_multipolygons']==uMP[iMP])[0]
        Crosswalk_sxy_to_mp[iMP]=d

    # Pull out project type and year for unique MPs
    ProjectType=np.zeros(uMP.size)
    ProjectArea=np.zeros(uMP.size)
    ProjectYear=np.zeros(uMP.size)
    Region=np.zeros(uMP.size)
    for iMP in range(uMP.size):
        ProjectType[iMP]=meta['Project']['ByMP']['Project Type'][uMP[iMP]]
        ProjectArea[iMP]=meta['Project']['ByMP']['Project Area'][uMP[iMP]]
        ProjectYear[iMP]=meta['Project']['ByMP']['Project Year'][uMP[iMP]]
        Region[iMP]=meta['Project']['ByMP']['Region ID'][uMP[iMP]]

    # Time series of saved results
    tv_saving=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)
    
    # Years of implementation
    uT=np.unique(ProjectYear[ProjectYear!=0])
    tv_Implementation=np.arange(np.min(uT),np.max(uT)+1,1)
    
    # Variables to include
    v2include=['Yield Lumber','Yield Plywood','Yield OSB','Yield MDF','Yield Paper','Yield Pellets','Yield PowerGrid',
               'Yield PowerFacilityDom','Yield FirewoodDom','Yield LogExport','Cost Roads','Cost Harvest Overhead',
               'Cost Harvest Felling and Piling','Cost Harvest Hauling','Cost Harvest Residuals',
               'Cost Milling','Cost Nutrient Management','Cost Planting','Cost Survey','Cost Knockdown','Cost Ripping','Cost PAS Deactivation',
               'Cost Slashpile Burn','Cost Total','Cost Silviculture Total','Revenue Lumber','Revenue Plywood',
               'Revenue OSB','Revenue MDF','Revenue Paper','Revenue PowerFacilityDom','Revenue PowerGrid','Revenue Pellets','Revenue FirewoodDom',
               'Revenue LogExport','Revenue Gross','Revenue Net','Revenue Net Disc','Revenue Gross Disc','Cost Total Disc','Cost Nutrient Management Disc',
               'Cost Silviculture Total Disc','Cost Total_cumu','Cost Silviculture Total_cumu','Cost Nutrient Management_cumu','Cost Total Disc_cumu',
               'Cost Silviculture Total Disc_cumu','Cost Nutrient Management Disc_cumu','Revenue Gross_cumu','Revenue Gross Disc_cumu','Revenue Net_cumu',
               'Revenue Net Disc_cumu']
    
    # Scale factor used to temporarily store data
    ScaleFactor=1.0
    
    # Loop through scenarios
    for iScn in range(meta['Project']['N Scenario']):
        
        #--------------------------------------------------------------------------
        # Initialize data structures        
        #--------------------------------------------------------------------------
        
        MoMeanByPT={}
        MoSumByPT={}
        MoMeanByPTAndYr={}
        MoSumByPTAndYr={}
        for k in v2include:
            MoMeanByPT[k]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
            MoSumByPT[k]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
        
            MoMeanByPTAndYr[k]={}
            MoSumByPTAndYr[k]={}
            for t in tv_Implementation:
                MoMeanByPTAndYr[k][t]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
                MoSumByPTAndYr[k][t]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],uPT.size,uReg.size) ,dtype=np.float)
        
        # Loop through ensembles
        for iEns in range(meta['Project']['N Ensemble']):
            
            #--------------------------------------------------------------------------
            # Initialize temporary data structure for full simulation            
            #--------------------------------------------------------------------------
            
            DataSXY={}
            for k in v2include:
                DataSXY[k]=np.zeros( (tv_saving.size,meta['Project']['N Stand']) ,dtype=np.float)
            
            #--------------------------------------------------------------------------
            # Import batches
            #--------------------------------------------------------------------------
            
            for iBat in range(meta['Project']['N Batch']):
        
                indBat=IndexToBatch(meta,iBat)
            
                d1=LoadSingleOutputFile(meta,iScn,iEns,iBat)
                
                if (meta['Scenario'][iScn]['Harvest Status Future']=='On') | (meta['Scenario'][iScn]['Breakup Status']=='On'):
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Modified_Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
                else:
                    ec=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
            
                # Uncompress event chronology if it has been compressed
                ec=EventChronologyDecompress(meta,ec,iScn,iEns,iBat)
            
                # Inventory
                inv=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + FixFileNum(iBat) + '.pkl')
                    
                # Cashflow
                econ1=econo.CalculateNetRevenue(meta,iScn,iEns,iBat,inv,ec,d1) 
                
                for k in v2include:
                    
                    DataSXY[k][:,indBat]=econ1[k].copy()
            
            del d1,ec,inv,econ1
            #garc.collect()
            
            #--------------------------------------------------------------------------
            # Convert from SXY results to MP averages
            #--------------------------------------------------------------------------
            
            MeanByMP={}
            SumByMP={}
            for k in v2include:
                MeanByMP[k]=np.zeros( (tv_saving.size,uMP.size) )
                SumByMP[k]=np.zeros( (tv_saving.size,uMP.size) )
            
            for iMP in range(uMP.size):                
                
                # Index to SXY for the ith multipolygon
                indSXY=Crosswalk_sxy_to_mp[iMP]['Index']
                
                for k in v2include:
                    MeanByMP[k][:,iMP]=MeanByMP[k][:,iMP]+np.mean(DataSXY[k][:,indSXY],axis=1)
                    SumByMP[k][:,iMP]=SumByMP[k][:,iMP]+np.sum(ProjectArea[iMP])*np.mean(DataSXY[k][:,indSXY],axis=1)
            
            #--------------------------------------------------------------------------
            # Summarize by project type, region, and time
            #--------------------------------------------------------------------------
            
            for iPT in range(uPT.size):
                
                for iReg in range(uReg.size):
                
                    ind1=np.where( (ProjectType==uPT[iPT]) & (Region==uReg[iReg]) )[0]
                
                    for k in v2include:
                    
                        MoMeanByPT[k][:,iEns,iPT,iReg]=ScaleFactor*np.mean(MeanByMP[k][:,ind1],axis=1)
                        MoSumByPT[k][:,iEns,iPT,iReg]=ScaleFactor*np.sum(SumByMP[k][:,ind1],axis=1)
                        
                        # Summmarize by PT and year
                        for t in tv_Implementation:
                        
                            ind2=np.where( (ProjectType==uPT[iPT]) & (Region==uReg[iReg]) & (ProjectYear==t) )[0]
                            
                            MoMeanByPTAndYr[k][t][:,iEns,iPT,iReg]=ScaleFactor*np.mean(MeanByMP[k][:,ind2],axis=1)
                            MoSumByPTAndYr[k][t][:,iEns,iPT,iReg]=ScaleFactor*np.sum(SumByMP[k][:,ind2],axis=1)

        #--------------------------------------------------------------------------
        # Save
        #--------------------------------------------------------------------------
        
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeReg_Mean_Scn' + str(iScn+1) + '_FromMPs.pkl',MoMeanByPT)
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeReg_Sum_Scn' + str(iScn+1) + '_FromMPs.pkl',MoSumByPT)
        
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeRegAndYear_Mean_Scn' + str(iScn+1) + '_FromMPs.pkl',MoMeanByPTAndYr)
        gu.opickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeRegAndYear_Sum_Scn' + str(iScn+1) + '_FromMPs.pkl',MoSumByPTAndYr)
    
    t1=time.time()
    print((t1-t0)/60)
    
    return

#%% Import completed multipolygon scenario data

def Import_Scenario_Data_FromPoints(meta):

    #ScaleFactor=1.0
    
    Data=[None]*meta['Project']['N Scenario']
    
    for iS in range(meta['Project']['N Scenario']):
        
        Data[iS]={}
        
        for oper in ['Mean','Sum']:
            
            Data[iS][oper]={}
            
            #------------------------------------------------------------------
            # GHG
            #------------------------------------------------------------------
            
            d=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeAndReg_' + oper + '_Scn' + str(iS+1) + '_FromPoints.pkl')
            
            # RH has an outlier 1 in 500 times
            ikp=np.where(np.mean(d['C_RH_Tot'],axis=0)>0)[0]
            
            for k in d.keys():
                
                #d[k]=d[k].astype(float)/ScaleFactor
                
                Data[iS][oper][k]={}
                Data[iS][oper][k]['Ensemble Mean']=np.mean(d[k][:,ikp],axis=1)
                Data[iS][oper][k]['Ensemble SD']=np.std(d[k][:,ikp],axis=1)
                Data[iS][oper][k]['Ensemble SE']=np.std(d[k][:,ikp],axis=1)/np.sqrt(meta['Project']['N Ensemble'])
                Data[iS][oper][k]['Ensemble P005']=np.percentile(d[k][:,ikp],0.5,axis=1)
                Data[iS][oper][k]['Ensemble P025']=np.percentile(d[k][:,ikp],2.5,axis=1)
                Data[iS][oper][k]['Ensemble P250']=np.percentile(d[k][:,ikp],25,axis=1)
                Data[iS][oper][k]['Ensemble P750']=np.percentile(d[k][:,ikp],75,axis=1)
                Data[iS][oper][k]['Ensemble P975']=np.percentile(d[k][:,ikp],97.5,axis=1)
                Data[iS][oper][k]['Ensemble P995']=np.percentile(d[k][:,ikp],99.5,axis=1)

            #------------------------------------------------------------------
            # Economics
            #------------------------------------------------------------------

            d=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeAndReg_' + oper + '_Scn' + str(iS+1) + '_FromPoints.pkl')
            
            for k in d.keys():
                
                #d[k][:,ikp]=d[k][:,ikp].astype(float)/ScaleFactor
                
                Data[iS][oper][k]={}
                Data[iS][oper][k]['Ensemble Mean']=np.mean(d[k][:,ikp],axis=1)
                Data[iS][oper][k]['Ensemble SD']=np.std(d[k][:,ikp],axis=1)
                Data[iS][oper][k]['Ensemble SE']=np.std(d[k][:,ikp],axis=1)/ikp.size
                Data[iS][oper][k]['Ensemble P005']=np.percentile(d[k][:,ikp],0.5,axis=1)
                Data[iS][oper][k]['Ensemble P025']=np.percentile(d[k][:,ikp],2.5,axis=1)
                Data[iS][oper][k]['Ensemble P250']=np.percentile(d[k][:,ikp],25,axis=1)
                Data[iS][oper][k]['Ensemble P750']=np.percentile(d[k][:,ikp],75,axis=1)
                Data[iS][oper][k]['Ensemble P975']=np.percentile(d[k][:,ikp],97.5,axis=1)
                Data[iS][oper][k]['Ensemble P995']=np.percentile(d[k][:,ikp],99.5,axis=1)
    
    mos={}
    mos['Scenarios']=Data

    return mos

#%% Import areas from points

def Import_Scenario_Areas_FromPoints(meta):
    
    Data=[None]*meta['Project']['N Scenario']
    
    for iS in range(meta['Project']['N Scenario']):
        
        Data[iS]={}
        
        for oper in ['Mean','Sum']:
            
            Data[iS][oper]={}
            
            #------------------------------------------------------------------
            # GHG
            #------------------------------------------------------------------
            
            d=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\Area_ByProjTypeAndReg_' + oper + '_Scn' + str(iS+1) + '_FromPoints.pkl')
            
            for k in d.keys():
                
                Data[iS][oper][k]={}
                Data[iS][oper][k]['Ensemble Mean']=np.mean(d[k],axis=1)
                Data[iS][oper][k]['Ensemble SD']=np.std(d[k],axis=1)

    return Data

#%% Import future scenario comparison

def Import_ScenarioComparisons_FromPoints(meta,mos):
    
    for sc in mos['Delta']:

        mos['Delta'][sc]['ByPT']={}
        
        for oper in ['Mean','Sum']:
            
            dC={}
            
            # GHG
            
            dB=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeAndReg_' + oper + '_Scn' + str(mos['Delta'][sc]['iB']+1) + '_FromPoints.pkl')
            dP=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeAndReg_' + oper + '_Scn' + str(mos['Delta'][sc]['iP']+1) + '_FromPoints.pkl')
            
            for k in dB.keys():
                
                #dB[k]=dB[k].astype(float)/ScaleFactor
                #dP[k]=dP[k].astype(float)/ScaleFactor
                
                dC[k]={}
                dC[k]['Ensemble Mean']=np.mean(dP[k]-dB[k],axis=1)
                dC[k]['Ensemble SD']=np.std(dP[k]-dB[k],axis=1)
                dC[k]['Ensemble SE']=np.std(dP[k]-dB[k],axis=1)/np.sqrt(meta['Project']['N Ensemble'])
                dC[k]['Ensemble P005']=np.percentile(dP[k]-dB[k],0.5,axis=1)
                dC[k]['Ensemble P025']=np.percentile(dP[k]-dB[k],2.5,axis=1)
                dC[k]['Ensemble P250']=np.percentile(dP[k]-dB[k],25,axis=1)
                dC[k]['Ensemble P750']=np.percentile(dP[k]-dB[k],75,axis=1)
                dC[k]['Ensemble P975']=np.percentile(dP[k]-dB[k],97.5,axis=1)                
                dC[k]['Ensemble P995']=np.percentile(dP[k]-dB[k],99.5,axis=1)
            
            # Economics
            
            dB=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeAndReg_' + oper + '_Scn' + str(mos['Delta'][sc]['iB']+1) + '_FromPoints.pkl')
            dP=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeAndReg_' + oper + '_Scn' + str(mos['Delta'][sc]['iP']+1) + '_FromPoints.pkl')
            
            for k in dB.keys():
                
                #dB[k]=dB[k].astype(float)/ScaleFactor
                #dP[k]=dP[k].astype(float)/ScaleFactor
                
                dC[k]={}
                dC[k]['Ensemble Mean']=np.mean(dP[k]-dB[k],axis=1)
                dC[k]['Ensemble SD']=np.std(dP[k]-dB[k],axis=1)
                dC[k]['Ensemble SE']=np.std(dP[k]-dB[k],axis=1)/np.sqrt(meta['Project']['N Ensemble'])
                dC[k]['Ensemble P005']=np.percentile(dP[k]-dB[k],0.5,axis=1)
                dC[k]['Ensemble P025']=np.percentile(dP[k]-dB[k],2.5,axis=1)
                dC[k]['Ensemble P250']=np.percentile(dP[k]-dB[k],25,axis=1)
                dC[k]['Ensemble P750']=np.percentile(dP[k]-dB[k],75,axis=1)
                dC[k]['Ensemble P975']=np.percentile(dP[k]-dB[k],97.5,axis=1)
                dC[k]['Ensemble P995']=np.percentile(dP[k]-dB[k],99.5,axis=1)
            mos['Delta'][sc]['ByPT'][oper]=copy.deepcopy(dC)

    return mos

#%% Import scenario data from MP data

def Import_Scenario_Data_FromMPs(meta,mos):
    
    Data=[None]*meta['Project']['N Scenario']
    
    for iS in range(meta['Project']['N Scenario']):
        
        Data[iS]={}
        
        for oper in ['Mean','Sum']:
            
            Data[iS][oper]={}
            
            #------------------------------------------------------------------
            # GHG
            #------------------------------------------------------------------
            
            d=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeReg_' + oper + '_Scn' + str(iS+1) + '_FromMPs.pkl')
            
            for k in d.keys():
                
                Data[iS][oper][k]={}
                Data[iS][oper][k]['Ensemble Mean']=np.mean(d[k],axis=1)
                Data[iS][oper][k]['Ensemble SD']=np.std(d[k],axis=1)
                Data[iS][oper][k]['Ensemble SE']=np.std(d[k],axis=1)/np.sqrt(meta['Project']['N Ensemble'])
                Data[iS][oper][k]['Ensemble P005']=np.percentile(d[k],0.5,axis=1)
                Data[iS][oper][k]['Ensemble P025']=np.percentile(d[k],2.5,axis=1)
                Data[iS][oper][k]['Ensemble P250']=np.percentile(d[k],25,axis=1)
                Data[iS][oper][k]['Ensemble P750']=np.percentile(d[k],75,axis=1)
                Data[iS][oper][k]['Ensemble P975']=np.percentile(d[k],97.5,axis=1)
                Data[iS][oper][k]['Ensemble P995']=np.percentile(d[k],99.5,axis=1)

            #------------------------------------------------------------------
            # Economics
            #------------------------------------------------------------------

            d=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeReg_' + oper + '_Scn' + str(iS+1) + '_FromMPs.pkl')
            
            for k in d.keys():
                
                Data[iS][oper][k]={}
                Data[iS][oper][k]['Ensemble Mean']=np.mean(d[k],axis=1)
                Data[iS][oper][k]['Ensemble SD']=np.std(d[k],axis=1)
                Data[iS][oper][k]['Ensemble SE']=np.std(d[k],axis=1)/np.sqrt(meta['Project']['N Ensemble'])
                Data[iS][oper][k]['Ensemble P005']=np.percentile(d[k],0.5,axis=1)
                Data[iS][oper][k]['Ensemble P025']=np.percentile(d[k],2.5,axis=1)
                Data[iS][oper][k]['Ensemble P250']=np.percentile(d[k],25,axis=1)
                Data[iS][oper][k]['Ensemble P750']=np.percentile(d[k],75,axis=1)
                Data[iS][oper][k]['Ensemble P975']=np.percentile(d[k],97.5,axis=1)
                Data[iS][oper][k]['Ensemble P995']=np.percentile(d[k],99.5,axis=1)

    # Add to MOS structure
    mos['Scenarios']=Data

    return mos

#%% Import scenario comparisons from MP data

def Import_ScenarioComparisons_FromMPs(meta,mos):
    
    for sc in mos['Delta']:

        mos['Delta'][sc]['ByPT']={}
        mos['Delta'][sc]['ByPTAndYear']={}
        
        for oper in ['Mean','Sum']:
            
            #------------------------------------------------------------------
            # By projet type and region
            #------------------------------------------------------------------
            
            dC={}
            
            # GHG 
            
            dB=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeReg_' + oper + '_Scn' + str(mos['Delta'][sc]['iB']+1) + '_FromMPs.pkl')
            dP=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeReg_' + oper + '_Scn' + str(mos['Delta'][sc]['iP']+1) + '_FromMPs.pkl')
            
            for k in dB.keys():
                
                dC[k]={}
                dC[k]['Ensemble Mean']=np.mean(dP[k]-dB[k],axis=1)
                dC[k]['Ensemble SD']=np.std(dP[k]-dB[k],axis=1)
                dC[k]['Ensemble SE']=np.std(dP[k]-dB[k],axis=1)/np.sqrt(meta['Project']['N Ensemble'])
                dC[k]['Ensemble P005']=np.percentile(dP[k]-dB[k],0.5,axis=1)
                dC[k]['Ensemble P025']=np.percentile(dP[k]-dB[k],2.5,axis=1)
                dC[k]['Ensemble P250']=np.percentile(dP[k]-dB[k],25,axis=1)
                dC[k]['Ensemble P750']=np.percentile(dP[k]-dB[k],75,axis=1)
                dC[k]['Ensemble P975']=np.percentile(dP[k]-dB[k],97.5,axis=1)
                dC[k]['Ensemble P995']=np.percentile(dP[k]-dB[k],99.5,axis=1)
            
            # Economics
            
            dB=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeReg_' + oper + '_Scn' + str(mos['Delta'][sc]['iB']+1) + '_FromMPs.pkl')
            dP=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeReg_' + oper + '_Scn' + str(mos['Delta'][sc]['iP']+1) + '_FromMPs.pkl')
            
            for k in dB.keys():
                
                dC[k]={}
                dC[k]['Ensemble Mean']=np.mean(dP[k]-dB[k],axis=1)
                dC[k]['Ensemble SD']=np.std(dP[k]-dB[k],axis=1)
                dC[k]['Ensemble SE']=np.std(dP[k]-dB[k],axis=1)/np.sqrt(meta['Project']['N Ensemble'])
                dC[k]['Ensemble P005']=np.percentile(dP[k]-dB[k],0.5,axis=1)
                dC[k]['Ensemble P025']=np.percentile(dP[k]-dB[k],2.5,axis=1)
                dC[k]['Ensemble P250']=np.percentile(dP[k]-dB[k],25,axis=1)
                dC[k]['Ensemble P750']=np.percentile(dP[k]-dB[k],75,axis=1)
                dC[k]['Ensemble P975']=np.percentile(dP[k]-dB[k],97.5,axis=1)
                dC[k]['Ensemble P995']=np.percentile(dP[k]-dB[k],99.5,axis=1)
            
            mos['Delta'][sc]['ByPT'][oper]=dC
            
            #------------------------------------------------------------------
            # By project type, region and year
            #------------------------------------------------------------------
            
            dC={}
            
            # GHG
            
            dB=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeRegAndYear_' + oper + '_Scn' + str(mos['Delta'][sc]['iB']+1) + '_FromMPs.pkl')
            dP=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\GHGB_ByProjTypeRegAndYear_' + oper + '_Scn' + str(mos['Delta'][sc]['iP']+1) + '_FromMPs.pkl')
            
            for k in dB.keys():
                
                dC[k]={}
                for t in dB[k].keys():
                    
                    dC[k][t]={}
                    dC[k][t]['Ensemble Mean']=np.mean(dP[k][t]-dB[k][t],axis=1)
                    dC[k][t]['Ensemble SD']=np.std(dP[k][t]-dB[k][t],axis=1)
                    dC[k][t]['Ensemble SE']=np.std(dP[k][t]-dB[k][t],axis=1)/np.sqrt(meta['Project']['N Ensemble'])
                    dC[k][t]['Ensemble P005']=np.percentile(dP[k][t]-dB[k][t],0.5,axis=1)
                    dC[k][t]['Ensemble P025']=np.percentile(dP[k][t]-dB[k][t],2.5,axis=1)
                    dC[k][t]['Ensemble P250']=np.percentile(dP[k][t]-dB[k][t],25,axis=1)
                    dC[k][t]['Ensemble P750']=np.percentile(dP[k][t]-dB[k][t],75,axis=1)
                    dC[k][t]['Ensemble P975']=np.percentile(dP[k][t]-dB[k][t],97.5,axis=1)
                    dC[k][t]['Ensemble P995']=np.percentile(dP[k][t]-dB[k][t],99.5,axis=1)
            
            # Economics
            
            dB=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeRegAndYear_' + oper + '_Scn' + str(mos['Delta'][sc]['iB']+1) + '_FromMPs.pkl')
            dP=gu.ipickle(meta['Paths']['Project'] + '\\Outputs\\Econ_ByProjTypeRegAndYear_' + oper + '_Scn' + str(mos['Delta'][sc]['iP']+1) + '_FromMPs.pkl')

            for k in dB.keys():
                dC[k]={}
                for t in dB[k].keys():
                    
                    #dB[k][t]=dB[k][t].astype(float)/ScaleFactor
                    #dP[k][t]=dP[k][t].astype(float)/ScaleFactor
                    
                    dC[k][t]={}
                    dC[k][t]['Ensemble Mean']=np.mean(dP[k][t]-dB[k][t],axis=1)
                    dC[k][t]['Ensemble SD']=np.std(dP[k][t]-dB[k][t],axis=1)
                    dC[k][t]['Ensemble SE']=np.std(dP[k][t]-dB[k][t],axis=1)/np.sqrt(meta['Project']['N Ensemble'])
                    dC[k][t]['Ensemble P005']=np.percentile(dP[k][t]-dB[k][t],0.5,axis=1)
                    dC[k][t]['Ensemble P025']=np.percentile(dP[k][t]-dB[k][t],2.5,axis=1)
                    dC[k][t]['Ensemble P250']=np.percentile(dP[k][t]-dB[k][t],25,axis=1)
                    dC[k][t]['Ensemble P750']=np.percentile(dP[k][t]-dB[k][t],75,axis=1)
                    dC[k][t]['Ensemble P975']=np.percentile(dP[k][t]-dB[k][t],97.5,axis=1)
                    dC[k][t]['Ensemble P995']=np.percentile(dP[k][t]-dB[k][t],99.5,axis=1)
            
            mos['Delta'][sc]['ByPTAndYear'][oper]=dC

    return mos

#%% Save MOS GHG output variables by multipolygon subset

#def MosByMPSubset_GHGB(meta,ListSubsetMP):
#
#    t0=time.time()
#    
#    # Error multiplier
#    sigma_multiplier=1.0
#    
#    # Import multipolygons
#    atu_multipolygons=gu.ipickle(meta['Paths']['Geospatial'] + '\\atu_multipolygons.pkl')
#
#    # Import sxy
#    sxy=gu.ipickle(meta['Paths']['Geospatial'] + '\\sxy.pkl')
#
#    # Create listed index (faster than indexing on the fly)
#    Crosswalk_sxy_to_mp=[None]*len(ListSubsetMP)
#    for iMP in range(len(ListSubsetMP)):
#        d={}
#        d['Index']=np.where(sxy['ID_atu_multipolygons']==ListSubsetMP[iMP])[0]
#        Crosswalk_sxy_to_mp[iMP]=d
#    
#    # Area
#    Area=np.zeros( len(ListSubsetMP) )
#    for iMP in range(len(ListSubsetMP)):                
#        Area0=atu_multipolygons[ListSubsetMP[iMP]]['ACTUAL_TREATMENT_AREA']
#        if Area0==None:
#            print('Encounterd no area, using zero')
#            Area0=0
#        Area[iMP]=Area0
#
#    # Time series of saved results
#    tv_saving=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)
#    
#    # Operations
#    oper=['Mean','Sum']
#    
#    # GHG balance variables
#    d0=LoadSingleOutputFile(meta,0,0,0)
#
#    # All (excluding 'C_M_ByAgent' and 'Year')
#    v2include=['A', 'V_Merch', 'V_MerchToMill', 'LogSizeEnhancement', 'C_Biomass_Tot', 'C_Piled_Tot', 'C_Litter_Tot', 'C_DeadWood_Tot', 'C_Soil_Tot', 
#               'C_InUse_Tot', 'C_DumpLandfill_Tot', 'C_M_Dist', 'C_G_Gross_Tot', 'C_G_Net_Tot', 'C_M_Reg_Tot', 'C_LF_Tot', 'C_RH_Tot', 'C_ToMillMerch', 
#               'C_ToMillNonMerch', 'C_ToMillSnagStem', 'C_ToSlashpileBurn', 'C_ToLumber', 'C_ToPlywood', 'C_ToOSB', 'C_ToMDF', 'C_ToPaper', 'C_ToPowerFacilityDom', 
#               'C_ToPowerFacilityFor', 'C_ToPowerGrid', 'C_ToPellets', 'C_ToFirewoodDom', 'C_ToFirewoodFor', 'C_ToLogExport', 'E_CO2e_LULUCF_NEE', 'E_CO2e_LULUCF_Wildfire', 
#               'E_CO2e_LULUCF_OpenBurning', 'E_CO2e_LULUCF_EcoOther', 'E_CO2e_LULUCF_HWP', 'E_CO2e_ESC_Comb', 'E_CO2e_ESC_SubE', 'E_CO2e_ESC_SubBM', 'E_CO2e_ET_Comb', 
#               'E_CO2e_IPPU_Comb', 'C_NPP_Tot', 'C_ToMill', 'E_CO2e_LULUCF_Fire', 'E_CO2e_AGHGB_WSub', 'E_CO2e_AGHGB_WOSub', 'E_CO2e_AGHGB_WSub_cumu', 'E_CO2e_AGHGB_WOSub_cumu', 
#               'E_CO2e_AGHGB_WSub_cumu_from_tref', 'E_CO2e_AGHGB_WOSub_cumu_from_tref']
#
#    #--------------------------------------------------------------------------
#    # Initialize data by multipolygon structure    
#    #--------------------------------------------------------------------------
#    
#    mos=[None]*meta['Project']['N Scenario']
#    
#    for iScn in range(meta['Project']['N Scenario']):
#        
#        d={}
#        for op in oper:
#            d[op]={}
#            
#        for k in d0.keys():
#            
#            if np.isin(k,v2include)==False:
#                continue
#            
#            for op in oper:
#                d[op][k]={}
#                d[op][k]['Ensemble Mean']=np.zeros((tv_saving.size,len(ListSubsetMP)))
#                d[op][k]['Ensemble CIL']=np.zeros((tv_saving.size,len(ListSubsetMP)))
#                d[op][k]['Ensemble CIH']=np.zeros((tv_saving.size,len(ListSubsetMP)))
#        
#        mos[iScn]=d
#    
#    #--------------------------------------------------------------------------
#    # Loop through scenarios
#    #--------------------------------------------------------------------------
#    
#    for iScn in range(meta['Project']['N Scenario']):
#        
#        # Initialize data matrix
#        Data={}
#        for op in oper:
#            Data[op]=np.zeros( (tv_saving.size,meta['Project']['N Ensemble'],len(ListSubsetMP),len(v2include)) ,dtype=np.float)
#        
#        #----------------------------------------------------------------------
#        # Loop through ensembles
#        #----------------------------------------------------------------------
#        
#        for iEns in range(meta['Project']['N Ensemble']):
#            print(iEns)
#            
#            #------------------------------------------------------------------
#            # Import batches
#            #------------------------------------------------------------------
#            
#            # Initialize temporary data structure for full simulation            
#            DataSXY=np.zeros( (tv_saving.size,meta['Project']['N Stand'],len(v2include)) ,dtype=np.float)
#            
#            # Import batches
#            for iBat in range(meta['Project']['N Batch']):
#        
#                indBat=IndexToBatch(meta,iBat)
#            
#                d1=LoadSingleOutputFile(meta,iScn,iEns,iBat)
#                
#                cnt_k=0
#                for k in d0.keys():
#                    
#                    if np.isin(k,v2include)==False:
#                        continue
#                    
#                    DataSXY[:,indBat,cnt_k]=d1[k].copy()
#                    
#                    cnt_k=cnt_k+1
#            
#            del d1
#            #garc.collect()
#            
#            #------------------------------------------------------------------
#            # Calculate mean for multipolygon subset
#            #------------------------------------------------------------------
#            
#            for iMP in range(len(ListSubsetMP)):
#                
#                indMP=Crosswalk_sxy_to_mp[iMP]['Index']
#                
#                mu=np.mean(DataSXY[:,indMP,:],axis=1)
#                
#                for op in oper:
#                    
#                    if op=='Mean':
#                        Data[op][:,iEns,iMP,:]=mu
#                    else:
#                        Data[op][:,iEns,iMP,:]=np.sum(Area[iMP])*mu
#           
#        #----------------------------------------------------------------------
#        # Add stats to MOS structure
#        #----------------------------------------------------------------------
#        
#        cnt_k=0    
#        for k in d0.keys():
#
#            if np.isin(k,v2include)==False:
#                continue
#                
#            for op in oper:
#            
#                mu=np.mean(Data[op][:,:,:,cnt_k],axis=1)
#                sd=np.std(Data[op][:,:,:,cnt_k],axis=1)
#                cil=mu-sigma_multiplier*sd/np.sqrt(meta['Project']['N Ensemble'])
#                cih=mu+sigma_multiplier*sd/np.sqrt(meta['Project']['N Ensemble'])
#            
#                mos[iScn][op][k]['Ensemble Mean']=mu
#                mos[iScn][op][k]['Ensemble CIL']=cil
#                mos[iScn][op][k]['Ensemble CIH']=cih
#            
#            cnt_k=cnt_k+1
#
#    #--------------------------------------------------------------------------
#    # Save
#    #--------------------------------------------------------------------------
#    
#    gu.opickle(meta['Paths']['Project'] + '\\Outputs\\MosByMPSubset_GHGB.pkl',mos)
#    
#    t1=time.time()
#    print((t1-t0)/60)
#    
#    return


#%% GET TASS GROWTH CURVES

def GetTASSCurves(meta,iScn,iGC,fny,fnc,fnm):
    
    # Assume just one stand per scenario (i.e., demo)
    iS=0
    
    dfY=pd.read_csv(fny,header=30)
    
    # Initialize age response of net growth
    G=np.zeros((N_Age,1,6),dtype=np.int16)
    
    G[:,iS,0]=G_StemMerch[:,indTIPSY[0]]/meta['GC']['Scale Factor']
    G[:,iS,1]=G_StemNonMerch[:,indTIPSY[0]]/meta['GC']['Scale Factor']
    G[:,iS,2]=G_Bark[:,indTIPSY[0]]/meta['GC']['Scale Factor']
    G[:,iS,3]=G_Branch[:,indTIPSY[0]]/meta['GC']['Scale Factor']
    G[:,iS,4]=G_Foliage[:,indTIPSY[0]]/meta['GC']['Scale Factor']
    G[:,iS,5]=G_VStemMerch[:,indTIPSY[0]]/meta['GC']['Scale Factor']
                    
    # Save data to file in input variables folder of project
    gu.opickle(meta['Paths']['Project'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve' + str(iGC+1) + '_Bat' + FixFileNum(1) + '.pkl',G)
    
    return    

#%% POST-PROCESS TIPSY GROWTH CURVES
# Nested list, gc[Scenario][Stand][Growth Curve]

def Import_BatchTIPSY_Output(meta):
        
    # Growth curve parameters and TIPSY outputs
    dfPar=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx',sheet_name='Sheet1',skiprows=6)
    txtDat=np.loadtxt(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Output.out',skiprows=4)
    
    # TIPSY saves to text file -> convert to dataframe (column names must match TIPSY output file design)
    dfDat=pd.DataFrame(txtDat,columns=meta['GC']['BatchTIPSY Column Names'])

    # Define age vector (must be consistent with how TIPSY was set up)
    Age=np.arange(0,meta['GC']['BatchTIPSY Maximum Age']+1,1)

    # Get dimensions of the TIPSY output file to reshape the data into Age x Stand
    N_Age=Age.size
    N_GC=int(dfDat.shape[0]/N_Age)

    gc=[None]*N_GC
    for i in range(N_GC): 
        gc[i]={}    
    for i in range(len(meta['GC']['BatchTIPSY Column Names'])):
        data=np.reshape(dfDat[meta['GC']['BatchTIPSY Column Names'][i]].values,(N_Age,N_GC),order='F')
        for j in range(N_GC):
            gc[j][meta['GC']['BatchTIPSY Column Names'][i]]=data[:,j]

    gc2=[]
    uScn=np.unique(dfPar['ID_Scenario'])
    for iScn in range(uScn.size):
        ind=np.where(dfPar['ID_Scenario']==uScn[iScn])[0]
        uStand=np.unique(dfPar.loc[ind,'ID_Stand'])
        gc1=[]
        for iS in range(uStand.size):
            ind=np.where( (dfPar['ID_Scenario']==uScn[iScn]) & (dfPar['ID_Stand']==uStand[iS]) )[0]
            uGC=np.unique(dfPar.loc[ind,'ID_GC'])
            gc0=[]
            for iGC in range(uGC.size):
                ind=np.where( (dfPar['ID_Scenario']==uScn[iScn]) & (dfPar['ID_Stand']==uStand[iS]) & (dfPar['ID_GC']==uGC[iGC]) )[0]
                d={}
                for i in range(len(meta['GC']['BatchTIPSY Column Names'])):
                    data=np.reshape(dfDat[meta['GC']['BatchTIPSY Column Names'][i]].values,(N_Age,N_GC),order='F')
                    d[meta['GC']['BatchTIPSY Column Names'][i]]=data[:,ind]
                gc0.append(d)
            gc1.append(gc0)
        gc2.append(gc1)
    return gc2

#%% Import growth curves

def Import_CompiledGrowthCurves(meta,scn):

    gc=[]
    for iScn in range(len(scn)): #range(meta['Project']['N Scenario']):
        gc0=[]
        
        gc1=[]
        for iBat in range(0,meta['Project']['N Batch']):            
            tmp=gu.ipickle(meta['Paths']['Project'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve1_Bat' + FixFileNum(iBat) + '.pkl')
            tmp=tmp[:,:,0].astype(float)
            for iS in range(tmp.shape[1]):
                gc1.append(tmp[:,iS].copy()*meta['GC']['Scale Factor'])
        gc0.append(gc1.copy())
        
        gc1=[]
        for iBat in range(0,meta['Project']['N Batch']):            
            tmp=gu.ipickle(meta['Paths']['Project'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve2_Bat' + FixFileNum(iBat) + '.pkl')
            tmp=tmp[:,:,0].astype(float)
            for iS in range(tmp.shape[1]):
                gc1.append(tmp[:,iS].copy()*meta['GC']['Scale Factor'])
        gc0.append(gc1.copy())
        
        gc1=[]
        for iBat in range(0,meta['Project']['N Batch']):            
            tmp=gu.ipickle(meta['Paths']['Project'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve3_Bat' + FixFileNum(iBat) + '.pkl')
            tmp=tmp[:,:,0].astype(float)
            for iS in range(tmp.shape[1]):
                gc1.append(tmp[:,iS].copy()*meta['GC']['Scale Factor'])
        gc0.append(gc1.copy())
        gc.append(gc0.copy())
    return gc

#%% WRITE SPREADSHEET OF BatchTIPSY PARAMTERS

def Write_BatchTIPSY_Input_Spreadsheet(meta,ugc):
    
    # Create a function that will return the column corresponding to a variable name
    fin=meta['Paths']['Model Code'] + '\\Parameters\\GrowthCurvesTIPSY_Parameters_Template.xlsx'
    df_frmt=pd.read_excel(fin,sheet_name='Sheet1')
    gy_labels=df_frmt.loc[5,:].values
    def GetColumn(lab):
        ind=np.where(gy_labels==lab)[0]
        return int(ind+1)

    # Open spreadsheet
    PathGrowthCurveParameters=meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx'
    xfile=openpyxl.load_workbook(PathGrowthCurveParameters)
    sheet=xfile.get_sheet_by_name('Sheet1')    
    N_headers=7

    # Overwrite existing data entries with empty cells
    # *** This is really important - failing to wipe it clean first will lead to 
    # weird parameters ***
    for i in range(int(1.5*ugc['Unique'].shape[0])):
        for j in range(len(gy_labels)):
            sheet.cell(row=i+1+N_headers,column=j+1).value=''

    # Initialize counter
    cnt=1

    # Loop through unique stand types
    for iUGC in range(ugc['Unique'].shape[0]):
    
        # It isn't necessary to populate these, as we will use the crosswalk in 
        # Python session instead            
        sheet.cell(row=cnt+N_headers,column=1).value=cnt
        sheet.cell(row=cnt+N_headers,column=2).value=0
        sheet.cell(row=cnt+N_headers,column=3).value=0
        sheet.cell(row=cnt+N_headers,column=4).value=0 
    
        # Regeneration type (N, C, P)   
        vnam='regeneration_method'
        id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        cd=lut_n2s(meta['LUT']['TIPSY'][vnam],id)[0]
        sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd      
    
        # Species 1
        vnam='s1'
        id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        cd=lut_n2s(meta['LUT']['VRI']['SPECIES_CD_1'],id)[0]
        sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd
    
        vnam='p1'
        dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
    
        vnam='i1'
        dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
    
        vnam='gain1'
        dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        if dat!=-999:
            sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
            sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(10)
    
        # Species 2
        vnam='s2'
        id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        if id!=-999:
            cd=lut_n2s(meta['LUT']['VRI']['SPECIES_CD_1'],id)[0]
            sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd
        
            vnam='p2'
            dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]        
            sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
        
            vnam='gain2'
            dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
            if dat!=-999:
                sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
                sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(10)
    
        # Species 3
        vnam='s3'
        id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        if id!=-999:
            cd=lut_n2s(meta['LUT']['VRI']['SPECIES_CD_1'],id)[0]
            sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd
        
            vnam='p3'
            dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]        
            sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
        
            vnam='gain3'
            dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
            if dat!=-999:
                sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
                sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(10)
                
        # Species 4
        vnam='s4'
        id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        if id!=-999:
            cd=lut_n2s(meta['LUT']['VRI']['SPECIES_CD_1'],id)[0]
            sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd
        
            vnam='p4'
            dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]        
            sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
        
            vnam='gain4'
            dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
            if dat!=-999:
                sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
                sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(10)
            
        # Species 5
        vnam='s5'
        id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        if id!=-999:
            cd=lut_n2s(meta['LUT']['VRI']['SPECIES_CD_1'],id)[0]
            sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd
        
            vnam='p5'
            dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]        
            sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
        
            vnam='gain5'
            dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
            if dat!=-999:
                sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
                sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(10)
    
        # Planting density
        vnam='init_density'
        dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat) 
    
        # Regeneration delay
        vnam='regen_delay'
        dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat)
    
        # OAF1 
        vnam='oaf1'
        dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=dat[0]
    
        # OAF2
        vnam='oaf2'
        dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=dat[0]
    
        # BEC zone
        vnam='bec_zone'
        id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        cd=lut_n2s(meta['LUT']['VRI']['BEC_ZONE_CODE'],id)[0]    
        sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd
    
        # FIZ
        vnam='FIZ'
        id=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        cd=lut_n2s(meta['LUT']['TIPSY']['FIZ'],id)[0]
        sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=cd
    
        # Age at fertilization
        # vnam='fert_age1'
        #dat=ugc['Unique'][iUGC,np.where(ugc['GC_Variable_List']==vnam)[0]]
        #if dat!=-999:
        #    sheet.cell(row=cnt+N_headers,column=GetColumn(vnam)).value=int(dat[0])
    
        # Update counter
        cnt=cnt+1
    
    #------------------------------------------------------------------------------
    # Save to spreadsheet
    #------------------------------------------------------------------------------
    
    xfile.save(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx')
    
#%% Write BatchTIPSY input file
# Notes:
#    if the input spreadsheet has nan's, all data will be converted to float

def Write_BatchTIPSY_Input_File(meta):
    
    # Import format info (number of designated positions per variable)
    fin=meta['Paths']['Model Code'] + '\\Parameters\\GrowthCurvesTIPSY_Parameters_Template.xlsx'
    df_frmt=pd.read_excel(fin,sheet_name='Sheet1')

    # Format array
    nfrmt=df_frmt.iloc[1,4:53]

    # Import input data   
    df=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx',sheet_name='Sheet1',skiprows=6)

    # Convert to dictionary
    Data=df.to_dict('list')
    for k in Data.keys():
        Data[k]=np.array(Data[k])
    
    varnams=list(df.columns[4:53])
    
    # Change to list
    DataList=[None]*Data['ID'].size
    for i in range(len(DataList)):
        d={}
        for k in Data.keys():
            d[k]=Data[k][i]
        DataList[i]=d      
        
    # Create a spacer
    Spacer=" "
    
    # Open file
    fid=open(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_InputVariables.dat','w')    
    
    # Loop through rows of table
    for iStand in range(len(DataList)):
        
        LineString=""
        
        for iV in range(len(varnams)):
              
            s0=str(DataList[iStand][varnams[iV]])
    
            # Make sure NANs are blank for TIPSY
            if s0=='nan': s0=''
            if s0=='NA': s0=''
            if s0=='NaN': s0=''
            if s0=='': s0=''
    
            n=nfrmt[iV]
            if len(s0)==n: 
                s1=s0
            else: 
                s1=s0
                d=n-len(s0)
                for k in range(d): 
                    s1=s1 + Spacer
                
            # Add individual column variables
            LineString=LineString + s1 + Spacer
                
        # Add end of line
        LineString=LineString + '\n'

        # Save    
        fid.write(LineString)
    
    # Close
    fid.close() 

#%% IMPORT DISTURBANCE HISTORY

def GetDisturbanceHistory(meta):
    dh=[]
    for iScn in range(meta['Project']['N Scenario']):
        dhB=[]    
        for iBat in range(meta['Project']['N Batch']):
            iEns=0
            dh0=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
            if iBat==0:
                dhB=dh0
            else:
                for i in range(len(dh0)):
                    dhB.append(dh0[i])
            dh.append(dhB)
    return dh

#%% GRAPHICS

# Look at variables in rcParams:
#plt.rcParams.keys()

def Import_GraphicsParameters(x):    
    
    params={}    
    
    if x=='FCI_Demo':
        
        fs1=6
        fs2=7
        
        params={'font.sans-serif':'Arial',
                'font.size':fs1,
                'figure.titlesize':fs2,
                'figure.dpi':150,
                'figure.constrained_layout.use':True,
                'axes.edgecolor':'black',
                'axes.labelsize':fs1,
                'axes.labelcolor':'black',
                'axes.titlesize':fs2,
                'axes.titlepad':2,
                'axes.linewidth':0.5,        
                'lines.linewidth':1,
                'text.color':'black',
                'xtick.color':'black',        
                'xtick.labelsize':fs1,
                'xtick.major.width':0.5,
                'xtick.major.size':3,
                'xtick.direction':'in',
                'ytick.color':'black',
                'ytick.labelsize':fs1,
                'ytick.major.width':0.5,
                'ytick.major.size':3,
                'ytick.direction':'in',
                'legend.fontsize':fs1,        
                'savefig.dpi':900,
                'savefig.transparent':True,
                'savefig.format':'png',
                'savefig.pad_inches':0.1,
                'savefig.bbox':'tight'}
    
    elif x=='bc1ha_1':
        
        params={'font.sans-serif':'Arial',
                'font.size':7,
                'axes.labelsize':7,
                'axes.titlesize':14,
                'axes.linewidth':0.5,        
                'xtick.labelsize':7,
                'xtick.major.width':0.5,
                'xtick.major.size':5,
                'xtick.direction':'in',
                'ytick.labelsize':7,
                'ytick.major.width':0.5,
                'ytick.major.size':5,
                'ytick.direction':'in',
                'legend.fontsize':10,
                'savefig.dpi':150}
    
    elif (x=='Article') | (x=='article'):
        
        fs=6
        
        params={'font.sans-serif':'Arial','font.size':fs,'axes.edgecolor':'black','axes.labelsize':fs,'axes.labelcolor':'black',
                'axes.titlesize':fs,'axes.linewidth':0.5,'lines.linewidth':0.5,'text.color':'black','xtick.color':'black',        
                'xtick.labelsize':fs,'xtick.major.width':0.5,'xtick.major.size':3,'xtick.direction':'in','ytick.color':'black',
                'ytick.labelsize':fs,'ytick.major.width':0.5,'ytick.major.size':3,'ytick.direction':'in','legend.fontsize':fs}
    
    else:        
        
        params={}
    
    return params

#%% Prepare inventory from spreadsheet

def PrepareInventoryFromSpreadsheet(meta):
    
    for iScn in range(0,meta['Project']['N Scenario']):
    
        # Loop through batches, saving inventory to file
        for iBat in range(0,meta['Project']['N Batch']):
      
            inv={}
    
            # Index to batch
            indBat=IndexToBatch(meta,iBat)    
            N_StandsInBatch=len(indBat)
    
            # Initialize inventory variables
            inv['Lat']=np.zeros((1,N_StandsInBatch))
            inv['Lon']=np.zeros((1,N_StandsInBatch))
            inv['X']=inv['Lat']
            inv['Y']=inv['Lon']
        
            # BEC zone
            inv['ID_BECZ']=np.zeros((1,N_StandsInBatch),dtype=np.int)
            inv['ID_BECZ'][0,:]=meta['LUT']['VRI']['BEC_ZONE_CODE'][meta['Scenario'][iScn]['BGC Zone Code']]
    
            # Timber harvesting landbase (1=yes, 0=no)
            inv['THLB']=meta['Scenario'][iScn]['THLB Status']*np.ones((meta['Year'].size,N_StandsInBatch))
        
            # Temperature will be updated automatically
            inv['MAT']=4*np.ones((1,N_StandsInBatch))
            
            if meta['Project']['Biomass Module']=='Sawtooth':
                
                ind=np.where( meta['Param']['BE']['Sawtooth']['SRS Key']['SRS_CD']==meta['Scenario'][iScn]['SRS1_CD'])[0]
                
                id=meta['Param']['BE']['Sawtooth']['SRS Key']['SRS_ID'][ind]
                
                inv['SRS1_ID']=id*np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['SRS1_PCT']=100*np.ones((1,N_StandsInBatch),dtype=np.int)
                
                inv['SRS2_ID']=np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['SRS2_PCT']=0*np.ones((1,N_StandsInBatch),dtype=np.int)
                
                inv['SRS3_ID']=np.ones((1,N_StandsInBatch),dtype=np.int)
                inv['SRS3_PCT']=0*np.ones((1,N_StandsInBatch),dtype=np.int)

            # Save
            gu.opickle(meta['Paths']['Input Scenario'][iScn] + '\\Inventory_Bat' + FixFileNum(iBat) + '.pkl',inv)
        
    return

#%% Compile events
    
def CompileEvents(ec,tv,iS,ID_Type,Year,MortalityFactor,GrowthFactor,ID_GrowthCurve):
    
    if Year.size>0:
        YearFloor=np.floor(Year)
        uYearFloor=np.unique(YearFloor)
        for iU in range(uYearFloor.size):
            iT=np.where(tv==uYearFloor[iU])[0]
            if iT.size==0:
                continue
            indYear=np.where(YearFloor==uYearFloor[iU])[0]                        
            for iY in range(indYear.size):
                indAvailable=np.where(ec['ID_Type'][iT,iS,:].flatten()==0)[0]
                if indAvailable.size==0:
                    print('Warning, more events per year than can be handled!')
                else:
                    iE=indAvailable[0]
                    ec['ID_Type'][iT,iS,iE]=ID_Type[indYear[iY]]
                    ec['MortalityFactor'][iT,iS,iE]=MortalityFactor[indYear[iY]]
                    ec['GrowthFactor'][iT,iS,iE]=GrowthFactor[indYear[iY]]
                    ec['ID_GrowthCurve'][iT,iS,iE]=ID_GrowthCurve[indYear[iY]]
    return ec

#%% Mortality frequency distribution

def GetMortalityFrequencyDistribution(meta):
    
    iEns=0
    
    M=[None]*meta['Project']['N Scenario']    
    for iScn in range(meta['Project']['N Scenario']):
        
        tv=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)
        M[iScn]={}
        M[iScn]['Ma']={}
        M[iScn]['Mr']={}
        for k in meta['LUT']['Dist'].keys():
            M[iScn]['Ma'][k]=np.zeros((tv.size,meta['Project']['N Stand']))
            M[iScn]['Mr'][k]=np.zeros((tv.size,meta['Project']['N Stand']))
        M[iScn]['Ma']['Reg']=np.zeros((tv.size,meta['Project']['N Stand']))
        M[iScn]['Mr']['Reg']=np.zeros((tv.size,meta['Project']['N Stand']))
        
        for iBat in range(meta['Project']['N Batch']): 
            
            indBat=IndexToBatch(meta,iBat)
            
            d1=LoadSingleOutputFile(meta,iScn,iEns,iBat)        
            dh=gu.ipickle(meta['Paths']['Input Scenario'][iScn] + '\\Events_Ens' + FixFileNum(iEns) + '_Bat' + FixFileNum(iBat) + '.pkl')
            
            for iStandInBat in range(len(indBat)):
                iStand=indBat[iStandInBat]
                for iYr in range(dh[iStandInBat]['Year'].size):
                    it=np.where(tv==int(dh[iStandInBat]['Year'][iYr]))[0]
                    if it.size==0:
                        continue
                    nam=lut_n2s(meta['LUT']['Dist'],dh[iStandInBat]['ID_Type'][iYr])[0]                     
                    M[iScn]['Ma'][nam][it,iStand]=d1[0]['C_M_Dist'][it,iStandInBat]
                    M[iScn]['Mr'][nam][it,iStand]=d1[0]['C_M_Dist'][it,iStandInBat]/np.maximum(0.0001,d1[0]['Eco_Biomass'][it-1,iStandInBat])
                    M[iScn]['Ma']['Reg'][it,iStand]=d1[0]['C_M_Reg'][it,iStandInBat]
                    M[iScn]['Mr']['Reg'][it,iStand]=d1[0]['C_M_Reg'][it,iStandInBat]/np.maximum(0.0001,d1[0]['Eco_Biomass'][it-1,iStandInBat])
            del d1,dh
            garc.collect()
    
    return M

#%% Summarize affected area due to natural disturbance and management

def SummarizeAreaAffected(meta,mos,tv,iScn,iPT,iRE,AEF,ivlT):
    
    A={}    
    A['Nat Dist']=[None]*10; c=-1
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Wildfire'; A['Nat Dist'][c]['Color']=[0.75,0,0]; A['Nat Dist'][c]['Data']=mos['Areas'][iScn]['Sum']['Wildfire']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Mountain Pine beetle'; A['Nat Dist'][c]['Color']=[0,0.8,0]; A['Nat Dist'][c]['Data']=mos['Areas'][iScn]['Sum']['IBM']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Douglas-fir beetle'; A['Nat Dist'][c]['Color']=[0.6,1,0]; A['Nat Dist'][c]['Data']=mos['Areas'][iScn]['Sum']['IBD']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Spruce beetle'; A['Nat Dist'][c]['Color']=[0.25,1,1]; A['Nat Dist'][c]['Data']=mos['Areas'][iScn]['Sum']['IBS']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='W. balsam beetle'; A['Nat Dist'][c]['Color']=[0,0.45,0]; A['Nat Dist'][c]['Data']=mos['Areas'][iScn]['Sum']['IBB']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Other pests'; A['Nat Dist'][c]['Color']=[0.8,1,0]; A['Nat Dist'][c]['Data']=mos['Areas'][iScn]['Sum']['Beetles']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='W. spruce budworm'; A['Nat Dist'][c]['Color']=[0,0.75,1]; A['Nat Dist'][c]['Data']=mos['Areas'][iScn]['Sum']['IDW']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Rust'; A['Nat Dist'][c]['Color']=[0.75,0.5,1]; A['Nat Dist'][c]['Data']=mos['Areas'][iScn]['Sum']['Rust Onset']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Dwarf Mistletoe'; A['Nat Dist'][c]['Color']=[1,0.5,0.25]; A['Nat Dist'][c]['Data']=mos['Areas'][iScn]['Sum']['Dwarf Mistletoe Onset']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Mechanical'; A['Nat Dist'][c]['Color']=[0,0,0.6]; A['Nat Dist'][c]['Data']=mos['Areas'][iScn]['Sum']['Mechanical']['Ensemble Mean'][:,iPT,iRE]*AEF
    A['Nat Dist']=A['Nat Dist'][0:c+1]

    A['Management']=[None]*10; c=-1
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Harvest'; A['Management'][c]['Color']=[0,0.75,1]; A['Management'][c]['Data']=(mos['Areas'][iScn]['Sum']['Harvest']['Ensemble Mean'][:,iPT,iRE]+mos['Areas'][iScn]['Sum']['Harvest Salvage']['Ensemble Mean'][:,iPT,iRE])*AEF
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Slashpile burn'; A['Management'][c]['Color']=[0.75,0,0]; A['Management'][c]['Data']=(mos['Areas'][iScn]['Sum']['Slashpile Burn']['Ensemble Mean'][:,iPT,iRE]+mos['Areas'][iScn]['Sum']['Harvest Salvage']['Ensemble Mean'][:,iPT,iRE])*AEF
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Thinning and knockdown'; A['Management'][c]['Color']=[0.2,0.4,0.7]; A['Management'][c]['Data']=(mos['Areas'][iScn]['Sum']['Knockdown']['Ensemble Mean'][:,iPT,iRE]+mos['Areas'][iScn]['Sum']['Thinning']['Ensemble Mean'][:,iPT,iRE])*AEF
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Site preparation'; A['Management'][c]['Color']=[1,0.7,0.7]; A['Management'][c]['Data']=(mos['Areas'][iScn]['Sum']['Disc Trenching']['Ensemble Mean'][:,iPT,iRE]+mos['Areas'][iScn]['Sum']['Ripping']['Ensemble Mean'][:,iPT,iRE])*AEF
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Prescribed burn'; A['Management'][c]['Color']=[0.5,0,0]; A['Management'][c]['Data']=mos['Areas'][iScn]['Sum']['Prescribed Burn']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Dwarf Mistletoe control'; A['Management'][c]['Color']=[1,0.5,0]; A['Management'][c]['Data']=mos['Areas'][iScn]['Sum']['Dwarf Mistletoe Control']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Planting'; A['Management'][c]['Color']=[0.3,0.8,0.2]; A['Management'][c]['Data']=(mos['Areas'][iScn]['Sum']['Planting']['Ensemble Mean'][:,iPT,iRE]+mos['Areas'][iScn]['Sum']['Direct Seeding']['Ensemble Mean'][:,iPT,iRE])*AEF
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Foliage protection'; A['Management'][c]['Color']=[1,0.7,0]; A['Management'][c]['Data']=mos['Areas'][iScn]['Sum']['IDW Btk Spray']['Ensemble Mean'][:,iPT,iRE]*AEF
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Aerial nutrient application'; A['Management'][c]['Color']=[0.75,0.55,1]; A['Management'][c]['Data']=mos['Areas'][iScn]['Sum']['Fertilization Aerial']['Ensemble Mean'][:,iPT,iRE]*AEF
    A['Management']=A['Management'][0:c+1] 
        
    # Convert to x-year intervals
    A['tv']=gu.BlockMean(tv,ivlT)
    for i in range(len(A['Nat Dist'])):
        A['Nat Dist'][i]['Data']=gu.BlockMean(A['Nat Dist'][i]['Data'],ivlT)
    for i in range(len(A['Management'])):
        A['Management'][i]['Data']=gu.BlockMean(A['Management'][i]['Data'],ivlT)        
    
    return A

def SummarizeAreaAffected_old(meta,iScn,iEns,AEF,ivlT,tv,mos):
    
    A={}
    if iEns>=0:
        # Individual ensemble        
        A['Nat Dist']=[None]*10; c=-1
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Wildfire'; A['Nat Dist'][c]['Color']=[0.75,0,0]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['Wildfire']['Ensembles'][:,iEns]
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Mountain Pine beetle'; A['Nat Dist'][c]['Color']=[0,0.8,0]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['IBM']['Ensembles'][:,iEns]
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Douglas-fir beetle'; A['Nat Dist'][c]['Color']=[0.6,1,0]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['IBD']['Ensembles'][:,iEns]
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Spruce beetle'; A['Nat Dist'][c]['Color']=[0.25,1,1]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['IBS']['Ensembles'][:,iEns]
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='W. balsam beetle'; A['Nat Dist'][c]['Color']=[0,0.45,0]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['IBB']['Ensembles'][:,iEns]
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Other pests'; A['Nat Dist'][c]['Color']=[0.8,1,0]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['Beetles']['Ensembles'][:,iEns]
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='W. spruce budworm'; A['Nat Dist'][c]['Color']=[0,0.75,1]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['IDW']['Ensembles'][:,iEns]
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Rust'; A['Nat Dist'][c]['Color']=[0.75,0.5,1]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['Rust Onset']['Ensembles'][:,iEns]
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Dwarf Mistletoe'; A['Nat Dist'][c]['Color']=[1,0.5,0.25]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['Dwarf Mistletoe Onset']['Ensembles'][:,iEns]
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Mechanical'; A['Nat Dist'][c]['Color']=[0,0,0.6]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['Mechanical']['Ensembles'][:,iEns]
        A['Nat Dist']=A['Nat Dist'][0:c+1]

        A['Management']=[None]*10; c=-1
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Harvest'; A['Management'][c]['Color']=[0,0.75,1]; A['Management'][c]['Data']=mos[iScn]['Area']['Harvest']['Ensembles'][:,iEns]+mos[iScn]['Area']['Harvest Salvage']['Ensembles'][:,iEns]
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Slashpile burn'; A['Management'][c]['Color']=[0.75,0,0]; A['Management'][c]['Data']=mos[iScn]['Area']['Slashpile Burn']['Ensembles'][:,iEns]+mos[iScn]['Area']['Harvest Salvage']['Ensembles'][:,iEns]
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Thinning and knockdown'; A['Management'][c]['Color']=[0.2,0.4,0.7]; A['Management'][c]['Data']=mos[iScn]['Area']['Knockdown']['Ensembles'][:,iEns]+mos[iScn]['Area']['Thinning']['Ensembles'][:,iEns]
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Site preparation'; A['Management'][c]['Color']=[1,0.7,0.7]; A['Management'][c]['Data']=mos[iScn]['Area']['Disc Trenching']['Ensembles'][:,iEns]+mos[iScn]['Area']['Ripping']['Ensembles'][:,iEns]
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Prescribed burn'; A['Management'][c]['Color']=[0.5,0,0]; A['Management'][c]['Data']=mos[iScn]['Area']['Prescribed Burn']['Ensembles'][:,iEns]
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Dwarf Mistletoe control'; A['Management'][c]['Color']=[1,0.5,0]; A['Management'][c]['Data']=mos[iScn]['Area']['Dwarf Mistletoe Control']['Ensembles'][:,iEns]
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Planting'; A['Management'][c]['Color']=[0.3,0.8,0.2]; A['Management'][c]['Data']=mos[iScn]['Area']['Planting']['Ensembles'][:,iEns]+mos[iScn]['Area']['Direct Seeding']['Ensembles'][:,iEns]
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Foliage protection'; A['Management'][c]['Color']=[1,0.7,0]; A['Management'][c]['Data']=mos[iScn]['Area']['IDW Btk Spray']['Ensembles'][:,iEns]
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Aerial nutrient application'; A['Management'][c]['Color']=[0.75,0.55,1]; A['Management'][c]['Data']=mos[iScn]['Area']['Fertilization Aerial']['Ensembles'][:,iEns]
        A['Management']=A['Management'][0:c+1]
        
    else:
        # Ensemble mean
        A['Nat Dist']=[None]*10; c=-1
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Wildfire'; A['Nat Dist'][c]['Color']=[0.75,0,0]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['Wildfire']['Ensemble Mean']
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Mountain Pine beetle'; A['Nat Dist'][c]['Color']=[0,0.8,0]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['IBM']['Ensemble Mean']
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Douglas-fir beetle'; A['Nat Dist'][c]['Color']=[0.6,1,0]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['IBD']['Ensemble Mean']
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Spruce beetle'; A['Nat Dist'][c]['Color']=[0.25,1,1]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['IBS']['Ensemble Mean']
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='W. balsam beetle'; A['Nat Dist'][c]['Color']=[0,0.45,0]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['IBB']['Ensemble Mean']
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Other pests'; A['Nat Dist'][c]['Color']=[0.8,1,0]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['Beetles']['Ensemble Mean']
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='W. spruce budworm'; A['Nat Dist'][c]['Color']=[0,0.75,1]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['IDW']['Ensemble Mean']
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Rust'; A['Nat Dist'][c]['Color']=[0.75,0.5,1]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['Rust Onset']['Ensemble Mean']
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Dwarf Mistletoe'; A['Nat Dist'][c]['Color']=[1,0.5,0.25]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['Dwarf Mistletoe Onset']['Ensemble Mean']
        c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Mechanical'; A['Nat Dist'][c]['Color']=[0,0,0.6]; A['Nat Dist'][c]['Data']=mos[iScn]['Area']['Mechanical']['Ensemble Mean']
        A['Nat Dist']=A['Nat Dist'][0:c+1]

        A['Management']=[None]*10; c=-1
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Harvest'; A['Management'][c]['Color']=[0,0.75,1]; A['Management'][c]['Data']=mos[iScn]['Area']['Harvest']['Ensemble Mean']+mos[iScn]['Area']['Harvest Salvage']['Ensemble Mean']
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Slashpile burn'; A['Management'][c]['Color']=[0.75,0,0]; A['Management'][c]['Data']=mos[iScn]['Area']['Slashpile Burn']['Ensemble Mean']+mos[iScn]['Area']['Harvest Salvage']['Ensemble Mean']
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Thinning and knockdown'; A['Management'][c]['Color']=[0.2,0.4,0.7]; A['Management'][c]['Data']=mos[iScn]['Area']['Knockdown']['Ensemble Mean']+mos[iScn]['Area']['Thinning']['Ensemble Mean']
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Site preparation'; A['Management'][c]['Color']=[1,0.7,0.7]; A['Management'][c]['Data']=mos[iScn]['Area']['Disc Trenching']['Ensemble Mean']+mos[iScn]['Area']['Ripping']['Ensemble Mean']
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Prescribed burn'; A['Management'][c]['Color']=[0.5,0,0]; A['Management'][c]['Data']=mos[iScn]['Area']['Prescribed Burn']['Ensemble Mean']
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Dwarf Mistletoe control'; A['Management'][c]['Color']=[1,0.5,0]; A['Management'][c]['Data']=mos[iScn]['Area']['Dwarf Mistletoe Control']['Ensemble Mean']
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Planting'; A['Management'][c]['Color']=[0.3,0.8,0.2]; A['Management'][c]['Data']=mos[iScn]['Area']['Planting']['Ensemble Mean']+mos[iScn]['Area']['Direct Seeding']['Ensemble Mean']
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Foliage protection'; A['Management'][c]['Color']=[1,0.7,0]; A['Management'][c]['Data']=mos[iScn]['Area']['IDW Btk Spray']['Ensemble Mean']
        c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Aerial nutrient application'; A['Management'][c]['Color']=[0.75,0.55,1]; A['Management'][c]['Data']=mos[iScn]['Area']['Fertilization Aerial']['Ensemble Mean']
        A['Management']=A['Management'][0:c+1]

    # Apply area expansion factor
    for i in range(len(A['Nat Dist'])):
        A['Nat Dist'][i]['Data']=AEF*A['Nat Dist'][i]['Data']
    for i in range(len(A['Management'])):
        A['Management'][i]['Data']=AEF*A['Management'][i]['Data']    
        
    # Convert to x-year intervals
    A['tv']=gu.BlockMean(tv,ivlT)
    for i in range(len(A['Nat Dist'])):
        A['Nat Dist'][i]['Data']=gu.BlockMean(A['Nat Dist'][i]['Data'],ivlT)
    for i in range(len(A['Management'])):
        A['Management'][i]['Data']=gu.BlockMean(A['Management'][i]['Data'],ivlT)        
    
    return A

#%% Area affected by individual multipolygon
    
def AreaAffectedInSingleMultipolygon(meta,iScn,ivlT,tv,MosByMP,iMP):
    
    A={}
    # Ensemble mean (currently not equipped to give individual ensembles)
    A['Nat Dist']=[None]*10; 
    c=-1
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Wildfire'; A['Nat Dist'][c]['Color']=[0.75,0,0]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Wildfire']['Ensemble Mean'][:,iMP]
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Mountain Pine beetle'; A['Nat Dist'][c]['Color']=[0,0.8,0]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['IBM']['Ensemble Mean'][:,iMP]
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Douglas-fir beetle'; A['Nat Dist'][c]['Color']=[0.6,1,0]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['IBD']['Ensemble Mean'][:,iMP]
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Spruce beetle'; A['Nat Dist'][c]['Color']=[0.25,1,1]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['IBS']['Ensemble Mean'][:,iMP]
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='W. balsam beetle'; A['Nat Dist'][c]['Color']=[0,0.45,0]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['IBB']['Ensemble Mean'][:,iMP]
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Other pests'; A['Nat Dist'][c]['Color']=[0.8,1,0]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Beetles']['Ensemble Mean'][:,iMP]
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='W. spruce budworm'; A['Nat Dist'][c]['Color']=[0,0.75,1]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['IDW']['Ensemble Mean'][:,iMP]
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Rust'; A['Nat Dist'][c]['Color']=[0.75,0.5,1]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Rust Onset']['Ensemble Mean'][:,iMP]
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Dwarf Mistletoe'; A['Nat Dist'][c]['Color']=[1,0.5,0.25]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Dwarf Mistletoe Onset']['Ensemble Mean'][:,iMP]
    c=c+1; A['Nat Dist'][c]={}; A['Nat Dist'][c]['Name']='Mechanical'; A['Nat Dist'][c]['Color']=[0,0,0.6]; A['Nat Dist'][c]['Data']=MosByMP[iScn]['Area']['Mechanical']['Ensemble Mean'][:,iMP]
    A['Nat Dist']=A['Nat Dist'][0:c+1]

    A['Management']=[None]*10; 
    c=-1
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Harvest'; A['Management'][c]['Color']=[0,0.75,1]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Harvest']['Ensemble Mean'][:,iMP]+MosByMP[iScn]['Area']['Harvest Salvage']['Ensemble Mean'][:,iMP]
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Slashpile burn'; A['Management'][c]['Color']=[0.75,0,0]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Slashpile Burn']['Ensemble Mean'][:,iMP]+MosByMP[iScn]['Area']['Harvest Salvage']['Ensemble Mean'][:,iMP]
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Thinning and knockdown'; A['Management'][c]['Color']=[0.2,0.4,0.7]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Knockdown']['Ensemble Mean'][:,iMP]+MosByMP[iScn]['Area']['Thinning']['Ensemble Mean'][:,iMP]
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Site preparation'; A['Management'][c]['Color']=[1,0.7,0.7]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Disc Trenching']['Ensemble Mean'][:,iMP]+MosByMP[iScn]['Area']['Ripping']['Ensemble Mean'][:,iMP]
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Prescribed burn'; A['Management'][c]['Color']=[0.5,0,0]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Prescribed Burn']['Ensemble Mean'][:,iMP]
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Dwarf Mistletoe control'; A['Management'][c]['Color']=[1,0.5,0]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Dwarf Mistletoe Control']['Ensemble Mean'][:,iMP]
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Planting'; A['Management'][c]['Color']=[0.3,0.8,0.2]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Planting']['Ensemble Mean'][:,iMP]+MosByMP[iScn]['Area']['Direct Seeding']['Ensemble Mean'][:,iMP]
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Foliage protection'; A['Management'][c]['Color']=[1,0.7,0]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['IDW Btk Spray']['Ensemble Mean'][:,iMP]
    c=c+1; A['Management'][c]={}; A['Management'][c]['Name']='Aerial nutrient application'; A['Management'][c]['Color']=[0.65,0,1]; A['Management'][c]['Data']=MosByMP[iScn]['Area']['Fertilization Aerial']['Ensemble Mean'][:,iMP]
    A['Management']=A['Management'][0:c+1]  
        
    # Convert to x-year intervals
    A['tv']=gu.BlockMean(tv,ivlT)
    for i in range(len(A['Nat Dist'])):
        A['Nat Dist'][i]['Data']=gu.BlockMean(A['Nat Dist'][i]['Data'],ivlT)
    for i in range(len(A['Management'])):
        A['Management'][i]['Data']=gu.BlockMean(A['Management'][i]['Data'],ivlT)        
    
    return A

#%% Post process BatchTIPSY output

def PrepGrowthCurvesForCBR(meta):

    # Function used to smooth curves
    def smooth(y, box_pts):
        box = np.ones(box_pts)/box_pts
        y_smooth = np.convolve(y, box, mode='same')
        return y_smooth
    
    # TIPSY exports curves as MgDM/ha/yr, CBRunner expects inputs of MgC/ha/yr - create conversion factor
    dm2c=0.5

    # Growth curve parameters and TIPSY outputs
    dfPar=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx',sheet_name='Sheet1',skiprows=6)
    txtDat=np.loadtxt(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Output.out',skiprows=4)

    # TIPSY saves to text file -> convert to dataframe (column names must match TIPSY output file design)
    dfDat=pd.DataFrame(txtDat,columns=meta['GC']['BatchTIPSY Column Names'])

    # Define age vector (must be consistent with how TIPSY was set up)
    Age=np.arange(0,meta['GC']['BatchTIPSY Maximum Age']+1,1)

    # Get dimensions of the TIPSY output file to reshape the data into Age x Stand
    N_Age=Age.size
    N_GC=int(dfDat.shape[0]/N_Age)

    # Stemwood

    # Merchantable stemwood volume
    V_Merch=np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')
    V_StemTot=np.reshape(dfDat['VolTot0'].values,(N_Age,N_GC),order='F')
    G_VStemMerch=np.append(np.zeros((1,N_GC)),np.diff(V_Merch,axis=0),axis=0)

    # Extract age responses for each biomass pool
    C_Stem=dm2c*np.reshape(dfDat['ODT_Stem'].values,(N_Age,N_GC),order='F')

    #import matplotlib.pyplot as plt
    #plt.close('all')
    #plt.plot(C_Stem[:,0])

    # Apply smoothing - it messes up the last ten years so don't smooth that part
    for j in range(C_Stem.shape[1]):
        a=smooth(C_Stem[:,j],10)
        C_Stem[:-10,j]=a[:-10]
        a=smooth(V_Merch[:,j],10)
        V_Merch[:-10,j]=a[:-10]
        a=smooth(V_StemTot[:,j],10)
        V_StemTot[:-10,j]=a[:-10]
    #plt.plot(C_Stem[:,0],'--')
    
    # Define the fraction of merchantable stemwood
    fMerch=np.nan_to_num(V_Merch/V_StemTot)
    fNonMerch=1-fMerch  

    # Calculate growth
    z=np.zeros((1,N_GC))
    G_Stem=np.append(z,np.diff(C_Stem,axis=0),axis=0)

    # Adjust early net growth, but don't change the total stock change
    A_th=30
    ind=np.where(Age<=A_th)[0]
    bin=np.arange(0.005,0.15,0.005)
    x=np.arange(0,A_th+1,1)
    for j in range(N_GC):
        Gtot=C_Stem[ind[-1],j]
        y_th=G_Stem[ind[-1],j]
        Gtot_hat=1000*np.ones(bin.size)
        for k in range(bin.size):
            Gtot_hat[k]=np.sum(y_th*np.exp(bin[k]*(x-A_th)))
        ind1=np.where(np.abs(Gtot_hat-Gtot)==np.min(np.abs(Gtot_hat-Gtot)))[0]
        G_Stem[ind,j]=y_th*np.exp(bin[ind1[0]]*(x-A_th))

    # Update merch and nonmerch growth
    C_Stem=np.cumsum(G_Stem,axis=0)
    C_StemMerch=fMerch*C_Stem
    C_StemNonMerch=fNonMerch*C_Stem
    
    G_StemMerch=np.append(z,np.diff(C_StemMerch,axis=0),axis=0)
    G_StemNonMerch=np.append(z,np.diff(C_StemNonMerch,axis=0),axis=0)
    
    # Fix growth of year zero
    G_StemMerch[0,:]=G_StemMerch[1,:]
    G_StemNonMerch[0,:]=G_StemNonMerch[1,:]
    
    # Add negative nonmerch to merch 
    ind=np.where(G_StemNonMerch<0)
    G_StemMerch[ind]=G_StemMerch[ind]+G_StemNonMerch[ind]
    G_StemNonMerch[ind]=0
    
    #plt.close('all')
    #plt.plot(np.maximum(-1,G_Stem[:,1]))
    #plt.plot(np.maximum(-1,G_StemMerch[:,1]),'--')
    #plt.plot(np.maximum(-1,G_StemNonMerch[:,1]),'-.')

    # Other pools - these are no longer being used. Net growth of non-stemwood
    # biomass is simulated in the annual loop based on allometric reatlionships
    # with stemwood net growth and stand age.

    # Foliage biomass is very low, revise
    #bF1=0.579
    #bF2=0.602
    #C_Foliage=np.maximum(0,bF1*C_Stem**bF2)
    
    C_Foliage=dm2c*np.reshape(dfDat['ODT_Foliage'].values,(N_Age,N_GC),order='F')    
    C_Branch=dm2c*np.reshape(dfDat['ODT_Branch'].values,(N_Age,N_GC),order='F')    
    C_Bark=dm2c*np.reshape(dfDat['ODT_Bark'].values,(N_Age,N_GC),order='F')

    G_Foliage=np.append(z,np.diff(C_Foliage,axis=0),axis=0)
    G_Branch=np.append(z,np.diff(C_Branch,axis=0),axis=0)
    G_Bark=np.append(z,np.diff(C_Bark,axis=0),axis=0)

    for iScn in range(meta['Project']['N Scenario']):
    
        # Define growth curves (from TIPSY output)
        for iBat in range(meta['Project']['N Batch']):
                
            # Index to batch
            indBat=IndexToBatch(meta,iBat)
                  
            for iGC in range(3):
                
                # Initialize age response of net growth
                G=np.zeros((N_Age,indBat.size,6),dtype=np.int32)
    
                # Populate the growth curve
                for iS in range(indBat.size):
               
                    #u=np.unique(ec['ID_GrowthCurve'][:,iS,:])
                    
                    if (meta['Project']['Scenario Source']=='Spreadsheet'):
                        
                        indTIPSY=np.where(
                                (dfPar['ID_Scenario']==iScn+1) &
                                (dfPar['ID_GC']==int(meta['GC']['ID GC Unique'][iGC])) )[0]                    
                    
                    elif (meta['Project']['Scenario Source']=='Script') | (meta['Project']['Scenario Source']=='Portfolio'): 
                        
                        indTIPSY=np.where(
                            (dfPar['ID_Stand']==indBat[iS]+1) & 
                            (dfPar['ID_Scenario']==iScn+1) &
                            (dfPar['ID_GC']==int(iGC+1)))[0]  
                    
                    if (indTIPSY.size==0):
                        # This can happen if only some stands have a third GC, for example
                        continue
                    
                    G[:,iS,0]=G_StemMerch[:,indTIPSY[0]]/meta['GC']['Scale Factor']   
                    G[:,iS,1]=G_StemNonMerch[:,indTIPSY[0]]/meta['GC']['Scale Factor']
                    G[:,iS,2]=G_Bark[:,indTIPSY[0]]/meta['GC']['Scale Factor']
                    G[:,iS,3]=G_Branch[:,indTIPSY[0]]/meta['GC']['Scale Factor']
                    G[:,iS,4]=G_Foliage[:,indTIPSY[0]]/meta['GC']['Scale Factor']
                    G[:,iS,5]=G_VStemMerch[:,indTIPSY[0]]/meta['GC']['Scale Factor']
                
                # Save data to file in input variables folder of project
                gu.opickle(meta['Paths']['Project'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve' + str(iGC+1) + '_Bat' + FixFileNum(iBat) + '.pkl',G)
                    
    return

#%% Prepare growth curves (with early correction)

def PrepGrowthCurvesUniqueForCBR(meta,ugc):

    # *** This adjusts early net growth so that it is not zero. There is a
    # second version of this function that excludes the correction. ***
    
    # TIPSY exports curves as MgDM/ha/yr, CBRunner expects inputs of MgC/ha/yr. Create
    # conversion factor.
    dm2c=0.5

    # Growth curve parameters and TIPSY outputs
    #dfPar=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx',sheet_name='Sheet1',skiprows=7)
    txtDat=np.loadtxt(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Output.out',skiprows=4)

    # TIPSY saves to text file -> convert to dataframe (column names must match TIPSY output file design)
    dfDat=pd.DataFrame(txtDat,columns=meta['GC']['BatchTIPSY Column Names'])

    del txtDat
    
    # Define age vector (must be consistent with how TIPSY was set up)
    Age=np.arange(0,meta['GC']['BatchTIPSY Maximum Age']+1,1)

    # Get dimensions of the TIPSY output file to reshape the data into Age x Stand
    N_Age=Age.size
    N_GC=int(dfDat.shape[0]/N_Age)

    # Define the fraction of merchantable stemwood
    fMerch=np.nan_to_num(np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')/np.reshape(dfDat['VolTot0'].values,(N_Age,N_GC),order='F'))
    fNonMerch=1-fMerch

    ## Merchantable stemwood volume
    #V_Merch=np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')
    #G_VStemMerch=np.append(np.zeros((1,N_GC)),np.diff(V_Merch,axis=0),axis=0)
    #
    ## Extract age responses for each biomass pool
    #C_Stem=dm2c*np.reshape(dfDat['ODT_Stem'].values,(N_Age,N_GC),order='F')
    #C_StemMerch=fMerch*C_Stem
    #C_StemNonMerch=fNonMerch*C_Stem
    #C_Foliage=dm2c*np.reshape(dfDat['ODT_Foliage'].values,(N_Age,N_GC),order='F')
    #C_Branch=dm2c*np.reshape(dfDat['ODT_Branch'].values,(N_Age,N_GC),order='F')
    #C_Bark=dm2c*np.reshape(dfDat['ODT_Bark'].values,(N_Age,N_GC),order='F')
    #
    ## Calculate growth
    #z=np.zeros((1,N_GC))
    #G_StemMerch=np.append(z,np.diff(C_StemMerch,axis=0),axis=0)
    #G_StemNonMerch=np.append(z,np.diff(C_StemNonMerch,axis=0),axis=0)
    #G_Foliage=np.append(z,np.diff(C_Foliage,axis=0),axis=0)
    #G_Branch=np.append(z,np.diff(C_Branch,axis=0),axis=0)
    #G_Bark=np.append(z,np.diff(C_Bark,axis=0),axis=0)
    #
    #plt.plot(np.mean(G_StemMerch+G_StemNonMerch,axis=1))
    ##plt.plot(np.percentile(G_StemMerch+G_StemNonMerch,20,axis=1))
    ##plt.plot(np.percentile(G_StemMerch+G_StemNonMerch,80,axis=1))
    
    # Function used to smooth curves
    def smooth(y, box_pts):
        box = np.ones(box_pts)/box_pts
        y_smooth = np.convolve(y, box, mode='same')
        return y_smooth

    # Stemwood

    # Merchantable stemwood volume
    V_Merch=np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')
    V_StemTot=np.reshape(dfDat['VolTot0'].values,(N_Age,N_GC),order='F')
    G_VStemMerch=np.append(np.zeros((1,N_GC)),np.diff(V_Merch,axis=0),axis=0)

    # Extract age responses for each biomass pool
    C_Stem=dm2c*np.reshape(dfDat['ODT_Stem'].values,(N_Age,N_GC),order='F')

    #import matplotlib.pyplot as plt
    #plt.close('all')
    #plt.plot(C_Stem[:,0])

    # Apply smoothing - it messes up the last ten years so don't smooth that part
    for j in range(C_Stem.shape[1]):
        a=smooth(C_Stem[:,j],10)
        C_Stem[:-10,j]=a[:-10]
        a=smooth(V_Merch[:,j],10)
        V_Merch[:-10,j]=a[:-10]
        a=smooth(V_StemTot[:,j],10)
        V_StemTot[:-10,j]=a[:-10]
        #plt.plot(C_Stem[:,0],'--')
    
    # Define the fraction of merchantable stemwood
    fMerch=np.nan_to_num(V_Merch/V_StemTot)
    fNonMerch=1-fMerch  

    # Calculate growth
    z=np.zeros((1,N_GC))
    G_Stem=np.append(z,np.diff(C_Stem,axis=0),axis=0)

    # Adjust early net growth, but don't change the total stock change
    A_th=30
    ind=np.where(Age<=A_th)[0]
    bin=np.arange(0.005,0.15,0.005)
    x=np.arange(0,A_th+1,1)
    for j in range(N_GC):
        Gtot=C_Stem[ind[-1],j]
        y_th=G_Stem[ind[-1],j]
        Gtot_hat=1000*np.ones(bin.size)
        for k in range(bin.size):
            Gtot_hat[k]=np.sum(y_th*np.exp(bin[k]*(x-A_th)))
        ind1=np.where(np.abs(Gtot_hat-Gtot)==np.min(np.abs(Gtot_hat-Gtot)))[0]
        G_Stem[ind,j]=y_th*np.exp(bin[ind1[0]]*(x-A_th))

    # Update merch and nonmerch growth
    C_Stem=np.cumsum(G_Stem,axis=0)
    C_StemMerch=fMerch*C_Stem
    C_StemNonMerch=fNonMerch*C_Stem
    G_StemMerch=np.append(z,np.diff(C_StemMerch,axis=0),axis=0)
    G_StemNonMerch=np.append(z,np.diff(C_StemNonMerch,axis=0),axis=0)
    
    # Fix growth of year zero
    G_StemMerch[0,:]=G_StemMerch[1,:]
    G_StemNonMerch[0,:]=G_StemNonMerch[1,:]
    
    # Add negative nonmerch to merch 
    ind=np.where(G_StemNonMerch<0)
    G_StemMerch[ind]=G_StemMerch[ind]+G_StemNonMerch[ind]
    G_StemNonMerch[ind]=0

    # Other pools - these are no longer being used. Net growth of non-stemwood
    # biomass is simulated in the annual loop based on allometric reatlionships
    # with stemwood net growth and stand age.

    # Foliage biomass is very low, revise
    #bF1=0.579
    #bF2=0.602
    #C_Foliage=np.maximum(0,bF1*C_Stem**bF2)   
    
    C_Foliage=dm2c*np.reshape(dfDat['ODT_Foliage'].values,(N_Age,N_GC),order='F')    
    C_Branch=dm2c*np.reshape(dfDat['ODT_Branch'].values,(N_Age,N_GC),order='F')
    C_Bark=dm2c*np.reshape(dfDat['ODT_Bark'].values,(N_Age,N_GC),order='F')

    G_Foliage=np.append(z,np.diff(C_Foliage,axis=0),axis=0)
    G_Branch=np.append(z,np.diff(C_Branch,axis=0),axis=0)
    G_Bark=np.append(z,np.diff(C_Bark,axis=0),axis=0)

    del C_Stem,C_StemMerch,C_StemNonMerch,C_Foliage,C_Branch,C_Bark,fMerch,fNonMerch

    for iScn in range(meta['Project']['N Scenario']):    
        for iGC in range(meta['GC']['N Growth Curves']):
            
            # Index to the full set of growth curves for scenario iScn and growth curve iGC
            ind_ugc_ScnAndGc=np.where( (ugc['Full'][:,1]==iScn) & (ugc['Full'][:,2]==meta['GC']['ID GC Unique'][iGC]) )[0]
        
            # Extract the unique growth curve ID for scenario iScn and growth curve iGC
            ID_ugc_ScnAndGc=ugc['Full'][ind_ugc_ScnAndGc,0]
        
            # Extract the inverse index for scenario iScn and growth curve iGC
            Inverse_ugc_ScnAndGc=ugc['Inverse'][ind_ugc_ScnAndGc]
        
            for iBat in range(0,meta['Project']['N Batch']):
            
                # Index to batch
                indBat=IndexToBatch(meta,iBat)
            
                # Intersect
                c,inda,indb=np.intersect1d(ID_ugc_ScnAndGc,indBat,return_indices=True)
            
                Inverse_ugc_ScnAndGcAndBat=Inverse_ugc_ScnAndGc[inda]
            
                # Initialize array of growth data
                G=np.zeros((N_Age,indBat.size,6),dtype=np.int16)
            
                for i in range(inda.size):
                
                    iStand=indb[i]
                    iGC_Unique=Inverse_ugc_ScnAndGcAndBat[i]
            
                    G[:,iStand,0]=G_StemMerch[:,iGC_Unique].T/meta['GC']['Scale Factor']
                    G[:,iStand,1]=G_StemNonMerch[:,iGC_Unique].T/meta['GC']['Scale Factor']
                    G[:,iStand,2]=G_Bark[:,iGC_Unique].T/meta['GC']['Scale Factor']
                    G[:,iStand,3]=G_Branch[:,iGC_Unique].T/meta['GC']['Scale Factor']
                    G[:,iStand,4]=G_Foliage[:,iGC_Unique].T/meta['GC']['Scale Factor']
                    G[:,iStand,5]=G_VStemMerch[:,iGC_Unique].T/meta['GC']['Scale Factor']
        
                # Save data to file in input variables folder of project
                gu.opickle(meta['Paths']['Project'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve' + str(meta['GC']['ID GC Unique'][iGC]) + '_Bat' + FixFileNum(iBat) + '.pkl',G)

    return

#%% Prepare growth curves (without early correction)

def PrepGrowthCurvesUniqueForCBR_WithoutEarlyCorrection(meta,ugc):

    # TIPSY exports curves as MgDM/ha/yr, CBRunner expects inputs of MgC/ha/yr. Create
    # conversion factor.
    dm2c=0.5

    # Growth curve parameters and TIPSY outputs
    #dfPar=pd.read_excel(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Parameters.xlsx',sheet_name='Sheet1',skiprows=7)
    txtDat=np.loadtxt(meta['Paths']['Project'] + '\\Inputs\\GrowthCurvesTIPSY_Output.out',skiprows=4)

    # TIPSY saves to text file -> convert to dataframe (column names must match TIPSY output file design)
    dfDat=pd.DataFrame(txtDat,columns=meta['GC']['BatchTIPSY Column Names'])

    del txtDat
    
    # Define age vector (must be consistent with how TIPSY was set up)
    Age=np.arange(0,meta['GC']['BatchTIPSY Maximum Age']+1,1)

    # Get dimensions of the TIPSY output file to reshape the data into Age x Stand
    N_Age=Age.size
    N_GC=int(dfDat.shape[0]/N_Age)
    
    # Define the fraction of merchantable stemwood
    fMerch=np.nan_to_num(np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')/np.reshape(dfDat['VolTot0'].values,(N_Age,N_GC),order='F'))
    fNonMerch=1-fMerch

    # Merchantable stemwood volume
    V_Merch=np.reshape(dfDat['VolMerch125'].values,(N_Age,N_GC),order='F')
    G_VStemMerch=np.append(np.zeros((1,N_GC)),np.diff(V_Merch,axis=0),axis=0)

    # Extract age responses for each biomass pool
    C_Stem=dm2c*np.reshape(dfDat['ODT_Stem'].values,(N_Age,N_GC),order='F')
    C_StemMerch=fMerch*C_Stem
    C_StemNonMerch=fNonMerch*C_Stem
    C_Foliage=dm2c*np.reshape(dfDat['ODT_Foliage'].values,(N_Age,N_GC),order='F')
    C_Branch=dm2c*np.reshape(dfDat['ODT_Branch'].values,(N_Age,N_GC),order='F')
    C_Bark=dm2c*np.reshape(dfDat['ODT_Bark'].values,(N_Age,N_GC),order='F')

    # Calculate growth
    z=np.zeros((1,N_GC))
    G_StemMerch=np.append(z,np.diff(C_StemMerch,axis=0),axis=0)
    G_StemNonMerch=np.append(z,np.diff(C_StemNonMerch,axis=0),axis=0)
    G_Foliage=np.append(z,np.diff(C_Foliage,axis=0),axis=0)
    G_Branch=np.append(z,np.diff(C_Branch,axis=0),axis=0)
    G_Bark=np.append(z,np.diff(C_Bark,axis=0),axis=0)

    # Fix growth of year zero
    G_StemMerch[0,:]=G_StemMerch[1,:]
    G_StemNonMerch[0,:]=G_StemNonMerch[1,:]
    G_Foliage[0,:]=G_Foliage[1,:]
    G_Branch[0,:]=G_Branch[1,:]
    G_Bark[0,:]=G_Bark[1,:]

    del C_Stem,C_StemMerch,C_StemNonMerch,C_Foliage,C_Branch,C_Bark,fMerch,fNonMerch

    for iScn in range(meta['Project']['N Scenario']):    
        for iGC in range(meta['GC']['N Growth Curves']):
            
            # Index to the full set of growth curves for scenario iScn and growth curve iGC
            ind_ugc_ScnAndGc=np.where( (ugc['Full'][:,1]==iScn) & (ugc['Full'][:,2]==meta['GC']['ID GC Unique'][iGC]) )[0]
        
            # Extract the unique growth curve ID for scenario iScn and growth curve iGC
            ID_ugc_ScnAndGc=ugc['Full'][ind_ugc_ScnAndGc,0]
        
            # Extract the inverse index for scenario iScn and growth curve iGC
            Inverse_ugc_ScnAndGc=ugc['Inverse'][ind_ugc_ScnAndGc]
        
            for iBat in range(0,meta['Project']['N Batch']):
            
                # Index to batch
                indBat=IndexToBatch(meta,iBat)
            
                # Intersect
                c,inda,indb=np.intersect1d(ID_ugc_ScnAndGc,indBat,return_indices=True)
            
                Inverse_ugc_ScnAndGcAndBat=Inverse_ugc_ScnAndGc[inda]
            
                # Initialize array of growth data
                G=np.zeros((N_Age,indBat.size,6),dtype=np.int16)
            
                for i in range(inda.size):
                    
                    iStand=indb[i]
                    iGC_Unique=Inverse_ugc_ScnAndGcAndBat[i]
            
                    G[:,iStand,0]=G_StemMerch[:,iGC_Unique].T/meta['GC']['Scale Factor']
                    G[:,iStand,1]=G_StemNonMerch[:,iGC_Unique].T/meta['GC']['Scale Factor']
                    G[:,iStand,2]=G_Bark[:,iGC_Unique].T/meta['GC']['Scale Factor']
                    G[:,iStand,3]=G_Branch[:,iGC_Unique].T/meta['GC']['Scale Factor']
                    G[:,iStand,4]=G_Foliage[:,iGC_Unique].T/meta['GC']['Scale Factor']
                    G[:,iStand,5]=G_VStemMerch[:,iGC_Unique].T/meta['GC']['Scale Factor']
        
                # Save data to file in input variables folder of project
                gu.opickle(meta['Paths']['Project'] + '\\Inputs\\Scenario' + FixFileNum(iScn) + '\\GrowthCurve' + str(meta['GC']['ID GC Unique'][iGC]) + '_Bat' + FixFileNum(iBat) + '.pkl',G)

    return

#%% Import parameters

def ImportParameters(meta):
      
    # Path to parameters    
    pthin=meta['Paths']['Model Code'] + '\\Parameters\\'
    
    # Initialize parameter structure    
    meta['Param']={}
    meta['Param']['BE']={}
    meta['Param']['Sigma']={}
    meta['Param']['BL']={}
    meta['Param']['BU']={}
    meta['Param']['BEV']={}
    
    #--------------------------------------------------------------------------
    # Biophysical
    #--------------------------------------------------------------------------
    
    df=pd.read_excel(pthin + '\Parameters_Biophysical.xlsx',sheet_name='Sheet1')    
    meta['Param']['BE']['Biophysical']={}
    for i in range(len(df)):
        Name=df['Name'][i]
        meta['Param']['BE']['Biophysical'][Name]=df['Value'][i]
    
    #--------------------------------------------------------------------------
    # Biomass - allometry
    # Values are location specific so no specific processing is done here (see cbrun.py)
    #--------------------------------------------------------------------------
    
    meta['Param']['BE']['Biomass Allometry']={}
    meta['Param']['BE']['Biomass Allometry']['Raw']=gu.ReadExcel(pthin + '\Parameters_BiomassAllometrySL.xlsx')
    
    #--------------------------------------------------------------------------
    # Biomass - turnover
    #--------------------------------------------------------------------------
    
    df=pd.read_excel(pthin + '\Parameters_BiomassTurnover.xlsx',sheet_name='Sheet1')    
    meta['Param']['BE']['Biomass Turnover']={}
    meta['Param']['Sigma']['Biomass Turnover']={}
    meta['Param']['BL']['Biomass Turnover']={}
    meta['Param']['BU']['Biomass Turnover']={}
    for i in range(len(df)):
        Name=df['Name'][i]
        meta['Param']['BE']['Biomass Turnover'][Name]=df['Best Estimate'][i]
        meta['Param']['Sigma']['Biomass Turnover'][Name]=df['Sigma'][i]
        meta['Param']['BL']['Biomass Turnover'][Name]=df['BL'][i]
        meta['Param']['BU']['Biomass Turnover'][Name]=df['BU'][i]
    
    #--------------------------------------------------------------------------
    # Inter-pool transfer parameters (Kurz et al. 2009)
    #--------------------------------------------------------------------------
    
    df=pd.read_excel(pthin + '\Parameters_InterPoolFluxes.xlsx',sheet_name='Sheet1')    
    meta['Param']['BE']['Inter Pool Fluxes']={}
    meta['Param']['Sigma']['Inter Pool Fluxes']={}
    for i in range(len(df)):
        Name=df['Name'][i]
        meta['Param']['BE']['Inter Pool Fluxes'][Name]=df['Best Estimate'][i]
        meta['Param']['Sigma']['Inter Pool Fluxes'][Name]=df['Sigma'][i]
    
    #--------------------------------------------------------------------------
    # Decomposition parameters (Kurz et al. 2009)
    #--------------------------------------------------------------------------

    df=pd.read_excel(pthin + '\Parameters_Decomposition.xlsx',sheet_name='Sheet1')    
    meta['Param']['BE']['Decomp']={}
    meta['Param']['Sigma']['Decomp']={}
    meta['Param']['BL']['Decomp']={}
    meta['Param']['BU']['Decomp']={}
    for i in range(len(df)):        
        Name=df['Name'].iloc[i]
        meta['Param']['BE']['Decomp'][Name]=df['Best Estimate'].iloc[i]
        meta['Param']['Sigma']['Decomp'][Name]=df['Sigma'].iloc[i]
        meta['Param']['BL']['Decomp'][Name]=df['BL'].iloc[i]
        meta['Param']['BU']['Decomp'][Name]=df['BU'].iloc[i]
        
    #--------------------------------------------------------------------------
    # Disturbance - static parameters
    #--------------------------------------------------------------------------
    
    p=gu.ReadExcel(pthin + '\Parameters_Disturbances.xlsx')
    
    str_to_exclude=['ID','Name','Species_CD','MortalityOccurs','GrowthFactor',
                    'GrowthFactor_Source','GrowthRecovery_HL','GrowthRecovery_HL_Source',
                    'QA1','QA2','QA3']    
    
    meta['Param']['BE']['Dist']={}
    for i in range(p['ID'].size):
        meta['Param']['BE']['Dist'][p['ID'][i]]={}
        meta['Param']['BE']['Dist'][p['ID'][i]]['BiomassMerch_Affected']=1
        meta['Param']['BE']['Dist'][p['ID'][i]]['BiomassNonMerch_Affected']=1
        meta['Param']['BE']['Dist'][p['ID'][i]]['Snags_Affected']=1
        for k in p.keys():
            # Exclude some legacy variables that may be re-implemented
            if np.isin(k,str_to_exclude)==True:
                continue
            meta['Param']['BE']['Dist'][p['ID'][i]][k]=np.nan_to_num(p[k][i])
    
    #--------------------------------------------------------------------------
    # Disutrbance - fate of felled material
    # Values are location specific so no specific processing is done here (see cbrun.py)
    #--------------------------------------------------------------------------

    meta['Param']['BE']['Felled Fate']={}
    meta['Param']['BE']['Felled Fate']=gu.ipickle(meta['Paths']['Model Code'] + '\\Parameters\\Variables_FelledFate.pkl')
    
    #--------------------------------------------------------------------------
    # Disturbance - Wildfire aspatial (Taz-AAO) parameters
    #--------------------------------------------------------------------------
     
    meta['Param']['BE']['Taz']={}
    
    wf=gu.ReadExcel(meta['Paths']['Model Code'] + '\\Parameters\\Parameters_WildfireStatsMod.xlsx')
    
    meta['Param']['BE']['Taz']['WF']={}    
    for i in range(wf['Name'].size):
        try:
            meta['Param']['BE']['Taz']['WF'][wf['Name'][i]]=wf['Value'][i].astype(float)
        except:
            meta['Param']['BE']['Taz']['WF'][wf['Name'][i]]=wf['Value'][i]
    
    #--------------------------------------------------------------------------
    # Disturbance - Mountain Pine Beetle (Taz-AAO) parameters
    #--------------------------------------------------------------------------
        
    ibm=gu.ReadExcel(meta['Paths']['Model Code'] + '\\Parameters\\Parameters_IBMStatsMod.xlsx')
    
    meta['Param']['BE']['Taz']['IBM']={}
    for i in range(ibm['Name'].size):
        try:
            meta['Param']['BE']['Taz']['IBM'][ibm['Name'][i]]=ibm['Value'][i].astype(float)
        except:
            meta['Param']['BE']['Taz']['IBM'][ibm['Name'][i]]=ibm['Value'][i]        
        
    #--------------------------------------------------------------------------
    # Disturbance - On the fly parameters
    #--------------------------------------------------------------------------
    
    df=pd.read_excel(pthin + '\Parameters_OnTheFly.xlsx',sheet_name='Sheet1')    
    meta['Param']['BE']['On The Fly']={}
    for i in range(len(df)):
        Name=df['Name'].iloc[i]
        meta['Param']['BE']['On The Fly'][Name]=df['Value'].iloc[i]
    
    #--------------------------------------------------------------------------     
    # Disturbance - harvesting - basic early historical reconstruction
    #--------------------------------------------------------------------------
    
    fin=meta['Paths']['Taz Datasets'] + '\\Harvest Stats and Scenarios\\HarvestHistoricalProbabilitySimple.xlsx'
    meta['Param']['BE']['Taz']['Ph_Simp']=gu.ReadExcel(fin)
    
    #--------------------------------------------------------------------------
    # Disturbance - By severity class
    #--------------------------------------------------------------------------
    
    meta['Param']['BE']['DistBySC']=gu.ReadExcel(meta['Paths']['Model Code'] + '\\Parameters\\Parameters_DisturbanceBySeverityClass.xlsx')
    
    #--------------------------------------------------------------------------
    # Removed fate scenarios
    # Values are location specific so no specific processing is done here (see cbrun.py)
    #--------------------------------------------------------------------------

    meta['Param']['BE']['Removed Fate']=gu.ipickle(meta['Paths']['Model Code'] + '\\Parameters\\Variables_RemovedFate.pkl')
    
    #--------------------------------------------------------------------------
    # HWP - static parameters
    # Values are location specific so no specific processing is done here (see cbrun.py)
    #--------------------------------------------------------------------------

    meta['Param']['BE']['HWP']={}
    meta['Param']['BE']['HWP']['Raw']=pd.read_excel(pthin + '\Parameters_HWP.xlsx',sheet_name='Default')
    
    #--------------------------------------------------------------------------
    # HWP - End Use parameters
    # Values are location specific so no specific processing is done here (see cbrun.py)
    #--------------------------------------------------------------------------
    
    meta['Param']['BE']['HWP End Use']={}
    meta['Param']['BE']['HWP End Use']=gu.ipickle(meta['Paths']['Model Code'] + '\\Parameters\\Variables_HWP_EndUses.pkl')
    
    #--------------------------------------------------------------------------
    # Nutrient application paramaters
    #--------------------------------------------------------------------------
    
    df=pd.read_excel(pthin + '\Parameters_NutrientApplication.xlsx',sheet_name='Sheet1')
    meta['Param']['BE']['Nutrient Management']={} 
    meta['Param']['Sigma']['Nutrient Management']={} 
    for i in range(len(df)):
        Name=df['Name'].iloc[i]
        meta['Param']['BE']['Nutrient Management'][Name]=df['Value'].iloc[i]
        meta['Param']['Sigma']['Nutrient Management'][Name]=df['Sigma'].iloc[i]
        
    #--------------------------------------------------------------------------
    # Economics
    #--------------------------------------------------------------------------
    
    df=pd.read_excel(pthin + '\Parameters_Economics.xlsx',sheet_name='Sheet1')
    meta['Param']['BE']['Econ']={}
    for i in range(len(df)):
        Name=df['Name'].iloc[i]
        meta['Param']['BE']['Econ'][Name]=df['Value'].iloc[i]
     
    #--------------------------------------------------------------------------
    # Substitution effects (use of basic displacement factors)
    #--------------------------------------------------------------------------

    df=pd.read_excel(pthin + '\\Parameters_Substitution.xlsx',sheet_name='Sheet1')
    meta['Param']['BE']['Substitution']={}
    meta['Param']['Sigma']['Substitution']={}
    for i in range(len(df)):
        Name=df['Name'].iloc[i]
        meta['Param']['BE']['Substitution'][Name]=df['Value'].iloc[i]
        meta['Param']['Sigma']['Substitution'][Name]=df['Sigma'].iloc[i]
    
    #--------------------------------------------------------------------------
    # Genetic worth
    #--------------------------------------------------------------------------

    meta['Param']['BE']['Genetic Worth']=gu.ReadExcel(pthin + '\\Parameters_Seedlot_GW.xlsx')

    #--------------------------------------------------------------------------
    # Sawtooth parameters
    #--------------------------------------------------------------------------
    
    if meta['Project']['Biomass Module']=='Sawtooth':
    
        ps={}
        #pthin=r'C:\Users\rhember\Documents\Code_Python\fcgadgets\cbrunner\Parameters'        
        
        # Core sawtooth parameters     
        d=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','Core')        
        ps['Core']={}
        for i in range(d['Name'].size):
            ps['Core'][d['Name'][i]]=d['Value'][i]            
        
        # Sawtooth key between provincial species codes adn species-region sample 
        # codes        
        ps['SRS Key']=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','SRS Key')
        
        # Tree-level allometry
        ps['Allom']=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','Allometry')
        ps['Eq R']={}
        ps['Eq R']['Def1']=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','R_Def1')
        
        ps['Eq M']={}
        ps['Eq M']['Def1']=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','M_Def1')
        
        ps['Eq G']={}
        ps['Eq G']['Def1']=gu.ReadExcel(pthin + '\Parameters_Sawtooth.xlsx','G_Def1')
        
        meta['Param']['BE']['Sawtooth']=ps

    return meta

#%% Delete all output files
    
def DeleteAllOutputFiles(meta):
    
    for iScn in range(meta['Project']['N Scenario']):
        files=glob.glob(meta['Paths']['Output Scenario'][iScn] + '\\*')
        for f in files:
            os.remove(f)

    for iBat in range(meta['Project']['N Batch']):
        try:
            pth=meta['Paths']['Project'] + '\\Outputs\\WorkingOnBatch_' + FixFileNum(iBat) + '.pkl'
            os.remove(pth)
        except:
            pass
    
    return

#%% Import custom harvest assumptions (optional)

def ImportCustomHarvestAssumptions(pthin):

    # Import parameters
    df=pd.read_excel(pthin,sheet_name='Sheet1',skiprows=1)
    
    d={}
    d['BiomassMerch_Affected']=df.iloc[1,1]
    d['BiomassMerch_Burned']=df.iloc[3,1]
    d['BiomassMerch_LeftOnSite']=df.iloc[4,1]    
    d['BiomassMerch_Removed']=df.iloc[5,1]
    
    d['BiomassNonMerch_Affected']=df.iloc[1,2]
    d['BiomassNonMerch_Burned']=df.iloc[3,2]
    d['BiomassNonMerch_LeftOnSite']=df.iloc[4,2]    
    d['BiomassNonMerch_Removed']=df.iloc[5,2]
    
    d['Snags_Affected']=df.iloc[1,3]
    d['Snags_Burned']=df.iloc[3,3]
    d['Snags_LeftOnSite']=df.iloc[4,3]    
    d['Snags_Removed']=df.iloc[5,3]
    
    d['RemovedMerchToSawMill']=df.iloc[7,1]
    d['RemovedMerchToPulpMill']=df.iloc[8,1]
    d['RemovedMerchToPelletMill']=df.iloc[9,1]
    d['RemovedMerchToPlywoodMill']=df.iloc[10,1]    
    d['RemovedMerchToOSBMill']=df.iloc[11,1]
    d['RemovedMerchToMDFMill']=df.iloc[12,1]
    d['RemovedMerchToFirewood']=df.iloc[13,1]
    d['RemovedMerchToIPP']=df.iloc[14,1]
    d['RemovedMerchToLogExport']=df.iloc[15,1]
    
    d['RemovedNonMerchToSawMill']=df.iloc[7,2]
    d['RemovedNonMerchToPulpMill']=df.iloc[8,2]
    d['RemovedNonMerchToPelletMill']=df.iloc[9,2]
    d['RemovedNonMerchToPlywoodMill']=df.iloc[10,2]    
    d['RemovedNonMerchToOSBMill']=df.iloc[11,2]
    d['RemovedNonMerchToMDFMill']=df.iloc[12,2]
    d['RemovedNonMerchToFirewood']=df.iloc[13,2]
    d['RemovedNonMerchToIPP']=df.iloc[14,2]
    d['RemovedNonMerchToLogExport']=df.iloc[15,2]
    
    d['RemovedSnagStemToSawMill']=df.iloc[7,3]
    d['RemovedSnagStemToPulpMill']=df.iloc[8,3]
    d['RemovedSnagStemToPelletMill']=df.iloc[9,3]
    d['RemovedSnagStemToPlywoodMill']=df.iloc[10,3]    
    d['RemovedSnagStemToOSBMill']=df.iloc[11,3]
    d['RemovedSnagStemToMDFMill']=df.iloc[12,3]
    d['RemovedSnagStemToFirewood']=df.iloc[13,3]
    d['RemovedSnagStemToIPP']=df.iloc[14,3]
    d['RemovedSnagStemToLogExport']=df.iloc[15,3]
    
    return d

#%% Unpack ensemble stats from MOS

def UnpackEnsembleStatsFromMos(meta,mos):
    
    tv=np.arange(meta['Project']['Year Start Saving'],meta['Project']['Year End']+1,1)
    
    #--------------------------------------------------------------------------
    # Unpack contents for easy use
    #--------------------------------------------------------------------------
    
    mu_mos=[]
    for iScn in range(meta['Project']['N Scenario']):        
        d={}
        for k in mos[iScn]['v1']['Mean'].keys():
            d[k]=mos[iScn]['v1']['Mean'][k]['Ensemble Mean']        
        for k in mos[iScn]['Cashflow']['Mean'].keys():
            d[k]=mos[iScn]['Cashflow']['Mean'][k]['Ensemble Mean']        
        mu_mos.append(d)
    
    cil_mos=[]
    for iScn in range(meta['Project']['N Scenario']):        
        d={}
        for k in mos[iScn]['v1']['Mean'].keys():
            d[k]=mos[iScn]['v1']['Mean'][k]['Ensemble CIL']
        for k in mos[iScn]['Cashflow']['Mean'].keys():
            d[k]=mos[iScn]['Cashflow']['Mean'][k]['Ensemble CIL']
        cil_mos.append(d)
    
    cih_mos=[]
    for iScn in range(meta['Project']['N Scenario']):        
        d={}
        for k in mos[iScn]['v1']['Mean'].keys():
            d[k]=mos[iScn]['v1']['Mean'][k]['Ensemble CIH']
        for k in mos[iScn]['Cashflow']['Mean'].keys():
            d[k]=mos[iScn]['Cashflow']['Mean'][k]['Ensemble CIH']
        cih_mos.append(d)
    
    sdl_mos=[]
    for iScn in range(meta['Project']['N Scenario']):        
        d={}
        for k in mos[iScn]['v1']['Mean'].keys():
            d[k]=mos[iScn]['v1']['Mean'][k]['Ensemble Mean']-mos[iScn]['v1']['Mean'][k]['Ensemble SD']
        for k in mos[iScn]['Cashflow']['Mean'].keys():
            d[k]=mos[iScn]['Cashflow']['Mean'][k]['Ensemble Mean']-mos[iScn]['Cashflow']['Mean'][k]['Ensemble SD']
        sdl_mos.append(d)
    
    sdh_mos=[]
    for iScn in range(meta['Project']['N Scenario']):        
        d={}
        for k in mos[iScn]['v1']['Mean'].keys():
            d[k]=mos[iScn]['v1']['Mean'][k]['Ensemble Mean']+mos[iScn]['v1']['Mean'][k]['Ensemble SD']
        for k in mos[iScn]['Cashflow']['Mean'].keys():
            d[k]=mos[iScn]['Cashflow']['Mean'][k]['Ensemble Mean']+mos[iScn]['Cashflow']['Mean'][k]['Ensemble SD']
        sdh_mos.append(d)
    
#    p1_mos=[]
#    for iScn in range(meta['Project']['N Scenario']):        
#        d={}
#        for k in mos[iScn]['v1']['Sum'].keys():
#            d[k]=mos[iScn]['v1']['Sum'][k]['Ensemble P1']
#        for k in mos[iScn]['Cashflow']['Sum'].keys():
#            d[k]=mos[iScn]['Cashflow']['Sum'][k]['Ensemble P1']
#        p1_mos.append(d)
#        
#    p10_mos=[]
#    for iScn in range(meta['Project']['N Scenario']):        
#        d={}
#        for k in mos[iScn]['v1']['Sum'].keys():
#            d[k]=mos[iScn]['v1']['Sum'][k]['Ensemble P10']
#        for k in mos[iScn]['Cashflow']['Sum'].keys():
#            d[k]=mos[iScn]['Cashflow']['Sum'][k]['Ensemble P10']
#        p10_mos.append(d)    
#
#        
#    p90_mos=[]
#    for iScn in range(meta['Project']['N Scenario']):        
#        d={}
#        for k in mos[iScn]['v1']['Sum'].keys():
#            d[k]=mos[iScn]['v1']['Sum'][k]['Ensemble P90']
#        for k in mos[iScn]['Cashflow']['Sum'].keys():
#            d[k]=mos[iScn]['Cashflow']['Sum'][k]['Ensemble P90']
#        p90_mos.append(d)
#     
#    p99_mos=[]
#    for iScn in range(meta['Project']['N Scenario']):        
#        d={}
#        for k in mos[iScn]['v1']['Sum'].keys():
#            d[k]=mos[iScn]['v1']['Sum'][k]['Ensemble P99']
#        for k in mos[iScn]['Cashflow']['Sum'].keys():
#            d[k]=mos[iScn]['Cashflow']['Sum'][k]['Ensemble P99']
#        p99_mos.append(d)
        
    return tv,mu_mos,cil_mos,cih_mos,sdl_mos,sdh_mos

